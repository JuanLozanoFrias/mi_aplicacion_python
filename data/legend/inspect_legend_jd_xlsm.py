"""
Inspector rápido del archivo: LEGEND jd.xlsm
- Extrae configuración (INFO)
- Lista equipos (BTU/hr/ft)
- Lista usos BT/MT
- Lista variadores FC102 y capacidades WCR
- Exporta un JSON para integrarlo en Calvo

Uso:
    python inspect_legend_jd_xlsm.py "LEGEND jd.xlsm"

Salida:
    legend_jd_analysis.json
"""
from __future__ import annotations

import json
import os
import re
import sys
from dataclasses import dataclass
from typing import Any, Dict, List

import openpyxl
from openpyxl.utils import get_column_letter

def extract_info_config(ws) -> Dict[str, Any]:
    config: Dict[str, Any] = {}
    for r in range(2, 40):
        key = ws[f"B{r}"].value
        val = ws[f"C{r}"].value
        if isinstance(key, str) and key.strip():
            config[key.strip()] = val
    return config

def extract_equipment_list(ws) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for r in range(1, 120):
        name = ws[f"F{r}"].value
        btu = ws[f"G{r}"].value
        kw_formula = ws[f"H{r}"].value
        if not (isinstance(name, str) and name.strip()):
            continue

        btu_val = None
        if isinstance(btu, (int, float)):
            btu_val = float(btu)
        elif isinstance(btu, str) and btu.startswith("="):
            # ejemplo: =+G8*0.6
            m = re.match(r"=+\+?G(\d+)\*([0-9.]+)", btu.replace(" ", ""))
            if m:
                ref = int(m.group(1))
                k = float(m.group(2))
                ref_val = ws[f"G{ref}"].value
                if isinstance(ref_val, (int, float)):
                    btu_val = float(ref_val) * k

        if btu_val is not None:
            out.append({
                "row": r,
                "equipo": name.strip(),
                "btu_hr_ft": btu_val,
                "kw_formula": kw_formula,
            })
    return out

def extract_uso_lists(ws) -> Dict[str, List[str]]:
    usos = {"BT": [], "MT": []}
    mode = None
    for r in range(40, 70):
        v = ws[f"F{r}"].value
        if not isinstance(v, str):
            continue
        vv = v.strip()
        if vv.upper() == "USO BT":
            mode = "BT"
            continue
        if vv.upper() == "USO MT":
            mode = "MT"
            continue
        if mode and vv and not vv.upper().startswith("USO"):
            usos[mode].append(vv)
    return usos

def extract_vfd_list(ws) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    for r in range(35, 60):
        model = ws[f"L{r}"].value
        power = ws[f"M{r}"].value
        if isinstance(model, str) and model.strip() and isinstance(power, (int, float)):
            items.append({"modelo": model.strip(), "potencia": float(power)})
    return items

def extract_wcr_list(ws) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    for r in range(60, 90):
        model = ws[f"L{r}"].value
        cap = ws[f"M{r}"].value
        if isinstance(model, str) and model.strip() and isinstance(cap, (int, float)):
            items.append({"modelo": model.strip(), "capacidad": float(cap)})
    return items

def extract_bom_templates(ws, max_rows: int = 60) -> Dict[str, List[Dict[str, Any]]]:
    templates = {"Rack Loop": [], "Minisistema": [], "Rack Americano": []}
    for r in range(3, max_rows + 1):
        ac, ad = ws[f"AC{r}"].value, ws[f"AD{r}"].value
        ae, af = ws[f"AE{r}"].value, ws[f"AF{r}"].value
        ag, ah = ws[f"AG{r}"].value, ws[f"AH{r}"].value

        if ad not in (None, ""):
            templates["Rack Loop"].append({"qty": ac, "descripcion": ad})
        if af not in (None, ""):
            templates["Minisistema"].append({"qty": ae, "descripcion": af})
        if ah not in (None, ""):
            templates["Rack Americano"].append({"qty": ag, "descripcion": ah})

    for k in list(templates.keys()):
        templates[k] = [it for it in templates[k] if it["descripcion"] not in (None, "")]
    return templates

def detect_perf_groups(ws, header_row: int = 4) -> List[Dict[str, int]]:
    starts = []
    for c in range(1, ws.max_column + 1):
        if ws.cell(header_row, c).value == "Refiferante":  # typo tal cual en el archivo
            starts.append(c)

    groups = []
    for i, sc in enumerate(starts):
        next_sc = starts[i + 1] if i + 1 < len(starts) else ws.max_column + 1
        groups.append({
            "start_col": sc,
            "end_col": next_sc - 1,
            "ref_col": sc,
            "tcond_col": sc + 1,
            "tevap_col": sc + 2,
            "metric_col": sc + 3,
            "models_start_col": sc + 4,
        })
    return groups

def detect_perf_blocks(ws, group: Dict[str, int], first_data_row: int = 5) -> List[int]:
    metric_col = group["metric_col"]
    ref_col = group["ref_col"]
    starts = []
    for r in range(first_data_row, ws.max_row + 1):
        if ws.cell(r, metric_col).value == "Capacidad Compresor":
            refv = ws.cell(r, ref_col).value
            if isinstance(refv, str) and refv.startswith("R-"):
                starts.append(r)
    return starts

def main(xlsm_path: str) -> int:
    wb = openpyxl.load_workbook(xlsm_path, data_only=False, keep_vba=True)
    info = wb["INFO"]

    summary = {
        "file": os.path.basename(xlsm_path),
        "sheets": wb.sheetnames,
        "INFO": {
            "config": extract_info_config(info),
            "equipos": extract_equipment_list(info),
            "usos": extract_uso_lists(info),
            "variadores_fc102": extract_vfd_list(info),
            "wcr_capacidades": extract_wcr_list(info),
            "plantillas_materiales": extract_bom_templates(info),
        },
        "PERFORMANCE_tables": {},
    }

    for perf_sheet in ("COPELAND", "BITZER"):
        ws = wb[perf_sheet]
        groups = detect_perf_groups(ws, header_row=4)
        group_summ = []
        for g in groups:
            starts = detect_perf_blocks(ws, g)
            model_cols = []
            for c in range(g["models_start_col"], g["end_col"] + 1):
                v = ws.cell(4, c).value
                if isinstance(v, str) and v.strip():
                    model_cols.append(v.strip())
            group_summ.append({
                "group_start_col": get_column_letter(g["start_col"]),
                "group_end_col": get_column_letter(g["end_col"]),
                "num_models": len(model_cols),
                "sample_models": model_cols[:10],
                "num_blocks_vertical": len(starts),
                "first_block_row": starts[0] if starts else None,
            })
        summary["PERFORMANCE_tables"][perf_sheet] = {
            "groups_detected": len(groups),
            "groups": group_summ,
        }

    out_json = os.path.join(os.path.dirname(xlsm_path), "legend_jd_analysis.json")
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print("OK ->", out_json)
    return 0

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python inspect_legend_jd_xlsm.py <ruta_archivo.xlsm>")
        raise SystemExit(2)
    raise SystemExit(main(sys.argv[1]))
