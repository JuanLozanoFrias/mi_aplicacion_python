from __future__ import annotations

from pathlib import Path
import json
import shutil
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Iterable, Tuple
import math
import pandas as pd

from PySide6.QtCore import Qt, QTimer, QRect, QEvent
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QFrame, QLabel, QLineEdit,
    QPushButton, QTableWidget, QTableWidgetItem, QMessageBox, QSpinBox,
    QComboBox, QHeaderView, QFileDialog, QDoubleSpinBox, QSizePolicy, QScrollArea, QAbstractScrollArea,
    QCheckBox, QMenu
)
from PySide6.QtGui import QKeySequence, QShortcut, QPainter, QFont, QColor

from logic.carga_electrica.cuadro_cargas import recalcular, ResultadoRamal, SeleccionRamal, CargaFase
from logic.legend_jd import LegendJDService
from .biblioteca_carga import BibliotecaCargaDialog

# Lista fija de usos (tomada de la columna H de DATOS, pero embebida para no depender del Excel).
USOS_DEFAULT = [
    "BEBIDAS", "C. FRIAS", "CARNE", "CC", "CERDO", "DELI", "FRUVER", "IC",
    "LACTEOS", "PANADERIA", "PESCADO", "POLLO", "POLLO FRIZADO", "PREP", "VARIOS",
    "PASILLO"
]


class GroupHeaderView(QHeaderView):
    """Cabecera de dos niveles sin bordes resaltados: solo texto de grupo arriba."""

    def __init__(self, groups: List[Tuple[str, int, int]], parent=None):
        super().__init__(Qt.Horizontal, parent)
        self._groups = groups
        self.setDefaultAlignment(Qt.AlignCenter)
        self.setMinimumHeight(46)

    def set_groups(self, groups: List[Tuple[str, int, int]]) -> None:
        self._groups = groups
        self.viewport().update()

    def paintEvent(self, event) -> None:
        # pinta normalmente las etiquetas base
        super().paintEvent(event)
        if not self._groups:
            return
        painter = QPainter(self.viewport())
        try:
            text_color = self.palette().color(self.foregroundRole())
            painter.setPen(text_color)
            font = painter.font()
            font.setBold(True)
            painter.setFont(font)
            height = self.height()
            top_h = max(18, height // 2 + 2)
            for title, start, span in self._groups:
                x = self.sectionViewportPosition(start)
                width = sum(self.sectionSize(start + i) for i in range(span))
                # titulo centrado arriba
                top_rect = QRect(x, 0, width, top_h)
                painter.drawText(top_rect, Qt.AlignCenter, title.upper())
        finally:
            painter.end()


class CargaElectricaPage(QWidget):
    """
    Nueva pestana: CARGA ELECTRICA.
    Permite recalcular el cuadro de cargas a partir del Excel base
    y muestra un resumen por ramal. Incluye copiar/pegar filas con Ctrl+C / Ctrl+V.
    """

    def __init__(self) -> None:
        super().__init__()
        self._equipos: List[str] = []
        self._usos: List[str] = []
        self._unidades: List[str] = []
        self._medidas_por_equipo: Dict[str, List[str]] = {}
        self._clipboard_row: Dict[int, str | int] = {}
        self._tabla4 = []
        self._metros_tension = {}
        self._metros_economico = {}
        self._preview_timer = QTimer(self)
        self._preview_timer.setSingleShot(True)
        self._preview_timer.timeout.connect(self._update_preview_now)
        self._build_ui()
        self._load_catalogos()
        self._group_spans: List[Tuple[str, int, int]] = []
        self._last_resultados: List[ResultadoRamal] = []
        self._seccion_resumen = []
        self._tabla4 = []
        self._metros_tension = {}
        self._metros_economico = {}

    # ------------------------------------------------------------------ economic analysis helpers
    @staticmethod
    def calcular_q(a: float, b: float, i: float, N: int) -> float:
        r = ((1 + a/100.0) ** 2 * (1 + b/100.0)) / (1 + i/100.0)
        if abs(1 - r) < 1e-9:
            return float(N)
        return (1 - r ** N) / (1 - r)

    @staticmethod
    def calcular_f(Np: int, Nc: int, horas_anuales: float, costo_kwh: float, i: float, Q: float) -> float:
        P_Wh = costo_kwh / 1000.0
        return Np * Nc * (horas_anuales * P_Wh) * Q / (1 + i/100.0)

    @staticmethod
    def numero_conductores(l1: float, l2: float, l3: float) -> int:
        fases = sum(1 for x in (l1, l2, l3) if abs(x) > 1e-3)
        if fases >= 3:
            return 5  # 3F + N + T
        if fases == 2:
            return 4  # 2F + N + T
        if fases == 1:
            return 3  # 1F + N + T
        return 0
    # ------------------------------------------------------------------ event filter
    def eventFilter(self, obj, event):
        if event.type() == QEvent.Wheel:
            return True
        if event.type() == QEvent.ContextMenu:
            # Reenviar clic derecho de los widgets de la tabla 1 al menú contextual de la tabla
            if hasattr(self, "sel_table") and self.sel_table and self.sel_table.isAncestorOf(obj):
                try:
                    pos = self.sel_table.viewport().mapFromGlobal(event.globalPos())
                    self._show_sel_context_menu(pos)
                    return True
                except Exception:
                    pass
        return super().eventFilter(obj, event)

    def _toggle_eco(self) -> None:
        enabled = self.chk_eco.isChecked()
        self.eco_frame.setVisible(enabled)
        self._schedule_preview()

    def _disable_wheel(self, widget):
        if widget is not None:
            widget.installEventFilter(self)

    def _force_upper_project(self, text: str) -> None:
        upper = (text or "").upper()
        if upper != text:
            pos = self.proj_edit.cursorPosition()
            self.proj_edit.blockSignals(True)
            self.proj_edit.setText(upper)
            self.proj_edit.setCursorPosition(pos)
            self.proj_edit.blockSignals(False)

    def _auto_height(self, table: QTableWidget, min_height: int = 0) -> None:
        """Ajusta altura del QTableWidget al contenido (sin scroll interno)."""
        if table is None:
            return
        rows = max(table.rowCount(), 1)
        vh = table.verticalHeader()
        row_h = vh.defaultSectionSize()
        vlen = vh.length()
        if vlen <= 0:
            vlen = rows * row_h
        header_h = table.horizontalHeader().height() if table.horizontalHeader() else 0
        frame = table.frameWidth() * 2
        total_h = header_h + vlen + frame + 4
        total_h = max(total_h, min_height)
        table.setMinimumHeight(total_h)
        # permitir crecer si hay mas filas
        table.setMaximumHeight(16777215)

    def _auto_height_table(self, table: QTableWidget) -> None:
        """Versión para mini tablas donde el header vertical ya está fijo."""
        if table is None:
            return
        header_h = table.horizontalHeader().height() if table.horizontalHeader() else 0
        row_h = table.verticalHeader().defaultSectionSize()
        rows = max(table.rowCount(), 1)
        total_h = header_h + row_h * rows + table.frameWidth() * 2 + 4
        table.setMinimumHeight(total_h)
        table.setMaximumHeight(total_h)

    def _adjust_medida_column(self) -> None:
        """Forzar un ancho mínimo en la columna MEDIDA para combos largos."""
        try:
            col = 2
            self.sel_table.resizeColumnsToContents()
            current = self.sel_table.columnWidth(col)
            min_w = 140
            if current < min_w:
                self.sel_table.setColumnWidth(col, min_w)
        except Exception:
            pass

    # ------------------------------------------------------------------ utils
    @staticmethod
    def _norm(text: str) -> str:
        import unicodedata
        s = str(text or "").strip().upper()
        s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
        return s

    @staticmethod
    def _slug(text: str) -> str:
        import re
        s = CargaElectricaPage._norm(text)
        s = re.sub(r"[^A-Z0-9]+", "_", s).strip("_")
        return s or "PROYECTO"

    @staticmethod
    def _unique_path(p: Path) -> Path:
        if not p.exists():
            return p
        for i in range(2, 10_000):
            cand = p.with_name(f"{p.stem}_{i}{p.suffix}")
            if not cand.exists():
                return cand
        return p

    def _apply_column_widths(self) -> None:
        hdr = self.sel_table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeToContents)
        hdr.setStretchLastSection(False)
        try:
            # MEDIDA debe ser ancho suficiente
            if self.sel_table.columnCount() > 2:
                self.sel_table.setColumnWidth(2, max(self.sel_table.columnWidth(2), 140))
            # distancia con ancho minimo
            if self.sel_table.columnCount() > 7:
                self.sel_table.setColumnWidth(7, max(self.sel_table.columnWidth(7), 80))
        except Exception:
            pass
        self._apply_result_widths()

    def _apply_result_widths(self) -> None:
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeToContents)
        hdr.setStretchLastSection(False)
        if isinstance(hdr, GroupHeaderView):
            hdr.set_groups(self._compute_group_spans())
        # Mantener las tres primeras columnas fijas en ancho para dar referencia
        self.table.setColumnWidth(0, 70)
        self.table.setColumnWidth(1, 200)
        self.table.setColumnWidth(2, 100)

    def _default_book(self) -> Path:
        base_dir = Path(__file__).resolve().parents[3] / "data"
        candidates = [
            base_dir / "CUADRO DE CARGAS -INDIVIDUAL JUAN LOZANO.xlsx",
            base_dir / "CUADRO DE CARGAS -INDIVIDUAL JUAN LOZANO.xlsm",
        ]
        for c in candidates:
            if c.exists():
                return c
        return candidates[0]

    # ------------------------------------------------------------------ UI
    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(10)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        root.addWidget(scroll)

        container = QWidget()
        scroll.setWidget(container)

        cl = QVBoxLayout(container)
        cl.setContentsMargins(0, 0, 0, 0)
        cl.setSpacing(10)

        card = QFrame()
        card.setStyleSheet("QFrame{background:#ffffff;border:1px solid #e2e8f5;border-radius:10px;}")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(16, 16, 16, 16)
        card_layout.setSpacing(10)

        primary_qss = (
            "QPushButton{"
            "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #1e3aef, stop:1 #14c9f5);"
            "color:#ffffff;font-weight:700;border:none;border-radius:10px;padding:10px 18px;"
            "}"
            "QPushButton:hover{"
            "background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #2749ff, stop:1 #29d6fa);"
            "}"
        )
        danger_qss = (
            "QPushButton{background:#fee2e2;color:#991b1b;font-weight:700;"
            "border-radius:6px;padding:6px 10px;border:none;}"
            "QPushButton:hover{background:#fecaca;}"
        )

        top_actions = QHBoxLayout()
        top_actions.setSpacing(8)
        self.btn_export = QPushButton("EXPORTAR PROYECTO"); self.btn_export.setStyleSheet(primary_qss); self.btn_export.clicked.connect(self._export_project)
        self.btn_legend = QPushButton("CARGAR LEGEND");    self.btn_legend.setStyleSheet(primary_qss);  self.btn_legend.clicked.connect(self._load_legend)
        self.btn_library = QPushButton("BIBLIOTECA");       self.btn_library.setStyleSheet(primary_qss); self.btn_library.clicked.connect(self._open_library)
        top_actions.addWidget(self.btn_export)
        top_actions.addWidget(self.btn_legend)
        top_actions.addWidget(self.btn_library)
        top_actions.addStretch(1)
        self.btn_reset = QPushButton("NUEVO PROYECTO")
        self.btn_reset.setStyleSheet(danger_qss)
        self.btn_reset.clicked.connect(self._reset_project)
        top_actions.addWidget(self.btn_reset)
        card_layout.addLayout(top_actions)

        title = QLabel("CARGA ELECTRICA")
        title.setStyleSheet("font-size:18px;font-weight:800;color:#0f172a;")
        card_layout.addWidget(title)

        top_form = QHBoxLayout()
        top_form.setSpacing(8)
        self.proj_edit = QLineEdit()
        self.proj_edit.setPlaceholderText("Nombre del proyecto")
        self.proj_edit.textChanged.connect(self._force_upper_project)
        self.ramales_spin = QSpinBox()
        self.ramales_spin.setRange(1, 200)
        self.ramales_spin.setValue(3)
        btn_gen = QPushButton("GENERAR TABLA")
        btn_gen.clicked.connect(self._gen_table)
        top_form.addWidget(QLabel("PROYECTO:"))
        top_form.addWidget(self.proj_edit, 1)
        top_form.addWidget(QLabel("# RAMALES:"))
        top_form.addWidget(self.ramales_spin)
        top_form.addWidget(btn_gen)
        card_layout.addLayout(top_form)

        # Factor de potencia para caida de tension
        fp_row = QHBoxLayout()
        fp_row.setSpacing(8)
        fp_row.addWidget(QLabel("FACTOR DE POTENCIA (FP):"))
        self.fp_spin = QDoubleSpinBox()
        self.fp_spin.setRange(0.5, 1.0)
        self.fp_spin.setSingleStep(0.01)
        self.fp_spin.setDecimals(2)
        self.fp_spin.setValue(0.95)
        self.fp_spin.valueChanged.connect(self._schedule_preview)
        fp_row.addWidget(self.fp_spin)

        fp_row.addWidget(QLabel("TEMPERATURA AMBIENTE (°C):"))
        self.temp_spin = QDoubleSpinBox()
        self.temp_spin.setRange(-20, 90)
        self.temp_spin.setDecimals(1)
        self.temp_spin.setSingleStep(1.0)
        self.temp_spin.setValue(30.0)
        self.temp_spin.valueChanged.connect(self._schedule_preview)
        fp_row.addWidget(self.temp_spin)

        # Check y parámetros de análisis económico RETIE
        self.chk_eco = QCheckBox("ANÁLISIS ECONÓMICO RETIE")
        self.chk_eco.stateChanged.connect(self._toggle_eco)
        fp_row.addWidget(self.chk_eco)

        fp_row.addStretch(1)
        card_layout.addLayout(fp_row)

        # Parámetros económicos (se muestran solo si el check está activo)
        self.eco_frame = QFrame()
        self.eco_frame.setStyleSheet("QFrame{background:#f8fafc;border:1px solid #e2e8f5;border-radius:8px;}")
        eco_layout = QHBoxLayout(self.eco_frame)
        eco_layout.setContentsMargins(8, 6, 8, 6)
        eco_layout.setSpacing(8)

        def add_param(label: str, widget: QDoubleSpinBox):
            box = QVBoxLayout()
            box.setSpacing(2)
            box.addWidget(QLabel(label))
            box.addWidget(widget)
            eco_layout.addLayout(box)

        self.eco_costo = QDoubleSpinBox(); self.eco_costo.setRange(0, 1e6); self.eco_costo.setDecimals(2); self.eco_costo.setValue(900.0)
        self.eco_horas = QDoubleSpinBox(); self.eco_horas.setRange(0, 20000); self.eco_horas.setDecimals(0); self.eco_horas.setValue(8760)
        self.eco_a = QDoubleSpinBox(); self.eco_a.setRange(0, 100); self.eco_a.setDecimals(2); self.eco_a.setValue(1.0)
        self.eco_b = QDoubleSpinBox(); self.eco_b.setRange(0, 100); self.eco_b.setDecimals(2); self.eco_b.setValue(3.0)
        self.eco_i = QDoubleSpinBox(); self.eco_i.setRange(0, 100); self.eco_i.setDecimals(2); self.eco_i.setValue(5.0)
        self.eco_n = QDoubleSpinBox(); self.eco_n.setRange(1, 100); self.eco_n.setDecimals(0); self.eco_n.setValue(20)

        for w in (self.eco_costo, self.eco_horas, self.eco_a, self.eco_b, self.eco_i, self.eco_n):
            w.valueChanged.connect(self._schedule_preview)

        add_param("Costo kWh ($)", self.eco_costo)
        add_param("Horas/año", self.eco_horas)
        add_param("Aum. carga (%)", self.eco_a)
        add_param("Aum. energía (%)", self.eco_b)
        add_param("Tasa descuento (%)", self.eco_i)
        add_param("Vida económica (años)", self.eco_n)
        card_layout.addWidget(self.eco_frame)
        self.eco_frame.setVisible(False)

        # Atajos de teclado para copiar/pegar filas
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self)
        self.copy_shortcut.activated.connect(self._copy_row)
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self)
        self.paste_shortcut.activated.connect(self._paste_row)

        self.status = QLabel("")  # se mantiene para mensajes internos, no se agrega al layout

        # Tabla de entrada (sin columna Ramal)
        self.sel_table = QTableWidget(0, 8)
        self.sel_table.setHorizontalHeaderLabels([
            "EQUIPO", "USO", "MEDIDA", "DESCONGELAMIENTO",
            "TIENE UNIDAD", "REF UNIDAD", "NIVELES", "DISTANCIA"
        ])
        hdr = self.sel_table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeToContents)
        hdr.setMinimumSectionSize(70)
        # Asegurar que la columna MEDIDA (idx 2) se expanda lo suficiente para mostrar modelos largos
        hdr.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        # Forzar un mínimo similar al aplicado en EQUIPO
        self.sel_table.setColumnWidth(2, 200)
        self.sel_table.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
        # Forzar actualización de anchos después de poblar
        self.sel_table.resizeColumnsToContents()
        self.sel_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.sel_table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.sel_table.verticalHeader().setVisible(True)
        self.sel_table.verticalHeader().setDefaultSectionSize(26)
        self.sel_table.setStyleSheet("QTableWidget{background:#ffffff;border:1px solid #e2e8f5;border-radius:8px;}")
        self.sel_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.sel_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.sel_table.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
        self.sel_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.sel_table.customContextMenuRequested.connect(self._show_sel_context_menu)
        self.sel_table.viewport().setAttribute(Qt.WA_TransparentForMouseEvents, False)
        card_layout.addWidget(self.sel_table)

        # Tabla de salida (cabecera agrupada)
        # Orden solicitado: ventiladores, iluminacion, resistencia principal, compresor,
        # condensador, y luego el resto de resistencias auxiliares.
        self._grouped_cols = [
            ("Ventiladores", ("VENTILADORES",)),
            ("Iluminacion", ("ILUMINACION",)),
            ("Resistencia", ("RESISTENCIAS",)),
            ("Compresor", ("COMPRESOR",)),
            ("Condensador", ("CONDENSADOR",)),
            ("Resistencia desempanante", ("RESISTENCIA DESEMPANANTE",)),
            ("Resistencia antisudor", ("RESISTENCIA ANTISUDOR",)),
            ("Resistencia desague", ("RESISTENCIA DESAGUE",)),
            ("Resistencia calefactora", ("RESISTENCIA CALEFACTORA",)),
        ]
        header_cols = ["RAMAL", "EQUIPO", "MEDIDA", "TOTAL L1 (A)", "TOTAL L2 (A)", "TOTAL L3 (A)"]
        for name, _ in self._grouped_cols:
            header_cols.extend(["L1 (A)", "L2 (A)", "L3 (A)"])
        header_cols.extend(["CALIBRE CAIDA", "%VD"])

        self.table = QTableWidget(0, len(header_cols))
        self.table.setHorizontalHeaderLabels(header_cols)
        group_header = GroupHeaderView(self._compute_group_spans(), self.table)
        group_header.setDefaultAlignment(Qt.AlignCenter)
        group_header.setStyleSheet("QHeaderView::section { padding:12px 12px 6px 12px; font-weight:600; }")
        self.table.setHorizontalHeader(group_header)
        self.table.verticalHeader().setVisible(False)
        self.table.setStyleSheet("QTableWidget{background:#ffffff;border:1px solid #e2e8f5;border-radius:8px;}")
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
        # Asegurar un alto visible y que se expanda en el contenedor
        self.table.setMinimumHeight(320)
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        card_layout.addWidget(self.table)
        card_layout.setAlignment(self.table, Qt.AlignLeft)
        self.table.setVisible(False)

        # Resumen global + Tabla 4
        summary_row = QHBoxLayout()
        summary_row.setSpacing(12)

        self.summary_card = QFrame()
        self.summary_card.setMaximumWidth(320)
        self.summary_card.setStyleSheet("QFrame{background:#ffffff;border:1px solid #d7e3f7;border-radius:10px;}")
        self.summary_card.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Preferred)
        grid = QGridLayout(self.summary_card)
        grid.setContentsMargins(8, 6, 8, 6)
        grid.setHorizontalSpacing(6)
        grid.setVerticalSpacing(6)

        # Factor arriba
        lbl_factor = QLabel("FACTOR RESISTENCIAS:")
        lbl_factor.setStyleSheet("font-weight:700;color:#0f172a;")
        self.factor_spin = QDoubleSpinBox()
        self.factor_spin.setDecimals(2)
        self.factor_spin.setRange(0, 10)
        self.factor_spin.setSingleStep(0.1)
        self.factor_spin.setValue(1.0)
        self.factor_spin.valueChanged.connect(self._recompute_summary)
        if self.factor_spin.lineEdit():
            self.factor_spin.lineEdit().editingFinished.connect(self._on_factor_commit)
        self.factor_spin.setFixedWidth(80)
        grid.addWidget(lbl_factor, 0, 0, 1, 1)
        grid.addWidget(self.factor_spin, 0, 1, 1, 1)

        # Mini tabla de resumen
        self.summary_table = QTableWidget(2, 3)
        self.summary_table.setHorizontalHeaderLabels(["L1 (A)", "L2 (A)", "L3 (A)"])
        self.summary_table.setVerticalHeaderLabels(["TOTAL GENERAL", "TOTAL AJUSTADO"])
        hh = self.summary_table.horizontalHeader()
        vh = self.summary_table.verticalHeader()
        hh.setSectionResizeMode(QHeaderView.Fixed)
        vh.setSectionResizeMode(QHeaderView.Fixed)
        hh.setFixedHeight(22)
        for i in range(3):
            self.summary_table.setColumnWidth(i, 70)
        width_needed = 70 * 3 + vh.sizeHint().width() + self.summary_table.frameWidth() * 2
        self.summary_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.summary_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.summary_table.verticalHeader().setDefaultSectionSize(24)
        self.summary_table.setSelectionMode(QTableWidget.NoSelection)
        self.summary_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.summary_table.setFixedWidth(width_needed)
        grid.addWidget(self.summary_table, 1, 0, 1, 2)

        summary_row.addWidget(self.summary_card)
        card_layout.addLayout(summary_row)
        self.summary_card.setVisible(False)

        # Tabla 4 (análisis económico RETIE) en su propia fila debajo de la tabla 3
        self.table4 = QTableWidget(0, 11)
        self.table4.setHorizontalHeaderLabels([
            "RAMAL", "CALIBRE CALCULADO", "CI CALCULADO ($)", "CJ CALCULADO ($)", "CT CALCULADO ($)",
            "CALIBRE SUPERIOR", "CI SUPERIOR ($)", "CJ SUPERIOR ($)", "CT SUPERIOR ($)", "RECOMENDADO", "AHORRO ($)"
        ])
        self.table4.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table4.verticalHeader().setVisible(False)
        self.table4.setStyleSheet("QTableWidget{background:#ffffff;border:1px solid #e2e8f5;border-radius:8px;}")
        self.table4.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.table4.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table4.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table4.setVisible(False)
        card_layout.addWidget(self.table4)
        self.table4.setVisible(False)

        # Resumen de metros por calibre (tensión vs económico)
        self.table_cables = QTableWidget(0, 3)
        self.table_cables.setHorizontalHeaderLabels(["CALIBRE", "CAIDA DE TENSION (m)", "CALCULO ECONOMICO (m)"])
        self.table_cables.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table_cables.verticalHeader().setVisible(False)
        self.table_cables.setStyleSheet("QTableWidget{background:#ffffff;border:1px solid #e2e8f5;border-radius:8px;}")
        self.table_cables.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.table_cables.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table_cables.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table_cables.setVisible(False)

        # fila para summary + resumen cables
        summary_row = QHBoxLayout()
        summary_row.setSpacing(12)
        summary_row.addWidget(self.summary_card)
        summary_row.addWidget(self.table_cables, 1)
        card_layout.addLayout(summary_row)

        # Notas económicas
        self.notes_label = QLabel()
        self.notes_label.setStyleSheet("font-size:12px; color:#0f172a;")
        self.notes_label.setVisible(False)
        card_layout.addWidget(self.notes_label)

        cl.addWidget(card)
        self._apply_result_widths()

    # ------------------------------------------------------------------ datos / catalogos
    def _load_catalogos(self) -> None:
        book = self._default_book()
        if not book.exists():
            return
        import pandas as pd
        try:
            df_muebles = pd.read_excel(book, sheet_name="MUEBLES")
            df_muebles.columns = [str(c).strip() for c in df_muebles.columns]

            def pick_col(df, target: str) -> str:
                cols = {self._norm(c): c for c in df.columns}
                return cols.get(self._norm(target), list(df.columns)[0])

            col_meq = pick_col(df_muebles, "NOMBRE")
            self._equipos = sorted(set(str(x).strip() for x in df_muebles[col_meq].dropna().unique() if str(x).strip()))

            self._usos = USOS_DEFAULT[:]

            col_dim = pick_col(df_muebles, "DIMENSION")
            for _, row in df_muebles.iterrows():
                eq = str(row.get(col_meq, "")).strip()
                dim = str(row.get(col_dim, "")).strip()
                if eq and dim:
                    self._medidas_por_equipo.setdefault(self._norm(eq), []).append(dim)
            for k, vals in list(self._medidas_por_equipo.items()):
                self._medidas_por_equipo[k] = sorted(set(vals))

            df_uni = pd.read_excel(book, sheet_name="UNIDADES", header=None)
            col_uni = df_uni.columns[2] if len(df_uni.columns) > 2 else df_uni.columns[0]
            self._unidades = sorted(set(str(x).strip() for x in df_uni[col_uni].dropna().unique() if str(x).strip()))

            if hasattr(self, "status"):
                self.status.setText(f"Catalogos: {len(self._equipos)} equipos, {len(self._usos)} usos, {len(self._unidades)} unidades.")
            self._apply_column_widths()
        except Exception as e:
            self._equipos = []
            self._usos = []
            self._unidades = []
            self._medidas_por_equipo = {}
            if hasattr(self, "status"):
                self.status.setText(f"Error cargando catalogos: {e}")

    # ------------------------------------------------------------------ tabla entrada
    def _gen_table(self) -> None:
        target = self.ramales_spin.value()
        current = self.sel_table.rowCount()
        for i in range(current, target):
            self.sel_table.insertRow(i)
            eq_combo = QComboBox()
            eq_combo.addItems([""] + self._equipos)
            uso_combo = QComboBox()
            uso_combo.addItems([""] + self._usos)
            med_combo = QComboBox()
            med_combo.addItem("")
            med_combo.setMinimumContentsLength(14)
            med_combo.setSizeAdjustPolicy(QComboBox.AdjustToContents)
            desc_combo = QComboBox()
            desc_combo.addItems(["", "RESISTENCIAS", "GAS CALIENTE", "TIEMPO"])
            tiene_combo = QComboBox()
            tiene_combo.addItems(["", "SI", "NO"])
            uni_combo = QComboBox()
            uni_combo.addItems([""] + self._unidades)
            niveles_spin = QSpinBox()
            niveles_spin.setRange(0, 10)
            dist_edit = QLineEdit()
            dist_edit.setPlaceholderText("m")
            dist_edit.textChanged.connect(self._schedule_preview)
            seccion_edit = QLineEdit()
            seccion_edit.setPlaceholderText("mm2")
            seccion_edit.textChanged.connect(self._schedule_preview)

            for w in (eq_combo, uso_combo, med_combo, desc_combo, tiene_combo, uni_combo, niveles_spin, dist_edit, seccion_edit):
                self._disable_wheel(w)

            eq_combo.currentTextChanged.connect(lambda txt, row=i, med_combo=med_combo: self._on_equipo_changed(row, txt, med_combo))
            def on_tiene_changed(text, uni_box=uni_combo):
                enabled = text.strip().upper() == "SI"
                uni_box.setEnabled(enabled)
                if not enabled:
                    uni_box.setCurrentIndex(0)
            tiene_combo.currentTextChanged.connect(on_tiene_changed)
            uni_combo.setEnabled(False)
            for combo in [eq_combo, uso_combo, med_combo, desc_combo, tiene_combo, uni_combo]:
                combo.currentIndexChanged.connect(self._schedule_preview)
            niveles_spin.valueChanged.connect(self._schedule_preview)

            self.sel_table.setCellWidget(i, 0, eq_combo)
            self.sel_table.setCellWidget(i, 1, uso_combo)
            self.sel_table.setCellWidget(i, 2, med_combo)
            self.sel_table.setCellWidget(i, 3, desc_combo)
            self.sel_table.setCellWidget(i, 4, tiene_combo)
            self.sel_table.setCellWidget(i, 5, uni_combo)
            self.sel_table.setCellWidget(i, 6, niveles_spin)
            self.sel_table.setCellWidget(i, 7, dist_edit)
            self.sel_table.setCellWidget(i, 8, seccion_edit)
        if target < current:
            for _ in range(current - target):
                self.sel_table.removeRow(self.sel_table.rowCount() - 1)
        self._apply_column_widths()
        self._auto_height(self.sel_table, min_height=300)
        self._adjust_medida_column()
        self._schedule_preview()

    def _on_equipo_changed(self, row: int, equipo: str, med_combo: QComboBox) -> None:
        med_combo.clear()
        med_combo.addItem("")
        medidas = self._medidas_por_equipo.get(self._norm(equipo), [])
        med_combo.addItems(medidas)

    def _choose_dimension(self, equipo: str, length_ft: float, ramales: int) -> str:
        """Dada una longitud en pies y un equipo, selecciona la dimension estandar en metros aproximada."""
        dims = self._medidas_por_equipo.get(self._norm(equipo), [])
        # convertir a float las dimensiones numericas
        numeric_dims = []
        for d in dims:
            try:
                numeric_dims.append(float(str(d).replace(",", ".")))
            except Exception:
                continue
        if not numeric_dims:
            # fallback: longitud en metros
            return f"{round(length_ft*0.3048/ max(1, ramales),2)}"
        target_m = (length_ft * 0.3048) / max(1, ramales)
        # preferir el tamaño minimo que cumpla target, si no el mayor disponible
        numeric_dims = sorted(set(numeric_dims))
        chosen = numeric_dims[-1]
        for d in numeric_dims:
            if d >= target_m:
                chosen = d
                break
        # devolver con formato original (manteniendo punto)
        return str(chosen)

    def _split_doors_modules(self, equipo: str, doors: int) -> list[str]:
        """Divide una cantidad de puertas en modulaciones existentes para ese equipo (preferir grandes)."""
        avail = []
        for d in self._medidas_por_equipo.get(self._norm(equipo), []):
            s = str(d).strip().upper()
            if s.endswith("P"):
                try:
                    avail.append(int("".join(ch for ch in s if ch.isdigit())))
                except Exception:
                    continue
        avail = sorted(set(avail), reverse=True)
        if not avail:
            return [f"{doors}P"]
        modules = []
        rem = doors
        smallest = avail[-1]
        while rem > 0:
            pick = None
            for a in avail:
                if a <= rem:
                    pick = a
                    break
            if pick is None:
                pick = smallest
            modules.append(f"{pick}P")
            rem -= pick
            if rem <= 0:
                break
            # evitar loop infinito
            if rem < smallest:
                modules.append(f"{smallest}P")
                rem = 0
        return modules

    def _split_autoservicio(self, length_ft: float) -> list[str]:
        """Divide longitud en metros en modulos preferidos 3.75, 2.5, 1.9."""
        preferred = [3.75, 2.5, 1.9]
        length_m = max(0.0, length_ft * 0.3048)
        modules = []
        rem = length_m
        while rem > 0.2:
            pick = None
            for p in preferred:
                if p <= rem + 0.05:
                    pick = p
                    break
            if pick is None:
                pick = preferred[-1]
            modules.append(str(pick))
            rem -= pick
            if len(modules) > 200:  # safety
                break
        if not modules:
            modules.append(str(preferred[0]))
        return modules

    def _split_length_generic(self, equipo: str, length_m: float) -> list[str]:
        """Divide longitud en metros en modulos del catalogo del equipo (greedy, prefiere grandes)."""
        dims = self._medidas_por_equipo.get(self._norm(equipo), [])
        numeric_dims = []
        for d in dims:
            try:
                numeric_dims.append(float(str(d).replace(",", ".")))
            except Exception:
                continue
        numeric_dims = sorted(set(numeric_dims), reverse=True)
        if not numeric_dims or length_m <= 0:
            return []
        modules = []
        rem = length_m
        smallest = numeric_dims[-1]
        limit = 0
        while rem > 0.05 and limit < 200:
            pick = None
            for d in numeric_dims:
                if d <= rem + 0.05:
                    pick = d
                    break
            if pick is None:
                pick = smallest
            modules.append(str(pick))
            rem -= pick
            limit += 1
            if rem < smallest:
                break
        if not modules:
            modules.append(str(numeric_dims[0]))
        return modules

    def _compute_group_spans(self) -> List[Tuple[str, int, int]]:
        groups: List[Tuple[str, int, int]] = []
        start = 6  # despues de Ramal, Equipo, Medida y Totales (3+3)
        for name, _ in getattr(self, "_grouped_cols", []):
            groups.append((name.upper(), start, 3))
            start += 3
        return groups

    def _collect_selection(self) -> List[SeleccionRamal]:
        selection: List[SeleccionRamal] = []
        rows = self.sel_table.rowCount()
        for r in range(rows):
            def widget_text(col: int) -> str:
                w = self.sel_table.cellWidget(r, col)
                if isinstance(w, QComboBox):
                    return w.currentText().strip()
                if isinstance(w, QLineEdit):
                    return w.text().strip()
                return ""

            ramal = r + 1
            niveles_val = 0
            w_niv = self.sel_table.cellWidget(r, 6)
            if isinstance(w_niv, QSpinBox):
                niveles_val = w_niv.value()
            dist_val = widget_text(7)
            selection.append(
                SeleccionRamal(
                    ramal=ramal,
                    equipo=widget_text(0),
                    uso=widget_text(1),
                    medida=widget_text(2),
                    descongelamiento=widget_text(3),
                    tiene_unidad=widget_text(4).upper() == "SI",
                    ref_unidad=widget_text(5),
                    niveles=niveles_val,
                )
            )
        return selection

    def _populate_row_from_data(self, row: int, data: Dict[str, object]) -> None:
        # asegura row existe
        while self.sel_table.rowCount() <= row:
            self.sel_table.insertRow(self.sel_table.rowCount())
        widgets = [
            self.sel_table.cellWidget(row, i) for i in range(9)
        ]
        eq_combo, uso_combo, med_combo, desc_combo, tiene_combo, uni_combo, niv_spin, dist_edit, sec_edit = widgets
        if isinstance(eq_combo, QComboBox):
            self._set_combo_text(eq_combo, str(data.get("equipo", "")))
            self._on_equipo_changed(row, eq_combo.currentText(), med_combo if isinstance(med_combo, QComboBox) else QComboBox())
        if isinstance(uso_combo, QComboBox):
            self._set_combo_text(uso_combo, str(data.get("uso", "")))
        if isinstance(med_combo, QComboBox):
            self._set_combo_text(med_combo, str(data.get("medida", "")))
        if isinstance(desc_combo, QComboBox):
            self._set_combo_text(desc_combo, str(data.get("descongelamiento", "")))
        if isinstance(tiene_combo, QComboBox):
            val = "SI" if str(data.get("tiene_unidad", "")).upper() == "SI" or data.get("tiene_unidad") is True else "NO" if data.get("tiene_unidad") is not None else ""
            self._set_combo_text(tiene_combo, val)
            if isinstance(uni_combo, QComboBox):
                uni_combo.setEnabled(val == "SI")
        if isinstance(uni_combo, QComboBox):
            self._set_combo_text(uni_combo, str(data.get("ref_unidad", "")))
        if isinstance(niv_spin, QSpinBox):
            try:
                niv_spin.setValue(int(data.get("niveles", 0)))
            except Exception:
                niv_spin.setValue(0)
        if isinstance(dist_edit, QLineEdit):
            dist_edit.setText(str(data.get("distancia", "")) if data.get("distancia", "") is not None else "")
        if isinstance(sec_edit, QLineEdit):
            sec_edit.setText(str(data.get("seccion_mm2", "")) if data.get("seccion_mm2", "") is not None else "")

    def _reset_project(self) -> None:
        """Limpia formulario y tablas para iniciar un proyecto nuevo."""
        self.proj_edit.clear()
        self.ramales_spin.setValue(3)
        self.sel_table.setRowCount(0)
        self._clipboard_row = {}
        self._gen_table()
        self.table.setRowCount(0)
        self.status.setText("Proyecto nuevo iniciado.")

    def _load_legend(self) -> None:
        svc = LegendJDService()
        try:
            data = svc.load_all()
        except Exception as e:
            QMessageBox.critical(self, "Legend", f"No se pudo cargar LEGEND: {e}")
            return

        cfg = data.get("config")
        if cfg and getattr(cfg, "proyecto", None):
            self.proj_edit.setText(str(cfg.proyecto).upper())

        usos = data.get("usos", {})
        usos_bt = usos.get("BT", [])
        usos_mt = usos.get("MT", [])
        summary = (
            f"Carpeta: {data.get('sources', {}).get('folder')}\n"
            f"Archivos: {', '.join(data.get('sources', {}).get('files_found', [])) or 'ninguno'}\n"
            f"Equipos: {len(data.get('equipos', []))} | Usos BT: {len(usos_bt)} | Usos MT: {len(usos_mt)}\n"
            f"Variadores: {len(data.get('variadores', []))} | WCR: {len(data.get('wcr', []))}\n"
            f"Plantillas: {{ {', '.join(f'{k}:{len(v)}' for k, v in data.get('plantillas', {}).items())} }}"
        )
        QMessageBox.information(self, "Legend", summary)

    # ------------------------------------------------------------------ copiar / pegar
    def _copy_row(self) -> None:
        sel = self.sel_table.selectionModel().selectedRows()
        if not sel:
            self.status.setText("Selecciona una fila (clic en la celda o en el numero de fila) y vuelve a copiar.")
            return
        r = sel[0].row()
        data: Dict[int, str | int] = {}
        for c in range(9):
            w = self.sel_table.cellWidget(r, c)
            if isinstance(w, QComboBox):
                data[c] = w.currentText()
            elif isinstance(w, QSpinBox):
                data[c] = w.value()
            elif isinstance(w, QLineEdit):
                data[c] = w.text()
        self._clipboard_row = data
        self.status.setText(f"Fila {r+1} copiada.")

    def _set_combo_text(self, combo: QComboBox, value: str) -> None:
        idx = combo.findText(value, Qt.MatchFixedString)
        if idx < 0:
            combo.addItem(value)
            idx = combo.count() - 1
        combo.setCurrentIndex(idx)

    def _paste_row(self) -> None:
        if not self._clipboard_row:
            self.status.setText("No hay fila copiada. Usa Ctrl+C primero.")
            return
        sel = self.sel_table.selectionModel().selectedRows()
        if not sel:
            self.status.setText("Selecciona una fila destino y pulsa Ctrl+V.")
            return
        data = self._clipboard_row
        rows = [idx.row() for idx in sel]
        for r in rows:
            for c in range(9):
                w = self.sel_table.cellWidget(r, c)
                if isinstance(w, QComboBox):
                    val = str(data.get(c, ""))
                    if c == 2:
                        eq_w = self.sel_table.cellWidget(r, 0)
                        if isinstance(eq_w, QComboBox):
                            self._on_equipo_changed(r, eq_w.currentText(), w)
                    if val:
                        self._set_combo_text(w, val)
                elif isinstance(w, QSpinBox):
                    try:
                        w.setValue(int(data.get(c, 0)))
                    except Exception:
                        w.setValue(0)
                elif isinstance(w, QLineEdit):
                    w.setText(str(data.get(c, "")))
        self.status.setText(f"Pegue en filas {', '.join(str(r+1) for r in rows)}.")
        self._schedule_preview()

    # ------------------------------------------------------------------ calculo / salida
    def _run(self) -> None:
        book = self._default_book()
        if not book.exists():
            QMessageBox.warning(self, "Archivo no encontrado", f"No se encontro:\n{book}\nColocalo en data/.")
            return
        try:
            selection = self._collect_selection()
            resultados = recalcular(book, write_sheet=True, sheet_name="CUADRO_GENERADO", selection=selection)
            self._fill_table(resultados)
            self.status.setText(f"Listo. Recalculado y escrito en {book.name} (CUADRO_GENERADO).")
        except Exception as e:
            self.status.setText(f"Error: {e}")
            QMessageBox.critical(self, "Error", str(e))

    def _export_project(self) -> None:
        book = self._default_book()
        if not book.exists():
            QMessageBox.warning(self, "Archivo no encontrado", f"No se encontro:\n{book}\nColocalo en data/.")
            return
        selection = self._collect_selection()
        proj_name = self.proj_edit.text().strip() or "PROYECTO"
        stamp = datetime.now().strftime("%Y%m%d")
        base_name = f"{stamp}_{self._slug(proj_name)}"
        base_dir = Path(r"C:\Users\ingmontajes4\Documents\PC-JUAN LOZANO 22032022\WESTON\JUAN LOZANO\CUADRO DE CARGAS")
        target_dir = QFileDialog.getExistingDirectory(self, "Selecciona carpeta destino", str(base_dir))
        if not target_dir:
            return
        target_dir = Path(target_dir)
        excel_out = target_dir / f"{base_name}.xlsx"
        try:
            resultados = recalcular(book, write_sheet=False, selection=selection)
            dist_list = [self._get_distance_for_row(i) for i in range(self.sel_table.rowCount())]
            self._write_export_excel(
                excel_out,
                proj_name,
                selection,
                resultados,
                self._tabla4 if self.chk_eco.isChecked() else None,
                self._metros_tension,
                self._metros_economico if self.chk_eco.isChecked() else None,
                dist_list,
            )
            data = {
                "proyecto": proj_name,
                "fecha": stamp,
                "analisis_economico": self.chk_eco.isChecked(),
                "eco_params": {
                    "costo_kwh": self.eco_costo.value(),
                    "horas_anuales": self.eco_horas.value(),
                    "aumento_carga": self.eco_a.value(),
                    "aumento_energia": self.eco_b.value(),
                    "tasa_descuento": self.eco_i.value(),
                    "vida_economica": self.eco_n.value(),
                },
                "ramales": [],
            }
            for idx, sel in enumerate(selection):
                dist = dist_list[idx] if idx < len(dist_list) else self._get_distance_for_row(idx)
                data["ramales"].append({
                    "ramal": sel.ramal,
                    "equipo": sel.equipo,
                    "uso": sel.uso,
                    "medida": sel.medida,
                    "descongelamiento": sel.descongelamiento,
                    "tiene_unidad": sel.tiene_unidad,
                    "ref_unidad": sel.ref_unidad,
                    "niveles": sel.niveles,
                    "distancia": dist,
                })
            # Guardar JSON solo en biblioteca interna
            lib_dir = Path(__file__).resolve().parents[3] / "data" / "proyectos" / "carga_electrica"
            lib_dir.mkdir(parents=True, exist_ok=True)
            lib_json = self._unique_path(lib_dir / f"{base_name}.json")
            with open(lib_json, "w", encoding="utf-8") as fh:
                json.dump(data, fh, ensure_ascii=False, indent=2)

            self.status.setText(f"Exportado: {excel_out.name} y JSON en biblioteca")
            QMessageBox.information(self, "Exportado", f"Archivos guardados:\nExcel: {excel_out}\nJSON (biblioteca): {lib_json}")
        except Exception as e:
            self.status.setText(f"Error al exportar: {e}")
            QMessageBox.critical(self, "Error", str(e))

    def _import_project(self) -> None:
        base_dir = Path(r"C:\Users\ingmontajes4\Documents\PC-JUAN LOZANO 22032022\WESTON\JUAN LOZANO\CUADRO DE CARGAS")
        path, _ = QFileDialog.getOpenFileName(self, "Cargar proyecto", str(base_dir), "JSON (*.json)")
        if not path:
            return
        self._load_project_file(path)

    def _load_project_file(self, path: str) -> None:
        """
        Carga un proyecto .json y lo aplica a la tabla (sin diálogos).
        Usado por el botón "CARGAR PROYECTO" y por la biblioteca local.
        """
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            proj = data.get("proyecto", "")
            ramales = data.get("ramales", [])
            self.proj_edit.setText(proj)
            self.ramales_spin.setValue(max(len(ramales), 1))
            self.sel_table.setRowCount(0)
            self._gen_table()
            for idx, row in enumerate(ramales):
                self._populate_row_from_data(idx, row)
            self._schedule_preview()
            self.status.setText(f"Proyecto cargado desde {Path(path).name}")
        except Exception as e:
            self.status.setText(f"Error al cargar proyecto: {e}")
            QMessageBox.critical(self, "Error", str(e))

    # ------------------------------------------------------------------ Excel export helper
    def _write_export_excel(self, path: Path, proyecto: str, seleccion: List[SeleccionRamal], resultados: List[ResultadoRamal], tabla4: List[Dict[str, object]] | None = None, metros_t: Dict[str, float] | None = None, metros_e: Dict[str, float] | None = None, dist_list: List[float] | None = None) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "CUADRO"

        header_fill = PatternFill("solid", fgColor="DCEAFB")
        header_font = Font(bold=True, color="0B1930")
        total_fill = PatternFill("solid", fgColor="FFE8D5")
        total_font = Font(bold=True, color="8B3A00")
        border = Border(
            left=Side(style="thin", color="B0BEC5"),
            right=Side(style="thin", color="B0BEC5"),
            top=Side(style="thin", color="B0BEC5"),
            bottom=Side(style="thin", color="B0BEC5"),
        )

        # Titulo
        title_text = f"CUADRO DE CARGAS - {proyecto.upper() or 'PROYECTO'}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
        tcell = ws.cell(row=1, column=1, value=title_text)
        tcell.font = Font(bold=True, size=14, color="0B1930")
        tcell.alignment = Alignment(horizontal="center", vertical="center")
        row_idx = 2

        # Parámetros usados
        ws.cell(row=row_idx, column=1, value="PARAMETROS").font = Font(bold=True)
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="FACTOR DE POTENCIA").value = "FACTOR DE POTENCIA"
        ws.cell(row=row_idx, column=2, value=self.fp_spin.value())
        ws.cell(row=row_idx, column=3, value="TEMP AMBIENTE (°C)")
        ws.cell(row=row_idx, column=4, value=self.temp_spin.value())
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="ANALISIS ECONOMICO RETIE")
        ws.cell(row=row_idx, column=2, value="SI" if self.chk_eco.isChecked() else "NO")
        if self.chk_eco.isChecked():
            ws.cell(row=row_idx, column=3, value="COSTO KWH ($)")
            ws.cell(row=row_idx, column=4, value=self.eco_costo.value())
            ws.cell(row=row_idx, column=5, value="HORAS/AÑO")
            ws.cell(row=row_idx, column=6, value=self.eco_horas.value())
            ws.cell(row=row_idx, column=7, value="AUM. CARGA (%)")
            ws.cell(row=row_idx, column=8, value=self.eco_a.value())
            ws.cell(row=row_idx, column=9, value="AUM. ENERGIA (%)")
            ws.cell(row=row_idx, column=10, value=self.eco_b.value())
            ws.cell(row=row_idx, column=11, value="TASA DESC. (%)")
            ws.cell(row=row_idx, column=12, value=self.eco_i.value())
            ws.cell(row=row_idx, column=13, value="VIDA ECO (AÑOS)")
            ws.cell(row=row_idx, column=14, value=self.eco_n.value())
        row_idx += 2

        # Tabla de entrada
        sel_headers = ["RAMAL", "EQUIPO", "USO", "MEDIDA", "DESCONGELAMIENTO", "TIENE UNIDAD", "REF UNIDAD", "NIVELES", "DISTANCIA (m)"]
        for c, text in enumerate(sel_headers, start=1):
            cell = ws.cell(row=row_idx, column=c, value=text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for idx_sel, sel in enumerate(seleccion):
            row_idx += 1
            dist_val = dist_list[idx_sel] if dist_list and idx_sel < len(dist_list) else self._get_distance_for_row(sel.ramal - 1)
            vals = [
                sel.ramal,
                (sel.equipo or "").upper(),
                (sel.uso or "").upper(),
                (sel.medida or "").upper(),
                (sel.descongelamiento or "").upper(),
                "SI" if sel.tiene_unidad else "NO",
                sel.ref_unidad,
                sel.niveles,
                dist_val,
            ]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row=row_idx, column=c, value=v)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        row_idx += 2  # espacio

        # Tabla de salida
        base_headers = ["RAMAL", "EQUIPO", "MEDIDA", "TOTAL L1 (A)", "TOTAL L2 (A)", "TOTAL L3 (A)"]
        group_names = [name for name, _ in self._grouped_cols]
        # fila superior (grupos)
        top_row = row_idx
        bottom_row = row_idx + 1
        col_idx = 1
        for bh in base_headers:
            ws.cell(row=top_row, column=col_idx, value=bh)
            ws.cell(row=bottom_row, column=col_idx, value="")
            ws.merge_cells(start_row=top_row, start_column=col_idx, end_row=bottom_row, end_column=col_idx)
            cell = ws.cell(row=top_row, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            col_idx += 1
        for name in group_names:
            ws.merge_cells(start_row=top_row, start_column=col_idx, end_row=top_row, end_column=col_idx + 2)
            for offset, label in enumerate(["L1 (A)", "L2 (A)", "L3 (A)"]):
                cell = ws.cell(row=bottom_row, column=col_idx + offset, value=label)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            cell_title = ws.cell(row=top_row, column=col_idx, value=name.upper())
            cell_title.fill = header_fill
            cell_title.font = header_font
            cell_title.alignment = Alignment(horizontal="center", vertical="center")
            cell_title.border = border
            col_idx += 3

        # Datos de salida
        for res in resultados:
            bottom_row += 1
            row_vals = [
                res.ramal.ramal,
                f"{(res.ramal.equipo or '').upper()} {(res.ramal.uso or '').upper()}".strip(),
                (res.ramal.medida or '').upper(),
                round(res.total.l1, 2),
                round(res.total.l2, 2),
                round(res.total.l3, 2),
            ]
            # sumas por grupo
            for _, prefijos in self._grouped_cols:
                sums = self._sum_por_prefijo(res.cargas, prefijos)
                row_vals.extend([round(sums[0], 2), round(sums[1], 2), round(sums[2], 2)])
            for c, v in enumerate(row_vals, start=1):
                cell = ws.cell(row=bottom_row, column=c, value=v)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if c in (4, 5, 6):  # columnas de totales
                    cell.fill = total_fill
                    cell.font = total_font
                if c == 1:  # ramal
                    cell.font = Font(bold=True)

        # Ajuste de anchos
        for idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 0
            col_letter = get_column_letter(idx)
            for cell in col_cells:
                if getattr(cell, "value", None) is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

        # Tabla 4 y notas en la misma hoja (si aplica)
        if tabla4:
            row_idx = bottom_row + 2
            row_idx += 2
            headers4 = ["RAMAL", "CALIBRE CALCULADO", "CI CALCULADO ($)", "CJ CALCULADO ($)", "CT CALCULADO ($)",
                        "CALIBRE SUPERIOR", "CI SUPERIOR ($)", "CJ SUPERIOR ($)", "CT SUPERIOR ($)", "RECOMENDADO", "AHORRO ($)"]
            for c, h in enumerate(headers4, start=1):
                cell = ws.cell(row=row_idx, column=c, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            currency_fmt = '"$"#,##0.00'
            for r_idx, row in enumerate(tabla4, start=row_idx+1):
                vals = [
                    row.get("Ramal", ""),
                    row.get("Calibre_calc", ""),
                    row.get("CI_calc", ""),
                    row.get("CJ_calc", ""),
                    row.get("CT_calc", ""),
                    row.get("Calibre_sup", ""),
                    row.get("CI_sup", ""),
                    row.get("CJ_sup", ""),
                    row.get("CT_sup", ""),
                    row.get("Recomendado", ""),
                    row.get("Ahorro", ""),
                ]
                for c, v in enumerate(vals, start=1):
                    cell = ws.cell(row=r_idx, column=c, value=v if not isinstance(v, float) else round(v, 2))
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if c in (3,4,5,7,8,9,11):
                        cell.number_format = currency_fmt
                rec = row.get("Recomendado", "")
                if rec == "CALCULADO":
                    ws.cell(row=r_idx, column=5).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws.cell(row=r_idx, column=9).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif rec == "SUPERIOR":
                    ws.cell(row=r_idx, column=9).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws.cell(row=r_idx, column=5).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            row_idx = r_idx + 2
            notes = [
                "Se considera un factor de carga de las pérdidas de 1 y un tiempo de operación de 8760 horas al año.",
                "Costo energía: tarifas reguladas CREG (Nov 2022), conservador mayor valor.",
                "Costo conductores por unidad: lista de precios fabricantes 2021.",
                "Metodología económica basada en IEC 60287-3-2.",
                "Criterio técnico RETIE salvo indicación en contra.",
                "Convenciones: verde = conductor óptimo, rojo = conductor descartado.",
            ]
            ws.cell(row=row_idx, column=1, value="NOTAS").font = Font(bold=True)
            for i, note in enumerate(notes, start=1):
                ws.cell(row=row_idx + i, column=1, value=f"{i}. {note}")

        # Resumen metros por calibre (una sola hoja)
        if metros_t or metros_e:
            row_idx = ws.max_row + 2
            ws.cell(row=row_idx, column=1, value="RESUMEN METROS POR CALIBRE").font = Font(bold=True)
            headers_res = ["CALIBRE", "CAIDA DE TENSION (m)", "CALCULO ECONOMICO (m)"]
            for c, h in enumerate(headers_res, start=1):
                cell = ws.cell(row=row_idx+1, column=c, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            # orden preferido
            orden_pref = ["12", "10", "8", "6", "4", "2", "1", "1/0", "2/0", "3/0", "4/0"]
            claves_set = set((metros_t or {}).keys()) | set((metros_e or {}).keys())
            claves = [c for c in orden_pref if c in claves_set] + sorted(claves_set - set(orden_pref), key=str)
            for r_off, cal in enumerate(claves, start=2):
                ws.cell(row=row_idx + r_off, column=1, value=cal)
                ws.cell(row=row_idx + r_off, column=2, value=round((metros_t or {}).get(cal, 0.0), 2))
                ws.cell(row=row_idx + r_off, column=3, value=round((metros_e or {}).get(cal, 0.0), 2))
                for c in range(1,4):
                    ws.cell(row=row_idx + r_off, column=c).border = border
                    ws.cell(row=row_idx + r_off, column=c).alignment = Alignment(horizontal="center", vertical="center")

        wb.save(path)

    def _schedule_preview(self) -> None:
        self._preview_timer.start(150)

    def _update_preview_now(self) -> None:
        book = self._default_book()
        if not book.exists():
            self.status.setText("Coloca el Excel base en data/ para previsualizar.")
            return
        try:
            selection = self._collect_selection()
            resultados = recalcular(book, write_sheet=False, selection=selection)
            self._fill_table(resultados)
            self.status.setText("Previsualizacion actualizada.")
        except Exception as e:
            self.status.setText(f"Error en previsualizacion: {e}")

    # ------------------------------------------------------------------ helpers y llenado
    def _sum_por_prefijo(self, cargas, prefijos: Tuple[str, ...]) -> tuple[float, float, float]:
        l1 = l2 = l3 = 0.0
        for c in cargas:
            lbl = c.etiqueta.upper()
            if any(lbl.startswith(p) for p in prefijos):
                l1 += c.l1
                l2 += c.l2
                l3 += c.l3
        return (l1, l2, l3)

    def _fill_table(self, resultados: List[ResultadoRamal]) -> None:
        self._last_resultados = resultados
        calibres_por_ramal: List[Tuple[str, float]] = []
        self._tabla4 = []
        # reset visibles segun datos
        has_data = bool(resultados)
        self.table.setVisible(has_data)
        self.summary_card.setVisible(has_data)
        self.table4.setVisible(False)
        self.table_cables.setVisible(False)
        try:
            try:
                df_caida = self._read_caida_df()
            except Exception:
                df_caida = self._default_caida_df()
            fp_val = self.fp_spin.value() if hasattr(self, "fp_spin") else 0.95
            temp_c = self.temp_spin.value() if hasattr(self, "temp_spin") else 20.0
            df_caida = self._compute_zef(df_caida, fp_val, temp_c)
            if df_caida.empty:
                raise ValueError("Hoja de caida vacia")
            try:
                pass
            except Exception:
                pass

            for idx, res in enumerate(resultados):
                imax = max(res.total.l1, res.total.l2, res.total.l3)
                dist_m = self._get_distance_for_row(idx)
                trif = not (math.isclose(res.total.l1, 0.0, abs_tol=1e-3) or math.isclose(res.total.l2, 0.0, abs_tol=1e-3) or math.isclose(res.total.l3, 0.0, abs_tol=1e-3))
                cal, vd = self._select_calibre(df_caida, imax, dist_m, trif)
                calibres_por_ramal.append((cal, vd))

            # Tabla 4 (análisis económico RETIE)
            if self.chk_eco.isChecked():
                self._tabla4 = self._build_tabla4(resultados, calibres_por_ramal, df_caida)
                self._render_tabla4()
            else:
                self.table4.setVisible(False)
                # aunque no haya tabla4, mostramos resumen de cables usando tension
                self._metros_tension = {}
                self._metros_economico = {}
                for idx, res in enumerate(resultados):
                    cal = calibres_por_ramal[idx][0] if idx < len(calibres_por_ramal) else ""
                    cal = self._clean_calibre(cal)
                    dist_m = self._get_distance_for_row(idx)
                    n_cond = self.numero_conductores(res.total.l1, res.total.l2, res.total.l3)
                    self._metros_tension[cal] = self._metros_tension.get(cal, 0.0) + dist_m * n_cond
                    self._metros_economico[cal] = self._metros_economico.get(cal, 0.0) + dist_m * n_cond
                self._render_resumen_cables()
        except Exception as e:
            if hasattr(self, "status"):
                self.status.setText(f"No se pudo calcular caida de tension (revisa hoja CAIDA TENSION): {e}")
            try:
                print("Error caida tension:", e)
            except Exception:
                pass
            calibres_por_ramal = [("", math.nan)] * len(resultados)
            self.table4.setVisible(False)

        self.table.setRowCount(0)
        for res in resultados:
            total = res.total
            sums: List[Tuple[float, float, float]] = []
            for _, prefijos in self._grouped_cols:
                sums.append(self._sum_por_prefijo(res.cargas, prefijos))

            row = self.table.rowCount()
            self.table.insertRow(row)
            equipo_nom = (res.ramal.equipo or "").upper()
            uso_nom = (res.ramal.uso or "").upper()
            equipo_display = f"{equipo_nom} {uso_nom}".strip()
            medida_display = (res.ramal.medida or "").upper()
            values = [
                str(res.ramal.ramal),
                equipo_display,
                medida_display,
                round(total.l1, 2),
                round(total.l2, 2),
                round(total.l3, 2),
            ]
            for l1, l2, l3 in sums:
                values.extend([round(l1, 2), round(l2, 2), round(l3, 2)])
            # agregar calibre/VD si calculado
            if len(calibres_por_ramal) > row:
                cal, vd = calibres_por_ramal[row]
            else:
                cal, vd = ("", math.nan)
            values.extend([cal, round(vd, 2) if not math.isnan(vd) else ""])

            for col, val in enumerate(values):
                item = QTableWidgetItem(str(val))
                self.table.setItem(row, col, item)
        self._recompute_summary()
        self._auto_height(self.table, min_height=320)

    def _calc_totals(self, resultados: List[ResultadoRamal], factor_res: float = 1.0):
        """Devuelve (totales balanceados sin factor, totales balanceados con factor aplicado a RESISTENCIAS)."""
        cargas_raw: List[CargaFase] = []
        cargas_fact: List[CargaFase] = []
        for res in resultados:
            for c in res.cargas:
                lbl = c.etiqueta.upper()
                cargas_raw.append(c)
                if lbl.startswith("RESISTENCIAS"):
                    cargas_fact.append(CargaFase(c.etiqueta, c.l1 * factor_res, c.l2 * factor_res, c.l3 * factor_res))
                else:
                    cargas_fact.append(c)

        def balance(cargas: List[CargaFase]) -> tuple[float, float, float]:
            tot = [0.0, 0.0, 0.0]
            for c in cargas:
                phases_present = [(c.l1, 0), (c.l2, 1), (c.l3, 2)]
                nonzero = [(val, idx) for val, idx in phases_present if val > 0]
                if len(nonzero) >= 3:
                    tot[0] += c.l1
                    tot[1] += c.l2
                    tot[2] += c.l3
                elif len(nonzero) == 2:
                    val = max(val for val, _ in nonzero)
                    pairs = [((0, 1), tot[0] + tot[1]), ((1, 2), tot[1] + tot[2]), ((0, 2), tot[0] + tot[2])]
                    pair = min(pairs, key=lambda x: x[1])[0]
                    tot[pair[0]] += val
                    tot[pair[1]] += val
                elif len(nonzero) == 1:
                    val = nonzero[0][0]
                    idx_min = min(range(3), key=lambda i: tot[i])
                    tot[idx_min] += val
            return tuple(tot)

        return balance(cargas_raw), balance(cargas_fact)

    # ------------------------------------------------------------------ caida de tension / calibre
    def _get_distance_for_row(self, row: int) -> float:
        w = self.sel_table.cellWidget(row, 7)
        if isinstance(w, QLineEdit):
            try:
                return float(str(w.text()).replace(",", "."))
            except Exception:
                return 0.0
        return 0.0

    def _create_row_widgets(self, row: int) -> None:
        """Crea widgets de la tabla 1 en la fila dada (asume que la fila ya existe)."""
        eq_combo = QComboBox()
        eq_combo.addItems([""] + self._equipos)
        uso_combo = QComboBox()
        uso_combo.addItems([""] + self._usos)
        med_combo = QComboBox()
        med_combo.addItem("")
        med_combo.setMinimumContentsLength(14)
        med_combo.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        desc_combo = QComboBox()
        desc_combo.addItems(["", "RESISTENCIAS", "GAS CALIENTE", "TIEMPO"])
        tiene_combo = QComboBox()
        tiene_combo.addItems(["", "SI", "NO"])
        uni_combo = QComboBox()
        uni_combo.addItems([""] + self._unidades)
        niveles_spin = QSpinBox()
        niveles_spin.setRange(0, 10)
        dist_edit = QLineEdit()
        dist_edit.setPlaceholderText("m")
        dist_edit.textChanged.connect(self._schedule_preview)

        for w in (eq_combo, uso_combo, med_combo, desc_combo, tiene_combo, uni_combo, niveles_spin, dist_edit):
            self._disable_wheel(w)

        eq_combo.currentTextChanged.connect(lambda txt, row=row, med_combo=med_combo: self._on_equipo_changed(row, txt, med_combo))

        def on_tiene_changed(text, uni_box=uni_combo):
            enabled = text.strip().upper() == "SI"
            uni_box.setEnabled(enabled)
            if not enabled:
                uni_box.setCurrentIndex(0)

        tiene_combo.currentTextChanged.connect(on_tiene_changed)
        uni_combo.setEnabled(False)
        for combo in [eq_combo, uso_combo, med_combo, desc_combo, tiene_combo, uni_combo]:
            combo.currentIndexChanged.connect(self._schedule_preview)
        niveles_spin.valueChanged.connect(self._schedule_preview)

        self.sel_table.setCellWidget(row, 0, eq_combo)
        self.sel_table.setCellWidget(row, 1, uso_combo)
        self.sel_table.setCellWidget(row, 2, med_combo)
        self.sel_table.setCellWidget(row, 3, desc_combo)
        self.sel_table.setCellWidget(row, 4, tiene_combo)
        self.sel_table.setCellWidget(row, 5, uni_combo)
        self.sel_table.setCellWidget(row, 6, niveles_spin)
        self.sel_table.setCellWidget(row, 7, dist_edit)

    def _show_sel_context_menu(self, pos) -> None:
        menu = QMenu(self.sel_table)
        act_insert = menu.addAction("Insertar fila")
        act_delete = menu.addAction("Eliminar fila")
        global_pos = self.sel_table.viewport().mapToGlobal(pos)
        action = menu.exec(global_pos)
        row = self.sel_table.indexAt(pos).row()
        if action == act_insert:
            target = row if row >= 0 else self.sel_table.rowCount()
            self._insert_row_at(target)
        elif action == act_delete and row >= 0:
            self._delete_row(row)

    def _insert_row_at(self, row: int) -> None:
        if row < 0:
            row = self.sel_table.rowCount()
        self.sel_table.insertRow(row)
        self._create_row_widgets(row)
        self.ramales_spin.setValue(self.sel_table.rowCount())
        self._apply_column_widths()
        self._auto_height(self.sel_table, min_height=300)
        self._schedule_preview()

    def _delete_row(self, row: int) -> None:
        if 0 <= row < self.sel_table.rowCount():
            self.sel_table.removeRow(row)
            self.ramales_spin.setValue(max(0, self.sel_table.rowCount()))
            self._schedule_preview()

    def _get_section_for_row(self, row: int) -> float:
        # Sección ahora se toma desde la hoja CAIDA TENSION; no se captura en GUI.
        return 0.0

    def _compute_zef(self, df_caida: pd.DataFrame, fp: float, temp_c: float) -> pd.DataFrame:
        sin_phi = math.sin(math.acos(max(min(fp, 1.0), 0.0)))
        rho20 = 1.72e-8  # ohm*m
        alpha = 0.00393  # 1/°C
        df = df_caida.copy()
        df.columns = [str(c).strip().upper() for c in df.columns]

        def pick(colname: str) -> str:
            for c in df.columns:
                if colname in c:
                    return c
            return colname

        col_cal = pick("CALIBRE")
        col_r = pick("RESISTENCIA (RC")
        if col_r not in df.columns:
            col_r = pick("RESISTENCIA (RC)")
        if col_r not in df.columns:
            col_r = pick("RESISTENCIA")
        col_x = pick("REACTANCIA (XC")
        if col_x not in df.columns:
            col_x = pick("REACTANCIA (XC)")
        if col_x not in df.columns:
            col_x = pick("REACTANCIA")
        col_amp = pick("AMPACIDAD A 75")
        if col_amp not in df.columns:
            col_amp = pick("AMPACIDAD")
        col_sec = pick("SECCION (MM")
        if col_sec not in df.columns:
            col_sec = pick("SECCION")
        col_cost = pick("COSTO INICIAL")

        missing = [c for c in (col_cal, col_r, col_x) if c not in df.columns]
        if missing:
            raise ValueError(f"No se encontraron columnas CALIBRE/R/X en CAIDA TENSION (faltan: {missing})")

        keep_cols = {col_cal: "CALIBRE", col_r: "R", col_x: "X"}
        if col_amp in df.columns:
            keep_cols[col_amp] = "AMPACIDAD"
        if col_sec in df.columns:
            keep_cols[col_sec] = "SECCION"
        if col_cost in df.columns:
            keep_cols[col_cost] = "COSTO_INICIAL"

        df = df[list(keep_cols.keys())].rename(columns=keep_cols)
        df["R"] = pd.to_numeric(df.get("R"), errors="coerce") / 1000.0  # ohm/km -> ohm/m
        df["X"] = pd.to_numeric(df.get("X"), errors="coerce") / 1000.0
        if "AMPACIDAD" in df.columns:
            df["AMPACIDAD"] = pd.to_numeric(df.get("AMPACIDAD"), errors="coerce")
        if "SECCION" in df.columns:
            df["SECCION"] = pd.to_numeric(df.get("SECCION"), errors="coerce")
        if "COSTO_INICIAL" in df.columns:
            def _parse_cost(val):
                if pd.isna(val):
                    return math.nan
                if isinstance(val, (int, float)):
                    return float(val)
                s = str(val).strip()
                s = s.replace(" ", "")
                # remove currency symbols
                s = "".join(ch for ch in s if ch.isdigit() or ch in ".,-")
                # caso miles + coma decimal
                if "." in s and "," in s:
                    s = s.replace(".", "").replace(",", ".")
                elif "," in s and "." not in s:
                    s = s.replace(",", ".")
                # else assume '.' decimal
                try:
                    return float(s)
                except Exception:
                    return math.nan
            df["COSTO_INICIAL"] = df.get("COSTO_INICIAL").apply(_parse_cost)

        # Si hay sección, recalcular R con la temperatura indicada
        if "SECCION" in df.columns:
            r_temp_list = []
            for _, row in df.iterrows():
                sec = row.get("SECCION")
                if pd.isna(sec) or sec <= 0:
                    r_temp_list.append(row.get("R"))
                    continue
                s_m2 = sec * 1e-6
                r20_ohm_km = rho20 * (1000.0 / s_m2)
                r_t_ohm_km = r20_ohm_km * (1 + alpha * (temp_c - 20.0))
                r_temp_list.append(r_t_ohm_km / 1000.0)  # ohm/m
            df["R"] = r_temp_list
            df["R_OHM_KM_T"] = [r * 1000.0 if pd.notna(r) else math.nan for r in df["R"]]
        else:
            df["R_OHM_KM_T"] = df["R"] * 1000.0

        df["ZEF"] = df.apply(
            lambda r: float(r["R"]) * fp + float(r["X"]) * sin_phi
            if not (pd.isna(r["R"]) or pd.isna(r["X"]))
            else math.nan,
            axis=1,
        )
        df = df.dropna(subset=["ZEF"])
        if df.empty:
            raise ValueError("Hoja CAIDA TENSION sin datos numéricos en R/X")
        return df

    def _read_caida_df(self) -> pd.DataFrame:
        book = self._default_book()
        sheet_names = ["CAIDA TENSION", "CAIDA DE TENSION", "CAIDA DE TENSIÓN", "CAIDA TENSIÓN"]
        last_exc = None
        for name in sheet_names:
            try:
                df = pd.read_excel(book, sheet_name=name)
                if df is not None and not df.empty:
                    return df
            except Exception as e:
                last_exc = e
                continue
        raise ValueError(f"No se pudo leer datos de caida de tension (probado {sheet_names}): {last_exc}")

    def _default_caida_df(self) -> pd.DataFrame:
        data = [
            ("12", 6.600/1000, 0.223/1000),
            ("10", 3.900/1000, 0.207/1000),
            ("8", 2.560/1000, 0.213/1000),
            ("6", 1.610/1000, 0.210/1000),
            ("4", 1.020/1000, 0.197/1000),
            ("3", 0.820/1000, 0.194/1000),
            ("2", 0.660/1000, 0.187/1000),
            ("1", 0.520/1000, 0.187/1000),
            ("1/0", 0.390/1000, 0.180/1000),
            ("2/0", 0.330/1000, 0.177/1000),
            ("3/0", 0.259/1000, 0.171/1000),
            ("4/0", 0.207/1000, 0.167/1000),
            ("250", 0.177/1000, 0.1717/1000),
            ("300", 0.148/1000, 0.167/1000),
            ("350", 0.128/1000, 0.164/1000),
            ("400", 0.115/1000, 0.161/1000),
            ("500", 0.095/1000, 0.157/1000),
            ("600", 0.082/1000, 0.157/1000),
            ("750", 0.069/1000, 0.157/1000),
            ("1000", 0.059/1000, 0.151/1000),
        ]
        return pd.DataFrame(data, columns=["CALIBRE", "RESISTENCIA (RC)", "REACTANCIA (XC)"])

    def _select_calibre(self, df_caida: pd.DataFrame, imax: float, dist_m: float, trifasico: bool, vd_limit: float = 3.0) -> tuple[str, float]:
        best_cal = None
        best_vd = None
        last_ok_amp = None
        required_amp = 1.25 * imax  # criterio termico
        # debug desactivado

        for _, row in df_caida.iterrows():
            cal = str(row["CALIBRE"])
            zef = float(row["ZEF"])
            ampacidad = float(row["AMPACIDAD"]) if "AMPACIDAD" in row and not pd.isna(row["AMPACIDAD"]) else None
            if ampacidad is not None and required_amp > ampacidad:
                print(f"  calibre={cal} salta por ampacidad {ampacidad}A < {required_amp}A")
                continue

            factor = math.sqrt(3) if trifasico else 2.0
            vd = (factor * imax * dist_m * zef / 208.0) * 100.0
            r_km = row["R_OHM_KM_T"] if "R_OHM_KM_T" in row else row["R"] * 1000.0
            # debug desactivado
            last_ok_amp = (cal, vd)
            if vd <= vd_limit:
                best_cal = cal
                best_vd = vd
                break

        if best_cal is None:
            # sin cumplir %VD, elegir el último que sí cumple ampacidad, si no el último del listado
            if last_ok_amp is not None:
                best_cal, best_vd = last_ok_amp
                best_cal += " (NO CUMPLE 3%)"
            else:
                last = df_caida.iloc[-1]
                best_cal = str(last["CALIBRE"]) + " (NO CUMPLE 3%)"
                factor = math.sqrt(3) if trifasico else 2.0
                best_vd = (factor * imax * float(last["ZEF"]) * dist_m / 208.0) * 100.0
        return best_cal, best_vd

    @staticmethod
    def _clean_calibre(cal: str) -> str:
        return cal.split()[0] if cal else ""

    def _build_tabla4(self, resultados: List[ResultadoRamal], calibres_por_ramal: List[Tuple[str, float]], df_caida: pd.DataFrame):
        tabla = []
        if df_caida is None or df_caida.empty:
            return tabla
        self._metros_tension = {}
        self._metros_economico = {}
        # Map calibre -> datos
        df_caida = df_caida.reset_index(drop=True)
        order = [str(c) for c in df_caida["CALIBRE"]]
        data_map = {str(row["CALIBRE"]): row for _, row in df_caida.iterrows()}

        # parametros economicos
        costo_kwh = self.eco_costo.value()
        horas_anuales = self.eco_horas.value()
        a = self.eco_a.value()
        b = self.eco_b.value()
        i = self.eco_i.value()
        N = int(self.eco_n.value())
        Q = self.calcular_q(a, b, i, N)
        F = self.calcular_f(1, 1, horas_anuales, costo_kwh, i, Q)

        for idx, res in enumerate(resultados):
            dist_m = self._get_distance_for_row(idx)
            L_km = dist_m / 1000.0
            imax_r = max(res.total.l1, res.total.l2, res.total.l3)
            cal_raw = calibres_por_ramal[idx][0] if idx < len(calibres_por_ramal) else ""
            cal_calc = self._clean_calibre(cal_raw)
            # calibre superior
            try:
                pos = order.index(cal_calc)
            except ValueError:
                pos = -1
            cal_up = order[pos+1] if pos >= 0 and pos+1 < len(order) else ""

            def datos_cal(cal: str):
                if not cal or cal not in data_map:
                    return None
                row = data_map[cal]
                rl_km = float(row["R_OHM_KM_T"]) if "R_OHM_KM_T" in row else float(row["R"]) * 1000.0
                costo_m = float(row["COSTO_INICIAL"]) if "COSTO_INICIAL" in row and not pd.isna(row["COSTO_INICIAL"]) else None
                return rl_km, costo_m

            d1 = datos_cal(cal_calc)
            d2 = datos_cal(cal_up) if cal_up else None
            if d1 is None:
                tabla.append({"Ramal": idx+1})
                continue
            rl1, costo_m1 = d1
            CI1 = (costo_m1 or 0) * dist_m
            CJ1 = (imax_r ** 2) * rl1 * L_km * F
            CT1 = CI1 + CJ1

            CI2 = CJ2 = CT2 = None
            if d2 is not None:
                rl2, costo_m2 = d2
                CI2 = (costo_m2 or 0) * dist_m
                CJ2 = (imax_r ** 2) * rl2 * L_km * F
                CT2 = CI2 + CJ2

            recomendado = "CALCULADO"
            ahorro = None
            if CT2 is not None:
                if CT2 <= CT1:
                    recomendado = "SUPERIOR"
                    ahorro = CT1 - CT2
                else:
                    recomendado = "CALCULADO"
                    ahorro = CT2 - CT1

            # metros por escenario
            n_cond = self.numero_conductores(res.total.l1, res.total.l2, res.total.l3)
            self._metros_tension[cal_calc] = self._metros_tension.get(cal_calc, 0.0) + dist_m * n_cond
            cal_econ = cal_up if recomendado == "SUPERIOR" and cal_up else cal_calc
            self._metros_economico[cal_econ] = self._metros_economico.get(cal_econ, 0.0) + dist_m * n_cond

            tabla.append({
                "Ramal": idx + 1,
                "Calibre_calc": cal_calc,
                "CI_calc": CI1,
                "CJ_calc": CJ1,
                "CT_calc": CT1,
                "Calibre_sup": cal_up,
                "CI_sup": CI2,
                "CJ_sup": CJ2,
                "CT_sup": CT2,
                "Recomendado": recomendado,
                "Ahorro": ahorro,
            })
            # Debug en consola
            # debug desactivado
        return tabla

    def _render_tabla4(self):
        tabla = self._tabla4
        if not tabla:
            self.table4.setVisible(False)
            self.table4.setRowCount(0)
            return
        self.table4.setVisible(True)
        self.table4.setRowCount(0)
        for row_data in tabla:
            r = self.table4.rowCount()
            self.table4.insertRow(r)
            vals = [
                row_data.get("Ramal", ""),
                row_data.get("Calibre_calc", ""),
                row_data.get("CI_calc", ""),
                row_data.get("CJ_calc", ""),
                row_data.get("CT_calc", ""),
                row_data.get("Calibre_sup", ""),
                row_data.get("CI_sup", ""),
                row_data.get("CJ_sup", ""),
                row_data.get("CT_sup", ""),
                row_data.get("Recomendado", ""),
                row_data.get("Ahorro", ""),
            ]
            for c, v in enumerate(vals):
                if isinstance(v, float):
                    disp = f"$ {v:,.2f}" if c in (2,3,4,6,7,8,10) else f"{v:.2f}"
                else:
                    disp = "" if v is None else str(v)
                item = QTableWidgetItem(disp)
                item.setTextAlignment(Qt.AlignCenter)
                self.table4.setItem(r, c, item)
            # Colorear
            rec = row_data.get("Recomendado", "")
            green = QColor(198, 239, 206)
            red = QColor(255, 199, 206)
            if rec == "CALCULADO":
                if self.table4.item(r, 4):
                    self.table4.item(r, 4).setBackground(green)
                if self.table4.item(r, 8):
                    self.table4.item(r, 8).setBackground(red)
            elif rec == "SUPERIOR":
                if self.table4.item(r, 8):
                    self.table4.item(r, 8).setBackground(green)
                if self.table4.item(r, 4):
                    self.table4.item(r, 4).setBackground(red)
        # Ajustar alto para mostrar todo sin scroll interno
        self._auto_height(self.table4, min_height=0)
        # Después de Tabla 4, mostrar resumen de cables
        self._render_resumen_cables()

    def _render_resumen_cables(self):
        metros_t = self._metros_tension or {}
        metros_e = self._metros_economico or {}
        if not metros_t and not metros_e:
            self.table_cables.setVisible(False)
            self.table_cables.setRowCount(0)
            return
        # fusionar claves con orden preferido
        orden_pref = ["12", "10", "8", "6", "4", "2", "1", "1/0", "2/0", "3/0", "4/0"]
        claves_set = set(list(metros_t.keys()) + list(metros_e.keys()))
        claves = [c for c in orden_pref if c in claves_set]
        restantes = sorted(claves_set - set(claves), key=lambda x: str(x))
        claves.extend(restantes)
        self.table_cables.setVisible(True)
        self.table_cables.setRowCount(0)
        for cal in claves:
            r = self.table_cables.rowCount()
            self.table_cables.insertRow(r)
            vals = [
                cal,
                metros_t.get(cal, 0.0),
                metros_e.get(cal, 0.0),
            ]
            for c, v in enumerate(vals):
                if isinstance(v, float):
                    disp = f"{v:,.2f}"
                else:
                    disp = str(v)
                item = QTableWidgetItem(disp)
                item.setTextAlignment(Qt.AlignCenter)
                self.table_cables.setItem(r, c, item)
        # Mostrar/ocultar columna económica según check
        show_eco = self.chk_eco.isChecked()
        self.table_cables.setColumnHidden(2, not show_eco)
        self._auto_height(self.table_cables, min_height=0)

    def _recompute_summary(self) -> None:
        tot_all, tot_adj = self._calc_totals(self._last_resultados or [], self.factor_spin.value() if hasattr(self, "factor_spin") else 1.0)
        def set_row(row: int, vals):
            for c, v in enumerate(vals):
                item = QTableWidgetItem(f"{round(v,2)}")
                item.setTextAlignment(Qt.AlignCenter)
                self.summary_table.setItem(row, c, item)
        set_row(0, tot_all)
        set_row(1, tot_adj)
        self._auto_height_table(self.summary_table)

    def _on_factor_commit(self) -> None:
        try:
            text = self.factor_spin.lineEdit().text()
            cleaned = text.replace(",", ".")
            val = float(cleaned)
            val = max(self.factor_spin.minimum(), min(self.factor_spin.maximum(), val))
            # actualizar sin disparar valueChanged infinitamente
            self.factor_spin.blockSignals(True)
            self.factor_spin.setValue(val)
            self.factor_spin.blockSignals(False)
            self._recompute_summary()
        except Exception:
            pass

    def _open_library(self) -> None:
        """
        Abre biblioteca local (data/proyectos/carga_electrica/) para cargar/duplicar/borrar proyectos .json.
        """
        lib_dir = Path(__file__).resolve().parents[3] / "data" / "proyectos" / "carga_electrica"
        dlg = BibliotecaCargaDialog(
            parent=self,
            library_dir=lib_dir,
            on_load=lambda p: self._load_project_file(str(p)),
            primary_qss=self.btn_export.styleSheet(),
            danger_qss=self.btn_reset.styleSheet(),
        )
        dlg.exec()
