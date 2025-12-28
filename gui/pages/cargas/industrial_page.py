from __future__ import annotations







import json







import math







import traceback







from dataclasses import asdict







from copy import deepcopy







from pathlib import Path







from typing import List







from datetime import datetime







from PySide6.QtCore import Qt, QTimer







from PySide6.QtWidgets import (







    QWidget,







    QVBoxLayout,







    QHBoxLayout,







    QLabel,







    QPushButton,







    QCheckBox,







    QSpinBox,







    QDoubleSpinBox,







    QLineEdit,







    QTableWidget,







    QTableWidgetItem,







    QMessageBox,







    QFileDialog,







    QComboBox,







    QFrame,







    QPlainTextEdit,







    QDialog,







    QTextEdit,







    QFormLayout,







    QScrollArea,







    QGroupBox,







    QTabWidget,







    QListWidget,







    QListWidgetItem,







    QStackedWidget,







    QSizePolicy,







    QHeaderView,







    QAbstractItemView,







    QTreeWidget,







    QTreeWidgetItem,







)







from logic.thermal_load import ThermalLoadCalculator, RoomInputs







from logic.thermal_load.validation import load_rules, validate_room







from .biblioteca_cuartos_ind import BibliotecaCuartosIndDialog, PROJECTS_DIR







class IndustrialPage(QWidget):







    def __init__(self, parent=None):







        super().__init__(parent)







        self.calc = ThermalLoadCalculator()







        self._loading_room = False







        self._updating_widgets = False







        self.rules = load_rules()







        self._last_results = []







        # Cargar perfiles desde el JSON (pueden ser varios perfiles con defaults)







        profiles_path = Path("data/cuartos_industriales/thermal_load_default_profiles.json")







        raw_profiles = json.loads(profiles_path.read_text(encoding="utf-8"))







        if isinstance(raw_profiles, dict) and "profiles" in raw_profiles:







            profiles_seq = raw_profiles.get("profiles", [])







        else:







            profiles_seq = raw_profiles







        self.profiles_by_id = {p["id"]: p for p in profiles_seq}







        self.profile_defs = self.profiles_by_id







        # Mostrar siempre en MAYSCULA y ordenar por temperatura interna si existe







        exclude_ids = {"picking_high_traffic_refrig", "frozen_high_traffic"}







        tmp_list = []







        for pid, data in self.profile_defs.items():







            if pid in exclude_ids:







                continue







            label = str(data.get("label", pid)).upper()







            ti = data.get("defaults", {}).get("internal_temp_C")







            tmp_list.append((pid, label, ti if ti is not None else 9999))







        # ordenar de menor a mayor temperatura interna







        self.lbl_state = QLabel("Listo")



        self.lbl_state.setVisible(False)



        self.log = None



        tmp_list.sort(key=lambda x: x[2])







        self.profile_list = [(pid, label) for pid, label, _ in tmp_list]







        self._build_ui()







        # Timer para clculo diferido







        self._calc_timer = QTimer(self)







        self._calc_timer.setSingleShot(True)







        self._calc_timer.timeout.connect(self._on_calc_timer)







    def _build_ui(self):







        outer = QVBoxLayout(self)







        outer.setContentsMargins(0, 0, 0, 0)







        main_scroll = QScrollArea()







        main_scroll.setWidgetResizable(True)







        main_scroll.setFrameShape(QFrame.NoFrame)







        main_widget = QWidget()







        root = QVBoxLayout(main_widget)







        root.setAlignment(Qt.AlignTop)







        title = QLabel('CALCULO DE CARGAS TERMICAS (CUARTOS INDUSTRIALES)')







        title.setStyleSheet("font-size:18px;font-weight:800;")







        root.addWidget(title)







        # Barra de acciones







        action_row = QHBoxLayout()







        self.btn_export = QPushButton("EXPORTAR PROYECTO")







        self.btn_export.clicked.connect(self._export_project)







        self.btn_library = QPushButton("BIBLIOTECA")







        self.btn_library.clicked.connect(self._open_library)







        self.btn_new = QPushButton("NUEVO PROYECTO")







        self.btn_new.setStyleSheet("QPushButton{background:#fdecec;color:#c53838;font-weight:800;border-radius:10px;padding:10px 16px;border:1px solid #f5c2c2;} QPushButton:hover{background:#fbd5d5;}")







        self.btn_new.clicked.connect(self._new_project)







        self.btn_calc_all = QPushButton("CALCULAR CARGA TÉRMICA")







        self.btn_calc_all.setStyleSheet(



            "QPushButton{background:#05a2e8;color:white;font-weight:800;border-radius:10px;padding:10px 18px;}"



            "QPushButton:hover{background:#078cc8;}"



        )







        self.btn_calc_all.clicked.connect(self._manual_calc)







        action_row.addWidget(self.btn_export)



        action_row.addWidget(self.btn_library)



        action_row.addWidget(self.btn_calc_all)



        action_row.addStretch()



        action_row.addWidget(self.btn_new)



        action_row.setAlignment(self.btn_new, Qt.AlignRight)







        root.addLayout(action_row)







        # Card superior







        card = QVBoxLayout()







        top = QWidget()







        top.setLayout(card)



        root.addWidget(top)







        top.setStyleSheet("QWidget{background:white;border:1px solid #e5e7eb;border-radius:12px;padding:8px;}")







        form_row1 = QHBoxLayout()







        self.ed_project = QLineEdit()







        self.ed_project.setPlaceholderText("Proyecto")



        self.ed_project.textChanged.connect(self._force_project_upper)







        form_row1.addWidget(QLabel("PROYECTO:"))







        form_row1.addWidget(self.ed_project, 2)



        self.lbl_warnings = QLabel("")



        self.lbl_warnings.setVisible(False)



        # Solo un cuarto: escondemos control de cantidad







        self.sp_rooms = QSpinBox()







        self.sp_rooms.setRange(1, 1)







        self.sp_rooms.setValue(1)







        self.sp_rooms.setVisible(False)







        self.sp_sf = QDoubleSpinBox()







        self.sp_sf.setRange(1.0, 1.5)







        self.sp_sf.setSingleStep(0.05)







        self.sp_sf.setValue(1.1)







        self.sp_sf.valueChanged.connect(lambda _=None: self._schedule_calc())







        form_row1.addWidget(QLabel("FACTOR DE SEGURIDAD:"))







        form_row1.addWidget(self.sp_sf)







        form_row1.addStretch()







        card.addLayout(form_row1)







        # Tabla de cuartos







        self.tbl = QTableWidget(0, 6)







        self.tbl.setSelectionMode(QAbstractItemView.NoSelection)







        self.tbl.setHorizontalHeaderLabels(







            ["CUARTO", "LARGO (m)", "ANCHO (m)", "ALTURA (m)", "PERFIL/USO", "PUERTAS"]







        )







        header = self.tbl.horizontalHeader()







        header.setSectionResizeMode(QHeaderView.Stretch)







        self.tbl.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)







        self.tbl.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)







        self.tbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)







        self.tbl.verticalHeader().setVisible(False)







        self.tbl.setSelectionBehavior(QTableWidget.SelectRows)







        self.tbl.setSelectionMode(QAbstractItemView.NoSelection)







        # evitar resaltado y ocultar columnas de uso/puertas en la vista







        self.tbl.setStyleSheet(self.tbl.styleSheet() + " QTableWidget::item:selected{background:transparent;color:#0f172a;}")







        self.tbl.setColumnHidden(4, True)







        self.tbl.setColumnHidden(5, True)







        root.addWidget(self.tbl)







        # init storage







        self.rooms_inputs: List[RoomInputs] = []







        self.current_row = 0







        # Seleccin por fila







        self.tbl.itemSelectionChanged.connect(self._select_row_from_selection)







        # Placeholder/editor (dentro de un scroll para evitar compresin al maximizar)







        self.editor_stack = QStackedWidget()







        self.editor_stack.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)







        self.editor_stack.setMinimumHeight(500)







        self.editor_placeholder = QLabel("Selecciona un cuarto y presiona DETALLE para editarlo.")







        self.editor_placeholder.setStyleSheet("padding:16px;color:#475569;")







        self.editor_stack.addWidget(self.editor_placeholder)







        self.editor_stack.addWidget(self._build_editor())







        # mostrar siempre el editor por defecto







        self.editor_stack.setCurrentIndex(1)







        # editor sin scroll interno; se deja que el scroll general de la pgina acte







        root.addWidget(self.editor_stack)







        # RESULTADOS







        self.lbl_total_btuh = QLabel("--")







        self.lbl_total_kw = QLabel("--")







        self.lbl_total_tr = QLabel("--")







        res_card = QWidget()







        res_card.setStyleSheet("QWidget{background:white;border:1px solid #e5e7eb;border-radius:12px;padding:10px;}")







        res_layout = QVBoxLayout(res_card)







        res_layout.addWidget(QLabel("RESULTADOS"))







        kpi_row = QHBoxLayout()







        for lbl, title in [







            (self.lbl_total_btuh, "CARGA TOTAL (BTU/H)"),







            (self.lbl_total_kw, "POTENCIA (kW)"),







            (self.lbl_total_tr, "TR"),







        ]:







            box = QVBoxLayout()







            t = QLabel(title); t.setStyleSheet("font-weight:700;")







            big = lbl; big.setStyleSheet("font-size:20px;font-weight:800;")







            box.addWidget(t); box.addWidget(big); box.addStretch()







            kpi_row.addLayout(box, 1)







        res_layout.addLayout(kpi_row)







        # Tablas ocultas para mantener compatibilidad con la lgica existente







        self.tbl_res = QTableWidget(0, 6)







        self.tbl_res.setVisible(False)







        self.tbl_break = QTableWidget(0, 4)







        self.tbl_break.setVisible(False)







        root.addWidget(res_card)







        self._build_rows(1)







        self.current_row = 0







        # Aadir al scroll principal







        main_scroll.setWidget(main_widget)







        outer.addWidget(main_scroll)







    def _adv_spin(self, label: str, minv: float, maxv: float, default: float, layout: QHBoxLayout):







        box = QVBoxLayout()







        t = QLabel(label); t.setStyleSheet("font-weight:600;")







        sp = QDoubleSpinBox()







        sp.setRange(minv, maxv); sp.setDecimals(2); sp.setValue(default)







        box.addWidget(t); box.addWidget(sp)







        layout.addLayout(box)







        return sp







    def _toggle_adv(self, checked: bool):







        self.adv_frame.setVisible(checked)







        self._set_adv_tabs(checked)







    def _set_adv_tabs(self, show: bool):







        if hasattr(self, "adv_tabs"):







            for idx in self.adv_tabs:







                self.tabs.setTabVisible(idx, show)







            # ensure current visible







            if not show and hasattr(self, "tabs"):







                self.tabs.setCurrentIndex(self.basic_tabs[0])







    def _schedule_calc(self, delay_ms: int = 400):







        """Lanza un clculo diferido para evitar recomputar en cada tick."""







        if hasattr(self, "_calc_timer"):







            self._calc_timer.stop()







            self._calc_timer.start(delay_ms)







    def _on_calc_timer(self):







        """Handler debounced: guarda editor y calcula."""







        try:







            if not getattr(self, "_loading_room", False):







                self._save_editor_to_room()







            self._calculate()







        except Exception as exc:







            self._append_log(f"Calc timer error: {exc}")







    def _ensure_rooms(self, n: int):







        while len(self.rooms_inputs) < n:







            idx = len(self.rooms_inputs)







            self.rooms_inputs.append(RoomInputs(nombre=f"CUARTO {idx+1}"))







        if len(self.rooms_inputs) > n:







            self.rooms_inputs = self.rooms_inputs[:n]







    def _select_row(self, idx: int):







        self.current_row = idx







        self._ensure_rooms(self.sp_rooms.value())







        # marcar seleccin en tabla







        self.tbl.selectRow(idx)







        self._load_room_to_editor(self.rooms_inputs[idx])







        self.editor_stack.setCurrentIndex(1)







        self.update_process_outputs(idx)







        self._refresh_infiltration_outputs(idx)







        if hasattr(self, "cb_detail_room") and self.cb_detail_room.count() > idx:







            self.cb_detail_room.blockSignals(True)







            self.cb_detail_room.setCurrentIndex(idx)







            self.cb_detail_room.blockSignals(False)







            self._populate_detail_view(idx)







    def _load_room_to_editor(self, room: RoomInputs):







        self._loading_room = True







        widgets_to_block = [







            self.cb_perfil_head, self.sp_largo, self.sp_ancho, self.sp_alto,







            self.sp_tint, self.sp_tef, self.sp_teb, self.sp_ted, self.sp_tei, self.sp_tet,







            self.sp_ground, self.sp_transfer, self.sp_run_hours,







            self.sp_taire, self.sp_rh, self.sp_recambios, self.sp_usefactor,







            self.sp_espesor, self.cb_AISLAMIENTO, self.sp_tin, self.sp_tout,







            self.sp_masa, self.sp_ciclo, self.cb_producto,







            self.sp_luz_w, self.sp_luz_h, self.sp_mont_hp, self.sp_mont_h,







            self.sp_motor_w, self.sp_motor_h, self.sp_people, self.sp_people_h,







            self.sp_people_btuh, self.sp_desh_w, self.sp_desh_n, self.sp_desh_min, self.sp_desh_frac,







        ]







        for w in widgets_to_block:







            try:







                w.blockSignals(True)







            except Exception:







                pass







        self.ed_nombre_cuarto.setText(room.nombre)







        self.cb_perfil_head.setCurrentIndex(max(self.cb_perfil_head.findData(room.perfil_id),0))







        self.sp_largo.setValue(room.largo_m)







        self.sp_ancho.setValue(room.ancho_m)







        self.sp_alto.setValue(room.altura_m)







        self.sp_tint.setValue(room.T_internal_C)







        self.sp_tef.setValue(room.T_ext_front_C)







        self.sp_teb.setValue(room.T_ext_back_C)







        self.sp_ted.setValue(room.T_ext_right_C)







        self.sp_tei.setValue(room.T_ext_left_C)







        self.sp_tet.setValue(room.T_ext_roof_C)







        self.sp_ground.setValue(room.ground_temp_C)







        self.sp_transfer.setValue(room.wall_transfer_factor)







        self.sp_run_hours.setValue(room.run_hours_supp)







        self.sp_taire.setValue(room.outside_air_temp_C)







        self.sp_rh.setValue(room.outside_RH*100)







        if hasattr(self, "sp_inside_rh"):







            self.sp_inside_rh.setValue((room.inside_RH if room.inside_RH is not None else 0.85)*100)







        if hasattr(self, "chk_infil_enabled"):







            self.chk_infil_enabled.setChecked(getattr(room, "infiltration_enabled", True))







        self.sp_recambios.setValue(room.air_changes_24h_override or 0)







        self.sp_usefactor.setValue(room.use_factor)







        self.sp_espesor.setValue(room.insulation_thickness_in)







        idxa = self.cb_AISLAMIENTO.findData(room.insulation_type)







        if idxa>=0:







            self.cb_AISLAMIENTO.setCurrentIndex(idxa)







        self.sp_tin.setValue(room.product_Tin_C or 0)







        self.sp_tout.setValue(room.product_Tout_C)







        self.sp_masa.setValue(room.product_mass_kg)







        self.sp_ciclo.setValue(room.product_cycle_h)







        self.cb_producto.setCurrentIndex(max(self.cb_producto.findData(room.product_name),0))







        self.sp_luz_w.setValue(room.lighting_W)







        self.sp_luz_h.setValue(room.lighting_hours)







        self.sp_mont_hp.setValue(room.forklift_hp)







        self.sp_mont_h.setValue(room.forklift_hours)







        self.sp_motor_w.setValue(room.motors_W)







        self.sp_motor_h.setValue(room.motors_hours)







        self.sp_people.setValue(room.people_count)







        self.sp_people_h.setValue(room.people_hours)







        self.sp_people_btuh.setValue(room.people_btuh)







        self.sp_desh_w.setValue(room.defrost_W)







        self.sp_desh_n.setValue(room.defrost_count)







        self.sp_desh_min.setValue(room.defrost_duration_min)







        self.sp_desh_frac.setValue(room.defrost_fraction_to_room)







        self._update_volume_labels()







        for w in widgets_to_block:







            try:







                w.blockSignals(False)







            except Exception:







                pass







        self._loading_room = False







        # Recalcular U24 una sola vez sin guardar ni agendar durante carga







        self.update_insulation_outputs(self.current_row, save_model=False, schedule_recalc=False)







        self._schedule_calc()







    def _on_editor_profile_changed(self):







        """Cuando cambia el perfil en el editor, aplicar defaults y sincronizar."""







        if self.current_row is None:







            return







        pid = self.cb_perfil_head.currentData()







        self.rooms_inputs[self.current_row].perfil_id = pid







        combo: QComboBox = self.tbl.cellWidget(self.current_row, 4)







        if combo is not None:







            combo.blockSignals(True)







            combo.setCurrentIndex(max(combo.findData(pid), 0))







            combo.blockSignals(False)







        self._apply_profile_defaults(self.current_row)







    def _save_editor_to_room(self):







        self._ensure_rooms(self.sp_rooms.value())







        r = self.rooms_inputs[self.current_row]







        r.nombre = self.ed_nombre_cuarto.text().strip() or f"CUARTO {self.current_row+1}"







        r.perfil_id = self.cb_perfil_head.currentData()







        r.largo_m = self.sp_largo.value()







        r.ancho_m = self.sp_ancho.value()







        r.altura_m = self.sp_alto.value()







        r.T_internal_C = self.sp_tint.value()







        r.T_ext_front_C = self.sp_tef.value()







        r.T_ext_back_C = self.sp_teb.value()







        r.T_ext_right_C = self.sp_ted.value()







        r.T_ext_left_C = self.sp_tei.value()







        r.T_ext_roof_C = self.sp_tet.value()







        r.ground_temp_C = self.sp_ground.value()







        r.wall_transfer_factor = self.sp_transfer.value()







        r.outside_air_temp_C = self.sp_taire.value()







        r.outside_RH = self.sp_rh.value()/100







        rec = self.sp_recambios.value()







        r.air_changes_24h_override = None if rec <= 0 else rec







        r.use_factor = self.sp_usefactor.value()







        r.inside_RH = self.sp_inside_rh.value()/100







        r.infiltration_enabled = self.chk_infil_enabled.isChecked()







        r.insulation_type = self.cb_AISLAMIENTO.currentData()







        r.insulation_thickness_in = self.sp_espesor.value()







        prod_data = self.cb_producto.currentData()







        if prod_data is None or prod_data == "":







            prod_data = (self.cb_producto.currentText() or "").strip()







        r.product_name = prod_data







        r.product_Tin_C = self.sp_tin.value()







        r.product_Tout_C = self.sp_tout.value()







        r.product_mass_kg = self.sp_masa.value()







        r.product_cycle_h = self.sp_ciclo.value()







        r.run_hours_supp = self.sp_run_hours.value()







        r.lighting_W = self.sp_luz_w.value()







        r.lighting_hours = self.sp_luz_h.value()







        r.forklift_hp = self.sp_mont_hp.value()







        r.forklift_hours = self.sp_mont_h.value()







        r.motors_W = self.sp_motor_w.value()







        r.motors_hours = self.sp_motor_h.value()







        r.people_count = self.sp_people.value()







        r.people_hours = self.sp_people_h.value()







        r.people_btuh = self.sp_people_btuh.value()







        r.defrost_W = self.sp_desh_w.value()







        r.defrost_count = self.sp_desh_n.value()







        r.defrost_duration_min = self.sp_desh_min.value()







        r.defrost_fraction_to_room = self.sp_desh_frac.value()







        # actualizar tabla resumen (L/A/H/perfil/puertas)







        self.tbl.item(self.current_row,0).setText(r.nombre)







        self.tbl.cellWidget(self.current_row,1).setValue(r.largo_m)







        self.tbl.cellWidget(self.current_row,2).setValue(r.ancho_m)







        self.tbl.cellWidget(self.current_row,3).setValue(r.altura_m)







        combo: QComboBox = self.tbl.cellWidget(self.current_row,4)







        combo.setCurrentIndex(max(combo.findData(r.perfil_id),0))







        puertas_val = self.tbl.cellWidget(self.current_row,5).value()







        r.puertas = puertas_val







        self.tbl.cellWidget(self.current_row,5).setValue(int(puertas_val or 1))







        self._update_volume_labels()







    def _build_rows(self, n: int):







        self._ensure_rooms(n)







        self.tbl.setRowCount(n)







        self._autosize_table(self.tbl)







        for i in range(n):







            room = self.rooms_inputs[i]







            name_item = QTableWidgetItem(room.nombre or f"CUARTO {i+1}")







            self.tbl.setItem(i, 0, name_item)







            # largo/ancho/altura como spin







            for col, val in zip((1, 2, 3), (room.largo_m, room.ancho_m, room.altura_m)):







                sp = QDoubleSpinBox()







                sp.setRange(0.0, 200.0)







                sp.setDecimals(2)







                sp.setSingleStep(0.1)







                sp.setValue(val or 0.0)







                sp.setAlignment(Qt.AlignCenter)







                self.tbl.setCellWidget(i, col, sp)







                def _make_handler(r=i, c=col, w=sp):







                    def _h(_=None):







                        if c == 1:







                            self.rooms_inputs[r].largo_m = w.value()







                        elif c == 2:







                            self.rooms_inputs[r].ancho_m = w.value()







                        elif c == 3:







                            self.rooms_inputs[r].altura_m = w.value()







                        if r == self.current_row:







                            # reflejar en el editor sin disparar signals







                            for spin, val in ((self.sp_largo, self.rooms_inputs[r].largo_m),







                                              (self.sp_ancho, self.rooms_inputs[r].ancho_m),







                                              (self.sp_alto, self.rooms_inputs[r].altura_m)):







                                spin.blockSignals(True)







                                spin.setValue(val)







                                spin.blockSignals(False)







                            self._update_volume_labels()



                            if c in (1, 2, 3):



                                self._apply_process_autofill(r, force=False)







                            self._schedule_calc()







                        else:







                            if c in (1, 2, 3):







                                self._apply_process_autofill(r, force=False)







                            self._schedule_calc()







                    return _h







                sp.valueChanged.connect(_make_handler())







            combo = QComboBox()







            # Placeholder vaco







            combo.addItem("", "")







            for pid, label in self.profile_list:







                combo.addItem(label, pid)







            combo.setCurrentIndex(max(combo.findData(room.perfil_id), 0))







            combo.currentIndexChanged.connect(self._make_table_profile_handler(i, combo))







            self.tbl.setCellWidget(i, 4, combo)







            sp_p = QSpinBox(); sp_p.setRange(0, 20); sp_p.setValue(int(getattr(room, "puertas", 1) or 1)); sp_p.setAlignment(Qt.AlignCenter)







            self.tbl.setCellWidget(i, 5, sp_p)







            # seleccin manejada por itemSelectionChanged







        # mostrar el editor de inmediato







        self.editor_stack.setCurrentIndex(1)







        if n:







            self._select_row(0)







    def _make_table_profile_handler(self, row: int, combo: QComboBox):







        def _h(_=None):







            pid = combo.currentData()







            self.rooms_inputs[row].perfil_id = pid







            if row == self.current_row:







                self.cb_perfil_head.blockSignals(True)







                self.cb_perfil_head.setCurrentIndex(max(self.cb_perfil_head.findData(pid), 0))







                self.cb_perfil_head.blockSignals(False)







            self._apply_profile_defaults(row)







            self._schedule_calc()







        return _h







    def _gather_inputs(self) -> List[RoomInputs]:







        """Recolecta los cuartos de la tabla. Ignora filas con L/A/H cero."""







        n = self.sp_rooms.value()







        if self.tbl.rowCount() != n:







            self._build_rows(n)







        self._ensure_rooms(n)







        rooms: List[RoomInputs] = []







        missing = []







        for i in range(n):







            room = self.rooms_inputs[i]







            if room.largo_m <= 0 or room.ancho_m <= 0 or room.altura_m <= 0:







                missing.append(room.nombre or f"CUARTO {i+1}")







                continue







            rooms.append(deepcopy(room))







        if missing:







            self._append_log(f"Se omitieron cuartos sin dimensiones: {', '.join(missing)}")







        return rooms







    # --------- acciones de editor --------- #







    def _autosize_table(self, table: QTableWidget):







        """Ajusta la altura de la tabla a sus filas para evitar scroll interno."""







        rows = table.rowCount()







        hh = table.horizontalHeader().height()







        rowh = table.verticalHeader().defaultSectionSize()







        height = hh + rows * rowh + 4







        table.setFixedHeight(height)







    def _validate_editor(self):







        """Valida campos bsicos del cuarto activo y muestra faltantes."""







        missing = []







        if self.sp_largo.value() <= 0:







            missing.append("Largo")







        if self.sp_ancho.value() <= 0:







            missing.append("Ancho")







        if self.sp_alto.value() <= 0:







            missing.append("Alto")







        if self.sp_tint.value() == 0:







            missing.append("Temperatura interna")







        if self.sp_taire.value() == 0:







            missing.append("T. externa aire")







        if self.sp_rh.value() == 0:







            missing.append("Humedad relativa")







        if self.sp_luz_w.value() == 0 and self.sp_motor_w.value() == 0:







            missing.append("Cargas internas (iluminacin/motores)")







        if self.sp_masa.value() == 0:







            missing.append("Cantidad producto")







        if missing:







            msg = "Faltan: " + ", ".join(missing)







            self.lbl_warnings.setText(msg)







            self.lbl_warnings.setStyleSheet("color:#b91c1c;font-weight:700;")







            self._append_log(msg)







        else:







            self.lbl_warnings.setText("Datos bsicos completos.")







            self.lbl_warnings.setStyleSheet("color:#16a34a;font-weight:700;")







            self._append_log("Validacin OK para el cuarto actual.")







    def _clear_editor(self):







        """Limpia el cuarto actual a valores base."""







        for sp in (







            self.sp_largo,







            self.sp_ancho,







            self.sp_alto,







            self.sp_tint,







            self.sp_tef,







            self.sp_teb,







            self.sp_ted,







            self.sp_tei,







            self.sp_tet,







            self.sp_taire,







        ):







            sp.setValue(0.0)







        self.sp_rh.setValue(60.0)







        self.sp_recambios.setValue(0.0)







        self.sp_usefactor.setValue(2.0)







        self.sp_espesor.setValue(4.0)







        self.cb_AISLAMIENTO.setCurrentIndex(0)







        self.cb_producto.setCurrentIndex(0)







        self.sp_tin.setValue(0.0)







        self.sp_tout.setValue(0.0)







        self.sp_masa.setValue(0.0)







        self.sp_ciclo.setValue(24.0)







        self.sp_luz_w.setValue(0.0)







        self.sp_luz_h.setValue(18.0)







        self.sp_motor_w.setValue(0.0)







        self.sp_motor_h.setValue(24.0)







        self.sp_mont_hp.setValue(0.0)







        self.sp_mont_h.setValue(2.0)







        self.sp_people.setValue(0)







        self.sp_people_h.setValue(8.0)







        self.sp_people_btuh.setValue(1350.0)







        self.sp_desh_w.setValue(0.0)







        self.sp_desh_n.setValue(0)







        self.sp_desh_min.setValue(0.0)







        self.sp_desh_frac.setValue(1.0)







        self.sp_run_hours.setValue(20.0)







        self.sp_ground.setValue(13.0)







        self.sp_transfer.setValue(0.96)







        self.sp_forklift.setValue(2.0)







        self._update_volume_labels()







        self._append_log("Cuarto limpiado.")







    def _duplicate_room(self):







        """Duplica el cuarto actual en una nueva fila."""







        if not self.rooms_inputs:







            return







        from copy import deepcopy







        src = self.rooms_inputs[self.current_row]







        try:







            new_room = RoomInputs(**asdict(src))







        except Exception:







            new_room = deepcopy(src)







        new_room.nombre = f"{src.nombre} copia"







        self.rooms_inputs.append(new_room)







        self.sp_rooms.setValue(len(self.rooms_inputs))







        self._build_rows(len(self.rooms_inputs))







        self._select_row(len(self.rooms_inputs) - 1)







    def _calc_current_room(self):







        """Calcula solo el cuarto activo para dar feedback rpido."""







        self._save_editor_to_room()







        rooms = [self.rooms_inputs[self.current_row]]







        if rooms[0].largo_m <= 0 or rooms[0].ancho_m <= 0 or rooms[0].altura_m <= 0:







            self._append_log("Dimensiones incompletas para calcular cuarto actual.")







            return







        proj = self.calc.compute_project(rooms, self.sp_sf.value())







        if not proj.rooms:







            return







        res = proj.rooms[0]







        self.lbl_total_btuh.setText(f"{res.total_btuh:,.2f}".replace(",", "."))







        self.lbl_total_kw.setText(f"{res.total_kw:,.2f}".replace(",", "."))







        self.lbl_total_tr.setText(f"{res.total_tr:,.2f}".replace(",", "."))







        self._populate_breakdown([res])







        self._last_results = [res]







        self._last_inputs = rooms







        self._append_log(f"Cuarto {self.current_row+1} calculado: {res.total_btuh:,.2f} BTU/h")







    def _calculate(self):







        if hasattr(self, "_calc_timer"):







            self._calc_timer.stop()







        # Guardar el editor activo en el modelo antes de calcular







        self._save_editor_to_room()







        rooms = self._gather_inputs()







        issues_all = []







        for r in rooms:







            issues_all.extend(validate_room(asdict(r), self.rules))







        if issues_all:







            self.lbl_warnings.setText(" ? ".join([i.message for i in issues_all]))







        else:







            self.lbl_warnings.setText("")







        if not rooms:







            self._append_log("No hay cuartos v?lidos para calcular.")







            self.lbl_state.setText("Warning")







            self.lbl_state.setStyleSheet("color:#b45309;font-weight:700;")







            self.lbl_total_btuh.setText("--")







            self.lbl_total_kw.setText("--")







            self.tbl_res.setRowCount(0)







            self.tbl_break.setRowCount(0)







            self.tbl_res.setRowCount(0)







            self.tbl_break.setRowCount(0)







            self._autosize_table(self.tbl_res)







            self._autosize_table(self.tbl_break)







            self.update_process_outputs(self.current_row)







            return







        # Aplicar valores globales del proyecto y loguear inputs clave







        for idx, room in enumerate(rooms):







            room.run_hours_supp = self.sp_run_hours.value()







            room.ground_temp_C = self.sp_ground.value()







            if hasattr(self, "sp_forklift"):







                room.forklift_hours = self.sp_forklift.value()







            log_line = (







                f"Calc room {idx+1}: Ti={room.T_internal_C}, Te_f={room.T_ext_front_C}, "







                f"esp={room.insulation_thickness_in}, ais={room.insulation_type}, "







                f"RH={room.outside_RH}, recambios={room.air_changes_24h_override}, "







                f"iluminacion_W={room.lighting_W}, motores_W={room.motors_W}, "







                f"montacargas_HP={room.forklift_hp}, personas={room.people_count}, "







                f"producto={room.product_name}, Tin={room.product_Tin_C}, Tout={room.product_Tout_C}, "







                f"kg={room.product_mass_kg}"







            )







            self._append_log(log_line)







        proj = self.calc.compute_project(rooms, self.sp_sf.value())







        if not proj:







            self._append_log("No hay cuartos v?lidos para calcular.")







            self.lbl_state.setText("Warning")







            self.lbl_state.setStyleSheet("color:#b45309;font-weight:700;")







            self.update_process_outputs(self.current_row)







            return







        total_btuh = proj.total_btuh







        total_kw = proj.total_kw







        total_tr = proj.total_tr







        # log post-clculo por cuarto







        for idx, res in enumerate(proj.rooms):







            comp = res.components







            u_used = self.calc._u_factor(res.inputs.insulation_type, res.inputs.insulation_thickness_in)







            self._append_log(







                f"[POST-CALC] room {idx+1}: U24={u_used:.3f} trans={comp.transmission_btuh:,.2f} "







                f"infil_tot={comp.infiltration_btuh:,.2f} sens={comp.infiltration_sensible_btuh:,.2f} "







                f"lat={comp.infiltration_latent_btuh:,.2f}"







            )







        self.lbl_total_btuh.setText(f"{total_btuh:,.2f}".replace(",", "."))







        self.lbl_total_kw.setText(f"{total_kw:,.2f}".replace(",", "."))







        self.lbl_total_tr.setText(f"{total_tr:,.2f}".replace(",", "."))







        self._populate_results(proj.rooms)







        self._last_results = proj.rooms







        self._last_inputs = rooms







        self._populate_breakdown(proj.rooms)







        self._refresh_detail_selector()







        if hasattr(self, "cb_detail_room") and self.cb_detail_room.count():







            self._populate_detail_view(self.cb_detail_room.currentIndex())







        self.lbl_state.setText("Listo")







        self.lbl_state.setStyleSheet("color:#16a34a;font-weight:700;")







        self._append_log(f"Proyecto calculado. Total BTU/h: {total_btuh:,.2f}")







        self.update_process_outputs(self.current_row)







        self._refresh_infiltration_outputs(self.current_row)







    def update_insulation_outputs(self, room_idx: int, save_model: bool = True, schedule_recalc: bool = True):







        """Actualiza U24 calculado al cambiar aislamiento/espesor."""







        try:







            if room_idx is None or room_idx < 0 or room_idx >= len(self.rooms_inputs):







                return







            room = self.rooms_inputs[room_idx]







            if room_idx == self.current_row and hasattr(self, "cb_AISLAMIENTO") and hasattr(self, "sp_espesor"):







                ins_type = self.cb_AISLAMIENTO.currentData()







                thickness = self.sp_espesor.value()







                if save_model and not self._loading_room and not self._updating_widgets:







                    room.insulation_type = ins_type







                    room.insulation_thickness_in = thickness







            else:







                ins_type = room.insulation_type







                thickness = room.insulation_thickness_in







            u_val = self.calc._u_factor(ins_type, thickness)







            self.ro_u24.setText(f"{u_val:.3f}")







            self._append_log(f"[AISLAMIENTO] room={room_idx} tipo={ins_type} esp={thickness} U24={u_val:.3f}")







            if save_model and not self._loading_room and not self._updating_widgets:







                room.insulation_type = ins_type







                room.insulation_thickness_in = thickness







            if schedule_recalc and not self._loading_room and not self._updating_widgets:







                self._schedule_calc()







        except Exception as e:







            self._append_log(f"U24 calc error: {e}")







    def _refresh_infiltration_outputs(self, room_idx: int):







        if not hasattr(self, "ro_infil_total"):







            return







        if room_idx is None or room_idx < 0:







            vals = ("--", "--", "--")







            mode_txt = ""







        elif getattr(self, "_last_results", None) and room_idx < len(self._last_results):







            res = self._last_results[room_idx]







            comp = getattr(res, "components", None)







            if comp:







                def fmt(v):







                    return f"{v:,.2f}".replace(",", ".")







                vals = (fmt(comp.infiltration_btuh), fmt(comp.infiltration_sensible_btuh), fmt(comp.infiltration_latent_btuh))







            else:







                vals = ("--", "--", "--")







            if getattr(res, "infiltration_sensible_only", False):







                mode_txt = "Sensible-only (sin RH ext)"







            else:







                mode_txt = "Normal"







            if not getattr(res.inputs, "infiltration_enabled", True):







                mode_txt = "Desactivada por usuario"







            notes = getattr(res, "infiltration_notes", None)







            if notes:







                mode_txt = f"{mode_txt} [{notes}]"







        else:







            vals = ("--", "--", "--")







            mode_txt = ""







        self.ro_infil_total.setText(vals[0])







        self.ro_infil_sensible.setText(vals[1])







        self.ro_infil_latente.setText(vals[2])







        if hasattr(self, "lbl_infil_mode"):







            self.lbl_infil_mode.setText(mode_txt)







    def _refresh_detail_selector(self):







        if not hasattr(self, "cb_detail_room"):







            return







        self.cb_detail_room.blockSignals(True)







        self.cb_detail_room.clear()







        results = getattr(self, "_last_results", []) or []







        for i, res in enumerate(results):







            name = getattr(res.inputs, "nombre", f"CUARTO {i+1}")







            self.cb_detail_room.addItem(name, i)







        target = self.current_row if self.cb_detail_room.count() else -1







        if target >= 0 and target < self.cb_detail_room.count():







            self.cb_detail_room.setCurrentIndex(target)







        elif self.cb_detail_room.count():







            self.cb_detail_room.setCurrentIndex(0)







        self.cb_detail_room.blockSignals(False)







        self._populate_detail_view(self.cb_detail_room.currentIndex() if self.cb_detail_room.count() else -1)







    def _populate_detail_view(self, idx: int):







        if not hasattr(self, "tree_detail"):







            return







        self.tree_detail.clear()







        results = getattr(self, "_last_results", []) or []







        if idx is None or idx < 0 or idx >= len(results):







            self.tree_detail.addTopLevelItem(QTreeWidgetItem(["SIN RESULTADOS", "", "", ""]))







            self.tree_detail.expandAll()







            return







        res = results[idx]







        comp = getattr(res, "components", None)







        total_btuh = comp.total_btuh if comp else 0.0







        pct_map = getattr(res, "percentages", {}) or {}







        def pct_of(val):







            return (val / total_btuh * 100) if total_btuh else None







        def fmt_btuh(v):







            return f"{v:,.2f}".replace(",", ".")







        def fmt_kw(v):







            return f"{(v/3412.142):,.2f}".replace(",", ".")







        def fmt_value(label: str, value):







            """Formatea valores de entrada (RH, ACH, CFM, etc.)."""







            try:







                num = float(value)







            except Exception:







                return str(value)







            label_up = label.upper()







            if "RH" in label_up and "%" in label_up:







                return f"{num*100:,.2f}%".replace(",", ".")







            if "ACH" in label_up:







                return f"{num:,.2f}".replace(",", ".")







            if "CFM" in label_up:







                return f"{num:,.2f}".replace(",", ".")







            if "RUN_HOURS" in label_up or "H/DIA" in label_up or "H/DÍA" in label_up:







                return f"{num:,.2f}".replace(",", ".")







            return f"{num:,.2f}".replace(",", ".")







        def up(txt):







            try:







                return str(txt).upper()







            except Exception:







                return str(txt)







        def add_row(parent, label, btuh=None, pct=None, value=None):







            btuh_txt = fmt_btuh(btuh) if btuh is not None else (fmt_value(label, value) if value is not None else "")







            kw_txt = fmt_kw(btuh) if btuh is not None else ""







            pct_txt = f"{pct:,.2f}%".replace(",", ".") if pct is not None else ""







            item = QTreeWidgetItem(parent, [up(label), btuh_txt, kw_txt, pct_txt])







            # Alinear: texto izquierda, números derecha







            item.setTextAlignment(0, Qt.AlignLeft | Qt.AlignVCenter)







            for col in (1, 2, 3):







                item.setTextAlignment(col, Qt.AlignRight | Qt.AlignVCenter)







            return item







        inputs_root = QTreeWidgetItem(["INPUTS", "", "", ""])







        add_row(inputs_root, "TI (C)", value=str(getattr(res.inputs, "T_internal_C", "")))







        add_row(inputs_root, "TE FRENTE (C)", value=str(getattr(res.inputs, "T_ext_front_C", "")))







        add_row(inputs_root, "TE TECHO (C)", value=str(getattr(res.inputs, "T_ext_roof_C", "")))







        add_row(inputs_root, "TE SUELO/GROUND (C)", value=str(getattr(res.inputs, "ground_temp_C", "")))







        add_row(inputs_root, "RH EXT (%)", value=str(getattr(res.inputs, "outside_RH", "")))







        add_row(inputs_root, "RH INT (%)", value=str(getattr(res.inputs, "inside_RH", "")))







        add_row(inputs_root, "ACH OVERRIDE", value=str(getattr(res, "infiltration_ach_24h", getattr(res.inputs, "air_changes_24h_override", ""))))







        add_row(inputs_root, "CFM", value=str(getattr(res, "infiltration_cfm", "")))







        add_row(inputs_root, "RUN_HOURS_SUPP (H/DIA)", value=str(getattr(res.inputs, "run_hours_supp", "")))







        add_row(inputs_root, "FACTOR USO", value=str(getattr(res.inputs, "use_factor", "")))







        add_row(inputs_root, "INFILTRACION HABILITADA", value="SI" if getattr(res.inputs, "infiltration_enabled", True) else "NO")







        comp_root = QTreeWidgetItem(["COMPONENTES", "", "", ""])







        if comp:







            add_row(comp_root, "TRANSMISIÓN", comp.transmission_btuh, pct_of(comp.transmission_btuh))







            add_row(comp_root, "INFILTRACIÓN (TOTAL)", comp.infiltration_btuh, pct_of(comp.infiltration_btuh))







            add_row(comp_root, "  INFILTRACIÓN SENSIBLE", comp.infiltration_sensible_btuh, pct_of(comp.infiltration_sensible_btuh))







            add_row(comp_root, "  INFILTRACIÓN LATENTE", comp.infiltration_latent_btuh, pct_of(comp.infiltration_latent_btuh))







            add_row(comp_root, "ILUMINACIÓN", comp.lighting_btuh, pct_of(comp.lighting_btuh))







            add_row(comp_root, "MOTORES", comp.motors_btuh, pct_of(comp.motors_btuh))







            add_row(comp_root, "MONTACARGAS", comp.forklift_btuh, pct_of(comp.forklift_btuh))







            add_row(comp_root, "PERSONAS", comp.people_btuh, pct_of(comp.people_btuh))







            add_row(comp_root, "DESHIELO", comp.defrost_btuh, pct_of(comp.defrost_btuh))







            add_row(comp_root, "PRODUCTO/PROCESO", comp.product_btuh, pct_of(comp.product_btuh))







            add_row(comp_root, "INTERNAS (SUMA)", comp.internal_btuh, pct_of(comp.internal_btuh))







            add_row(comp_root, "TOTAL", total_btuh, 100.0)







        notes_root = QTreeWidgetItem(["NOTAS", "", "", ""])







        if getattr(res.inputs, "infiltration_enabled", True) is False:







            add_row(notes_root, "INFILTRACIÓN", value="DESACTIVADA POR USUARIO")







        if getattr(res, "infiltration_sensible_only", False):







            add_row(notes_root, "INFILTRACIÓN", value="MODO SENSIBLE-ONLY (SIN RH EXT)")







        if getattr(res, "infiltration_notes", None):







            add_row(notes_root, "DETALLE", value=str(getattr(res, "infiltration_notes")))







        if getattr(res, "product", None) and getattr(res.product, "total_btu_cycle", None) is None:







            add_row(notes_root, "PRODUCTO", value="SIN BREAKDOWN DE PRODUCTO")







        self.tree_detail.addTopLevelItem(inputs_root)







        self.tree_detail.addTopLevelItem(comp_root)







        self.tree_detail.addTopLevelItem(notes_root)







        self.tree_detail.expandAll()







    def _populate_results(self, results):







        self.tbl_res.setRowCount(len(results))







        self._autosize_table(self.tbl_res)







        self._autosize_table(self.tbl_res)







        for i, res in enumerate(results):







            nombre = getattr(res.inputs, "nombre", f"CUARTO {i+1}")







            perfil = self._perfil_label_from_id(getattr(res.inputs, "perfil_id", ""))







            self.tbl_res.setItem(i, 0, QTableWidgetItem(nombre))







            self.tbl_res.setItem(i, 1, QTableWidgetItem(perfil))







            self.tbl_res.setItem(i, 2, QTableWidgetItem(f"{res.total_btuh:,.2f}".replace(",", ".")))







            self.tbl_res.setItem(i, 3, QTableWidgetItem(f"{res.total_kw:,.2f}".replace(",", ".")))







            self.tbl_res.setItem(i, 4, QTableWidgetItem(f"{res.total_tr:,.2f}".replace(",", ".")))







            btn = QPushButton("DETALLE")







            btn.clicked.connect(lambda _, r=i: self._show_detail(r))







            self.tbl_res.setCellWidget(i, 5, btn)







        self._refresh_detail_selector()







    def _show_detail(self, idx: int):







        try:







            res = self._last_results[idx]







        except Exception:







            return







        nombre = getattr(res.inputs, "nombre", f"CUARTO {idx+1}")







        perfil = self._perfil_label_from_id(getattr(res.inputs, "perfil_id", ""))







        comp = res.components







        meta = getattr(res, "metadata", {}) or {}







        run_hours = getattr(res.inputs, "run_hours_supp", None)







        if run_hours is None:







            run_hours = getattr(res.inputs, "run_hours", None)







        run_hours = run_hours if run_hours is not None else 0







        ach_display = "Auto (sin datos)"







        try:







            ach_override = getattr(res.inputs, "air_changes_24h_override", None)







            if ach_override is not None:







                ach_display = ach_override







            else:







                L_ft = res.inputs.largo_m * 3.28084







                W_ft = res.inputs.ancho_m * 3.28084







                H_ft = res.inputs.altura_m * 3.28084







                if L_ft > 0 and W_ft > 0 and H_ft > 0:







                    volume_ft3 = L_ft * W_ft * H_ft







                    freezing = (res.inputs.T_internal_C or 0) < -5







                    ach_display = self.calc._air_changes(volume_ft3, freezing)







                else:







                    self._append_log("Detalle: Auto (sin datos) por volumen cero")







        except Exception as e:







            self._append_log(f"Detalle: no se pudo calcular ACH auto: {e}")







        msg = (







            f"<b>{nombre}</b><br>"







            f"Perfil: {perfil}<br>"







            f"TRANSMISIN: {comp.transmission_btuh:,.0f} BTU/h<br>"







            f"Infiltracin: {comp.infiltration_btuh:,.0f} BTU/h<br>"







            f"Internas: {comp.internal_btuh:,.0f} BTU/h<br>"







            f"Total: {res.total_btuh:,.0f} BTU/h<br>"







            f"run_hours_supp: {run_hours} h/da<br>"







            f"air_changes_24h: {ach_display}"







        )







        QMessageBox.information(self, "Detalle", msg)







    def _export_project(self):
        if not hasattr(self, "_last_results") or not hasattr(self, "_last_inputs"):
            QMessageBox.information(self, "Exportar", "Primero calcule el proyecto.")
            return

        proj = self.ed_project.text().strip().upper() or "PROYECTO"
        proj_safe = proj.replace(" ", "_")

        data = {
            "project_name": proj,
            "safety_factor": self.sp_sf.value(),
            "rooms": [asdict(r) for r in self._last_inputs],
            "results": [asdict(r) for r in self._last_results],
        }

        PROJECTS_DIR.mkdir(parents=True, exist_ok=True)
        json_name = f"{datetime.now():%Y%m%d_%H%M%S}_{proj_safe}.json"
        json_path = PROJECTS_DIR / json_name
        with open(json_path, "w", encoding="utf-8") as fh:
            json.dump(data, fh, ensure_ascii=False, indent=2)

        xls_default = f"{datetime.now():%Y%m%d_%H%M%S}_{proj_safe}.xlsx"
        save_to, _ = QFileDialog.getSaveFileName(self, "Guardar Excel", xls_default, "Excel (*.xlsx)")

        if save_to:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
                from openpyxl.utils import get_column_letter

                wb = Workbook()
                ws = wb.active
                ws.title = "RESULTADOS"

                title_font = Font(bold=True, size=14)
                header_font = Font(bold=True)
                group_fill = PatternFill(start_color="FFE8F1FF", end_color="FFE8F1FF", fill_type="solid")
                section_fill = PatternFill(start_color="FFF5F7FA", end_color="FFF5F7FA", fill_type="solid")
                entry_fill = PatternFill(start_color="FFF9FBFD", end_color="FFF9FBFD", fill_type="solid")
                border = Border(
                    left=Side(style="thin", color="FFCCCCCC"),
                    right=Side(style="thin", color="FFCCCCCC"),
                    top=Side(style="thin", color="FFCCCCCC"),
                    bottom=Side(style="thin", color="FFCCCCCC"),
                )
                align_left = Alignment(horizontal="left")
                align_right = Alignment(horizontal="right")
                align_center = Alignment(horizontal="center")

                def write_kv_block(title, rows, start_row):
                    r = start_row
                    ws.cell(row=r, column=1, value=title).font = header_font
                    ws.cell(row=r, column=1).fill = section_fill
                    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
                    r += 1
                    for row in rows:
                        if len(row) == 2:
                            label, value = row
                            fmt = None
                        else:
                            label, value, fmt = row
                        ws.cell(row=r, column=1, value=label).font = header_font
                        ws.cell(row=r, column=1).alignment = align_left
                        ws.cell(row=r, column=1).border = border
                        ws.cell(row=r, column=1).fill = entry_fill
                        cell = ws.cell(row=r, column=2, value=value)
                        cell.border = border
                        cell.fill = entry_fill
                        cell.alignment = align_right if isinstance(value, (int, float)) else align_left
                        if fmt:
                            cell.number_format = fmt
                        r += 1
                    return r

                ws.cell(row=1, column=1, value="RESUMEN DEL PROYECTO").font = title_font
                ws.cell(row=1, column=1).alignment = align_left

                meta_rows = [
                    ("PROYECTO", proj, None),
                    ("FECHA", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), None),
                    ("FACTOR SEGURIDAD", self.sp_sf.value(), "0.00"),
                ]
                next_row = write_kv_block("METADATOS", meta_rows, start_row=3) + 1

                for idx, result in enumerate(self._last_results):
                    inputs = getattr(result, "inputs", None)
                    name = getattr(inputs, "nombre", None) or f"CUARTO {idx+1}"
                    perfil = getattr(result, "perfil_label", None) or getattr(inputs, "perfil_id", "")

                    ws.cell(row=next_row, column=1, value=f"{name} ({perfil})").font = title_font
                    ws.cell(row=next_row, column=1).fill = group_fill
                    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=4)
                    next_row += 1

                    ins = inputs
                    gv = lambda attr, default=0: getattr(ins, attr, default)
                    inputs_rows = [
                        ("PERFIL/USO", gv("perfil_id", ""), None),
                        ("LARGO (m)", gv("largo_m"), "0.00"),
                        ("ANCHO (m)", gv("ancho_m"), "0.00"),
                        ("ALTURA (m)", gv("altura_m"), "0.00"),
                        ("PUERTAS", gv("puertas"), "0"),
                        ("TI INT (C)", gv("T_internal_C"), "0.00"),
                        ("T EXT FRENTE (C)", gv("T_ext_front_C"), "0.00"),
                        ("T EXT ATRAS (C)", gv("T_ext_back_C"), "0.00"),
                        ("T EXT DER (C)", gv("T_ext_right_C"), "0.00"),
                        ("T EXT IZQ (C)", gv("T_ext_left_C"), "0.00"),
                        ("T EXT TECHO (C)", gv("T_ext_roof_C"), "0.00"),
                        ("TEMP SUELO (C)", gv("ground_temp_C"), "0.00"),
                        ("FACTOR PARED", gv("wall_transfer_factor"), "0.00"),
                        ("AISLAMIENTO TIPO", gv("insulation_type", ""), None),
                        ("AISLAMIENTO ESP (in)", gv("insulation_thickness_in"), "0.00"),
                        ("T EXT AIRE (C)", gv("outside_air_temp_C"), "0.00"),
                        ("RH EXT (0-1)", gv("outside_RH"), "0.00"),
                        ("RH INT (0-1)", gv("inside_RH"), "0.00"),
                        ("ACH OVERRIDE", gv("air_changes_24h_override"), "0.00"),
                        ("INFILTRACION HABILITADA", "SI" if gv("infiltration_enabled", True) else "NO", None),
                        ("FACTOR USO", gv("use_factor", 1), "0.00"),
                        ("HORAS EFECTIVAS (h/dia)", gv("run_hours_supp", 24), "0.00"),
                        ("ILUMINACION W", gv("lighting_W"), "0.00"),
                        ("ILUMINACION h/dia", gv("lighting_hours"), "0.00"),
                        ("MOTORES W", gv("motors_W"), "0.00"),
                        ("MOTORES h/dia", gv("motors_hours"), "0.00"),
                        ("MONTACARGAS HP", gv("forklift_hp"), "0.00"),
                        ("MONTACARGAS h/dia", gv("forklift_hours"), "0.00"),
                        ("PERSONAS", gv("people_count"), "0.00"),
                        ("PERSONAS h/dia", gv("people_hours"), "0.00"),
                        ("PERSONAS BTU/H", gv("people_btuh"), "0.00"),
                        ("DESHIELO W", gv("defrost_W"), "0.00"),
                        ("DESHIELO CANT", gv("defrost_count"), "0"),
                        ("DESHIELO MIN", gv("defrost_duration_min"), "0.00"),
                        ("DESHIELO FRACCION", gv("defrost_fraction_to_room"), "0.00"),
                        ("PRODUCTO", gv("product_name", "") or gv("perfil_id", ""), None),
                        ("PACKAGING MULT", gv("product_packaging_multiplier", 1), "0.00"),
                        ("PRODUCT METHOD", gv("product_method", ""), None),
                        ("KG", gv("product_mass_kg"), "0.00"),
                        ("TIN PROCESO (C)", gv("product_Tin_C"), "0.00"),
                        ("TOUT PROCESO (C)", gv("product_Tout_C"), "0.00"),
                        ("CICLO (h)", gv("product_cycle_h"), "0.00"),
                    ]

                    next_row = write_kv_block("ENTRADAS", inputs_rows, next_row) + 1

                    comp = getattr(result, "components", None)
                    total_btuh = getattr(result, "total_btuh", 0.0) or 0.0
                    total_kw = getattr(result, "total_kw", 0.0) or 0.0
                    rows_res = [
                        ("TRANSMISIÓN", getattr(comp, "transmission_btuh", 0.0)),
                        ("INFILTRACIÓN (TOTAL)", getattr(comp, "infiltration_btuh", 0.0)),
                        ("INF SENSIBLE", getattr(comp, "infiltration_sensible_btuh", 0.0)),
                        ("INF LATENTE", getattr(comp, "infiltration_latent_btuh", 0.0)),
                        ("ILUMINACIÓN", getattr(comp, "lighting_btuh", 0.0)),
                        ("MOTORES", getattr(comp, "motors_btuh", 0.0)),
                        ("MONTACARGAS", getattr(comp, "forklift_btuh", 0.0)),
                        ("PERSONAS", getattr(comp, "people_btuh", 0.0)),
                        ("DESHIELO", getattr(comp, "defrost_btuh", 0.0)),
                        ("PRODUCTO/PROCESO", getattr(comp, "product_btuh", 0.0)),
                    ]

                    ws.cell(row=next_row, column=1, value="RESULTADOS").font = header_font
                    ws.cell(row=next_row, column=1).fill = section_fill
                    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=4)
                    next_row += 1

                    headers = ["COMPONENTE", "BTU/H", "kW", "% DEL TOTAL"]
                    for c, h in enumerate(headers, start=1):
                        cell = ws.cell(row=next_row, column=c, value=h)
                        cell.font = header_font
                        cell.fill = section_fill
                        cell.alignment = align_center
                        cell.border = border
                    next_row += 1

                    for label, btuh in rows_res:
                        kw_val = btuh / 3412.142 if btuh else 0.0
                        pct = (btuh / total_btuh) if total_btuh else 0.0
                        row_vals = [label, btuh, kw_val, pct]
                        fmts = [None, "#,##0.00", "#,##0.00", "0.00%"]
                        for c, (val, fmt) in enumerate(zip(row_vals, fmts), start=1):
                            cell = ws.cell(row=next_row, column=c, value=val)
                            cell.border = border
                            cell.alignment = align_left if c == 1 else align_right
                            if fmt:
                                cell.number_format = fmt
                        next_row += 1

                    for c, val in enumerate(["TOTAL CUARTO", total_btuh, total_kw, 1], start=1):
                        cell = ws.cell(row=next_row, column=c, value=val)
                        cell.font = header_font
                        if c == 1:
                            cell.fill = group_fill
                        cell.border = border
                        cell.alignment = align_left if c == 1 else align_right
                        if c > 1:
                            cell.number_format = "#,##0.00" if c < 4 else "0.00%"
                    next_row += 2

                for col in range(1, ws.max_column + 1):
                    max_len = 0
                    for row in range(1, ws.max_row + 1):
                        val = ws.cell(row=row, column=col).value
                        if val is None:
                            continue
                        max_len = max(max_len, len(str(val)))
                    ws.column_dimensions[get_column_letter(col)].width = max(12, min(40, max_len + 2))

                wb.save(save_to)
            except Exception as e:
                QMessageBox.critical(self, "Error Excel", str(e))

        QMessageBox.information(self, "Exportar", f"Guardado en biblioteca: {json_path}")

    def _open_library(self):







        dlg = BibliotecaCuartosIndDialog(self, on_load=self._load_project)







        dlg.exec()







    def _load_project(self, data: dict):







        self.ed_project.setText(data.get("project_name", ""))







        rooms = data.get("rooms", [])







        self.rooms_inputs = [RoomInputs(**r) for r in rooms] if rooms else []







        self.sp_rooms.setValue(len(rooms) or 1)







        self._build_rows(len(rooms) or 1)







        for i, r in enumerate(rooms):







            self.tbl.item(i,0).setText(r.get("nombre", f"CUARTO {i+1}"))







            self.tbl.cellWidget(i,1).setValue(r.get("largo_m",0))







            self.tbl.cellWidget(i,2).setValue(r.get("ancho_m",0))







            self.tbl.cellWidget(i,3).setValue(r.get("altura_m",0))







            # perfil







            combo: QComboBox = self.tbl.cellWidget(i,4)







            pid = r.get("perfil_id", "")







            idx = combo.findData(pid)







            if idx >= 0: combo.setCurrentIndex(idx)







            # puertas







            self.tbl.cellWidget(i,5).setValue(int(r.get("puertas",1)))







        # recalcular si hay RESULTADOS almacenados







        self._calculate()







    def _new_project(self):







        self.ed_project.clear()







        self.sp_rooms.setValue(1)







        self.rooms_inputs = []







        self._build_rows(1)







        self.tbl_res.setRowCount(0)







        self.lbl_total_btuh.setText("--")







        self.lbl_total_kw.setText("--")







        self.lbl_total_tr.setText("--")







        self._last_results = []







        self._last_inputs = []







        self.tbl_break.setRowCount(0)







        self._reset_log_safe()







        self.lbl_state.setText("Listo")







        self.lbl_state.setStyleSheet("color:#16a34a;font-weight:700;")







    # ------------------ HELPERS ------------------ #







    def _append_log(self, text: str):



        # Log deshabilitado (panel removido); evitamos salidas a consola
        return







    def _clear_log(self):



        """Compatibilidad cuando se llama a limpiar el log."""



        return







    def _reset_log_safe(self):



        """Usado por botones heredados; evita acceder a log inexistente."""



        return







    def _force_project_upper(self, text: str):



        """Fuerza el nombre de proyecto a mayúsculas."""



        if text is None:



            return



        up = text.upper()



        if up != text:



            self.ed_project.blockSignals(True)



            self.ed_project.setText(up)



            self.ed_project.blockSignals(False)



    def _perfil_label_from_id(self, pid: str) -> str:







        for pid0, label in self.profile_list:







            if pid0 == pid:







                return label







        return pid or "--"







    def _update_volume_labels(self):







        vol_m3 = self.sp_largo.value() * self.sp_ancho.value() * self.sp_alto.value()







        vol_ft3 = vol_m3 * 35.3147







        self.ro_vol_m3.setText(f"{vol_m3:.2f}")







        self.ro_vol_ft3.setText(f"{vol_ft3:.2f}")







        # Recalcular kg sugeridos si el perfil usa densidad por volumen







        if 0 <= self.current_row < len(self.rooms_inputs):







            room = self.rooms_inputs[self.current_row]







            profile = self.profiles_by_id.get(room.perfil_id or "", {})







            proc = profile.get("autofill", {}).get("process", {})







            if proc.get("kg_mode") == "volume_density":







                self._apply_process_autofill(self.current_row, force=False)







    def _apply_profile_defaults(self, row: int):



        pid = self.rooms_inputs[row].perfil_id



        if not pid:



            return







        defaults = self.profile_defs.get(pid, {}).get("defaults", {})







        ti = defaults.get("internal_temp_C")



        if ti is None:



            return







        ti = float(ti)



        self.rooms_inputs[row].T_internal_C = ti







        if row == self.current_row:



            self.sp_tint.blockSignals(True)



            self.sp_tint.setValue(ti)



            self.sp_tint.blockSignals(False)







        # No disparamos autofill internos/proceso automticamente; solo con botn







        self.rooms_inputs[row].T_internal_C = ti







        if row == self.current_row:







            self.sp_tint.blockSignals(True)







            self.sp_tint.setValue(ti)







            self.sp_tint.blockSignals(False)







        # No aplicar tﾃｭpicos internos/proceso automﾃ｡tico al cambiar perfil







    def _apply_internals_autofill(self, row: int, force: bool = False):







        """Aplica valores tpicos de la pestaa INTERNAS segn el perfil y el rea."""







        if row < 0 or row >= len(self.rooms_inputs):







            return







        room = self.rooms_inputs[row]







        profile = self.profile_defs.get(room.perfil_id or "", {})







        internals = profile.get("autofill", {}).get("internals", {})







        if not internals:







            return







        largo = room.largo_m or 0







        ancho = room.ancho_m or 0







        area_m2 = largo * ancho if largo > 0 and ancho > 0 else 0







        sug_luz = area_m2 * float(internals.get("lighting_w_per_m2", 0))







        sug_motor = area_m2 * float(internals.get("motors_w_per_m2", 0))







        sug_fork = math.ceil(float(internals.get("forklift_hp", 0)))







        sug_people = (area_m2 / 100.0) * float(internals.get("people_per_100m2", 0))







        sug_people = math.ceil(sug_people)







        sug_people = max(0, min(25, sug_people))







        sug_people_h = float(internals.get("people_hours_per_day", 0))







        def maybe_set_spin(spin: QDoubleSpinBox, val: float):







            if force or spin.value() == 0:







                spin.blockSignals(True)







                spin.setValue(val)







                spin.blockSignals(False)







        if row == self.current_row and hasattr(self, "sp_luz_w"):







            maybe_set_spin(self.sp_luz_w, sug_luz)







            maybe_set_spin(self.sp_motor_w, sug_motor)







            maybe_set_spin(self.sp_mont_hp, sug_fork)







            maybe_set_spin(self.sp_people, sug_people)







            maybe_set_spin(self.sp_people_h, sug_people_h)







            # Persistir en el modelo







            self._save_editor_to_room()







        else:







            if force or room.lighting_W == 0:







                room.lighting_W = sug_luz







            if force or room.motors_W == 0:







                room.motors_W = sug_motor







            if force or room.forklift_hp == 0:







                room.forklift_hp = sug_fork







            if force or room.people_count == 0:







                room.people_count = sug_people







            if force or room.people_hours == 0:







                room.people_hours = sug_people_h







    def _apply_process_autofill(self, row: int, force: bool = False):







        """Aplica valores tpicos de la pestaa PROCESO segn perfil y volumen."""







        if row < 0 or row >= len(self.rooms_inputs):







            return







        room = self.rooms_inputs[row]







        profile_id = None







        if hasattr(self, "cb_perfil_head") and row == self.current_row:







            profile_id = self.cb_perfil_head.currentData()







        if not profile_id and hasattr(self, "tbl"):







            combo = self.tbl.cellWidget(row, 4)







            if combo is not None:







                profile_id = combo.currentData()







        if not profile_id:







            profile_id = room.perfil_id







        profile = self.profiles_by_id.get(profile_id or "")







        if not profile:







            return







        auto = profile.get("autofill", {}).get("process", {}) or {}







        if not auto:







            return







        # Lecturas actuales







        if row == self.current_row and hasattr(self, "sp_tint"):







            ti_current = self.sp_tint.value()







        else:







            ti_current = room.T_internal_C or 0.0







        if row == self.current_row and hasattr(self, "sp_masa"):







            kg_current = self.sp_masa.value()







        else:







            kg_current = room.product_mass_kg or 0.0







        if row == self.current_row and hasattr(self, "sp_ciclo"):







            cycle_current = self.sp_ciclo.value()







        else:







            cycle_current = room.product_cycle_h or 0.0







        # Sugerencias con lgica kg_gate







        tin_proc_def = auto.get("tin_process_default_C")







        tout_proc_def = auto.get("tout_process_default_C")







        if kg_current > 0:







            tin_sug = float(tin_proc_def if tin_proc_def is not None else ti_current)







            tout_sug = float(tout_proc_def if tout_proc_def is not None else ti_current)







        else:







            tin_sug = float(ti_current)







            tout_sug = float(ti_current)







        cycle_sug = float(auto.get("cycle_hours_default") or cycle_current or 0.0)







        updated = False







        def maybe_set(spin: QDoubleSpinBox, val: float):







            nonlocal updated







            if force or spin.value() == 0:







                spin.blockSignals(True)







                spin.setValue(val)







                spin.blockSignals(False)







                updated = True







        if row == self.current_row and hasattr(self, "sp_tin"):







            maybe_set(self.sp_tin, tin_sug)







            maybe_set(self.sp_tout, tout_sug)







            maybe_set(self.sp_ciclo, cycle_sug)







            if updated:







                self._save_editor_to_room()







                self.update_process_outputs(row)







        else:







            if force or room.product_Tin_C in (None, 0):







                room.product_Tin_C = tin_sug







            if force or room.product_Tout_C in (None, 0):







                room.product_Tout_C = tout_sug







            if force or room.product_cycle_h in (None, 0):







                room.product_cycle_h = cycle_sug







        self.update_process_outputs(row)







    def _format_num(self, val: float) -> str:







        try:







            return f"{val:.2f}"







        except Exception:







            return str(val)







    def _set_process_outputs(self, ref, cong, post, total_btu, total_kw):







        self._append_log(







            f"[PROCESO][set] ref={ref} cong={cong} post={post} total_btu={total_btu} total_kw={total_kw}"







        )







        for widget, value in (







            (self.ro_carga_ref, ref),







            (self.ro_carga_cong, cong),







            (self.ro_carga_post, post),







            (self.ro_carga_total_btuc, total_btu),







            (self.ro_carga_total_kwc, total_kw),







        ):







            if isinstance(value, str):







                widget.setText(value)







            else:







                widget.setText(self._format_num(value))







    def update_process_outputs(self, room_idx: int) -> None:







        if room_idx is None or room_idx < 0 or room_idx >= len(self.rooms_inputs):







            return







        # Asegurar que el modelo refleja la UI







        self._save_editor_to_room()







        tin = self.sp_tin.value()







        tout = self.sp_tout.value()







        kg = self.sp_masa.value()







        ciclo = self.sp_ciclo.value()







        prod_name = self.cb_producto.currentData() or (self.cb_producto.currentText() or "").lower()







        profile_id = self.cb_perfil_head.currentData()







        block_kg = kg <= 0







        block_dt = abs(tin - tout) < 0.01







        self._append_log(







            f"[PROCESO] room={room_idx} Tin={tin:.2f} Tout={tout:.2f} kg={kg:.2f} ciclo_h={ciclo:.2f} "







            f"block_kg={block_kg} block_dt={block_dt}"







        )







        model_room = self.rooms_inputs[room_idx]







        self._append_log(







            f"[PROCESO][modelo] Tin={model_room.product_Tin_C} Tout={model_room.product_Tout_C} "







            f"kg={model_room.product_mass_kg} ciclo_h={model_room.product_cycle_h} "







            f"product={model_room.product_name} profile={model_room.perfil_id}"







        )







        if block_kg or ciclo <= 0 or block_dt:







            self._append_log("Sin proceso (kg<=0 o Tin==Tout)")







            self._set_process_outputs("--", "--", "--", "--", "--")







            return







        try:







            base_room = RoomInputs(**asdict(self.rooms_inputs[room_idx]))







            base_room.product_Tin_C = tin







            base_room.product_Tout_C = tout







            base_room.product_mass_kg = kg







            base_room.product_cycle_h = ciclo







            base_room.product_name = prod_name or base_room.product_name







            if not base_room.product_name:







                base_room.product_name = (self.cb_producto.currentText() or "").lower()







            self._append_log(







                f"[PROCESO] base_room producto={base_room.product_name} kg={base_room.product_mass_kg} ciclo={base_room.product_cycle_h}"







            )







            result = self.calc.compute_room(base_room, self.sp_sf.value())







            try:







                self._append_log(f"[PROCESO] result type={type(result)} keys={list(getattr(result, '__dict__', {}).keys())}")







                if hasattr(result, "components"):







                    self._append_log(f"[PROCESO] components keys={list(getattr(result.components, '__dict__', {}).keys())}")







                if hasattr(result, "product"):







                    self._append_log(f"[PROCESO] product keys={list(getattr(result.product, '__dict__', {}).keys())}")







                if hasattr(result, "product_result"):







                    self._append_log(f"[PROCESO] product_result keys={list(getattr(result.product_result, '__dict__', {}).keys())}")







                try:







                    self._append_log(







                        f"[PROCESO] values product_btuh={getattr(getattr(result, 'components', None), 'product_btuh', None)} "







                        f"product={getattr(result, 'product', None)}"







                    )







                except Exception:







                    pass







            except Exception:







                pass







        except Exception as e:







            self._append_log(f"update_process_outputs error: {e}")







            self._append_log(traceback.format_exc())







            self._set_process_outputs(0.0, 0.0, 0.0, 0.0, 0.0)







            return







        ref_btu = getattr(getattr(result, "product", None), "ref_btu_cycle", None)







        cong_btu = getattr(getattr(result, "product", None), "cong_btu_cycle", None)







        post_btu = getattr(getattr(result, "product", None), "post_btu_cycle", None)







        self._append_log(







            f"[PROCESO] values ref_btu_cycle={ref_btu} cong_btu_cycle={cong_btu} post_btu_cycle={post_btu} "







            f"product_btuh={getattr(result.components, 'product_btuh', None)}"







        )







        if ref_btu is None:







            ref_btu = getattr(getattr(result, "product", None), "ref_btuh", None)







            if ref_btu is not None:







                ref_btu *= ciclo







        if cong_btu is None:







            cong_btu = getattr(getattr(result, "product", None), "cong_btuh", None)







            if cong_btu is not None:







                cong_btu *= ciclo







        if post_btu is None:







            post_btu = getattr(getattr(result, "product", None), "post_btuh", None)







            if post_btu is not None:







                post_btu *= ciclo







        total_cycle_btu = getattr(getattr(result, "product", None), "total_btu_cycle", None)







        total_cycle_kw = None







        product_btuh = getattr(result.components, "product_btuh", None)







        if total_cycle_btu is None and product_btuh is not None:







            total_cycle_btu = product_btuh * ciclo







        if total_cycle_btu is None:







            self._append_log("[PROCESO] Motor no devuelve carga de producto: revisar campos esperados por RoomInputs o l?gica del motor.")







            self._set_process_outputs("--", "--", "--", "--", "--")







            return







        total_cycle_kw = total_cycle_btu / 3412.142  # kWh/ciclo







        if all(v is not None for v in (ref_btu, cong_btu, post_btu)):







            self._set_process_outputs(ref_btu, cong_btu, post_btu, total_cycle_btu, total_cycle_kw)







        else:







            self._append_log("[PROCESO] No hay desglose ref/cong/post en motor; mostrando --")







            self._set_process_outputs("--", "--", "--", total_cycle_btu, total_cycle_kw)







        self._append_log(







            f"process outputs room_idx={room_idx} Ti={base_room.T_internal_C} "







            f"run_hours_supp={base_room.run_hours_supp} ach={base_room.air_changes_24h_override} "







            f"Tin={tin} Tout={tout} kg={kg} total_btuh_cycle={total_cycle_btu} kw_cycle={total_cycle_kw}"







        )







    def estimate_process_kg(self, row: int):







        """Estima kg segun volumen y perfil sin pisar valores existentes."""







        if row < 0 or row >= len(self.rooms_inputs):







            return







        room = self.rooms_inputs[row]







        profile_id = None







        if hasattr(self, "cb_perfil_head") and row == self.current_row:







            profile_id = self.cb_perfil_head.currentData()







        if not profile_id and hasattr(self, "tbl"):







            combo = self.tbl.cellWidget(row, 4)







            if combo is not None:







                profile_id = combo.currentData()







        if not profile_id:







            profile_id = room.perfil_id







        profile = self.profiles_by_id.get(profile_id or "")







        if not profile:







            return







        cfg = profile.get("autofill", {}).get("process", {}).get("kg_estimate") or {}







        if not cfg.get("enabled"):







            return







        if row == self.current_row and all(hasattr(self, attr) for attr in ("sp_largo", "sp_ancho", "sp_alto")):







            largo_m = self.sp_largo.value()







            ancho_m = self.sp_ancho.value()







            alto_m = self.sp_alto.value()







        else:







            largo_m = room.largo_m or 0.0







            ancho_m = room.ancho_m or 0.0







            alto_m = room.altura_m or 0.0







        if largo_m <= 0 or ancho_m <= 0 or alto_m <= 0:







            if row == self.current_row and hasattr(self, "sp_masa"):







                self.sp_masa.setToolTip("Ingrese L/W/H para estimar kg")







            return







        volume_m3 = largo_m * ancho_m * alto_m







        density = float(cfg.get("density_kg_m3", 0.0) or 0.0)







        rotation = float(cfg.get("rotation_daily", 0.0) or 0.0)







        cap_frac = float(cfg.get("cap_fill_fraction", 1.0) or 1.0)







        round_to = float(cfg.get("round_to_kg", 1.0) or 1.0)







        density = max(50.0, min(600.0, density))







        rotation = max(0.0, min(1.0, rotation))







        kg_est = volume_m3 * density * rotation







        kg_cap = volume_m3 * density * cap_frac







        kg_est = min(kg_est, kg_cap)







        if round_to > 0:







            kg_est = round(kg_est / round_to) * round_to







        applied = False







        if row == self.current_row and hasattr(self, "sp_masa"):







            self.sp_masa.setValue(kg_est)







            applied = True







            self._save_editor_to_room()







        else:







            room.product_mass_kg = kg_est







            applied = True







        if applied:



            self._apply_process_autofill(row, force=True)



            self.update_process_outputs(row)







    def _make_card(self, title: str):







        card = QWidget()







        card.setStyleSheet("QWidget{background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:10px;}")







        layout = QVBoxLayout(card)







        lbl = QLabel(title)







        lbl.setStyleSheet("font-weight:800;")







        layout.addWidget(lbl)







        body = QWidget()







        body_layout = QVBoxLayout(body)







        body_layout.setContentsMargins(0,0,0,0)







        body_layout.setSpacing(6)







        layout.addWidget(body)







        return {"card": card, "body": body}







    def _build_editor(self):







        container = QWidget()







        container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)







        vbox = QVBoxLayout(container)







        tabs = QTabWidget()







        tabs.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)







        tabs.setMinimumHeight(500)







        tabs.setStyleSheet("QTabBar::tab{padding:6px 12px;font-weight:700;} QTabBar::tab:selected{background:#e0f2fe;border:1px solid #bae6fd;border-radius:8px;}")







        # Bsico: Producto







        tab_producto = QWidget(); form_prod = QFormLayout(tab_producto)







        # Datos de producto







        self.cb_producto = QComboBox()







        for name in self.calc.foods.keys():







            self.cb_producto.addItem(name.upper(), name)







        # campos readonly







        def _ro():







            le = QLineEdit(); le.setReadOnly(True); le.setPlaceholderText("--"); return le







        self.ro_agua = _ro(); self.ro_prot = _ro(); self.ro_grasa = _ro()







        self.ro_carb = _ro(); self.ro_fibra = _ro(); self.ro_ceniza = _ro()







        # handler producto -> compuestos







        def _on_producto_changed(idx: int):







            key = self.cb_producto.currentData()







            datos = self.calc.foods.get(key, {})







            comp = datos.get("composition_pct", {})







            self.ro_agua.setText(str(comp.get("moisture", "")))







            self.ro_prot.setText(str(comp.get("protein", "")))







            self.ro_grasa.setText(str(comp.get("fat", "")))







            self.ro_carb.setText(str(comp.get("carbohydrate", "")))







            self.ro_fibra.setText(str(comp.get("fiber", "")))







            self.ro_ceniza.setText(str(comp.get("ash", "")))







            self._schedule_calc()







            self.update_process_outputs(self.current_row)







        self.cb_producto.currentIndexChanged.connect(_on_producto_changed)







        self.cb_producto.currentIndexChanged.connect(lambda _=None: (self._save_editor_to_room(), self.update_process_outputs(self.current_row), self._schedule_calc()))







        # inicial







        _on_producto_changed(self.cb_producto.currentIndex())







        form_prod.addRow("PRODUCTO:", self.cb_producto)







        form_prod.addRow("AGUA (%):", self.ro_agua)







        form_prod.addRow("PROTENA (%):", self.ro_prot)







        form_prod.addRow("GRASA (%):", self.ro_grasa)







        form_prod.addRow("CARBOHIDRATO (%):", self.ro_carb)







        form_prod.addRow("FIBRA (%):", self.ro_fibra)







        form_prod.addRow("CENIZA (%):", self.ro_ceniza)







        tabs.addTab(tab_producto, "PRODUCTO")







        # Bsico: Cuarto







        tab_cuarto = QWidget(); form_cuarto = QFormLayout(tab_cuarto)







        form_cuarto.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)







        def _spin(minv,maxv,step=0.1,dec=2,val=0,ph=""):







            s = QDoubleSpinBox()







            s.setRange(minv,maxv); s.setSingleStep(step); s.setDecimals(dec); s.setValue(val)







            if ph:







                s.lineEdit().setPlaceholderText(ph)







            return s







        self.ed_nombre_cuarto = QLineEdit(); self.ed_nombre_cuarto.setPlaceholderText("Nombre del cuarto"); self.ed_nombre_cuarto.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)







        self.cb_perfil_head = QComboBox(); self.cb_perfil_head.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)







        # Placeholder vaco para que no parezca un perfil activo al iniciar







        self.cb_perfil_head.addItem("", "")







        for pid,label in self.profile_list:







            self.cb_perfil_head.addItem(label,pid)







        self.sp_largo = _spin(0,200,0.1,2,0,"Largo m"); self.sp_largo.setToolTip("Largo del cuarto")







        self.sp_ancho = _spin(0,200,0.1,2,0,"Ancho m"); self.sp_alto = _spin(0,50,0.1,2,0,"Alto m")







        self.sp_tint = _spin(-60,50,0.5,2,0,"Temperatura interna")







        # Temperaturas externas por cara







        self.sp_tef = _spin(-60,60,0.5,2,0,"T ext. frente")







        self.sp_teb = _spin(-60,60,0.5,2,0,"T ext. fondo")







        self.sp_ted = _spin(-60,60,0.5,2,0,"T ext. derecha")







        self.sp_tei = _spin(-60,60,0.5,2,0,"T ext. izquierda")







        self.sp_tet = _spin(-60,60,0.5,2,0,"T ext. techo")







        # Volumen (solo lectura)







        self.ro_vol_m3 = _ro(); self.ro_vol_ft3 = _ro()







        for _ro_vol in (self.ro_vol_m3, self.ro_vol_ft3):







            _ro_vol.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)







        form_cuarto.addRow("NOMBRE:", self.ed_nombre_cuarto)







        form_cuarto.addRow("PERFIL/USO:", self.cb_perfil_head)







        self.cb_perfil_head.currentIndexChanged.connect(self._on_editor_profile_changed)







        self.cb_perfil_head.currentIndexChanged.connect(lambda _=None: self._schedule_calc())







        form_cuarto.addRow("TEMPERATURA INTERNA (C):", self.sp_tint)







        form_cuarto.addRow("T EXT. FRENTE (C):", self.sp_tef)







        form_cuarto.addRow("T EXT. FONDO (C):", self.sp_teb)







        form_cuarto.addRow("T EXT. DERECHA (C):", self.sp_ted)







        form_cuarto.addRow("T EXT. IZQUIERDA (C):", self.sp_tei)







        form_cuarto.addRow("T EXT. TECHO (C):", self.sp_tet)







        # Campos avanzados integrados







        self.sp_ground = _spin(-20, 40, 0.5, 2, 13, "Temp. suelo")







        self.sp_transfer = _spin(0, 5, 0.01, 2, 0.96, "Factor pared")







        self.sp_run_hours = _spin(0, 24, 0.5, 2, 20, "Horas efectivas")







        form_cuarto.addRow("TEMP. SUELO (C):", self.sp_ground)







        form_cuarto.addRow("FACTOR TRANSFERENCIA PARED:", self.sp_transfer)







        form_cuarto.addRow("HORAS EFECTIVAS (H/DA):", self.sp_run_hours)







        form_cuarto.addRow("VOLUMEN (M):", self.ro_vol_m3)







        form_cuarto.addRow("VOLUMEN (FT):", self.ro_vol_ft3)







        for sp in (self.sp_largo, self.sp_ancho, self.sp_alto):







            sp.valueChanged.connect(self._update_volume_labels)







        tabs.addTab(tab_cuarto, "CUARTO")







        # Bsico: Internas







        tab_int = QWidget(); form_int = QFormLayout(tab_int)







        # Botn para aplicar valores tpicos







        btn_auto_int = QPushButton("APLICAR TÍPICOS")







        btn_auto_int.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)







        btn_auto_int.setStyleSheet("QPushButton{background:#e0f2fe;color:#0f172a;border-radius:8px;padding:6px 10px;}")







        btn_auto_int.clicked.connect(lambda: self._apply_internals_autofill(self.current_row, force=True))







        form_int.addRow(btn_auto_int)







        self.sp_luz_w = _spin(0,100000,100,2,0,"W")







        self.sp_luz_h = _spin(0,24,0.5,2,18,"h/dia")







        self.sp_motor_w = _spin(0,100000,100,2,0,"W")







        self.sp_motor_h = _spin(0,24,0.5,2,24,"h/dia")







        self.sp_mont_hp = _spin(0,200,0.1,2,0,"HP")







        self.sp_mont_h = _spin(0,24,0.5,2,2,"h/dia")







        self.sp_forklift = self.sp_mont_h







        self.sp_people = _spin(0,50,1,0,0,"#")







        self.sp_people_h = _spin(0,24,0.5,2,8,"h/dia")







        self.sp_people_btuh = _spin(0,5000,50,0,1350,"Btuh/persona")







        self.sp_desh_w = _spin(0,100000,100,2,0,"W")







        self.sp_desh_n = QSpinBox(); self.sp_desh_n.setRange(0,24); self.sp_desh_n.setValue(0)







        self.sp_desh_min = _spin(0,180,1,0,0,"min")







        self.sp_desh_frac = _spin(0,1,0.05,2,1.0,"")







        form_int.addRow("ILUMINACION (W):", self.sp_luz_w)







        form_int.addRow("HORAS ILUMINACION (H/DIA):", self.sp_luz_h)







        form_int.addRow("MOTORES (W):", self.sp_motor_w)







        form_int.addRow("HORAS MOTORES (H/DIA):", self.sp_motor_h)







        form_int.addRow("MONTACARGAS (HP):", self.sp_mont_hp)







        form_int.addRow("HORAS MONTACARGAS (H/DIA):", self.sp_mont_h)







        form_int.addRow("PERSONAS (#):", self.sp_people)







        form_int.addRow("HORAS PERSONAS (H/DIA):", self.sp_people_h)







        form_int.addRow("BTUH POR PERSONA:", self.sp_people_btuh)







        form_int.addRow("DESHIELO ELECTRICO (W):", self.sp_desh_w)







        form_int.addRow("N DESHIELOS/DIA:", self.sp_desh_n)







        form_int.addRow("DURACION DESHIELO (MIN):", self.sp_desh_min)







        form_int.addRow("FRACCION CALOR AL CUARTO:", self.sp_desh_frac)







        tabs.addTab(tab_int, "INTERNAS")







        # Avanzado: PROCESO







        tab_proc = QWidget()







        lay_proc = QVBoxLayout(tab_proc)







        lay_proc.setContentsMargins(0, 0, 0, 0)







        btn_auto_proc = QPushButton("APLICAR TÍPICOS")







        btn_auto_proc.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)







        btn_auto_proc.setStyleSheet("QPushButton{background:#e0f2fe;color:#0f172a;border-radius:8px;padding:6px 10px;}")







        btn_auto_proc.clicked.connect(lambda: self._apply_process_autofill(self.current_row, force=True))







        lay_proc.addWidget(btn_auto_proc)







        form_proc = QFormLayout()







        lay_proc.addLayout(form_proc)







        self.sp_tin = _spin(-60,50,0.5,2,0); self.sp_tin.setToolTip("Temperatura ingreso producto")







        self.sp_tout = _spin(-60,50,0.5,2,0); self.sp_masa = _spin(0,100000,10,2,0); self.sp_ciclo = _spin(0.1,168,0.5,2,24)







        for sp in (self.sp_tin, self.sp_tout, self.sp_masa, self.sp_ciclo):







            sp.valueChanged.connect(lambda _=None: self.update_process_outputs(self.current_row))







            sp.valueChanged.connect(lambda _=None: self._schedule_calc())







        kg_row = QWidget()







        kg_row_layout = QHBoxLayout(kg_row)







        kg_row_layout.setContentsMargins(0, 0, 0, 0)







        kg_row_layout.setSpacing(6)







        kg_row_layout.addWidget(self.sp_masa)







        self.btn_estimar_kg = QPushButton("ESTIMAR KG")







        self.btn_estimar_kg.clicked.connect(lambda: self.estimate_process_kg(self.current_row))







        kg_row_layout.addWidget(self.btn_estimar_kg)







        kg_row_layout.addStretch()







        self.ro_carga_ref=_ro(); self.ro_carga_cong=_ro(); self.ro_carga_post=_ro(); self.ro_carga_total_btuc=_ro(); self.ro_carga_total_kwc=_ro()







        form_proc.addRow("TEMPERATURA INGRESO (C):", self.sp_tin)







        form_proc.addRow("TEMPERATURA SALIDA (C):", self.sp_tout)







        form_proc.addRow("CANTIDAD PRODUCTO (KG):", kg_row)







        form_proc.addRow("TIEMPO CICLO (H):", self.sp_ciclo)







        form_proc.addRow("CARGA REF (BTU/CICLO):", self.ro_carga_ref)







        form_proc.addRow("CARGA CONG (BTU/CICLO):", self.ro_carga_cong)







        form_proc.addRow("CARGA POST (BTU/CICLO):", self.ro_carga_post)







        form_proc.addRow("CARGA TRMICA (BTU/CICLO):", self.ro_carga_total_btuc)







        form_proc.addRow("CARGA TRMICA (KW/CICLO):", self.ro_carga_total_kwc)







        tabs.addTab(tab_proc, "PROCESO")







        # Avanzado: Infiltracin







        tab_inf = QWidget(); form_inf = QFormLayout(tab_inf)







        self.chk_infil_enabled = QCheckBox("Infiltracin habilitada"); self.chk_infil_enabled.setChecked(True)







        self.sp_taire = _spin(-60,60,0.5,2,30); self.sp_rh = _spin(0,100,1,1,60); self.sp_inside_rh=_spin(0,100,1,1,85); self.sp_recambios=_spin(0,50,0.5,2,0); self.sp_recambios.setSpecialValueText("Auto"); self.sp_usefactor=_spin(0,5,0.5,2,2)







        form_inf.addRow("T EXTERNA AIRE (C):", self.sp_taire)







        form_inf.addRow("HUMEDAD RELATIVA EXT (%):", self.sp_rh)







        form_inf.addRow("HUMEDAD RELATIVA INT (%):", self.sp_inside_rh)







        form_inf.addRow("RECAMBIOS 24H (OVERRIDE):", self.sp_recambios)







        form_inf.addRow("FACTOR DE USO:", self.sp_usefactor)







        form_inf.addRow(self.chk_infil_enabled)







        self.chk_infil_enabled.stateChanged.connect(lambda _=None: (self._refresh_infiltration_outputs(self.current_row), self._schedule_calc()))







        self.ro_infil_total = _ro(); self.ro_infil_sensible = _ro(); self.ro_infil_latente = _ro()







        self.lbl_infil_mode = QLabel(""); self.lbl_infil_mode.setStyleSheet("color:#475569;")







        form_inf.addRow("INF TOTAL (BTU/H):", self.ro_infil_total)







        form_inf.addRow("INF SENSIBLE (BTU/H):", self.ro_infil_sensible)







        form_inf.addRow("INF LATENTE (BTU/H):", self.ro_infil_latente)







        form_inf.addRow("MODO INF:", self.lbl_infil_mode)







        tabs.addTab(tab_inf, "INFILTRACIÓN")







        # Avanzado: AISLAMIENTO







        tab_ais = QWidget(); form_ais = QFormLayout(tab_ais)







        self.cb_AISLAMIENTO = QComboBox()







        for k in self.calc.insulation.keys():







            self.cb_AISLAMIENTO.addItem(k,k)







        self.sp_espesor = _spin(1,20,0.5,2,4)







        self.ro_u24 = _ro()







        form_ais.addRow("TIPO AISLAMIENTO:", self.cb_AISLAMIENTO)







        form_ais.addRow("ESPESOR (IN):", self.sp_espesor)







        form_ais.addRow("U24 (CALC):", self.ro_u24)







        self.cb_AISLAMIENTO.currentIndexChanged.connect(lambda _=None: self.update_insulation_outputs(self.current_row))







        self.sp_espesor.valueChanged.connect(lambda _=None: self.update_insulation_outputs(self.current_row))







        tabs.addTab(tab_ais, "AISLAMIENTO")







        # Avanzado: RESULTADOS







        tab_res = QWidget(); res_form = QVBoxLayout(tab_res)







        res_form.setContentsMargins(4, 4, 4, 4)







        res_form.setSpacing(8)







        sel_row = QHBoxLayout()







        sel_row.addWidget(QLabel("Cuarto:"))







        self.cb_detail_room = QComboBox()







        self.cb_detail_room.currentIndexChanged.connect(self._populate_detail_view)







        sel_row.addWidget(self.cb_detail_room)







        sel_row.addStretch()







        res_form.addLayout(sel_row)







        self.tree_detail = QTreeWidget()







        self.tree_detail.setColumnCount(4)







        self.tree_detail.setHeaderLabels(["CONCEPTO", "BTU/H", "KW", "%/VALOR"])







        header = self.tree_detail.header()







        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)







        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)







        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)







        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)







        header.setStretchLastSection(False)







        self.tree_detail.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)







        self.tree_detail.setMinimumHeight(420)







        self.tree_detail.setStyleSheet("QTreeWidget::item:selected{background:#e0f2fe;color:#0f172a;} QTreeView::item{text-align: right;}")







        self.tree_detail.setHeaderHidden(False)







        res_form.addWidget(self.tree_detail)







        tabs.addTab(tab_res, "RESULTADOS")







        # Conexiones de reclculo diferido para spinners clave







        auto_calc_spins = [







            self.sp_largo, self.sp_ancho, self.sp_alto, self.sp_tint,







            self.sp_tef, self.sp_teb, self.sp_ted, self.sp_tei, self.sp_tet,







            self.sp_ground, self.sp_transfer, self.sp_run_hours,







            self.sp_taire, self.sp_rh, self.sp_inside_rh, self.sp_recambios, self.sp_usefactor,







            self.sp_espesor,







            self.sp_luz_w, self.sp_luz_h, self.sp_motor_w, self.sp_motor_h,







            self.sp_mont_hp, self.sp_mont_h, self.sp_people, self.sp_people_h,







            self.sp_people_btuh, self.sp_desh_w, self.sp_desh_n,







            self.sp_desh_min, self.sp_desh_frac,







            self.sp_tin, self.sp_tout, self.sp_masa, self.sp_ciclo,







        ]







        for sp in auto_calc_spins:







            try:







                sp.valueChanged.connect(lambda _=None: self._schedule_calc())







            except Exception:







                pass







        if hasattr(self, "cb_AISLAMIENTO"):







            self.cb_AISLAMIENTO.currentIndexChanged.connect(lambda _=None: self._schedule_calc())







        # tabs visibles







        self.tabs = tabs







        vbox.addWidget(tabs)







        # Barra inferior (se eliminan botones de validar/calcular)







        bottom_btns = QHBoxLayout()







        bottom_btns.addStretch()







        vbox.addLayout(bottom_btns)







        return container







    def _populate_breakdown(self, rooms):







        comps = {







            "TRANSMISIN": 0.0,







            "INFILTRACIN": 0.0,







            "ILUMINACIN": 0.0,







            "Motores": 0.0,







            "Montacargas": 0.0,







            "Personas": 0.0,







            "Deshielo": 0.0,







            "Producto": 0.0,







        }







        for r in rooms:







            c = r.components







            comps["TRANSMISIN"] += c.transmission_btuh







            comps["INFILTRACIN"] += c.infiltration_btuh







            comps["ILUMINACIN"] += c.lighting_btuh







            comps["Motores"] += c.motors_btuh







            comps["Montacargas"] += c.forklift_btuh







            comps["Personas"] += c.people_btuh







            comps["Deshielo"] += c.defrost_btuh







            comps["Producto"] += c.product_btuh







        total_btuh = sum(comps.values())







        run_hours = rooms[0].inputs.run_hours_supp if rooms else 24







        self.tbl_break.setRowCount(len(comps)+1)







        for i, (name, val) in enumerate(comps.items()):







            btuday = val * 24 / (run_hours if run_hours else 24)







            pct = (val/total_btuh*100) if total_btuh>0 else 0







            self.tbl_break.setItem(i,0,QTableWidgetItem(name))







            self.tbl_break.setItem(i,1,QTableWidgetItem(f"{btuday:,.2f}".replace(",", ".")))







            self.tbl_break.setItem(i,2,QTableWidgetItem(f"{val:,.2f}".replace(",", ".")))







            self.tbl_break.setItem(i,3,QTableWidgetItem(f"{pct:,.2f}".replace(",", ".")))







        # total







        self.tbl_break.setItem(len(comps),0,QTableWidgetItem("TOTAL"))







        self.tbl_break.setItem(len(comps),1,QTableWidgetItem(""))







        self.tbl_break.setItem(len(comps),2,QTableWidgetItem(f"{total_btuh:,.2f}".replace(",", ".")))







        self.tbl_break.setItem(len(comps),3,QTableWidgetItem("100.00"))







    def _calc_room1_only(self):







        rooms = self._gather_inputs()







        if not rooms:







            self._append_log("No hay cuartos vlidos para calcular.")







            return







        proj = self.calc.compute_project(rooms[:1], self.sp_sf.value())







        self._append_log(f"Cuarto 1 BTU/h: {proj.total_btuh:,.2f}")







        self.lbl_state.setText("Listo")







        self.lbl_state.setStyleSheet("color:#16a34a;font-weight:700;")







    def _load_golden(self):







        try:







            res, exp = self.calc.run_golden_case()







        except Exception as e:







            QMessageBox.critical(self, "Golden", str(e)); return







        # poner golden en cuarto 1







        self.sp_rooms.setValue(1)







        self._build_rows(1)







        r = res.inputs







        self.tbl.item(0,0).setText("GOLDEN")







        self.tbl.cellWidget(0,1).setValue(r.largo_m)







        self.tbl.cellWidget(0,2).setValue(r.ancho_m)







        self.tbl.cellWidget(0,3).setValue(r.altura_m)







        combo: QComboBox = self.tbl.cellWidget(0,4)







        idx = combo.findData(r.perfil_id)







        if idx>=0: combo.setCurrentIndex(idx)







        self.tbl.cellWidget(0,5).setValue(r.puertas)







        self._last_inputs = [r]







        self._last_results = [res]







        self._append_log(f"Golden cargado. Target BTU/h {exp['total_Btuh_excel']:.2f}")







        self._populate_breakdown([res])







        self.lbl_total_btuh.setText(f"{res.total_btuh:,.2f}".replace(",", "."))







        self.lbl_total_kw.setText(f"{res.total_kw:,.2f}".replace(",", "."))







        self.lbl_total_tr.setText(f"{res.total_tr:,.2f}".replace(",", "."))







        self.lbl_state.setText("Golden")







        self.lbl_state.setStyleSheet("color:#2563eb;font-weight:700;")







    def _select_row_from_selection(self):







        idxs = self.tbl.selectionModel().selectedRows()







        if not idxs:







            return







        self._select_row(idxs[0].row())







    def _show_json_viewer(self, data_obj):







        dlg = QDialog(self)







        dlg.setWindowTitle("JSON")







        layout = QVBoxLayout(dlg)







        txt = QTextEdit()







        txt.setReadOnly(True)







        try:







            import json







            txt.setPlainText(json.dumps(data_obj, default=lambda o: o.__dict__, ensure_ascii=False, indent=2))







        except Exception as e:







            txt.setPlainText(str(e))







        layout.addWidget(txt)







        dlg.resize(600,400)







        dlg.exec()







    def _manual_calc(self):







        try:







            self._save_editor_to_room()







            self._calculate()







        except Exception as exc:







            self._append_log(f"Manual calc error: {exc}")
