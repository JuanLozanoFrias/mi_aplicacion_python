from __future__ import annotations

from pathlib import Path
import json
import os
import tempfile
from datetime import datetime
from typing import Any, Callable, Dict, List
import traceback

from PySide6.QtCore import Qt, QAbstractTableModel, QModelIndex, QTimer
from PySide6.QtGui import QGuiApplication, QKeySequence, QShortcut, QIntValidator, QColor
from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QLabel,
    QPushButton,
    QMessageBox,
    QHBoxLayout,
    QSizePolicy,
    QTableView,
    QAbstractScrollArea,
    QAbstractItemView,
    QStyledItemDelegate,
    QComboBox,
    QLineEdit,
    QGridLayout,
    QFileDialog,
    QSpinBox,
    QDoubleSpinBox,
    QInputDialog,
    QScrollArea,
    QHeaderView,
    QMenu,
    QDialog,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QGroupBox,
)

try:
    from logic.legend_jd import LegendJDService
    from logic.legend_jd.project_service import LegendProjectService
    from logic.legend_jd.legend_exporter import build_legend_workbook, restore_template_assets
    from logic.legend.eev_calc import compute_eev
except Exception as exc:  # pragma: no cover
    LegendJDService = None  # type: ignore
    LegendProjectService = None  # type: ignore
    build_legend_workbook = None  # type: ignore
    restore_template_assets = None  # type: ignore
    compute_eev = None  # type: ignore
    _IMPORT_ERROR = exc
else:
    _IMPORT_ERROR = None

try:
    from logic.cuartos_frios_engine import ColdRoomEngine, ColdRoomInputs
except Exception:
    ColdRoomEngine = None  # type: ignore
    ColdRoomInputs = None  # type: ignore

from gui.pages.cargas.biblioteca_legend import BibliotecaLegendDialog


PROJ_COLUMNS = [
    "loop",
    "ramal",
    "equipo",
    "uso",
    "largo_m",
    "ancho_m",
    "alto_m",
    "dim_ft",
    "btu_ft",
    "btu_hr",
    "evap_modelo",
    "evap_qty",
    "familia",
]


class LegendPreviewDialog(QDialog):
    def __init__(self, wb, parent=None):
        super().__init__(parent)
        self.setWindowTitle("VISTA PREVIA LEGEND")
        self.resize(1100, 700)
        layout = QVBoxLayout(self)
        self.table = QTableWidget()
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionMode(QAbstractItemView.NoSelection)
        self.table.setShowGrid(True)
        layout.addWidget(self.table, 1)

        ws = wb.active
        try:
            from openpyxl.utils import range_boundaries, get_column_letter
        except Exception:
            return
        dim = ws.calculate_dimension()
        min_col, min_row, max_col, max_row = range_boundaries(dim)
        rows = max_row - min_row + 1
        cols = max_col - min_col + 1
        self.table.setRowCount(rows)
        self.table.setColumnCount(cols)

        # tamaños de columnas/filas
        for c in range(min_col, max_col + 1):
            letter = get_column_letter(c)
            width = ws.column_dimensions[letter].width
            if width:
                self.table.setColumnWidth(c - min_col, int(width * 7 + 12))
        for r in range(min_row, max_row + 1):
            height = ws.row_dimensions[r].height
            if height:
                self.table.setRowHeight(r - min_row, int(height * 1.33))

        # merges
        for merge in ws.merged_cells.ranges:
            if merge.min_row < min_row or merge.max_row > max_row:
                continue
            r = merge.min_row - min_row
            c = merge.min_col - min_col
            self.table.setSpan(r, c, merge.max_row - merge.min_row + 1, merge.max_col - merge.min_col + 1)

        # celdas
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(r, c)
                text = "" if cell.value is None else str(cell.value)
                item = QTableWidgetItem(text)

                # fuente
                if cell.font:
                    f = item.font()
                    f.setBold(bool(cell.font.bold))
                    f.setItalic(bool(cell.font.italic))
                    if cell.font.sz:
                        try:
                            f.setPointSize(int(cell.font.sz))
                        except Exception:
                            pass
                    item.setFont(f)
                    if cell.font.color and cell.font.color.type == "rgb":
                        try:
                            item.setForeground(QColor(f"#{cell.font.color.rgb[-6:]}"))
                        except Exception:
                            pass

                # fondo
                if cell.fill and cell.fill.patternType == "solid" and cell.fill.fgColor and cell.fill.fgColor.type == "rgb":
                    try:
                        item.setBackground(QColor(f"#{cell.fill.fgColor.rgb[-6:]}"))
                    except Exception:
                        pass

                # alineación
                align = cell.alignment
                if align and align.horizontal:
                    h = align.horizontal
                    if h == "center":
                        item.setTextAlignment(Qt.AlignCenter)
                    elif h == "right":
                        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    else:
                        item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)

                self.table.setItem(r - min_row, c - min_col, item)


class LegendPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.service = LegendJDService() if LegendJDService else None
        self.project_dir: Path | None = None
        self.project_data: Dict[str, Any] = {}
        self._dirty = False
        self._suspend_updates = False
        self._loading_project = False
        self.total_bt = 0.0
        self.total_mt = 0.0
        self._equipos_bt: List[str] = []
        self._equipos_mt: List[str] = []
        self._usos_bt: List[str] = []
        self._usos_mt: List[str] = []
        self._equipos_btu_ft: Dict[str, float] = {}
        self._familia_options = ["AUTO", "FRONTAL BAJA", "FRONTAL MEDIA", "DUAL"]
        self._cold_engine = None
        self._comp_perf = self._load_compresores_perf()
        self._comp_brands = sorted(self._comp_perf.get("brands", {}).keys())

        outer = QVBoxLayout(self)
        outer.setContentsMargins(12, 12, 12, 12)
        outer.setSpacing(6)

        title = QLabel("LEGEND - PROYECTO (EDITABLE)")
        title.setStyleSheet("font-size:18px;font-weight:800;color:#0f172a;")
        outer.addWidget(title)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        outer.addWidget(scroll, 1)

        content = QWidget()
        content.setMinimumWidth(0)
        content.setSizePolicy(QSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred))
        scroll.setWidget(content)

        proj_outer = QVBoxLayout(content)
        proj_outer.setContentsMargins(4, 4, 4, 4)
        proj_outer.setSpacing(6)

        self.btn_export_project = QPushButton("EXPORTAR PROYECTO")
        self.btn_library = QPushButton("BIBLIOTECA")
        self._proj_folder_full = ""
        self.lbl_proj_folder = QLabel("CARPETA: --")
        self.lbl_proj_folder.setToolTip("")
        self.lbl_proj_folder.setMinimumWidth(0)
        self.lbl_proj_folder.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.lbl_total_general = QLabel("TOTAL GENERAL: 0.0")
        proj_hdr = QHBoxLayout()
        proj_hdr.addWidget(self.btn_export_project)
        proj_hdr.addWidget(self.btn_library)
        proj_hdr.addWidget(self.lbl_proj_folder, 1)
        proj_hdr.addWidget(self.lbl_total_general)
        proj_outer.addLayout(proj_hdr)

        # Specs form
        self.spec_fields: Dict[str, QWidget] = {}
        specs_layout = QGridLayout()
        specs_layout.setHorizontalSpacing(8)
        specs_layout.setVerticalSpacing(4)
        label_width = 220
        specs_layout.setColumnMinimumWidth(0, label_width)
        specs_layout.setColumnMinimumWidth(2, label_width)
        specs_layout.setColumnStretch(1, 1)
        specs_layout.setColumnStretch(3, 1)

        row_idx = 0

        def _label(text: str) -> QLabel:
            lbl = QLabel(text)
            lbl.setWordWrap(False)
            lbl.setFixedWidth(label_width)
            lbl.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Preferred)
            return lbl

        def _expand(widget: QWidget) -> None:
            widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        def _add_single(label_text: str, widget: QWidget) -> None:
            nonlocal row_idx
            _expand(widget)
            specs_layout.addWidget(_label(label_text), row_idx, 0)
            specs_layout.addWidget(widget, row_idx, 1, 1, 3)
            row_idx += 1

        def _add_double(label1: str, w1: QWidget, label2: str, w2: QWidget) -> None:
            nonlocal row_idx
            _expand(w1)
            _expand(w2)
            specs_layout.addWidget(_label(label1), row_idx, 0)
            specs_layout.addWidget(w1, row_idx, 1)
            lbl2 = _label(label2)
            specs_layout.addWidget(lbl2, row_idx, 2)
            specs_layout.addWidget(w2, row_idx, 3)
            row_idx += 1

        # PROYECTO
        proj_edit = QLineEdit()
        proj_edit.textEdited.connect(lambda txt, w=proj_edit: self._on_spec_changed_upper("proyecto", w, txt))
        self.spec_fields["proyecto"] = proj_edit
        _add_single("PROYECTO", proj_edit)

        # CIUDAD
        ciudad_edit = QLineEdit()
        ciudad_edit.textEdited.connect(lambda txt, w=ciudad_edit: self._on_spec_changed_upper("ciudad", w, txt))
        self.spec_fields["ciudad"] = ciudad_edit
        _add_single("CIUDAD", ciudad_edit)

        # VENDEDOR
        vendedor_edit = QLineEdit()
        vendedor_edit.textEdited.connect(lambda txt, w=vendedor_edit: self._on_spec_changed_upper("vendedor", w, txt))
        self.spec_fields["vendedor"] = vendedor_edit
        _add_single("VENDEDOR", vendedor_edit)

        # TIPO DE SISTEMA + DISTRIBUCION TUBERIA
        tipo_cb = QComboBox()
        tipo_cb.addItems(["", "RACK", "WATERLOOP"])
        tipo_cb.setEditable(False)
        tipo_cb.currentTextChanged.connect(lambda _t: self._on_spec_changed("tipo_sistema"))
        self.spec_fields["tipo_sistema"] = tipo_cb
        distrib_cb = QComboBox()
        distrib_cb.addItems(["", "AMERICANA", "LOOP"])
        distrib_cb.setEditable(False)
        distrib_cb.currentTextChanged.connect(lambda _t: self._on_spec_changed("distribucion_tuberia"))
        self.spec_fields["distribucion_tuberia"] = distrib_cb
        _add_double("TIPO DE SISTEMA", tipo_cb, "DISTRIBUCION TUBERIA", distrib_cb)

        # TCOND F/C
        self.tcond_f_edit = QLineEdit()
        self.tcond_c_edit = QLineEdit()
        self.tcond_f_edit.setPlaceholderText("F")
        self.tcond_c_edit.setPlaceholderText("C")
        self.tcond_f_edit.textEdited.connect(lambda txt: self._on_temp_changed("f", txt))
        self.tcond_c_edit.textEdited.connect(lambda txt: self._on_temp_changed("c", txt))
        self.spec_fields["tcond_f"] = self.tcond_f_edit
        self.spec_fields["tcond_c"] = self.tcond_c_edit
        _add_double("TCOND (F)", self.tcond_f_edit, "TCOND (C)", self.tcond_c_edit)

        # Voltajes en la misma fila
        self.voltaje_principal_cb = QComboBox()
        self.voltaje_principal_cb.addItems(["", "460", "400", "220", "120"])
        self.voltaje_principal_cb.setEditable(False)
        self.voltaje_principal_cb.currentTextChanged.connect(lambda _t: self._on_spec_changed("voltaje_principal"))
        self.voltaje_control_cb = QComboBox()
        self.voltaje_control_cb.addItems(["", "220", "120", "24"])
        self.voltaje_control_cb.setEditable(False)
        self.voltaje_control_cb.currentTextChanged.connect(lambda _t: self._on_spec_changed("voltaje_control"))
        self.spec_fields["voltaje_principal"] = self.voltaje_principal_cb
        self.spec_fields["voltaje_control"] = self.voltaje_control_cb
        _add_double("VOLTAJE PRINCIPAL (V AC)", self.voltaje_principal_cb, "VOLTAJE CONTROL (V AC)", self.voltaje_control_cb)

        # Refrigerante + Controlador
        self.refrigerante_cb = QComboBox()
        self.refrigerante_cb.addItems(["", "R22", "R404", "R449", "R507N", "R290", "R744"])
        self.refrigerante_cb.setEditable(False)
        self.controlador_cb = QComboBox()
        self.controlador_cb.addItems(["", "AKSM 800", "AKPC 782", "URACK", "PRACK", "CPC 300", "COPELAND 1015"])
        self.controlador_cb.setEditable(False)
        self.refrigerante_cb.currentTextChanged.connect(lambda _t: self._on_spec_changed("refrigerante"))
        self.controlador_cb.currentTextChanged.connect(lambda _t: self._on_spec_changed("controlador"))
        self.spec_fields["refrigerante"] = self.refrigerante_cb
        self.spec_fields["controlador"] = self.controlador_cb
        _add_double("REFRIGERANTE", self.refrigerante_cb, "CONTROLADOR", self.controlador_cb)

        # Deshielos + Expansion
        self.deshielos_cb = QComboBox()
        self.deshielos_cb.addItems(["", "ELECTRICO", "GAS CALIENTE", "GAS TIBIO"])
        self.deshielos_cb.setEditable(False)
        self.expansion_cb = QComboBox()
        self.expansion_cb.addItems(["", "TERMOSTATICA", "ELECTRONICA"])
        self.expansion_cb.setEditable(False)
        self.deshielos_cb.currentTextChanged.connect(lambda _t: self._on_spec_changed("deshielos"))
        self.expansion_cb.currentTextChanged.connect(lambda _t: self._on_spec_changed("expansion"))
        self.spec_fields["deshielos"] = self.deshielos_cb
        self.spec_fields["expansion"] = self.expansion_cb
        _add_double("DESHIELOS", self.deshielos_cb, "EXPANSION", self.expansion_cb)

        specs_widget = QWidget()
        specs_widget.setLayout(specs_layout)
        proj_outer.addWidget(QLabel("ESPECIFICACIONES TECNICAS"))
        proj_outer.addWidget(specs_widget)

        self.btn_export_project.clicked.connect(self._export_project)
        self.btn_library.clicked.connect(self._open_library)

        self.bt_model = LegendItemsTableModel([])
        self.mt_model = LegendItemsTableModel([], ramal_offset_cb=self._get_mt_ramal_offset)
        self.bt_model.set_limit_callback(self._on_dim_limit)
        self.mt_model.set_limit_callback(self._on_dim_limit)
        self._init_cold_engine()
        self._apply_usage_map()

        self.bt_view = QTableView()
        self.bt_view.setModel(self.bt_model)
        self.bt_view.setMinimumWidth(0)
        self.bt_view.setSizePolicy(QSizePolicy(QSizePolicy.Ignored, QSizePolicy.Fixed))
        self.bt_view.setSizeAdjustPolicy(QAbstractScrollArea.AdjustIgnored)
        self.bt_view.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.bt_view.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.bt_view.setFocusPolicy(Qt.StrongFocus)
        self.bt_view.setEditTriggers(QAbstractItemView.SelectedClicked | QAbstractItemView.EditKeyPressed | QAbstractItemView.AnyKeyPressed)
        self.bt_view.setStyleSheet(
            "QTableView::item:selected { background: #eaf2ff; color: #0f172a; }"
        )
        self.bt_view.horizontalHeader().setStretchLastSection(True)
        self.bt_view.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.bt_view.verticalHeader().setVisible(False)
        self.bt_view.verticalHeader().setDefaultSectionSize(28)
        self.bt_view.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.bt_view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.bt_view.customContextMenuRequested.connect(
            lambda pos, v=self.bt_view, m=self.bt_model: self._show_table_menu(v, m, "bt", pos)
        )
        self.bt_model.modelReset.connect(self._recalc_totals)
        self.bt_model.dataChanged.connect(self._mark_dirty_and_totals)
        self.bt_model.rowsInserted.connect(self._mark_dirty_and_totals)
        self.bt_model.rowsRemoved.connect(self._mark_dirty_and_totals)
        self.bt_model.modelReset.connect(lambda *_: self._apply_dim_spans(self.bt_view, self.bt_model))
        self.bt_model.rowsInserted.connect(lambda *_: self._apply_dim_spans(self.bt_view, self.bt_model))
        self.bt_model.rowsRemoved.connect(lambda *_: self._apply_dim_spans(self.bt_view, self.bt_model))
        self.bt_model.dataChanged.connect(lambda *_: self._apply_dim_spans(self.bt_view, self.bt_model))
        self.bt_view.keyPressEvent = lambda event, v=self.bt_view: self._table_key_press(event, self.bt_model, v)
        self._install_shortcuts(self.bt_view, self.bt_model)
        self.bt_model.rowsInserted.connect(lambda *_: self._ensure_combo_editors(self.bt_view))
        self.bt_model.modelReset.connect(lambda *_: self._ensure_combo_editors(self.bt_view))

        self.mt_view = QTableView()
        self.mt_view.setModel(self.mt_model)
        self.mt_view.setMinimumWidth(0)
        self.mt_view.setSizePolicy(QSizePolicy(QSizePolicy.Ignored, QSizePolicy.Fixed))
        self.mt_view.setSizeAdjustPolicy(QAbstractScrollArea.AdjustIgnored)
        self.mt_view.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.mt_view.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.mt_view.setFocusPolicy(Qt.StrongFocus)
        self.mt_view.setEditTriggers(QAbstractItemView.SelectedClicked | QAbstractItemView.EditKeyPressed | QAbstractItemView.AnyKeyPressed)
        self.mt_view.setStyleSheet(
            "QTableView::item:selected { background: #eaf2ff; color: #0f172a; }"
        )
        self.mt_view.horizontalHeader().setStretchLastSection(True)
        self.mt_view.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.mt_view.verticalHeader().setVisible(False)
        self.mt_view.verticalHeader().setDefaultSectionSize(28)
        self.mt_view.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.mt_view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.mt_view.customContextMenuRequested.connect(
            lambda pos, v=self.mt_view, m=self.mt_model: self._show_table_menu(v, m, "mt", pos)
        )
        self.mt_model.modelReset.connect(self._recalc_totals)
        self.mt_model.dataChanged.connect(self._mark_dirty_and_totals)
        self.mt_model.rowsInserted.connect(self._mark_dirty_and_totals)
        self.mt_model.rowsRemoved.connect(self._mark_dirty_and_totals)
        self.mt_model.modelReset.connect(lambda *_: self._apply_dim_spans(self.mt_view, self.mt_model))
        self.mt_model.rowsInserted.connect(lambda *_: self._apply_dim_spans(self.mt_view, self.mt_model))
        self.mt_model.rowsRemoved.connect(lambda *_: self._apply_dim_spans(self.mt_view, self.mt_model))
        self.mt_model.dataChanged.connect(lambda *_: self._apply_dim_spans(self.mt_view, self.mt_model))
        self.mt_view.keyPressEvent = lambda event, v=self.mt_view: self._table_key_press(event, self.mt_model, v)
        self._install_shortcuts(self.mt_view, self.mt_model)
        self.mt_model.rowsInserted.connect(lambda *_: self._ensure_combo_editors(self.mt_view))
        self.mt_model.modelReset.connect(lambda *_: self._ensure_combo_editors(self.mt_view))

        # Delegates para equipos/usos desde data/LEGEND
        self._equipos_bt_delegate = ComboDelegate(lambda: self._equipos_bt, editable=True, parent=self)
        self._equipos_mt_delegate = ComboDelegate(lambda: self._equipos_mt, editable=True, parent=self)
        self._usos_bt_delegate = ComboDelegate(lambda: self._usos_bt, editable=True, parent=self)
        self._usos_mt_delegate = ComboDelegate(lambda: self._usos_mt, editable=True, parent=self)
        self._familia_delegate = ComboDelegate(lambda: self._familia_options, editable=False, parent=self)
        self._dim_delegate = DimDelegate(parent=self)
        col_equipo = PROJ_COLUMNS.index("equipo")
        col_uso = PROJ_COLUMNS.index("uso")
        col_largo = PROJ_COLUMNS.index("largo_m")
        col_familia = PROJ_COLUMNS.index("familia")
        self.bt_view.setItemDelegateForColumn(col_equipo, self._equipos_bt_delegate)
        self.mt_view.setItemDelegateForColumn(col_equipo, self._equipos_mt_delegate)
        self.bt_view.setItemDelegateForColumn(col_uso, self._usos_bt_delegate)
        self.mt_view.setItemDelegateForColumn(col_uso, self._usos_mt_delegate)
        self.bt_view.setItemDelegateForColumn(col_familia, self._familia_delegate)
        self.mt_view.setItemDelegateForColumn(col_familia, self._familia_delegate)
        self.bt_view.setItemDelegateForColumn(col_largo, self._dim_delegate)
        self.mt_view.setItemDelegateForColumn(col_largo, self._dim_delegate)
        self._ensure_combo_editors(self.bt_view)
        self._ensure_combo_editors(self.mt_view)
        self._apply_dim_spans(self.bt_view, self.bt_model)
        self._apply_dim_spans(self.mt_view, self.mt_model)
        # Ramales y tablas BT
        self.spin_bt = QSpinBox()
        self.spin_bt.setRange(1, 20)
        self.spin_bt.setEnabled(True)
        self.btn_apply_bt = QPushButton("APLICAR")
        self.btn_apply_bt.clicked.connect(lambda: self._on_bt_ramales_changed(self.spin_bt.value()))

        proj_outer.addWidget(QLabel("BAJA (BT)"))
        ram_bt_layout = QHBoxLayout()
        ram_bt_layout.addWidget(QLabel("RAMALES BT:"))
        ram_bt_layout.addWidget(self.spin_bt)
        ram_bt_layout.addWidget(self.btn_apply_bt)
        ram_bt_layout.addSpacing(12)
        ram_bt_layout.addStretch(1)
        proj_outer.addLayout(ram_bt_layout)

        # BT fila controls removed per request
        proj_outer.addWidget(self.bt_view)
        self.lbl_total_bt = QLabel("TOTAL BT: 0.0")
        proj_outer.addWidget(self.lbl_total_bt)

        # Ramales y tablas MT
        self.spin_mt = QSpinBox()
        self.spin_mt.setRange(1, 20)
        self.spin_mt.setEnabled(True)
        self.btn_apply_mt = QPushButton("APLICAR")
        self.btn_apply_mt.clicked.connect(lambda: self._on_mt_ramales_changed(self.spin_mt.value()))

        proj_outer.addWidget(QLabel("MEDIA (MT)"))
        ram_mt_layout = QHBoxLayout()
        ram_mt_layout.addWidget(QLabel("RAMALES MT:"))
        ram_mt_layout.addWidget(self.spin_mt)
        ram_mt_layout.addWidget(self.btn_apply_mt)
        ram_mt_layout.addSpacing(12)
        ram_mt_layout.addStretch(1)
        proj_outer.addLayout(ram_mt_layout)

        proj_outer.addWidget(self.mt_view)
        self.lbl_total_mt = QLabel("TOTAL MT: 0.0")
        proj_outer.addWidget(self.lbl_total_mt)

        # COMPRESORES
        self.comp_group = QGroupBox("COMPRESORES")
        comp_layout = QVBoxLayout()
        comp_layout.setContentsMargins(8, 8, 8, 8)
        comp_layout.setSpacing(6)
        brand_row = QHBoxLayout()
        brand_row.addWidget(QLabel("MARCA COMPRESORES"))
        self.comp_brand_cb = QComboBox()
        self.comp_brand_cb.addItem("")
        self.comp_brand_cb.addItems(self._comp_brands)
        brand_row.addWidget(self.comp_brand_cb)
        brand_row.addStretch(1)
        comp_layout.addLayout(brand_row)
        self.comp_bt = self._build_compressor_block("BAJA (BT)")
        self.comp_mt = self._build_compressor_block("MEDIA (MT)")
        comp_layout.addWidget(self.comp_bt["box"])
        comp_layout.addWidget(self.comp_mt["box"])
        self.comp_group.setLayout(comp_layout)
        proj_outer.addWidget(self.comp_group)
        self.comp_brand_cb.currentTextChanged.connect(self._on_comp_brand_changed)

        self.eev_group = QGroupBox()
        self.eev_group.setTitle("")
        eev_layout = QVBoxLayout()
        eev_layout.setContentsMargins(8, 8, 8, 8)
        eev_layout.setSpacing(6)

        top_bar = QHBoxLayout()
        self.lbl_eev_title = QLabel("EEV - EXPANSION ELECTRONICA")
        self.lbl_eev_title.setStyleSheet("font-weight:600;")
        top_bar.addWidget(self.lbl_eev_title)
        top_bar.addStretch(1)
        self.lbl_eev_valves = QLabel("VALVULAS: 0")
        self.lbl_eev_ramales = QLabel("RAMALES CON EEV: 0")
        for _lbl in (self.lbl_eev_valves, self.lbl_eev_ramales):
            _lbl.setStyleSheet("padding:2px 8px; border-radius:8px; background:#eef5ff; color:#1f3b6d;")
        top_bar.addWidget(self.lbl_eev_valves)
        top_bar.addWidget(self.lbl_eev_ramales)
        top_bar.addStretch(1)
        self.lbl_eev_total_cost = QLabel("TOTAL COSTO EEV: --")
        self.lbl_eev_total_cost.setStyleSheet(
            "padding:4px 10px; border-radius:10px; background:#e8f5e9; color:#1b5e20; font-weight:600;"
        )
        top_bar.addWidget(self.lbl_eev_total_cost)
        eev_layout.addLayout(top_bar)

        action_row = QHBoxLayout()
        self.btn_eev_costs = QPushButton("COSTOS EEV...")
        self.btn_eev_costs.clicked.connect(self._open_eev_costs)
        self.lbl_eev_missing = QLabel("ITEMS SIN PRECIO: 0")
        self.lbl_eev_missing.setStyleSheet("color:#b45309;")
        self.lbl_eev_missing.setCursor(Qt.PointingHandCursor)
        self.lbl_eev_missing.mousePressEvent = self._on_eev_missing_clicked
        action_row.addWidget(self.btn_eev_costs)
        action_row.addSpacing(8)
        action_row.addWidget(self.lbl_eev_missing)
        action_row.addStretch(1)
        self.eev_search = QLineEdit()
        self.eev_search.setPlaceholderText("BUSCAR...")
        self.eev_search.textChanged.connect(self._apply_eev_filters)
        action_row.addWidget(self.eev_search)
        eev_layout.addLayout(action_row)

        self.eev_tabs = QTabWidget()
        self.eev_detail = QTableWidget(0, 10)
        self.eev_bom = QTableWidget(0, 6)
        self.eev_sets = QTableWidget(0, 5)
        self._setup_eev_table(
            self.eev_detail,
            [
                "SUCCION",
                "LOOP",
                "RAMAL",
                "EQUIPO",
                "USO",
                "CARGA (BTU/HR)",
                "TEVAP (F)",
                "FAMILIA (EEV)",
                "ORIFICIO",
                "MODELO",
            ],
        )
        self._setup_eev_table(
            self.eev_bom,
            ["MODELO", "DESCRIPCION", "CANTIDAD", "COSTO UNITARIO", "COSTO TOTAL", "MONEDA"],
        )
        self._setup_eev_table(
            self.eev_sets,
            ["CATEGORIA", "CANTIDAD", "COSTO UNITARIO", "COSTO TOTAL", "MONEDA"],
        )
        self.eev_tabs.addTab(self.eev_detail, "DETALLE")
        self.eev_tabs.addTab(self.eev_bom, "RESUMEN")
        self.eev_tabs.addTab(self.eev_sets, "PAQUETES")
        self.lbl_eev_warn = QLabel("")
        self.lbl_eev_warn.setStyleSheet("color:#C00;")
        eev_layout.addWidget(self.eev_tabs)
        eev_layout.addWidget(self.lbl_eev_warn)
        self.eev_group.setLayout(eev_layout)
        self.eev_group.setVisible(False)
        proj_outer.addWidget(self.eev_group)
        self._eev_missing_count = 0
        self._eev_filter_missing = False
        self._eev_highlight_model = ""
        try:
            self.eev_detail.itemSelectionChanged.connect(self._on_eev_detail_select)
            self.eev_bom.itemSelectionChanged.connect(self._on_eev_bom_select)
        except Exception:
            pass

        if _IMPORT_ERROR:
            QMessageBox.critical(self, "LEGEND", f"No se pudo importar LegendJDService:\n{_IMPORT_ERROR}")
        else:
            self.refresh()

    def _on_open_project(self) -> None:
        if self._dirty and not self._confirm_discard():
            return
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Abrir proyecto Legend",
            str(Path.cwd()),
            "Legend Project (*.json)",
        )
        if not file_path or not LegendProjectService:
            return
        self._load_project(Path(file_path))

    def _on_new_project(self) -> None:
        if self._dirty and not self._confirm_discard():
            return
        dir_path = QFileDialog.getExistingDirectory(self, "Selecciona carpeta para nuevo proyecto", str(Path.cwd()))
        if not dir_path or not LegendProjectService:
            return
        self.project_dir = Path(dir_path)
        self.project_data = LegendProjectService.load(self.project_dir)
        self._render_project()
        self._set_folder_label(self.project_dir)
        self._set_dirty(True)

    def _init_cold_engine(self) -> None:
        if not ColdRoomEngine:
            return
        try:
            data_path = Path("data/cuartos_frios/cuartos_frios_data.json")
            self._cold_engine = ColdRoomEngine(data_path)
            self.bt_model.set_cold_engine(self._cold_engine)
            self.mt_model.set_cold_engine(self._cold_engine)
        except Exception:
            self._cold_engine = None

    def _apply_usage_map(self) -> None:
        usage_map = {
            "IC": "HELADOS",
            "HELADO": "HELADOS",
            "HELADOS": "HELADOS",
            "CC": "COMIDA CONGELADA",
            "COMIDA CONGELADA": "COMIDA CONGELADA",
            "PESCADO": "COMIDA CONGELADA",
            "POLLO FRIZADO": "COMIDA CONGELADA",
            "CARNE": "CARNES",
            "CERDO": "CARNES",
            "POLLO": "CARNES",
            "CARNES": "CARNES",
            "BEBIDAS": "LACTEOS",
            "C. FRIAS": "LACTEOS",
            "C.FRIAS": "LACTEOS",
            "DELI": "LACTEOS",
            "LACTEOS": "LACTEOS",
            "PANADERIA": "LACTEOS",
            "VARIOS": "LACTEOS",
            "PREP": "PROCESO",
            "FRUVER": "PROCESO",
            "PROCESO": "PROCESO",
        }
        self.bt_model.set_usage_map(usage_map)
        self.mt_model.set_usage_map(usage_map)

    def _on_save_project(self) -> None:
        if not LegendProjectService:
            QMessageBox.warning(self, "LEGEND", "Servicio de proyecto no disponible.")
            return
        if not self.project_dir:
            dir_path = QFileDialog.getExistingDirectory(self, "Selecciona carpeta para guardar", str(Path.cwd()))
            if not dir_path:
                return
            self.project_dir = Path(dir_path)
            self._set_folder_label(self.project_dir)
        try:
            self._update_compressors()
            self.project_data = self._collect_project_data()
            LegendProjectService.save(self.project_dir, self.project_data)
            QMessageBox.information(self, "LEGEND", "GUARDADO OK.")
            self._set_dirty(False)
        except Exception as exc:
            QMessageBox.critical(self, "LEGEND", f"No se pudo guardar:\n{exc}")

    def _load_project(self, dir_path: Path) -> None:
        if not LegendProjectService:
            return
        try:
            if dir_path.is_file():
                self.project_dir = dir_path.parent
                self.project_data = LegendProjectService.load_file(dir_path)
                self._set_folder_label(self.project_dir)
            else:
                self.project_dir = dir_path
                self.project_data = LegendProjectService.load(dir_path)
                self._set_folder_label(dir_path)
            self._render_project()
            self._set_dirty(False)
        except Exception as exc:
            QMessageBox.critical(self, "LEGEND", f"No se pudo cargar proyecto:\n{exc}")

    def _collect_project_data(self) -> Dict[str, Any]:
        data = dict(self.project_data) if isinstance(self.project_data, dict) else {}
        data["bt_items"] = list(self.bt_model.items)
        data["mt_items"] = list(self.mt_model.items)
        specs = data.get("specs", {}) if isinstance(data, dict) else {}
        for k, widget in self.spec_fields.items():
            if isinstance(widget, QComboBox):
                specs[k] = widget.currentText()
            else:
                specs[k] = widget.text()
        data["specs"] = specs
        # compresores: leer directo de la UI para no perder selección
        brand = self.comp_brand_cb.currentText() if hasattr(self, "comp_brand_cb") else ""
        comp = dict(self.project_data.get("compressors", {}) or {}) if isinstance(self.project_data, dict) else {}
        bt_items = self._read_comp_model_rows(self.comp_bt) if hasattr(self, "comp_bt") else []
        mt_items = self._read_comp_model_rows(self.comp_mt) if hasattr(self, "comp_mt") else []
        bt_target = float(self.comp_bt["target"].value()) if hasattr(self, "comp_bt") else 25.0
        mt_target = float(self.comp_mt["target"].value()) if hasattr(self, "comp_mt") else 25.0
        comp["brand"] = brand
        comp["bt"] = {"items": bt_items, "target_reserva_pct": bt_target}
        comp["mt"] = {"items": mt_items, "target_reserva_pct": mt_target}
        data["compressors"] = comp
        return data

    def _set_folder_label(self, path: Path | str | None) -> None:
        self._proj_folder_full = str(path) if path else ""
        self._update_folder_label()

    def _update_folder_label(self) -> None:
        text = f"CARPETA: {self._proj_folder_full}" if self._proj_folder_full else "CARPETA: --"
        if not self.lbl_proj_folder:
            return
        fm = self.lbl_proj_folder.fontMetrics()
        available = self.lbl_proj_folder.width()
        if available and available > 30:
            elided = fm.elidedText(text, Qt.ElideMiddle, available)
            self.lbl_proj_folder.setText(elided)
        else:
            self.lbl_proj_folder.setText(text)
        self.lbl_proj_folder.setToolTip(self._proj_folder_full or "")
    def _slug(self, text: str) -> str:
        import re

        t = (text or "").strip().lower()
        t = re.sub(r"[^a-z0-9]+", "_", t)
        t = re.sub(r"_+", "_", t).strip("_")
        return t or "proyecto"

    def _on_preview_legend(self) -> None:
        if not build_legend_workbook:
            QMessageBox.warning(self, "LEGEND", "Exportador no disponible.")
            return
        if not self._has_any_legend_data():
            QMessageBox.information(self, "LEGEND", "NO HAY DATOS PARA VISUALIZAR.")
            return
        template_path = self._resolve_legend_template()
        if not template_path:
            QMessageBox.warning(self, "LEGEND", "No se encontró la plantilla legend_template.xlsx en data/legend/plantillas/")
            return
        try:
            wb = build_legend_workbook(template_path, self._collect_project_data())
            dlg = LegendPreviewDialog(wb, self)
            dlg.setWindowModality(Qt.ApplicationModal)
            dlg.raise_()
            dlg.activateWindow()
            dlg.exec()
        except Exception as exc:
            QMessageBox.critical(self, "LEGEND", f"No se pudo generar vista previa:\n{exc}")

    def _export_project(self) -> None:
        if not build_legend_workbook:
            QMessageBox.warning(self, "LEGEND", "Exportador no disponible.")
            return
        if not self._has_any_legend_data():
            QMessageBox.information(self, "LEGEND", "NO HAY DATOS PARA EXPORTAR.")
            return
        template_path = self._resolve_legend_template()
        if not template_path:
            QMessageBox.warning(self, "LEGEND", "No se encontr? la plantilla legend_template.xlsx en data/legend/plantillas/")
            return
        specs = self.project_data.get("specs", {}) if isinstance(self.project_data, dict) else {}
        proj_name = str(specs.get("proyecto", "") or "PROYECTO")
        stamp = datetime.now().strftime("%Y%m%d")
        base_name = f"{stamp}_{self._slug(proj_name)}"
        target_dir = QFileDialog.getExistingDirectory(self, "Selecciona carpeta destino", str(Path.cwd()))
        if not target_dir:
            return
        try:
            self._update_compressors()
            data = self._collect_project_data()
            wb = build_legend_workbook(template_path, data)
            tmp_fd, tmp_name = tempfile.mkstemp(suffix=".xlsx")
            os.close(tmp_fd)
            Path(tmp_name).unlink(missing_ok=True)
            tmp_path = Path(tmp_name)
            wb.save(tmp_path)
            if restore_template_assets:
                logo_path = template_path.parent / "logo.png"
                ok = restore_template_assets(template_path, tmp_path, logo_path if logo_path.exists() else None)
                if not ok:
                    QMessageBox.warning(self, "LEGEND", "No se pudo restaurar el logo del template.")
            excel_path = Path(target_dir) / f"{base_name}.xlsx"
            excel_path.write_bytes(tmp_path.read_bytes())

            lib_dir = Path(__file__).resolve().parents[3] / "data" / "proyectos" / "legend"
            lib_dir.mkdir(parents=True, exist_ok=True)
            json_path = lib_dir / f"{base_name}.json"
            json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

            QMessageBox.information(
                self,
                "LEGEND",
                f"EXPORTACION LISTA.\nExcel: {excel_path}\nJSON (biblioteca): {json_path}",
            )
        except Exception as exc:
            QMessageBox.critical(self, "LEGEND", f"No se pudo exportar:\n{exc}")

    def _open_library(self) -> None:
        lib_dir = Path(__file__).resolve().parents[3] / "data" / "proyectos" / "legend"
        danger_qss = (
            "QPushButton{background:#fdecec;color:#c53030;font-weight:700;border:none;border-radius:10px;padding:8px 12px;}"
            "QPushButton:hover{background:#fbd5d5;}"
        )
        dlg = BibliotecaLegendDialog(
            parent=self,
            library_dir=lib_dir,
            on_load=lambda p: self._load_project(p),
            primary_qss=self.btn_export_project.styleSheet(),
            danger_qss=danger_qss,
        )
        dlg.exec()

    def _resolve_legend_template(self) -> Path | None:
        candidates = [
            Path("data/LEGEND/plantillas/legend_template.xlsx"),
            Path("data/legend/plantillas/legend_template.xlsx"),
        ]
        for p in candidates:
            if p.exists():
                return p
        return None

    def _has_any_legend_data(self) -> bool:
        try:
            if self.bt_model.items or self.mt_model.items:
                return True
        except Exception:
            pass
        try:
            specs = self.project_data.get("specs", {}) if isinstance(self.project_data, dict) else {}
            return any(str(v).strip() for v in specs.values())
        except Exception:
            return False
    def _on_spec_changed(self, key: str) -> None:
        if not isinstance(self.project_data, dict):
            self.project_data = {}
        specs = self.project_data.get("specs", {}) if isinstance(self.project_data, dict) else {}
        widget = self.spec_fields.get(key)
        if isinstance(widget, QComboBox):
            specs[key] = widget.currentText()
        else:
            specs[key] = widget.text() if widget else ""
        self.project_data["specs"] = specs
        if key == "distribucion_tuberia":
            self._update_loop_column_visibility()
        if key in ("expansion", "refrigerante"):
            self._update_eev()
        if key in ("refrigerante", "tcond_f", "tcond_c"):
            self._update_compressors()
        self._set_dirty(True)

    def _on_spec_changed_upper(self, key: str, widget: QLineEdit, text: str) -> None:
        upper = text.upper()
        if upper != text:
            cursor = widget.cursorPosition()
            widget.setText(upper)
            widget.setCursorPosition(cursor)
        self._on_spec_changed(key)

    def _on_temp_changed(self, scale: str, text: str) -> None:
        try:
            val = float(text)
        except Exception:
            val = None
        if scale == "f":
            if val is not None:
                c = (val - 32) * 5.0 / 9.0
                self.tcond_c_edit.blockSignals(True)
                self.tcond_c_edit.setText(f"{c:.2f}")
                self.tcond_c_edit.blockSignals(False)
                self.project_data.setdefault("specs", {})["tcond_c"] = f"{c:.2f}"
            self.project_data.setdefault("specs", {})["tcond_f"] = text
        else:
            if val is not None:
                f = val * 9.0 / 5.0 + 32
                self.tcond_f_edit.blockSignals(True)
                self.tcond_f_edit.setText(f"{f:.2f}")
                self.tcond_f_edit.blockSignals(False)
                self.project_data.setdefault("specs", {})["tcond_f"] = f"{f:.2f}"
            self.project_data.setdefault("specs", {})["tcond_c"] = text
        self._set_dirty(True)

    def _render_project(self) -> None:
        self._loading_project = True
        bt_items = self.project_data.get("bt_items", []) if isinstance(self.project_data, dict) else []
        mt_items = self.project_data.get("mt_items", []) if isinstance(self.project_data, dict) else []
        specs = self.project_data.get("specs", {}) if isinstance(self.project_data, dict) else {}
        for k, widget in self.spec_fields.items():
            widget.blockSignals(True)
            val = str(specs.get(k, ""))
            if isinstance(widget, QComboBox):
                idx = widget.findText(val)
                widget.setCurrentIndex(idx if idx >= 0 else 0)
            else:
                widget.setText(val)
            widget.blockSignals(False)
        self._update_loop_column_visibility()
        f_val = specs.get("tcond_f")
        c_val = specs.get("tcond_c")
        if f_val and not c_val:
            try:
                f_num = float(f_val)
                c_calc = (f_num - 32) * 5.0 / 9.0
                self.tcond_c_edit.setText(f"{c_calc:.2f}")
            except Exception:
                pass
        if c_val and not f_val:
            try:
                c_num = float(c_val)
                f_calc = c_num * 9.0 / 5.0 + 32
                self.tcond_f_edit.setText(f"{f_calc:.2f}")
            except Exception:
                pass
        self.spin_bt.blockSignals(True)
        self.spin_mt.blockSignals(True)
        self.spin_bt.setValue(self._to_int(self.project_data.get("bt_ramales", 1)))
        self.spin_mt.setValue(self._to_int(self.project_data.get("mt_ramales", 1)))
        self.spin_bt.setEnabled(True)
        self.spin_mt.setEnabled(True)
        self.spin_bt.blockSignals(False)
        self.spin_mt.blockSignals(False)
        self.bt_model.set_items(bt_items)
        self.mt_model.set_items(mt_items)
        self.bt_model.recompute_all()
        self.mt_model.recompute_all()
        self._fit_table_to_contents_view(self.bt_view)
        self._fit_table_to_contents_view(self.mt_view)
        self._recalc_totals()
        self._render_compressors()
        self._loading_project = False
        self._update_compressors()

    def _add_row(self, model, view: QTableView) -> None:
        model.add_row({})
        self._fit_table_to_contents_view(view)
        bloque = self._block_for_model(model)
        if bloque:
            self._renumber_ramales(bloque)
        else:
            self._set_dirty(True)

    def _del_row(self, view: QTableView, model) -> None:
        sel = view.selectionModel().selectedRows()
        if not sel:
            QMessageBox.information(self, "LEGEND", "SELECCIONA UNA FILA PRIMERO.")
            return
        model.del_row(sel[0].row())
        self._fit_table_to_contents_view(view)
        self._set_dirty(True)

    def _dup_row(self, view: QTableView, model) -> None:
        sel = view.selectionModel().selectedRows()
        if not sel:
            QMessageBox.information(self, "LEGEND", "SELECCIONA UNA FILA PRIMERO.")
            return
        model.dup_row(sel[0].row())
        self._fit_table_to_contents_view(view)
        self._set_dirty(True)

    def _show_table_menu(self, view: QTableView, model: "LegendItemsTableModel", bloque: str, pos) -> None:
        index = view.indexAt(pos)
        if index.isValid():
            if not view.selectionModel().isRowSelected(index.row(), QModelIndex()):
                view.selectRow(index.row())
        menu = QMenu(view)
        act_add = menu.addAction("AGREGAR FILA")
        act_del = menu.addAction("ELIMINAR FILA")
        menu.addSeparator()
        act_copy = menu.addAction("COPIAR")
        act_paste = menu.addAction("PEGAR")
        action = menu.exec(view.mapToGlobal(pos))
        if action == act_add:
            self._add_row_after(view, model, bloque)
            self._renumber_ramales(bloque)
        elif action == act_del:
            self._delete_selected_rows(view, model, bloque)
            self._renumber_ramales(bloque)
        elif action == act_copy:
            self._copy_selection(model, view)
        elif action == act_paste:
            self._paste_selection_at(model, view)
            self._renumber_ramales(bloque)

    def _install_shortcuts(self, view: QTableView, model: "LegendItemsTableModel") -> None:
        if not hasattr(self, "_shortcuts"):
            self._shortcuts = []
        sc_copy = QShortcut(QKeySequence.Copy, view)
        sc_copy.setContext(Qt.WidgetWithChildrenShortcut)
        sc_copy.activated.connect(lambda m=model, v=view: self._copy_selection(m, v))
        self._shortcuts.append(sc_copy)
        sc_paste = QShortcut(QKeySequence.Paste, view)
        sc_paste.setContext(Qt.WidgetWithChildrenShortcut)
        sc_paste.activated.connect(lambda m=model, v=view: self._paste_selection(m, v))
        self._shortcuts.append(sc_paste)
        sc_dup = QShortcut(QKeySequence(Qt.CTRL | Qt.Key_D), view)
        sc_dup.setContext(Qt.WidgetWithChildrenShortcut)
        sc_dup.activated.connect(lambda m=model, v=view: self._dup_selection(m, v))
        self._shortcuts.append(sc_dup)

    def _add_row_after(self, view: QTableView, model: "LegendItemsTableModel", bloque: str) -> None:
        sel = view.selectionModel().selectedRows()
        insert_at = model.rowCount()
        ramal_val = 1
        if sel:
            row_idx = sel[0].row()
            insert_at = row_idx + 1
            try:
                ramal_val = self._to_int(model.items[row_idx].get("ramal", 1))
            except Exception:
                ramal_val = 1
        row = self._default_row(bloque)
        row["ramal"] = ramal_val
        model.insert_rows_at(insert_at, [row])
        self._fit_table_to_contents_view(view)
        self._set_dirty(True)

    def _delete_selected_rows(self, view: QTableView, model: "LegendItemsTableModel", bloque: str) -> None:
        sel = view.selectionModel().selectedRows()
        if not sel:
            QMessageBox.information(self, "LEGEND", "SELECCIONA UNA FILA PRIMERO.")
            return
        rows = sorted((i.row() for i in sel), reverse=True)
        for r in rows:
            model.del_row(r)
        self._fit_table_to_contents_view(view)
        self._set_dirty(True)

    def _paste_selection_at(self, model: "LegendItemsTableModel", view: QTableView) -> None:
        text = QGuiApplication.clipboard().text()
        if not text:
            return
        sel = view.selectionModel().selectedRows()
        insert_at = sel[0].row() if sel else model.rowCount()
        rows: List[dict] = []
        for line in text.splitlines():
            parts = line.split("\t")
            row = model.default_row.copy()
            for idx, key in enumerate(model.headers):
                if idx < len(parts):
                    row[key] = parts[idx]
            rows.append(row)
        model.insert_rows_at(insert_at, rows)
        self._fit_table_to_contents_view(view)
        bloque = self._block_for_model(model)
        if bloque:
            self._renumber_ramales(bloque)
        else:
            self._set_dirty(True)

    def _selected_row_indices(self, view: QTableView) -> List[int]:
        rows = sorted({i.row() for i in view.selectionModel().selectedRows()})
        if rows:
            return rows
        rows = sorted({i.row() for i in view.selectionModel().selectedIndexes()})
        if rows:
            return rows
        current = view.currentIndex()
        if current.isValid():
            return [current.row()]
        return []

    def _renumber_ramales(self, bloque: str) -> None:
        model = self.bt_model if bloque == "bt" else self.mt_model
        view = self.bt_view if bloque == "bt" else self.mt_view
        key_items = "bt_items" if bloque == "bt" else "mt_items"
        key_ramales = "bt_ramales" if bloque == "bt" else "mt_ramales"
        items = model.items
        row_count = model.rowCount()
        for idx, row in enumerate(items):
            if isinstance(row, dict):
                row["ramal"] = idx + 1
        if row_count > 0:
            col = PROJ_COLUMNS.index("ramal")
            top = model.index(0, col)
            bottom = model.index(row_count - 1, col)
            model.dataChanged.emit(top, bottom, [Qt.DisplayRole, Qt.EditRole])
        self.project_data[key_items] = model.items
        self.project_data[key_ramales] = max(row_count, 1)
        if bloque == "bt":
            self.spin_bt.blockSignals(True)
            self.spin_bt.setValue(self.project_data[key_ramales])
            self.spin_bt.blockSignals(False)
            self._refresh_mt_ramal_display()
        else:
            self.spin_mt.blockSignals(True)
            self.spin_mt.setValue(self.project_data[key_ramales])
            self.spin_mt.blockSignals(False)
        self._fit_table_to_contents_view(view)
        self._set_dirty(True)

    def _fit_table_to_contents_view(self, view: QTableView) -> None:
        model = view.model()
        header = view.horizontalHeader()
        if model:
            for col in range(model.columnCount()):
                header.setSectionResizeMode(col, QHeaderView.Interactive)
        view.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        view.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        view.setMinimumWidth(0)
        view.setSizePolicy(QSizePolicy(QSizePolicy.Ignored, QSizePolicy.Fixed))
        view.setAlternatingRowColors(True)
        self._apply_combo_column_widths(view)
        self._apply_dim_spans(view, model)
        try:
            view.resizeRowsToContents()
        except Exception:
            pass
        row_count = model.rowCount() if model else 0
        header_h = view.horizontalHeader().height()
        if model and row_count > 0:
            total_rows_h = sum(view.rowHeight(i) for i in range(row_count))
        else:
            total_rows_h = view.verticalHeader().defaultSectionSize()
        scroll_h = view.horizontalScrollBar().sizeHint().height()
        h = header_h + total_rows_h + view.frameWidth() * 2 + scroll_h
        view.setFixedHeight(h)
        try:
            view.scrollToTop()
        except Exception:
            pass

    def _setup_eev_table(self, table: QTableWidget, headers: List[str]) -> None:
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.verticalHeader().setVisible(False)
        table.verticalHeader().setDefaultSectionSize(26)
        table.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.setAlternatingRowColors(True)
        table.setTextElideMode(Qt.ElideRight)
        table.setWordWrap(False)
        table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        header = table.horizontalHeader()
        for idx in range(len(headers)):
            header.setSectionResizeMode(idx, QHeaderView.ResizeToContents)
        header.setStretchLastSection(True)
        table.setStyleSheet(
            "QTableWidget::item:selected{background:#e6f2ff; color:#000;}"
            "QTableWidget::item:hover{background:#f2f7ff;}"
            "QTableWidget{alternate-background-color:#fafbfd;}"
        )

    def _fit_eev_table_to_contents(self, table: QTableWidget) -> None:
        row_count = table.rowCount()
        header_h = table.horizontalHeader().height()
        row_h = table.verticalHeader().defaultSectionSize()
        total_rows_h = row_h * max(row_count, 1)
        h = header_h + total_rows_h + table.frameWidth() * 2 + 6
        table.setMinimumHeight(h)
        table.setMaximumHeight(h)

    def _fmt_int(self, val: Any) -> str:
        try:
            return f"{int(round(float(val))):,}"
        except Exception:
            return ""

    def _fmt_temp(self, val: Any) -> str:
        try:
            return f"{float(val):.1f}"
        except Exception:
            return ""

    def _fmt_cost(self, val: Any) -> str:
        try:
            return f"{float(val):,.2f}"
        except Exception:
            return ""

    def _apply_eev_filters(self) -> None:
        term = self.eev_search.text().strip().upper() if hasattr(self, "eev_search") else ""
        self._filter_table(self.eev_detail, term, [3, 4, 9], False)
        self._filter_table(self.eev_sets, term, [0], False)
        self._filter_table(self.eev_bom, term, [0, 1], self._eev_filter_missing)

    def _filter_table(self, table: QTableWidget, term: str, cols: List[int], missing_only: bool) -> None:
        for r in range(table.rowCount()):
            if missing_only:
                flag = False
                item0 = table.item(r, 0)
                if item0:
                    flag = bool(item0.data(Qt.UserRole + 1))
                if not flag:
                    table.setRowHidden(r, True)
                    continue
            if not term:
                table.setRowHidden(r, False)
                continue
            row_text = ""
            for c in cols:
                item = table.item(r, c)
                if item:
                    row_text += " " + item.text().upper()
            table.setRowHidden(r, term not in row_text)
        self._fit_eev_table_to_contents(table)

    def _on_eev_missing_clicked(self, event=None) -> None:
        if self._eev_missing_count <= 0:
            return
        self._eev_filter_missing = not self._eev_filter_missing
        if self._eev_filter_missing:
            self.lbl_eev_missing.setStyleSheet("color:#b45309; font-weight:600;")
        else:
            self.lbl_eev_missing.setStyleSheet("color:#b45309;")
        self._apply_eev_filters()
        self._open_eev_costs()

    def _on_eev_detail_select(self) -> None:
        model = ""
        row = self.eev_detail.currentRow()
        if row >= 0:
            item = self.eev_detail.item(row, 9)
            if item:
                model = item.text()
        self._highlight_eev_by_model(model)

    def _on_eev_bom_select(self) -> None:
        model = ""
        row = self.eev_bom.currentRow()
        if row >= 0:
            item = self.eev_bom.item(row, 0)
            if item:
                model = item.text()
        self._highlight_eev_by_model(model)

    def _highlight_eev_by_model(self, model: str) -> None:
        target = self._norm_text(model)
        self._apply_highlight_table(self.eev_detail, 9, target)
        self._apply_highlight_table(self.eev_bom, 0, target)

    def _apply_highlight_table(self, table: QTableWidget, col_model: int, target: str) -> None:
        for r in range(table.rowCount()):
            item_model = table.item(r, col_model)
            match = False
            if target and item_model:
                match = self._norm_text(item_model.text()) == target
            for c in range(table.columnCount()):
                item = table.item(r, c)
                if not item:
                    continue
                font = item.font()
                font.setBold(match)
                item.setFont(font)
    def _load_eev_cost_profile(self) -> Dict[str, Any]:
        candidates = [
            Path("data/LEGEND/eev_cost_profile.json"),
            Path("data/legend/eev_cost_profile.json"),
        ]
        for path in candidates:
            if path.exists():
                try:
                    return json.loads(path.read_text(encoding="utf-8"))
                except Exception:
                    return {}
        return {}

    def _open_eev_costs(self) -> None:
        profile = self._load_eev_cost_profile()
        if not profile:
            QMessageBox.warning(self, "LEGEND", "No se encontro eev_cost_profile.json")
            return
        overrides = {}
        if isinstance(self.project_data, dict):
            overrides = dict(self.project_data.get("eev_cost_overrides", {}) or {})
        factor_default = float(profile.get("factor_default", 0.0) or 0.0)
        factor_val = overrides.get("factor")
        if factor_val is None:
            factor_val = factor_default
        parts = profile.get("parts", {}) if isinstance(profile, dict) else {}
        parts_override = overrides.get("parts_base_cost", {}) if isinstance(overrides, dict) else {}
        models_override = overrides.get("models_unit_cost", {}) if isinstance(overrides, dict) else {}
        sets_cfg = profile.get("sets", {}) if isinstance(profile, dict) else {}

        dlg = QDialog(self)
        dlg.setWindowTitle("COSTOS EEV")
        dlg.setMinimumSize(900, 700)
        layout = QVBoxLayout()

        factor_row = QHBoxLayout()
        factor_row.addWidget(QLabel("FACTOR"))
        factor_spin = QDoubleSpinBox()
        factor_spin.setDecimals(4)
        factor_spin.setRange(0.0, 5.0)
        factor_spin.setSingleStep(0.01)
        factor_spin.setValue(float(factor_val))
        factor_row.addWidget(factor_spin)
        factor_row.addStretch(1)
        layout.addLayout(factor_row)

        parts_table = QTableWidget(0, 3)
        parts_table.setHorizontalHeaderLabels(["PART_KEY", "LABEL", "BASE_COST"])
        parts_table.verticalHeader().setVisible(False)
        parts_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        parts_table.setSelectionMode(QAbstractItemView.SingleSelection)
        parts_table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked)
        parts_table.setRowCount(len(parts))
        for r_idx, key in enumerate(sorted(parts.keys())):
            label = str(parts.get(key, {}).get("label", ""))
            base_cost = float(parts.get(key, {}).get("base_cost", 0.0) or 0.0)
            if key in parts_override:
                base_cost = float(parts_override.get(key, base_cost) or 0.0)
            key_item = QTableWidgetItem(str(key))
            key_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
            label_item = QTableWidgetItem(str(label))
            label_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
            cost_item = QTableWidgetItem(f"{base_cost:.2f}")
            parts_table.setItem(r_idx, 0, key_item)
            parts_table.setItem(r_idx, 1, label_item)
            parts_table.setItem(r_idx, 2, cost_item)
        parts_table.resizeColumnsToContents()
        parts_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(QLabel("PARTES (BASE COST)"))
        layout.addWidget(parts_table)

        sets_table = QTableWidget(0, 2)
        sets_table.setHorizontalHeaderLabels(["SET", "COSTO UNITARIO"])
        sets_table.verticalHeader().setVisible(False)
        sets_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        sets_table.setSelectionMode(QAbstractItemView.NoSelection)
        sets_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        sets_table.horizontalHeader().setStretchLastSection(True)

        def _unit_cost_for_part(part_key: str, factor_val: float) -> float:
            base_cost = float(parts.get(part_key, {}).get("base_cost", 0.0) or 0.0)
            if part_key in parts_override:
                try:
                    base_cost = float(parts_override.get(part_key, base_cost) or 0.0)
                except Exception:
                    base_cost = float(base_cost or 0.0)
            return base_cost * (1.0 + factor_val)

        def _refresh_sets_table() -> None:
            factor_val_local = float(factor_spin.value())
            items = []
            for key in sorted(sets_cfg.keys()):
                cfg = sets_cfg.get(key, {}) if isinstance(sets_cfg, dict) else {}
                label = str(cfg.get("label", key))
                parts_list = cfg.get("parts", []) if isinstance(cfg, dict) else []
                total = 0.0
                for part_key in parts_list:
                    total += _unit_cost_for_part(str(part_key), factor_val_local)
                items.append((label, total))
            sets_table.setRowCount(len(items))
            for r_idx, (label, total) in enumerate(items):
                item_label = QTableWidgetItem(str(label).upper())
                item_val = QTableWidgetItem(f"{total:,.2f}")
                item_val.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                sets_table.setItem(r_idx, 0, item_label)
                sets_table.setItem(r_idx, 1, item_val)
            sets_table.resizeColumnsToContents()
            sets_table.horizontalHeader().setStretchLastSection(True)

        if sets_cfg:
            layout.addWidget(QLabel("SETS (COSTO UNITARIO)"))
            layout.addWidget(sets_table)
            _refresh_sets_table()
            factor_spin.valueChanged.connect(_refresh_sets_table)

        models_table = QTableWidget(0, 2)
        models_table.setHorizontalHeaderLabels(["MODELO", "UNIT_COST"])
        models_table.verticalHeader().setVisible(False)
        models_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        models_table.setSelectionMode(QAbstractItemView.SingleSelection)
        models_table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked)
        rows = list(models_override.items())
        models_table.setRowCount(len(rows))
        for r_idx, (model, unit_cost) in enumerate(rows):
            models_table.setItem(r_idx, 0, QTableWidgetItem(str(model)))
            models_table.setItem(r_idx, 1, QTableWidgetItem(f"{float(unit_cost):.2f}"))
        models_table.resizeColumnsToContents()
        models_table.horizontalHeader().setStretchLastSection(True)

        btn_row = QHBoxLayout()
        btn_add = QPushButton("AGREGAR FILA")
        btn_del = QPushButton("ELIMINAR FILA")
        btn_row.addWidget(btn_add)
        btn_row.addWidget(btn_del)
        btn_row.addStretch(1)
        layout.addWidget(QLabel("OVERRIDE POR MODELO"))
        layout.addLayout(btn_row)
        layout.addWidget(models_table)

        def _add_model_row() -> None:
            r = models_table.rowCount()
            models_table.insertRow(r)
            models_table.setItem(r, 0, QTableWidgetItem(""))
            models_table.setItem(r, 1, QTableWidgetItem("0.00"))

        def _del_model_row() -> None:
            rows = sorted({i.row() for i in models_table.selectionModel().selectedRows()})
            if not rows:
                return
            for r in reversed(rows):
                models_table.removeRow(r)

        btn_add.clicked.connect(_add_model_row)
        btn_del.clicked.connect(_del_model_row)

        btns = QHBoxLayout()
        btn_save = QPushButton("GUARDAR")
        btn_cancel = QPushButton("CANCELAR")
        btns.addStretch(1)
        btns.addWidget(btn_cancel)
        btns.addWidget(btn_save)
        layout.addLayout(btns)

        dlg.setLayout(layout)

        def _save_costs() -> None:
            new_overrides = {}
            new_overrides["factor"] = float(factor_spin.value())
            parts_out: Dict[str, float] = {}
            for r in range(parts_table.rowCount()):
                key = parts_table.item(r, 0).text().strip()
                if not key:
                    continue
                try:
                    val = float(parts_table.item(r, 2).text().replace(",", "."))
                except Exception:
                    continue
                base_cost = float(parts.get(key, {}).get("base_cost", 0.0) or 0.0)
                if abs(val - base_cost) > 1e-6:
                    parts_out[key] = val
            models_out: Dict[str, float] = {}
            for r in range(models_table.rowCount()):
                model = models_table.item(r, 0)
                cost = models_table.item(r, 1)
                if not model or not cost:
                    continue
                m_text = model.text().strip()
                if not m_text:
                    continue
                try:
                    c_val = float(cost.text().replace(",", "."))
                except Exception:
                    continue
                models_out[m_text] = c_val
            new_overrides["parts_base_cost"] = parts_out
            new_overrides["models_unit_cost"] = models_out
            if not isinstance(self.project_data, dict):
                self.project_data = {}
            self.project_data["eev_cost_overrides"] = new_overrides
            self._set_dirty(True)
            self._update_eev()
            self._on_save_project()
            dlg.accept()

        btn_save.clicked.connect(_save_costs)
        btn_cancel.clicked.connect(dlg.reject)

        dlg.exec()

    def _refresh_table_view(self, view: QTableView) -> None:
        try:
            view.doItemsLayout()
            view.viewport().update()
        except Exception:
            pass

    def _notify_model_changed(self, model: "LegendItemsTableModel") -> None:
        try:
            model.layoutChanged.emit()
            if model.rowCount() > 0:
                top = model.index(0, 0)
                bottom = model.index(model.rowCount() - 1, model.columnCount() - 1)
                model.dataChanged.emit(top, bottom, [Qt.DisplayRole, Qt.EditRole])
        except Exception:
            pass

    def _schedule_after_insert(self, view: QTableView) -> None:
        def _apply() -> None:
            try:
                model = view.model()
                if model:
                    try:
                        model.beginResetModel()
                        model.endResetModel()
                    except Exception:
                        pass
            except Exception:
                pass
            try:
                view.viewport().update()
            except Exception:
                pass
            try:
                view.resizeRowsToContents()
            except Exception:
                pass
            try:
                self._fit_table_to_contents_view(view)
            except Exception:
                pass
        from PySide6.QtCore import QTimer
        QTimer.singleShot(0, _apply)

    def _mark_dirty_and_totals(self, *args, **kwargs) -> None:
        if self._suspend_updates:
            return
        self._set_dirty(True)

    def _set_dirty(self, value: bool) -> None:
        self._dirty = value
        self._recalc_totals()

    def _on_dim_limit(self, label: str, max_val: float) -> None:
        QMessageBox.warning(
            self,
            "Dimensión fuera de rango",
            f"{label} máximo {max_val:.2f} m.\nUSE CUADRO DE CARGAS CUARTOS INDUSTRIALES.",
        )

    def _update_loop_column_visibility(self) -> None:
        try:
            distrib = ""
            widget = self.spec_fields.get("distribucion_tuberia")
            if isinstance(widget, QComboBox):
                distrib = widget.currentText().strip().upper()
            col_loop = PROJ_COLUMNS.index("loop")
            show_loop = distrib == "LOOP"
            self.bt_view.setColumnHidden(col_loop, not show_loop)
            self.mt_view.setColumnHidden(col_loop, not show_loop)
        except Exception:
            pass

    def _confirm_discard(self) -> bool:
        resp = QMessageBox.question(
            self,
            "LEGEND",
            "HAY CAMBIOS SIN GUARDAR. ¿DESEAS DESCARTARLOS?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        return resp == QMessageBox.Yes

    def _to_int(self, val, default: int = 1) -> int:
        try:
            return int(val)
        except Exception:
            return default

    def _to_float(self, val, default: float | None = 0.0) -> float | None:
        try:
            if isinstance(val, str):
                val = val.replace(",", ".").strip()
            return float(val)
        except Exception:
            return default

    def _norm_text(self, text: str) -> str:
        try:
            import unicodedata

            norm = unicodedata.normalize("NFKD", str(text))
            norm = "".join(c for c in norm if not unicodedata.combining(c))
            return " ".join(norm.strip().upper().split())
        except Exception:
            return " ".join(str(text or "").strip().upper().split())

    def _get_mt_ramal_offset(self) -> int:
        bt_count = 0
        if isinstance(self.project_data, dict):
            bt_count = self._to_int(self.project_data.get("bt_ramales", 0), 0)
        if bt_count <= 0 and self.bt_model:
            bt_count = max(
                (self._to_int(r.get("ramal", 0), 0) for r in self.bt_model.items if isinstance(r, dict)),
                default=0,
            )
        return max(bt_count, 0)

    def _refresh_mt_ramal_display(self) -> None:
        try:
            if not self.mt_model:
                return
            col = PROJ_COLUMNS.index("ramal")
            if self.mt_model.rowCount() > 0:
                top = self.mt_model.index(0, col)
                bottom = self.mt_model.index(self.mt_model.rowCount() - 1, col)
                self.mt_model.dataChanged.emit(top, bottom, [Qt.DisplayRole, Qt.EditRole])
            self._fit_table_to_contents_view(self.mt_view)
        except Exception:
            pass

    def _block_for_model(self, model: "LegendItemsTableModel") -> str:
        if model is self.bt_model:
            return "bt"
        if model is self.mt_model:
            return "mt"
        return ""

    def _default_row(self, bloque: str) -> dict:
        return dict(self.bt_model.default_row if bloque == "bt" else self.mt_model.default_row)

    def _write_error_log(self, text: str) -> None:
        try:
            log_path = Path.cwd() / "legend_error.log"
            with open(log_path, "a", encoding="utf-8") as fh:
                fh.write(text + "\n")
        except Exception:
            pass

    def _on_bt_ramales_changed(self, value: int) -> None:
        if not isinstance(self.project_data, dict):
            return
        current = self._to_int(self.project_data.get("bt_ramales", 1))
        if value == current:
            if self.bt_model.rowCount() == 0 and value >= 1:
                row = self._default_row("bt")
                row["ramal"] = 1
                self._suspend_updates = True
                self.bt_model.add_rows([row])
                self._suspend_updates = False
                self.project_data["bt_items"] = self.bt_model.items
                self._fit_table_to_contents_view(self.bt_view)
                self._set_dirty(True)
            return
        if value > current and self.bt_model.rowCount() == 0:
            row = self._default_row("bt")
            row["ramal"] = 1
            self._suspend_updates = True
            self.bt_model.add_rows([row])
            self._suspend_updates = False
            self.project_data["bt_items"] = self.bt_model.items
            current = 1
        if value > current:
            self._add_ramales_bulk("bt", value - current)
        else:
            self._trim_ramales_to("bt", value)
        self._refresh_mt_ramal_display()

    def _on_mt_ramales_changed(self, value: int) -> None:
        if not isinstance(self.project_data, dict):
            return
        current = self._to_int(self.project_data.get("mt_ramales", 1))
        if value == current:
            if self.mt_model.rowCount() == 0 and value >= 1:
                row = self._default_row("mt")
                row["ramal"] = 1
                self._suspend_updates = True
                self.mt_model.add_rows([row])
                self._suspend_updates = False
                self.project_data["mt_items"] = self.mt_model.items
                self._fit_table_to_contents_view(self.mt_view)
                self._set_dirty(True)
            return
        if value > current and self.mt_model.rowCount() == 0:
            row = self._default_row("mt")
            row["ramal"] = 1
            self._suspend_updates = True
            self.mt_model.add_rows([row])
            self._suspend_updates = False
            self.project_data["mt_items"] = self.mt_model.items
            current = 1
        if value > current:
            self._add_ramales_bulk("mt", value - current)
        else:
            self._trim_ramales_to("mt", value)

    def _add_ramales_bulk(self, bloque: str, count: int) -> None:
        try:
            self._write_error_log(f"[ADD_RAMALES] START bloque={bloque} count={count}")
            if count <= 0:
                return
            if not isinstance(self.project_data, dict):
                self.project_data = {}
            key_items = "bt_items" if bloque == "bt" else "mt_items"
            key_ramales = "bt_ramales" if bloque == "bt" else "mt_ramales"
            model = self.bt_model if bloque == "bt" else self.mt_model
            view = self.bt_view if bloque == "bt" else self.mt_view
            items = model.items
            current = self._to_int(self.project_data.get(key_ramales, 1))
            if current >= 20:
                QMessageBox.information(self, "LEGEND", "MAXIMO 20 RAMALES.")
                return
            if current + count > 20:
                count = 20 - current
            n = current
            temp_items = [r for r in items if isinstance(r, dict)]
            new_rows_all: List[dict] = []
            for _ in range(count):
                base_rows = [dict(r) for r in temp_items if self._to_int(r.get("ramal", 1)) == n]
                new_rows = []
                if not base_rows:
                    new_row = self._default_row(bloque)
                    new_row["ramal"] = n + 1
                    new_rows.append(new_row)
                else:
                    for r in base_rows:
                        nr = dict(r)
                        nr["ramal"] = n + 1
                        nr["largo_m"] = ""
                        nr["ancho_m"] = ""
                        nr["alto_m"] = ""
                        nr["btu_hr"] = 0
                        nr["evap_qty"] = 0
                        nr["evap_modelo"] = ""
                        nr["familia"] = r.get("familia", "AUTO")
                        new_rows.append(nr)
                new_rows_all.extend(new_rows)
                temp_items.extend(new_rows)
                n += 1
            self._suspend_updates = True
            model.add_rows(new_rows_all)
            self._suspend_updates = False
            self._write_error_log(f"[ADD_RAMALES] rows={model.rowCount()} items={len(model.items)}")
            self.project_data[key_items] = model.items
            self.project_data[key_ramales] = n
            if bloque == "bt":
                self.spin_bt.setValue(n)
                self._write_error_log("[ADD_RAMALES] BT add_rows OK")
                self._fit_table_to_contents_view(self.bt_view)
            else:
                self.spin_mt.setValue(n)
                self._write_error_log("[ADD_RAMALES] MT add_rows OK")
                self._fit_table_to_contents_view(self.mt_view)
            self._set_dirty(True)
            self._write_error_log("[ADD_RAMALES] DONE")
        except Exception:
            msg = traceback.format_exc()
            self._write_error_log(msg)
            QMessageBox.critical(self, "LEGEND", f"ERROR INTERNO:\n{msg}")

    def _add_ramal(self, bloque: str) -> None:
        if not isinstance(self.project_data, dict):
            self.project_data = {}
        key_items = "bt_items" if bloque == "bt" else "mt_items"
        key_ramales = "bt_ramales" if bloque == "bt" else "mt_ramales"
        model = self.bt_model if bloque == "bt" else self.mt_model
        view = self.bt_view if bloque == "bt" else self.mt_view
        items = [r for r in model.items if isinstance(r, dict)]
        n = self._to_int(self.project_data.get(key_ramales, 1))
        n2 = n + 1
        base_rows = [dict(r) for r in items if isinstance(r, dict) and self._to_int(r.get("ramal", 1)) == n]
        new_rows = []
        if not base_rows:
            new_row = self._default_row(bloque)
            new_row["ramal"] = n2
            new_rows.append(new_row)
        else:
            for r in base_rows:
                nr = dict(r)
                nr["ramal"] = n2
                nr["largo_m"] = ""
                nr["ancho_m"] = ""
                nr["alto_m"] = ""
                nr["btu_hr"] = 0
                nr["evap_qty"] = 0
                nr["evap_modelo"] = ""
                nr["familia"] = r.get("familia", "AUTO")
                new_rows.append(nr)
        self._suspend_updates = True
        model.add_rows(new_rows)
        self._suspend_updates = False
        self._write_error_log(f"[ADD_RAMALES] rows={model.rowCount()} items={len(model.items)}")
        self.project_data[key_items] = model.items
        self.project_data[key_ramales] = n2
        if bloque == "bt":
            self.spin_bt.setValue(n2)
            self._fit_table_to_contents_view(self.bt_view)
        else:
            self.spin_mt.setValue(n2)
            self._fit_table_to_contents_view(self.mt_view)
        self._set_dirty(True)

    def _del_ramal(self, bloque: str) -> None:
        if not isinstance(self.project_data, dict):
            return
        key_items = "bt_items" if bloque == "bt" else "mt_items"
        key_ramales = "bt_ramales" if bloque == "bt" else "mt_ramales"
        items = self.project_data.get(key_items, [])
        if not isinstance(items, list):
            items = []
        items = [r for r in items if isinstance(r, dict)]
        n = self._to_int(self.project_data.get(key_ramales, 1))
        if n <= 1:
            return
        resp = QMessageBox.question(
            self,
            "LEGEND",
            f"SE ELIMINARAN TODAS LAS FILAS DEL RAMAL {n}. ¿CONTINUAR?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if resp != QMessageBox.Yes:
            return
        items = [dict(r) for r in items if isinstance(r, dict) and self._to_int(r.get("ramal", 1)) != n]
        self.project_data[key_items] = items
        self.project_data[key_ramales] = n - 1
        if bloque == "bt":
            self.bt_view.setUpdatesEnabled(False)
            self.bt_model.blockSignals(True)
            self._suspend_updates = True
            self.bt_model.set_items(items)
            self.bt_model.blockSignals(False)
            self.bt_view.setUpdatesEnabled(True)
            self._suspend_updates = False
            self.spin_bt.setValue(n - 1)
            self._fit_table_to_contents_view(self.bt_view)
        else:
            self.mt_view.setUpdatesEnabled(False)
            self.mt_model.blockSignals(True)
            self._suspend_updates = True
            self.mt_model.set_items(items)
            self.mt_model.blockSignals(False)
            self.mt_view.setUpdatesEnabled(True)
            self._suspend_updates = False
            self.spin_mt.setValue(n - 1)
            self._fit_table_to_contents_view(self.mt_view)
        self._set_dirty(True)

    def _trim_ramales_to(self, bloque: str, target: int) -> None:
        if not isinstance(self.project_data, dict):
            return
        key_items = "bt_items" if bloque == "bt" else "mt_items"
        key_ramales = "bt_ramales" if bloque == "bt" else "mt_ramales"
        model = self.bt_model if bloque == "bt" else self.mt_model
        view = self.bt_view if bloque == "bt" else self.mt_view
        items = [r for r in model.items if isinstance(r, dict)]
        target = max(1, min(20, self._to_int(target)))
        if items:
            new_items = items[:target]
        else:
            new_items = []
        if not new_items and target >= 1:
            row = self._default_row(bloque)
            row["ramal"] = 1
            new_items = [row]
        model.set_items(new_items)
        self.project_data[key_items] = model.items
        self.project_data[key_ramales] = target
        if bloque == "bt":
            self.spin_bt.blockSignals(True)
            self.spin_bt.setValue(target)
            self.spin_bt.blockSignals(False)
            self._fit_table_to_contents_view(self.bt_view)
        else:
            self.spin_mt.blockSignals(True)
            self.spin_mt.setValue(target)
            self.spin_mt.blockSignals(False)
            self._fit_table_to_contents_view(self.mt_view)
        self._renumber_ramales(bloque)

    def _recalc_totals(self) -> None:
        def _sum(model: "LegendItemsTableModel") -> float:
            total = 0.0
            for r in model.items:
                try:
                    total += float(r.get("btu_hr", r.get("carga_btu_h", 0)) or 0)
                except Exception:
                    pass
            return total

        self.total_bt = _sum(self.bt_model)
        self.total_mt = _sum(self.mt_model)
        fmt = lambda v: f"{int(round(v)):,}"
        self.lbl_total_bt.setText(f"TOTAL BT (BTU/H): {fmt(self.total_bt)}")
        self.lbl_total_mt.setText(f"TOTAL MT (BTU/H): {fmt(self.total_mt)}")
        self.lbl_total_general.setText(f"TOTAL GENERAL: {fmt(self.total_bt + self.total_mt)}")
        self._update_eev()

    def resizeEvent(self, event) -> None:
        super().resizeEvent(event)
        self._update_folder_label()
        self._update_compressors()

    def _load_compresores_perf(self) -> Dict[str, Any]:
        for path in (
            Path("data/LEGEND/compresores_perf.json"),
            Path("data/legend/compresores_perf.json"),
        ):
            if path.exists():
                try:
                    return json.loads(path.read_text(encoding="utf-8"))
                except Exception:
                    return {}
        return {}

    def _get_comp_models(self, brand: str) -> List[str]:
        brands = self._comp_perf.get("brands", {}) if isinstance(self._comp_perf, dict) else {}
        models = brands.get(brand, {}).get("models", {}) if isinstance(brands, dict) else {}
        return sorted(models.keys())

    def _parse_perf_points(self, points: Dict[str, Any]) -> List[Dict[str, Any]]:
        parsed = []
        for key, val in points.items():
            if not isinstance(key, str):
                continue
            parts = key.split("|")
            if len(parts) < 3:
                continue
            try:
                ref = parts[0].strip()
                tcond = float(parts[1])
                tevap = float(parts[2])
            except Exception:
                continue
            parsed.append({"ref": ref, "tcond_f": tcond, "tevap_f": tevap, "data": val})
        return parsed

    def _get_comp_perf(
        self, brand: str, model: str, tcond_f: float, tevap_f: float, refrigerante: str
    ) -> tuple[Dict[str, Any] | None, bool]:
        brands = self._comp_perf.get("brands", {}) if isinstance(self._comp_perf, dict) else {}
        brand_data = brands.get(brand, {}) if isinstance(brands, dict) else {}
        models = brand_data.get("models", {}) if isinstance(brand_data, dict) else {}
        model_data = models.get(model, {}) if isinstance(models, dict) else {}
        points = model_data.get("points", {}) if isinstance(model_data, dict) else {}
        parsed = self._parse_perf_points(points) if isinstance(points, dict) else []
        if not parsed:
            return None, False
        ref_norm = self._norm_text(refrigerante).replace("-", "")
        best = None
        best_score = None
        ref_mismatch = False
        for p in parsed:
            pref = self._norm_text(p.get("ref", "")).replace("-", "")
            mismatch = bool(ref_norm) and pref and ref_norm not in pref
            score = abs(p["tcond_f"] - tcond_f) + abs(p["tevap_f"] - tevap_f)
            if best is None or (not mismatch and best_score is not None and score < best_score):
                best = p
                best_score = score
                ref_mismatch = mismatch
            elif best is not None and best_score is not None:
                if score < best_score and not mismatch:
                    best = p
                    best_score = score
                    ref_mismatch = mismatch
                elif best is None:
                    best = p
                    best_score = score
                    ref_mismatch = mismatch
        if best is None:
            return None, False
        return best.get("data", {}), ref_mismatch

    def _compute_tevap_design(self, items: List[Dict[str, Any]]) -> float | None:
        vals = []
        for it in items:
            try:
                load = float(it.get("btu_hr", it.get("carga_btu_h", 0)) or 0)
            except Exception:
                load = 0.0
            if load <= 0:
                continue
            try:
                tev = float(it.get("tevap_f", it.get("tevap", 0)) or 0)
            except Exception:
                continue
            vals.append(tev)
        if not vals:
            return None
        return min(vals)

    def _build_compressor_block(self, title: str) -> Dict[str, Any]:
        box = QGroupBox(title)
        grid = QGridLayout()
        grid.setContentsMargins(6, 6, 6, 6)
        grid.setSpacing(4)

        n_spin = QSpinBox()
        n_spin.setRange(0, 99)
        n_spin.setEnabled(False)
        target_spin = QSpinBox()
        target_spin.setRange(0, 200)
        target_spin.setValue(25)
        btn_auto = QPushButton("AUTOSELECCIONAR")
        btn_add_model = QPushButton("AGREGAR MODELO")
        btn_del_model = QPushButton("ELIMINAR MODELO")

        models_tbl = QTableWidget(0, 2)
        models_tbl.setHorizontalHeaderLabels(["MODELO", "CANTIDAD"])
        models_tbl.horizontalHeader().setStretchLastSection(False)
        models_tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        models_tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        models_tbl.verticalHeader().setVisible(False)
        models_tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        models_tbl.setSelectionMode(QAbstractItemView.SingleSelection)
        models_tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        models_tbl.setMinimumHeight(90)
        models_tbl.setStyleSheet(
            "QComboBox { background: #ffffff; color: #111827; }"
            "QComboBox QAbstractItemView { background: #ffffff; color: #111827;"
            " selection-background-color: #dbeafe; selection-color: #111827; }"
        )

        lbl_tevap = QLabel("--")
        lbl_tcond = QLabel("--")
        lbl_carga = QLabel("--")
        lbl_cap = QLabel("--")
        lbl_res = QLabel("--")
        lbl_state = QLabel("SIN DATA")
        lbl_state.setStyleSheet("color:#9ca3af;")

        grid.addWidget(QLabel("TARGET RESERVA %"), 0, 0)
        grid.addWidget(target_spin, 0, 1)
        grid.addWidget(btn_auto, 0, 2)

        grid.addWidget(QLabel("MODELOS"), 1, 0)
        grid.addWidget(models_tbl, 1, 1, 1, 2)
        btns_row = QHBoxLayout()
        btns_row.addWidget(btn_add_model)
        btns_row.addWidget(btn_del_model)
        btns_row.addStretch(1)
        grid.addLayout(btns_row, 2, 1, 1, 2)

        grid.addWidget(QLabel("TOTAL COMPRESORES"), 3, 0)
        grid.addWidget(n_spin, 3, 1)

        grid.addWidget(QLabel("TEVAP DISEÑO (F)"), 4, 0)
        grid.addWidget(lbl_tevap, 4, 1)
        grid.addWidget(QLabel("TCOND (F)"), 4, 2)
        grid.addWidget(lbl_tcond, 4, 3)

        grid.addWidget(QLabel("CARGA (BTU/H)"), 5, 0)
        grid.addWidget(lbl_carga, 5, 1)
        grid.addWidget(QLabel("CAPACIDAD (BTU/H)"), 5, 2)
        grid.addWidget(lbl_cap, 5, 3)

        grid.addWidget(QLabel("RESERVA %"), 6, 0)
        grid.addWidget(lbl_res, 6, 1)
        grid.addWidget(QLabel("ESTADO"), 6, 2)
        grid.addWidget(lbl_state, 6, 3)

        grid.setColumnStretch(3, 1)
        box.setLayout(grid)

        block = {
            "box": box,
            "n": n_spin,
            "target": target_spin,
            "btn_auto": btn_auto,
            "btn_add_model": btn_add_model,
            "btn_del_model": btn_del_model,
            "models_tbl": models_tbl,
            "lbl_tevap": lbl_tevap,
            "lbl_tcond": lbl_tcond,
            "lbl_carga": lbl_carga,
            "lbl_cap": lbl_cap,
            "lbl_res": lbl_res,
            "lbl_state": lbl_state,
        }

        target_spin.valueChanged.connect(lambda _v, b=block: self._on_comp_changed(b))
        btn_auto.clicked.connect(lambda _=None, b=block: self._auto_select_compressor(b))
        btn_add_model.clicked.connect(lambda _=None, b=block: self._add_comp_model_row(b))
        btn_del_model.clicked.connect(lambda _=None, b=block: self._del_comp_model_row(b))
        return block

    def _on_comp_brand_changed(self) -> None:
        brand = self.comp_brand_cb.currentText() if hasattr(self, "comp_brand_cb") else ""
        models = self._get_comp_models(brand)
        for block in (self.comp_bt, self.comp_mt):
            tbl: QTableWidget = block["models_tbl"]
            for r in range(tbl.rowCount()):
                combo = tbl.cellWidget(r, 0)
                if isinstance(combo, QComboBox):
                    current = combo.currentText()
                    combo.blockSignals(True)
                    combo.clear()
                    combo.addItem("")
                    combo.addItems(models)
                    if current and current not in models:
                        combo.addItem(current)
                    combo.setCurrentText(current if current else "")
                    combo.blockSignals(False)
            self._on_comp_changed(block)

    def _on_comp_changed(self, block: Dict[str, Any]) -> None:
        self._update_compressors()
        if not self._suspend_updates:
            self._set_dirty(True)

    def _add_comp_model_row(self, block: Dict[str, Any], model: str = "", qty: int = 1, trigger_update: bool = True) -> None:
        tbl: QTableWidget = block["models_tbl"]
        row = tbl.rowCount()
        tbl.insertRow(row)

        combo = QComboBox()
        combo.addItem("")
        brand = self.comp_brand_cb.currentText() if hasattr(self, "comp_brand_cb") else ""
        models = self._get_comp_models(brand)
        combo.addItems(models)
        combo.blockSignals(True)
        if model:
            if model not in models:
                combo.addItem(model)
            combo.setCurrentText(model)
        combo.blockSignals(False)
        combo.currentTextChanged.connect(lambda _t, b=block: self._on_comp_changed(b))

        spin = QSpinBox()
        spin.setRange(1, 10)
        spin.setValue(max(1, int(qty or 1)))
        spin.valueChanged.connect(lambda _v, b=block: self._on_comp_changed(b))

        tbl.setCellWidget(row, 0, combo)
        tbl.setCellWidget(row, 1, spin)
        tbl.setRowHeight(row, 26)

        if trigger_update:
            self._on_comp_changed(block)

    def _del_comp_model_row(self, block: Dict[str, Any]) -> None:
        tbl: QTableWidget = block["models_tbl"]
        row = tbl.currentRow()
        if row < 0:
            return
        tbl.removeRow(row)
        self._on_comp_changed(block)

    def _read_comp_model_rows(self, block: Dict[str, Any]) -> List[Dict[str, Any]]:
        tbl: QTableWidget = block["models_tbl"]
        rows: List[Dict[str, Any]] = []
        for r in range(tbl.rowCount()):
            combo = tbl.cellWidget(r, 0)
            spin = tbl.cellWidget(r, 1)
            if not isinstance(combo, QComboBox) or not isinstance(spin, QSpinBox):
                continue
            model = combo.currentText().strip()
            if not model:
                continue
            rows.append({"model": model, "n": int(spin.value())})
        return rows

    def _sync_comp_total(self, block: Dict[str, Any], total_n: int) -> None:
        n_spin: QSpinBox = block["n"]
        n_spin.blockSignals(True)
        n_spin.setValue(max(0, int(total_n or 0)))
        n_spin.blockSignals(False)

    def _auto_select_compressor(self, block: Dict[str, Any]) -> None:
        brand = self.comp_brand_cb.currentText() if hasattr(self, "comp_brand_cb") else ""
        if not brand:
            return
        models = self._get_comp_models(brand)
        if not models:
            return
        block_key = "bt" if block is self.comp_bt else "mt"
        items = self.bt_model.items if block_key == "bt" else self.mt_model.items
        load = self.total_bt if block_key == "bt" else self.total_mt
        tevap_design = self._compute_tevap_design(items)
        tcond_f = self._to_float(self.tcond_f_edit.text(), None)
        refrigerante = self.spec_fields.get("refrigerante").currentText() if isinstance(self.spec_fields.get("refrigerante"), QComboBox) else ""
        target = float(block["target"].value())

        tbl: QTableWidget = block["models_tbl"]
        if load <= 0:
            tbl.setRowCount(0)
            self._sync_comp_total(block, 0)
            block["lbl_cap"].setText("--")
            block["lbl_res"].setText("--")
            block["lbl_state"].setText("SIN CARGA")
            block["lbl_state"].setStyleSheet("color:#6b7280;")
            return
        if tevap_design is None or tcond_f is None:
            return

        required = load * (1.0 + target / 100.0)
        n_min = 1
        n_max = 8
        if load >= 400000:
            n_pref = 4
        elif load >= 250000:
            n_pref = 3
        else:
            n_pref = 2

        best = None
        best_score = None
        best_ref_warn = False
        max_choice = None
        max_cap = None

        for n in range(n_min, n_max + 1):
            for m in models:
                perf, ref_warn = self._get_comp_perf(brand, m, tcond_f, tevap_design, refrigerante)
                if not perf:
                    continue
                cap = self._to_float(
                    perf.get("capacity_btu_h", perf.get("evaporator_capacity_btu_h", 0.0)), 0.0
                )
                cap_total = n * cap
                if cap_total >= required:
                    oversize = cap_total / required - 1.0
                    penalty_n = (n - n_pref) ** 2
                    score = oversize * 100.0 + penalty_n * 3.0
                    if best_score is None or score < best_score:
                        best_score = score
                        best = (m, n, cap_total)
                        best_ref_warn = ref_warn
                if max_cap is None or cap_total > max_cap:
                    max_cap = cap_total
                    max_choice = (m, n, cap_total, ref_warn)

        if best is not None:
            chosen_model, chosen_n, _cap_total = best
            tbl.setRowCount(0)
            self._add_comp_model_row(block, chosen_model, chosen_n, trigger_update=False)
            self._on_comp_changed(block)
        elif max_choice is not None:
            chosen_model, chosen_n, _cap_total, _ref_warn = max_choice
            tbl.setRowCount(0)
            self._add_comp_model_row(block, chosen_model, chosen_n, trigger_update=False)
            self._on_comp_changed(block)
        else:
            tbl.setRowCount(0)
            self._on_comp_changed(block)

    def _update_compressors(self) -> None:
        if self._loading_project:
            return
        try:
            self._update_compressors_block("bt", self.comp_bt)
            self._update_compressors_block("mt", self.comp_mt)
        except Exception:
            pass

    def _update_compressors_block(self, key: str, block: Dict[str, Any]) -> None:
        brand = self.comp_brand_cb.currentText() if hasattr(self, "comp_brand_cb") else ""
        target = float(block["target"].value())
        items = self.bt_model.items if key == "bt" else self.mt_model.items
        load = self.total_bt if key == "bt" else self.total_mt
        tevap_design = self._compute_tevap_design(items)
        tcond_f = self._to_float(self.tcond_f_edit.text(), None)
        refrigerante = self.spec_fields.get("refrigerante").currentText() if isinstance(self.spec_fields.get("refrigerante"), QComboBox) else ""

        model_rows = self._read_comp_model_rows(block)
        total_n = sum(int(r.get("n", 0)) for r in model_rows)
        self._sync_comp_total(block, total_n)

        def _set_label(lbl: QLabel, val: str) -> None:
            lbl.setText(val)

        _set_label(block["lbl_carga"], f"{int(round(load)):,}" if load else "0")
        _set_label(block["lbl_tcond"], f"{tcond_f:.1f}" if tcond_f is not None else "--")
        _set_label(block["lbl_tevap"], f"{tevap_design:.1f}" if tevap_design is not None else "--")

        if not brand or not model_rows or tevap_design is None or tcond_f is None:
            _set_label(block["lbl_cap"], "--")
            _set_label(block["lbl_res"], "--")
            block["lbl_state"].setText("SIN DATA")
            block["lbl_state"].setStyleSheet("color:#9ca3af;")
        else:
            cap_total = 0.0
            has_perf = False
            missing_perf = False
            ref_warn_any = False
            for entry in model_rows:
                model = entry.get("model", "")
                qty = int(entry.get("n", 0) or 0)
                if not model or qty <= 0:
                    continue
                perf, ref_warn = self._get_comp_perf(brand, model, tcond_f, tevap_design, refrigerante)
                if ref_warn:
                    ref_warn_any = True
                if not perf:
                    missing_perf = True
                    continue
                has_perf = True
                cap = self._to_float(
                    perf.get("capacity_btu_h", perf.get("evaporator_capacity_btu_h", 0.0)), 0.0
                )
                cap_total += qty * cap

            if not has_perf:
                _set_label(block["lbl_cap"], "--")
                _set_label(block["lbl_res"], "--")
                block["lbl_state"].setText("SIN DATA")
                block["lbl_state"].setStyleSheet("color:#9ca3af;")
            else:
                _set_label(block["lbl_cap"], f"{int(round(cap_total)):,}")
                if load > 0:
                    reserva = (cap_total - load) / load * 100.0
                    _set_label(block["lbl_res"], f"{reserva:.1f}")
                    if reserva >= target and not missing_perf:
                        block["lbl_state"].setText("OK" + (" (REF?)" if ref_warn_any else ""))
                        block["lbl_state"].setStyleSheet("color:#16a34a;")
                    else:
                        block["lbl_state"].setText("WARNING" + (" (REF?)" if ref_warn_any else ""))
                        block["lbl_state"].setStyleSheet("color:#d97706;")
                else:
                    _set_label(block["lbl_res"], "0.0")
                    block["lbl_state"].setText("SIN CARGA")
                    block["lbl_state"].setStyleSheet("color:#6b7280;")

        if not isinstance(self.project_data, dict):
            self.project_data = {}
        comp = self.project_data.get("compressors", {}) if isinstance(self.project_data, dict) else {}
        comp.setdefault(key, {})
        comp["brand"] = brand
        comp[key] = {
            "items": model_rows,
            "target_reserva_pct": target,
        }
        self.project_data["compressors"] = comp

    def _render_compressors(self) -> None:
        comp = {}
        if isinstance(self.project_data, dict):
            comp = self.project_data.get("compressors", {}) or {}
        brand = str(comp.get("brand", "") or "") if isinstance(comp, dict) else ""
        self._suspend_updates = True
        if hasattr(self, "comp_brand_cb"):
            self.comp_brand_cb.blockSignals(True)
            if brand and brand not in self._comp_brands:
                self.comp_brand_cb.addItem(brand)
            self.comp_brand_cb.setCurrentText(brand if brand else "")
            self.comp_brand_cb.blockSignals(False)
            self._on_comp_brand_changed()
        for key, block in (("bt", self.comp_bt), ("mt", self.comp_mt)):
            data = comp.get(key, {}) if isinstance(comp, dict) else {}
            target = float(data.get("target_reserva_pct", 25) or 25)
            models_list = data.get("items", []) if isinstance(data, dict) else []
            if not isinstance(models_list, list):
                models_list = []
            if not models_list:
                legacy_model = str(data.get("model", "") or "")
                legacy_n = int(data.get("n", 1) or 1)
                if legacy_model:
                    models_list = [{"model": legacy_model, "n": legacy_n}]

            block["target"].blockSignals(True)
            block["target"].setValue(target)
            block["target"].blockSignals(False)

            tbl: QTableWidget = block["models_tbl"]
            tbl.setRowCount(0)
            for entry in models_list:
                self._add_comp_model_row(block, str(entry.get("model", "") or ""), int(entry.get("n", 1) or 1), trigger_update=False)

        self._suspend_updates = False
        self._update_compressors()

    def _update_eev(self) -> None:
        try:
            if not compute_eev:
                self.eev_group.setVisible(True)
                self._fill_eev_detail([])
                self._fill_eev_bom([])
                self._fill_eev_sets([])
                self.lbl_eev_warn.setText("EEV NO DISPONIBLE (MODULO NO CARGADO)")
                return
            exp_widget = self.spec_fields.get("expansion")
            exp_text = exp_widget.currentText() if isinstance(exp_widget, QComboBox) else ""
            exp_norm = self._norm_text(exp_text)
            if "ELECTRON" not in exp_norm:
                self.eev_group.setVisible(False)
                return
            self.eev_group.setVisible(True)
            project_data = self._collect_project_data()
            bt_items = list(self.bt_model.items)
            mt_items = list(self.mt_model.items)
            mt_offset = self._get_mt_ramal_offset()
            result = compute_eev(project_data, bt_items, mt_items, mt_ramal_offset=mt_offset)
            detail_rows = result.get("detail_rows", [])
            bom_rows = result.get("bom_rows", [])
            warnings = result.get("warnings", [])
            self._fill_eev_detail(detail_rows)
            self._fill_eev_bom(bom_rows)
            self._fill_eev_sets(self._compute_eev_set_rows(project_data, result))
            valves_count = len(detail_rows)
            ramal_keys = set()
            for r in detail_rows:
                ramal_keys.add(
                    (
                        self._norm_text(r.get("suction", "")),
                        self._norm_text(r.get("loop", "")),
                        self._norm_text(r.get("ramal", "")),
                    )
                )
            self.lbl_eev_valves.setText(f"VALVULAS: {valves_count}")
            self.lbl_eev_ramales.setText(f"RAMALES CON EEV: {len(ramal_keys)}")
            currency = result.get("cost_currency", "")
            total_cost = result.get("cost_total")
            if isinstance(total_cost, (int, float)):
                self.lbl_eev_total_cost.setText(
                    f"TOTAL COSTO EEV: {currency} {total_cost:,.2f}".strip()
                )
            else:
                self.lbl_eev_total_cost.setText("TOTAL COSTO EEV: --")
            if warnings:
                self.lbl_eev_warn.setText(" | ".join(sorted(set(warnings))))
            else:
                self.lbl_eev_warn.setText("")
        except Exception:
            self.eev_group.setVisible(True)
            self._fill_eev_detail([])
            self._fill_eev_bom([])
            self._fill_eev_sets([])
            self.lbl_eev_warn.setText("ERROR AL CALCULAR EEV")

    def _fill_eev_detail(self, rows: List[Dict[str, Any]]) -> None:
        self.eev_detail.setRowCount(len(rows))
        for r_idx, row in enumerate(rows):
            btu_text = self._fmt_int(row.get("btu_hr", 0))
            tevap_text = self._fmt_temp(row.get("tevap_f", ""))
            values = [
                row.get("suction", ""),
                self._fmt_int(row.get("loop", "")),
                self._fmt_int(row.get("ramal", "")),
                row.get("equipo", ""),
                row.get("uso", ""),
                btu_text,
                tevap_text,
                row.get("familia", ""),
                row.get("orifice", ""),
                row.get("model", ""),
            ]
            for c_idx, val in enumerate(values):
                text = str(val)
                if c_idx not in (5, 6) and not isinstance(val, (int, float)):
                    text = text.upper()
                item = QTableWidgetItem(text)
                if c_idx in (1, 2, 5, 6):
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if c_idx == 8:
                    item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                if c_idx == 3:
                    item.setToolTip(str(val))
                self.eev_detail.setItem(r_idx, c_idx, item)
        self.eev_detail.resizeColumnsToContents()
        header = self.eev_detail.horizontalHeader()
        header.setSectionResizeMode(3, QHeaderView.Stretch)
        header.setSectionResizeMode(8, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(9, QHeaderView.ResizeToContents)
        self._fit_eev_table_to_contents(self.eev_detail)
        self._apply_eev_filters()

    def _fill_eev_bom(self, rows: List[Dict[str, Any]]) -> None:
        self.eev_bom.setRowCount(len(rows))
        missing_count = 0
        for r_idx, row in enumerate(rows):
            unit_cost = row.get("unit_cost")
            total_cost = row.get("total_cost")
            missing = bool(row.get("cost_missing"))
            if missing:
                missing_count += 1
            unit_text = "?" if missing else (self._fmt_cost(unit_cost) if isinstance(unit_cost, (int, float)) else "")
            total_text = "?" if missing else (self._fmt_cost(total_cost) if isinstance(total_cost, (int, float)) else "")
            values = [
                row.get("model", ""),
                row.get("description", ""),
                self._fmt_int(row.get("qty", 0)),
                unit_text,
                total_text,
                row.get("currency", ""),
            ]
            for c_idx, val in enumerate(values):
                text = str(val)
                if c_idx not in (2, 3, 4) and not isinstance(val, (int, float)):
                    text = text.upper()
                item = QTableWidgetItem(text)
                if c_idx in (2, 3, 4):
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if c_idx == 1:
                    item.setToolTip(str(values[1]))
                if missing and c_idx in (3, 4):
                    item.setBackground(QColor(255, 248, 220))
                if c_idx == 0:
                    item.setData(Qt.UserRole + 1, missing)
                self.eev_bom.setItem(r_idx, c_idx, item)
        self.eev_bom.resizeColumnsToContents()
        header = self.eev_bom.horizontalHeader()
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        self._fit_eev_table_to_contents(self.eev_bom)
        self._eev_missing_count = missing_count
        self.lbl_eev_missing.setText(f"ITEMS SIN PRECIO: {missing_count}")
        self._apply_eev_filters()

    def _compute_eev_set_rows(
        self, project_data: Dict[str, Any], result: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        profile = self._load_eev_cost_profile()
        if not profile:
            return []
        overrides = {}
        if isinstance(project_data, dict):
            overrides = project_data.get("eev_cost_overrides", {}) or {}
        parts = profile.get("parts", {}) if isinstance(profile, dict) else {}
        sets_cfg = profile.get("sets", {}) if isinstance(profile, dict) else {}
        package_counts = result.get("package_counts", {}) if isinstance(result, dict) else {}
        parts_override = overrides.get("parts_base_cost", {}) if isinstance(overrides, dict) else {}
        factor_default = float(profile.get("factor_default", 0.0) or 0.0)
        factor_override = overrides.get("factor")
        factor_final = float(factor_override) if factor_override is not None else factor_default
        currency = profile.get("currency", "") if isinstance(profile, dict) else ""

        def _unit_cost_for_part(part_key: str) -> float:
            base_cost = float(parts.get(part_key, {}).get("base_cost", 0.0) or 0.0)
            if isinstance(parts_override, dict) and part_key in parts_override:
                try:
                    base_cost = float(parts_override.get(part_key, base_cost) or 0.0)
                except Exception:
                    base_cost = float(base_cost or 0.0)
            return base_cost * (1.0 + factor_final)

        order = ["VALVULA", "CONTROL", "SENSOR", "TRANSDUCTOR", "CAJAS", "SENSORES CO2"]
        rows: List[Dict[str, Any]] = []
        for key in order:
            if key in ("CAJAS", "SENSORES CO2"):
                part_key = "CAJAS_ELECTRICAS" if key == "CAJAS" else "DGS_IR_CO2"
                qty = int(package_counts.get(key.upper(), 0) or 0)
                if qty <= 0:
                    continue
                unit_cost = _unit_cost_for_part(part_key)
                rows.append(
                    {
                        "label": key,
                        "qty": qty,
                        "unit_cost": unit_cost,
                        "total_cost": unit_cost * qty if qty else 0.0,
                        "currency": currency,
                    }
                )
                continue
            cfg = sets_cfg.get(key, {}) if isinstance(sets_cfg, dict) else {}
            label = str(cfg.get("label", key))
            parts_list = cfg.get("parts", []) if isinstance(cfg, dict) else []
            total = 0.0
            for part_key in parts_list:
                total += _unit_cost_for_part(str(part_key))
            qty = int(package_counts.get(key.upper(), 0) or 0)
            rows.append(
                {
                    "label": label,
                    "qty": qty,
                    "unit_cost": total,
                    "total_cost": total * qty if qty else 0.0,
                    "currency": currency,
                }
            )
        return rows

    def _fill_eev_sets(self, rows: List[Dict[str, Any]]) -> None:
        total_sum = 0.0
        self.eev_sets.setRowCount(len(rows) + (1 if rows else 0))
        for r_idx, row in enumerate(rows):
            unit_cost = row.get("unit_cost")
            total_cost = row.get("total_cost")
            if isinstance(total_cost, (int, float)):
                total_sum += float(total_cost)
            values = [
                row.get("label", ""),
                self._fmt_int(row.get("qty", 0)),
                self._fmt_cost(unit_cost) if isinstance(unit_cost, (int, float)) else "",
                self._fmt_cost(total_cost) if isinstance(total_cost, (int, float)) else "",
                row.get("currency", ""),
            ]
            for c_idx, val in enumerate(values):
                text = str(val)
                if c_idx == 0:
                    text = text.upper()
                item = QTableWidgetItem(text)
                if c_idx in (1, 2, 3):
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.eev_sets.setItem(r_idx, c_idx, item)
        if rows:
            r_total = len(rows)
            total_item = QTableWidgetItem("TOTAL")
            total_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            total_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
            self.eev_sets.setItem(r_total, 0, total_item)
            self.eev_sets.setItem(r_total, 3, QTableWidgetItem(self._fmt_cost(total_sum)))
            self.eev_sets.setItem(r_total, 4, QTableWidgetItem(rows[0].get("currency", "")))
            for c_idx in (3, 4):
                item = self.eev_sets.item(r_total, c_idx)
                if item:
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.eev_sets.resizeColumnsToContents()
        self._fit_eev_table_to_contents(self.eev_sets)
        self._apply_eev_filters()

    def _table_key_press(self, event, model: "LegendItemsTableModel", view: QTableView) -> None:
        if event.matches(QKeySequence.StandardKey.Copy):
            self._copy_selection(model, view)
            event.accept()
            return
        if event.matches(QKeySequence.StandardKey.Paste):
            self._paste_selection(model, view)
            event.accept()
            return
        QTableView.keyPressEvent(view, event)

    def _copy_selection(self, model: "LegendItemsTableModel", view: QTableView) -> None:
        rows = self._selected_row_indices(view)
        if not rows:
            return
        lines = []
        for r in rows:
            item = model.items[r]
            vals = [str(item.get(k, "")) for k in model.headers]
            lines.append("\t".join(vals))
        QGuiApplication.clipboard().setText("\n".join(lines))

    def _paste_selection(self, model: "LegendItemsTableModel", view: QTableView) -> None:
        self._paste_selection_at(model, view)

    def _dup_selection(self, model: "LegendItemsTableModel", view: QTableView) -> None:
        rows = self._selected_row_indices(view)
        if not rows:
            return
        for r in rows:
            model.dup_row(r)
        self._fit_table_to_contents_view(view)
        bloque = self._block_for_model(model)
        if bloque:
            self._renumber_ramales(bloque)
        else:
            self._set_dirty(True)

    def _duplicate_ramal_dialog(self, bloque: str) -> None:
        key_ramales = "bt_ramales" if bloque == "bt" else "mt_ramales"
        max_val = self._to_int(self.project_data.get(key_ramales, 1))
        src, ok = QInputDialog.getInt(self, "Ramal origen", "Ramal origen:", 1, 1, max(max_val, 20))
        if not ok:
            return
        dest, ok = QInputDialog.getInt(self, "Ramal destino", "Ramal destino:", src + 1, 1, 20)
        if not ok:
            return
        self._duplicate_ramal(bloque, src, dest)

    def _duplicate_ramal(self, bloque: str, src: int, dest: int) -> None:
        key_items = "bt_items" if bloque == "bt" else "mt_items"
        key_ramales = "bt_ramales" if bloque == "bt" else "mt_ramales"
        items = self.project_data.get(key_items, [])
        if not isinstance(items, list):
            items = []
        items = [r for r in items if isinstance(r, dict)]
        base_rows = [dict(r) for r in items if isinstance(r, dict) and self._to_int(r.get("ramal", 1)) == src]
        new_rows = []
        for r in base_rows:
            nr = dict(r)
            nr["ramal"] = dest
            nr["largo_m"] = ""
            nr["ancho_m"] = ""
            nr["alto_m"] = ""
            nr["btu_hr"] = 0
            nr["evap_qty"] = 0
            nr["evap_modelo"] = ""
            nr["familia"] = r.get("familia", "AUTO")
            new_rows.append(nr)
        if not new_rows:
            base = self._default_row(bloque)
            base["ramal"] = dest
            new_rows.append(base)
        items = [r for r in items if isinstance(r, dict) and self._to_int(r.get("ramal", 1)) != dest] + new_rows
        self.project_data[key_items] = items
        current = self._to_int(self.project_data.get(key_ramales, 1))
        self.project_data[key_ramales] = max(current, dest)
        if bloque == "bt":
            self.bt_view.setUpdatesEnabled(False)
            self.bt_model.blockSignals(True)
            self._suspend_updates = True
            self.bt_model.set_items(items)
            self.bt_model.blockSignals(False)
            self.bt_view.setUpdatesEnabled(True)
            self._suspend_updates = False
            self.spin_bt.setValue(self.project_data[key_ramales])
            self._fit_table_to_contents_view(self.bt_view)
        else:
            self.mt_view.setUpdatesEnabled(False)
            self.mt_model.blockSignals(True)
            self._suspend_updates = True
            self.mt_model.set_items(items)
            self.mt_model.blockSignals(False)
            self.mt_view.setUpdatesEnabled(True)
            self._suspend_updates = False
            self.spin_mt.setValue(self.project_data[key_ramales])
            self._fit_table_to_contents_view(self.mt_view)
        self._set_dirty(True)

    def refresh(self) -> None:
        if not self.service:
            return
        try:
            data = self.service.load_all()
        except Exception:
            return
        usos = data.get("usos", {}) if isinstance(data, dict) else {}
        equipos = data.get("equipos", []) if isinstance(data, dict) else []
        equipos_bt = data.get("equipos_bt", []) if isinstance(data, dict) else []
        equipos_mt = data.get("equipos_mt", []) if isinstance(data, dict) else []
        btu_map: Dict[str, float] = {}
        for e in equipos:
            try:
                name = getattr(e, "equipo", "")
                val = float(getattr(e, "btu_hr_ft", 0.0) or 0.0)
                key = self._normalize_equipo_name(name)
                if key:
                    btu_map[key] = val
            except Exception:
                continue
        self._equipos_btu_ft = btu_map
        self.bt_model.set_btu_map(self._equipos_btu_ft)
        self.mt_model.set_btu_map(self._equipos_btu_ft)
        if equipos_bt:
            self._equipos_bt = sorted(
                [getattr(e, "equipo", "") for e in equipos_bt if getattr(e, "equipo", "")],
                key=lambda x: x.upper(),
            )
        else:
            self._equipos_bt = sorted(
                [getattr(e, "equipo", "") for e in equipos if getattr(e, "equipo", "")],
                key=lambda x: x.upper(),
            )
        if equipos_mt:
            self._equipos_mt = sorted(
                [getattr(e, "equipo", "") for e in equipos_mt if getattr(e, "equipo", "")],
                key=lambda x: x.upper(),
            )
        else:
            self._equipos_mt = list(self._equipos_bt)
        self._usos_bt = sorted(
            usos.get("BT", []) if isinstance(usos, dict) else [],
            key=lambda x: str(x).upper(),
        )
        self._usos_mt = sorted(
            usos.get("MT", []) if isinstance(usos, dict) else [],
            key=lambda x: str(x).upper(),
        )
        tevap_bt, tevap_mt, def_bt, def_mt = self._load_tevap_maps()
        self.bt_model.set_tevap_map(tevap_bt, def_bt)
        self.mt_model.set_tevap_map(tevap_mt, def_mt)
        self._apply_combo_column_widths(self.bt_view)
        self._apply_combo_column_widths(self.mt_view)
        self._refresh_combo_editors(self.bt_view)
        self._refresh_combo_editors(self.mt_view)

    def _load_tevap_maps(self) -> tuple[Dict[str, float], Dict[str, float], float, float]:
        tevap_bt: Dict[str, float] = {}
        tevap_mt: Dict[str, float] = {}
        def_bt = -22.0
        def_mt = 15.0
        config_path = Path("data/legend/config.json")
        if config_path.exists():
            try:
                cfg = json.loads(config_path.read_text(encoding="utf-8"))
                def_bt = float(cfg.get("tevap_bt", def_bt))
                def_mt = float(cfg.get("tevap_mt", def_mt))
            except Exception:
                pass
        tevap_path = Path("data/legend/tevap_por_uso.json")
        if tevap_path.exists():
            try:
                data = json.loads(tevap_path.read_text(encoding="utf-8"))
                tevap_bt = data.get("BT", {}) if isinstance(data, dict) else {}
                tevap_mt = data.get("MT", {}) if isinstance(data, dict) else {}
            except Exception:
                pass
        return tevap_bt, tevap_mt, def_bt, def_mt

    def _apply_combo_column_widths(self, view: QTableView) -> None:
        # Mantener mismas columnas y tamaños en BT/MT, basado en el mayor texto de ambos
        self._sync_table_column_widths()
        try:
            self._refresh_combo_editors(view)
        except Exception:
            pass

    def _sync_table_column_widths(self) -> None:
        if not self.bt_view or not self.mt_view:
            return
        bt_model = self.bt_view.model()
        mt_model = self.mt_view.model()
        if not bt_model or not mt_model:
            return
        fm = self.bt_view.fontMetrics()

        equipos_all = list(self._equipos_bt) + list(self._equipos_mt)
        usos_all = list(self._usos_bt) + list(self._usos_mt)

        min_widths = {
            "equipo": 140,
            "uso": 120,
            "familia": 140,
        }
        combo_cols = {"equipo", "uso", "familia"}

        for col in range(len(PROJ_COLUMNS)):
            key = PROJ_COLUMNS[col]
            header_text = bt_model.headerData(col, Qt.Horizontal, Qt.DisplayRole) or ""
            max_w = fm.horizontalAdvance(str(header_text))

            if key == "equipo":
                for t in equipos_all:
                    max_w = max(max_w, fm.horizontalAdvance(str(t)))
            elif key == "uso":
                for t in usos_all:
                    max_w = max(max_w, fm.horizontalAdvance(str(t)))
            elif key == "familia":
                for t in self._familia_options:
                    max_w = max(max_w, fm.horizontalAdvance(str(t)))

            for model in (bt_model, mt_model):
                for row in range(model.rowCount()):
                    try:
                        val = model.index(row, col).data(Qt.DisplayRole)
                    except Exception:
                        val = None
                    if val:
                        max_w = max(max_w, fm.horizontalAdvance(str(val)))

            pad = 70 if key in combo_cols else 24
            width = max_w + pad
            if key in min_widths:
                width = max(width, min_widths[key])
            self.bt_view.setColumnWidth(col, width)
            self.mt_view.setColumnWidth(col, width)

    def _normalize_equipo_name(self, text: str) -> str:
        return " ".join(str(text or "").strip().upper().split())

    def _is_cuarto(self, equipo: str) -> bool:
        return "CUARTO" in (equipo or "").upper()

    def _apply_dim_spans(self, view: QTableView, model: "LegendItemsTableModel" | None) -> None:
        if not view or not model:
            return
        try:
            col_largo = PROJ_COLUMNS.index("largo_m")
        except ValueError:
            return
        view.clearSpans()
        for row in range(model.rowCount()):
            try:
                equipo = str(model.items[row].get("equipo", ""))
            except Exception:
                equipo = ""
            if not self._is_cuarto(equipo):
                view.setSpan(row, col_largo, 1, 3)
        try:
            self._refresh_combo_editors(view)
        except Exception:
            pass

    def _ensure_combo_editors(self, view: QTableView) -> None:
        model = view.model()
        if not model:
            return
        col_equipo = PROJ_COLUMNS.index("equipo")
        col_uso = PROJ_COLUMNS.index("uso")
        col_familia = PROJ_COLUMNS.index("familia")
        for row in range(model.rowCount()):
            for col in (col_equipo, col_uso):
                idx = model.index(row, col)
                view.openPersistentEditor(idx)
            try:
                if hasattr(model, "_row_is_cuarto") and model._row_is_cuarto(row):
                    idx = model.index(row, col_familia)
                    view.openPersistentEditor(idx)
            except Exception:
                pass

    def _refresh_combo_editors(self, view: QTableView) -> None:
        model = view.model()
        if not model:
            return
        col_equipo = PROJ_COLUMNS.index("equipo")
        col_uso = PROJ_COLUMNS.index("uso")
        col_familia = PROJ_COLUMNS.index("familia")
        for row in range(model.rowCount()):
            for col in (col_equipo, col_uso):
                idx = model.index(row, col)
                view.closePersistentEditor(idx)
                view.openPersistentEditor(idx)
            try:
                idx_f = model.index(row, col_familia)
                view.closePersistentEditor(idx_f)
                if hasattr(model, "_row_is_cuarto") and model._row_is_cuarto(row):
                    view.openPersistentEditor(idx_f)
            except Exception:
                pass


class ComboDelegate(QStyledItemDelegate):
    def __init__(self, items_fn: Callable[[], List[str]], editable: bool = True, parent=None):
        super().__init__(parent)
        self.items_fn = items_fn
        self.editable = editable

    def createEditor(self, parent, option, index):
        items = self.items_fn() or []
        if not items:
            return super().createEditor(parent, option, index)
        cb = QComboBox(parent)
        cb.setEditable(self.editable)
        cb.addItems(items)
        cb.setInsertPolicy(QComboBox.NoInsert)
        cb.view().setStyleSheet(
            "background: #ffffff; color: #0f172a; selection-background-color: #dbeafe; selection-color: #0f172a;"
        )
        cb.setMaxVisibleItems(max(len(items), 10))
        return cb

    def setEditorData(self, editor, index):
        if isinstance(editor, QComboBox):
            val = str(index.data(Qt.EditRole) or "")
            idx = editor.findText(val)
            if idx >= 0:
                editor.setCurrentIndex(idx)
            else:
                if self.editable:
                    editor.setEditText(val)
            return
        super().setEditorData(editor, index)

    def setModelData(self, editor, model, index):
        if isinstance(editor, QComboBox):
            model.setData(index, editor.currentText(), Qt.EditRole)
            return
        super().setModelData(editor, model, index)

    def updateEditorGeometry(self, editor, option, index):
        if editor:
            editor.setGeometry(option.rect)
        else:
            super().updateEditorGeometry(editor, option, index)


class DimDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        model = index.model()
        editor = QLineEdit(parent)
        editor.setAlignment(Qt.AlignCenter)
        try:
            if hasattr(model, "_row_is_multipuerta") and model._row_is_multipuerta(index.row()):
                editor.setValidator(QIntValidator(0, 9999, editor))
        except Exception:
            pass
        return editor

    def updateEditorGeometry(self, editor, option, index):
        if editor:
            editor.setGeometry(option.rect)
        else:
            super().updateEditorGeometry(editor, option, index)


class LegendItemsTableModel(QAbstractTableModel):
    headers = PROJ_COLUMNS
    MAX_LW_M = 12.8
    MAX_H_M = 3.7

    def __init__(
        self,
        items: List[dict] | None = None,
        ramal_offset_cb: Callable[[], int] | None = None,
    ):
        super().__init__()
        self.items: List[dict] = items or []
        self._btu_map: Dict[str, float] = {}
        self._cold_engine = None
        self._usage_map: Dict[str, str] = {}
        self._tevap_map: Dict[str, float] = {}
        self._tevap_default: float = 0.0
        self._limit_cb: Callable[[str, float], None] | None = None
        self._ramal_offset_cb: Callable[[], int] | None = ramal_offset_cb
        self.default_row = {
            "loop": 1,
            "ramal": 1,
            "largo_m": "",
            "ancho_m": "",
            "alto_m": "",
            "dim_ft": "",
            "equipo": "",
            "uso": "",
            "btu_ft": 0.0,
            "btu_hr": 0.0,
            "evap_qty": 0,
            "evap_modelo": "",
            "familia": "AUTO",
        }

    def set_limit_callback(self, cb: Callable[[str, float], None] | None) -> None:
        self._limit_cb = cb

    def set_ramal_offset_cb(self, cb: Callable[[], int] | None) -> None:
        self._ramal_offset_cb = cb

    def _ramal_offset(self) -> int:
        if not self._ramal_offset_cb:
            return 0
        try:
            return int(self._ramal_offset_cb() or 0)
        except Exception:
            return 0

    def rowCount(self, parent=QModelIndex()):
        return len(self.items)

    def columnCount(self, parent=QModelIndex()):
        return len(self.headers)

    def data(self, index: QModelIndex, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        item = self.items[index.row()]
        key = self.headers[index.column()]
        if role in (Qt.DisplayRole, Qt.EditRole):
            if key == "ramal":
                val = item.get("ramal", "")
                try:
                    val_int = int(val)
                except Exception:
                    return val
                offset = self._ramal_offset()
                if offset > 0:
                    return val_int + offset
                return val_int
            if key == "largo_m":
                return item.get("largo_m", item.get("dim_m", ""))
            if key == "btu_ft":
                val = item.get("btu_ft", item.get("btu_hr_ft", ""))
                try:
                    return int(round(float(val)))
                except Exception:
                    return val
            if key == "btu_hr":
                val = item.get("btu_hr", item.get("carga_btu_h", ""))
                try:
                    return int(round(float(val)))
                except Exception:
                    return val
            if key == "dim_ft":
                return self._dim_ft_text(index.row())
            if key == "familia" and not self._row_is_cuarto(index.row()):
                return ""
            if key in ("ancho_m", "alto_m") and not self._row_is_cuarto(index.row()):
                return ""
            return item.get(key, "")
        if role == Qt.TextAlignmentRole:
            if key in ("largo_m", "ancho_m", "alto_m") and not self._row_is_cuarto(index.row()):
                return Qt.AlignHCenter | Qt.AlignVCenter
            if key == "dim_ft":
                return Qt.AlignHCenter | Qt.AlignVCenter
            if key in ("loop", "ramal", "largo_m", "ancho_m", "alto_m", "btu_ft", "btu_hr", "evap_qty"):
                return Qt.AlignRight | Qt.AlignVCenter
            return Qt.AlignLeft | Qt.AlignVCenter
        return None

    def setData(self, index: QModelIndex, value, role=Qt.EditRole):
        if role == Qt.EditRole and index.isValid():
            key = self.headers[index.column()]
            if key == "ramal":
                val = self._parse_float(value)
                try:
                    val_int = int(round(val))
                except Exception:
                    val_int = 1
                offset = self._ramal_offset()
                if offset > 0:
                    val_int = max(val_int - offset, 1)
                item = self.items[index.row()]
                item["ramal"] = val_int
                self.dataChanged.emit(index, index, [Qt.DisplayRole, Qt.EditRole])
                return True
            if key in ("largo_m", "ancho_m", "alto_m"):
                val = self._parse_float(value)
                if self._row_is_cuarto(index.row()):
                    max_val = self.MAX_LW_M if key in ("largo_m", "ancho_m") else self.MAX_H_M
                    if val > max_val:
                        val = max_val
                        if self._limit_cb:
                            label = "LARGO" if key == "largo_m" else "ANCHO" if key == "ancho_m" else "ALTO"
                            self._limit_cb(label, max_val)
                if key == "largo_m":
                    self.items[index.row()]["largo_m"] = val
                    self.items[index.row()]["dim_m"] = val
                else:
                    self.items[index.row()][key] = val
                self.dataChanged.emit(index, index, [Qt.DisplayRole, Qt.EditRole])
                self._recompute_row(index.row())
                return True
            if key in ("ancho_m", "alto_m", "uso", "evap_qty", "familia"):
                self.items[index.row()][key] = value
                self.dataChanged.emit(index, index, [Qt.DisplayRole, Qt.EditRole])
                self._recompute_row(index.row())
                return True
            if key == "equipo":
                self.items[index.row()][key] = value
                self.dataChanged.emit(index, index, [Qt.DisplayRole, Qt.EditRole])
                self._recompute_row(index.row())
                return True
            elif key == "btu_hr":
                self.items[index.row()]["btu_hr"] = value
                self.items[index.row()]["carga_btu_h"] = value
            elif key == "btu_ft":
                self.items[index.row()]["btu_ft"] = value
                self.items[index.row()]["btu_hr_ft"] = value
            else:
                self.items[index.row()][key] = value
            self.dataChanged.emit(index, index, [Qt.DisplayRole, Qt.EditRole])
            return True
        return False

    def flags(self, index: QModelIndex):
        if not index.isValid():
            return Qt.ItemIsEnabled
        key = self.headers[index.column()]
        if key in ("btu_ft", "btu_hr", "dim_ft"):
            return Qt.ItemIsEnabled | Qt.ItemIsSelectable
        if key == "familia" and not self._row_is_cuarto(index.row()):
            return Qt.ItemIsEnabled | Qt.ItemIsSelectable
        if key in ("ancho_m", "alto_m") and not self._row_is_cuarto(index.row()):
            return Qt.ItemIsEnabled | Qt.ItemIsSelectable
        return Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable

    def headerData(self, section: int, orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole and orientation == Qt.Horizontal:
            key = self.headers[section]
            if key == "largo_m":
                return "LARGO"
            if key == "ancho_m":
                return "ANCHO"
            if key == "alto_m":
                return "ALTO"
            if key == "btu_ft":
                return "BTU/FT"
            if key == "btu_hr":
                return "BTU/HR"
            if key == "dim_ft":
                return "DIMENSIONES FT"
            if key == "evap_modelo":
                return "MODELO"
            if key == "evap_qty":
                return "# EVAPORADORES"
            if key == "familia":
                return "FAMILIA"
            return key.upper()
        return None

    def _row_is_cuarto(self, row: int) -> bool:
        try:
            equipo = str(self.items[row].get("equipo", "")).upper()
            return "CUARTO" in equipo
        except Exception:
            return False

    def _row_is_multipuerta(self, row: int) -> bool:
        try:
            equipo = str(self.items[row].get("equipo", "")).upper()
            if "MULTIPUERTA" not in equipo:
                return False
            return "CONGEL" in equipo or "REFRIG" in equipo or "CONSERV" in equipo
        except Exception:
            return False

    def _m_to_ft_round(self, m: float) -> int:
        try:
            if self._cold_engine and hasattr(self._cold_engine, "m_to_ft_round"):
                return int(self._cold_engine.m_to_ft_round(m))
        except Exception:
            pass
        return int(round(m * 3.28))

    def _auto_evap_qty_furniture(self, row: int, largo_m: float) -> int:
        if largo_m <= 0:
            return 0
        if self._row_is_multipuerta(row):
            doors = int(round(largo_m))
            if doors <= 0:
                return 0
            if doors == 1:
                return 1
            # Preferir juegos de 4, luego 3 y 2 (evitar 1 si se puede)
            max4 = doors // 4
            for n4 in range(max4, -1, -1):
                rem4 = doors - (4 * n4)
                if rem4 == 0:
                    return n4
                max3 = rem4 // 3
                for n3 in range(max3, -1, -1):
                    rem = rem4 - (3 * n3)
                    if rem == 0:
                        return n4 + n3
                    if rem % 2 == 0:
                        n2 = rem // 2
                        return n4 + n3 + n2
            return doors
        # Preferencias en metros (misma lógica que Carga Eléctrica)
        preferred = [3.75, 2.5, 1.9]
        rem = largo_m
        modules = 0
        limit = 0
        while rem > 0.2 and limit < 200:
            pick = None
            for p in preferred:
                if p <= rem + 0.05:
                    pick = p
                    break
            if pick is None:
                pick = preferred[-1]
            rem -= pick
            modules += 1
            limit += 1
        if modules == 0:
            modules = 1
        return modules

    def _dim_ft_text(self, row: int) -> str:
        if row < 0 or row >= len(self.items):
            return ""
        item = self.items[row]
        largo = self._parse_float(item.get("largo_m", item.get("dim_m", 0)))
        if self._row_is_cuarto(row):
            ancho = self._parse_float(item.get("ancho_m", 0))
            alto = self._parse_float(item.get("alto_m", 0))
            if largo <= 0 or ancho <= 0 or alto <= 0:
                return ""
            l_ft = self._m_to_ft_round(largo)
            w_ft = self._m_to_ft_round(ancho)
            h_ft = self._m_to_ft_round(alto)
            return f"{l_ft}x{w_ft}x{h_ft}"
        if self._row_is_multipuerta(row):
            puertas = int(round(largo)) if largo > 0 else 0
            return f"{puertas} PUERTAS" if puertas > 0 else ""
        if largo <= 0:
            return ""
        ft = largo * 3.28084
        return f"{ft:.2f}"

    def add_row(self, row: dict) -> None:
        self.beginInsertRows(QModelIndex(), len(self.items), len(self.items))
        if not row:
            row = dict(self.default_row)
        else:
            norm = dict(self.default_row)
            norm.update(row)
            row = norm
        self.items.append(row)
        self.endInsertRows()
        self._recompute_row(len(self.items) - 1)

    def add_rows(self, rows: List[dict]) -> None:
        if not rows:
            return
        start = len(self.items)
        end = start + len(rows) - 1
        self.beginInsertRows(QModelIndex(), start, end)
        for row in rows:
            if not row:
                row = dict(self.default_row)
            else:
                norm = dict(self.default_row)
                norm.update(row)
                row = norm
            self.items.append(row)
        self.endInsertRows()
        for r in range(start, end + 1):
            self._recompute_row(r)

    def insert_rows_at(self, index: int, rows: List[dict]) -> None:
        if not rows:
            return
        index = max(0, min(index, len(self.items)))
        start = index
        end = start + len(rows) - 1
        self.beginInsertRows(QModelIndex(), start, end)
        insert_list: List[dict] = []
        for row in rows:
            if not row:
                row = dict(self.default_row)
            else:
                norm = dict(self.default_row)
                norm.update(row)
                row = norm
            insert_list.append(row)
        self.items[index:index] = insert_list
        self.endInsertRows()
        for r in range(start, end + 1):
            self._recompute_row(r)

    def del_row(self, idx: int) -> None:
        if 0 <= idx < len(self.items):
            self.beginRemoveRows(QModelIndex(), idx, idx)
            self.items.pop(idx)
            self.endRemoveRows()

    def dup_row(self, idx: int) -> None:
        if 0 <= idx < len(self.items):
            row = dict(self.items[idx])
            self.add_row(row)

    def set_items(self, items: List[dict]) -> None:
        self.beginResetModel()
        if not isinstance(items, list):
            items = []
        self.items = [r for r in items if isinstance(r, dict)]
        self.endResetModel()
        self.recompute_all()

    def set_btu_map(self, btu_map: Dict[str, float]) -> None:
        self._btu_map = btu_map or {}
        self.recompute_all()

    def set_cold_engine(self, engine) -> None:
        self._cold_engine = engine
        self.recompute_all()

    def set_usage_map(self, usage_map: Dict[str, str]) -> None:
        self._usage_map = usage_map or {}
        self.recompute_all()

    def set_tevap_map(self, tevap_map: Dict[str, float], default_val: float) -> None:
        norm: Dict[str, float] = {}
        for k, v in (tevap_map or {}).items():
            key = " ".join(str(k or "").strip().upper().split())
            try:
                norm[key] = float(v)
            except Exception:
                continue
        self._tevap_map = norm
        try:
            self._tevap_default = float(default_val)
        except Exception:
            self._tevap_default = 0.0
        self.recompute_all()

    def recompute_all(self) -> None:
        if not self.items:
            return
        for i in range(len(self.items)):
            self._recompute_row(i, emit=False)
        try:
            col_ft = self.headers.index("btu_ft")
            col_hr = self.headers.index("btu_hr")
            col_model = self.headers.index("evap_modelo")
            col_qty = self.headers.index("evap_qty")
            col_dim = self.headers.index("dim_ft")
            col_start = min(col_ft, col_hr, col_model, col_qty, col_dim)
            col_end = max(col_ft, col_hr, col_model, col_qty, col_dim)
            top = self.index(0, col_start)
            bottom = self.index(len(self.items) - 1, col_end)
            self.dataChanged.emit(top, bottom, [Qt.DisplayRole, Qt.EditRole])
        except Exception:
            pass

    def _normalize_equipo(self, text: str) -> str:
        return " ".join(str(text or "").strip().upper().split())

    def _parse_float(self, val) -> float:
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        try:
            s = str(val).strip().replace(",", ".")
            return float(s) if s else 0.0
        except Exception:
            return 0.0

    def _map_usage(self, usage: str) -> str:
        u = " ".join(str(usage or "").strip().upper().split())
        return self._usage_map.get(u, u)

    def _tevap_for_usage(self, usage: str) -> float:
        u = self._map_usage(usage)
        return float(self._tevap_map.get(u, self._tevap_default))

    def _recompute_row(self, row: int, emit: bool = True) -> None:
        if row < 0 or row >= len(self.items):
            return
        item = self.items[row]
        equipo = self._normalize_equipo(item.get("equipo", ""))
        btu_ft = float(self._btu_map.get(equipo, 0.0))
        raw_largo = self._parse_float(item.get("largo_m", item.get("dim_m", 0)))
        if self._row_is_cuarto(row):
            largo = min(raw_largo, self.MAX_LW_M)
        else:
            largo = raw_largo
        item["largo_m"] = largo
        item["dim_m"] = largo
        if self._row_is_cuarto(row) and self._cold_engine and ColdRoomInputs:
            ancho = min(self._parse_float(item.get("ancho_m", 0)), self.MAX_LW_M)
            alto = min(self._parse_float(item.get("alto_m", 0)), self.MAX_H_M)
            item["ancho_m"] = ancho
            item["alto_m"] = alto
            evap_model = ""
            evap_qty_out = 0
            if largo > 0 and ancho > 0 and alto > 0:
                uso = self._map_usage(item.get("uso", ""))
                item["tevap_f"] = self._tevap_for_usage(uso)
                evap_qty = int(self._parse_float(item.get("evap_qty", 0)))
                n_evap = evap_qty if evap_qty > 0 else None
                fam_raw = " ".join(str(item.get("familia", "")).strip().upper().split())
                fam_override = None
                if fam_raw and fam_raw != "AUTO":
                    if "BAJA" in fam_raw:
                        fam_override = "frontal_wef"
                    elif "MEDIA" in fam_raw:
                        fam_override = "frontal_wefm"
                    elif "DUAL" in fam_raw:
                        fam_override = "dual"
                try:
                    inp = ColdRoomInputs(
                        length_m=largo,
                        width_m=ancho,
                        height_m=alto,
                        usage=uso,
                        n_evaporators=n_evap,
                        family_override=fam_override,
                    )
                    res = self._cold_engine.compute(inp)
                    if res and res.valid:
                        btu_hr = float(res.load_btu_hr or 0.0)
                        evap_model = str(res.evap_model or "")
                        evap_qty_out = int(res.n_used or n_evap or 0)
                    else:
                        btu_hr = 0.0
                except Exception:
                    btu_hr = 0.0
            else:
                btu_hr = 0.0
            btu_ft = 0.0
            item["evap_modelo"] = evap_model
            item["evap_qty"] = evap_qty_out
        elif self._row_is_multipuerta(row):
            largo_int = int(round(largo))
            item["largo_m"] = largo_int
            item["dim_m"] = largo_int
            btu_hr = btu_ft * largo_int
            item["tevap_f"] = self._tevap_for_usage(item.get("uso", ""))
            item["evap_modelo"] = ""
            if int(self._parse_float(item.get("evap_qty", 0))) <= 0:
                item["evap_qty"] = self._auto_evap_qty_furniture(row, largo_int)
        else:
            btu_hr = btu_ft * (largo * 3.28084)
            item["tevap_f"] = self._tevap_for_usage(item.get("uso", ""))
            item["evap_modelo"] = ""
            if int(self._parse_float(item.get("evap_qty", 0))) <= 0:
                item["evap_qty"] = self._auto_evap_qty_furniture(row, largo)
        item["btu_ft"] = btu_ft
        item["btu_hr"] = btu_hr
        item["btu_hr_ft"] = btu_ft
        item["carga_btu_h"] = btu_hr
        item["dim_ft"] = self._dim_ft_text(row)
        if emit:
            try:
                col_ft = self.headers.index("btu_ft")
                col_hr = self.headers.index("btu_hr")
                col_model = self.headers.index("evap_modelo")
                col_qty = self.headers.index("evap_qty")
                col_l = self.headers.index("largo_m")
                col_dim = self.headers.index("dim_ft")
                col_start = min(col_l, col_ft, col_hr, col_model, col_qty, col_dim)
                col_end = max(col_l, col_ft, col_hr, col_model, col_qty, col_dim)
                top = self.index(row, col_start)
                bottom = self.index(row, col_end)
                self.dataChanged.emit(top, bottom, [Qt.DisplayRole, Qt.EditRole])
            except Exception:
                pass


if __name__ == "__main__":
    from PySide6.QtWidgets import QApplication
    import sys

    app = QApplication(sys.argv)
    w = LegendPage()
    w.resize(1000, 700)
    w.show()
    sys.exit(app.exec())
