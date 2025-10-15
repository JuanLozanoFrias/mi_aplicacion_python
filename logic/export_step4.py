# logic/export_step4.py
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple
import json

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .step4_engine import Step4Engine
from .opciones_co2_engine import ResumenTable
from .resumen_cables import build_cable_badges


# ----------------------------- helpers de estilo -----------------------------
_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

F_WHITE = PatternFill("solid", fgColor="FFFFFF")
F_GREY = PatternFill("solid", fgColor="F2F4F7")      # gris muy claro
F_GREY2 = PatternFill("solid", fgColor="E9EDF3")     # gris claro para cabeceras

FONT_TITLE = Font(name="Calibri", bold=True, size=12, color="000000")
FONT_HEADER = Font(name="Calibri", bold=True, size=11, color="000000")
FONT_CELL = Font(name="Calibri", size=11, color="000000")


def _autosize(ws) -> None:
    """Autoajuste simple por ancho de texto."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = max(10, min(90, max_len + 2))


def _write_meta(ws, meta: List[Tuple[str, str]], start_row: int = 1) -> int:
    """Escribe bloques clave/valor (PROYECTO, CIUDAD, …) con estilo claro."""
    r = start_row
    for label, value in meta:
        c1 = ws.cell(row=r, column=1, value=label)
        c2 = ws.cell(row=r, column=2, value=value)
        for c in (c1, c2):
            c.fill = F_GREY
            c.font = FONT_CELL if c is c2 else FONT_HEADER
            c.alignment = Alignment(vertical="center")
            c.border = _BORDER
        r += 1
    return r + 1  # deja una fila en blanco


def _write_section_header(ws, title: str, row: int) -> int:
    c = ws.cell(row=row, column=1, value=title)
    c.font = FONT_TITLE
    c.fill = F_GREY2
    c.alignment = Alignment(vertical="center")
    c.border = _BORDER
    # banda hasta la J (10 columnas)
    for j in range(2, 11):
        cc = ws.cell(row=row, column=j, value=None)
        cc.fill = F_GREY2
        cc.border = _BORDER
    return row + 1


def _write_subtitle(ws, parts: List[str], row: int) -> int:
    """Línea de subtítulo (compresor / corriente / cable / cap / ICOND) en negro."""
    text = "  |  ".join([p for p in parts if p])
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_CELL  # NEGRO
    c.alignment = Alignment(vertical="center")
    c.border = _BORDER
    for j in range(2, 11):
        cc = ws.cell(row=row, column=j)
        cc.border = _BORDER
    return row + 1


def _write_table(ws, headers: List[str], rows: List[List[str]], row: int) -> int:
    # cabecera
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.font = FONT_HEADER
        cell.fill = F_GREY2
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
    row += 1

    # filas
    for rdata in rows:
        for j, val in enumerate(rdata, start=1):
            cell = ws.cell(row=row, column=j, value=val)
            cell.font = FONT_CELL
            cell.fill = F_WHITE
            cell.alignment = Alignment(vertical="center")
            cell.border = _BORDER
        row += 1
    return row + 1  # espacio


def _sanitize_name(name: str) -> str:
    keep = "".join(ch for ch in (name or "") if ch.isalnum() or ch in (" ", "-", "_"))
    return keep.strip().replace(" ", "_") or "Proyecto"


def _ensure_item_col(rows: List[List[str]]) -> List[List[str]]:
    """
    Si una fila trae 8 columnas (sin ITEM), usa la REFERENCIA como ITEM.
    Formato de 8 cols esperado:
      [CÓDIGO, MODELO, NOMBRE, DESCRIPCIÓN, ICC240, ICC480, REFERENCIA, TORQUE]
    """
    out: List[List[str]] = []
    for r in rows or []:
        if len(r) == 8:
            item = r[6] if len(r) > 6 else ""
            out.append([item] + r)
        else:
            out.append(r)
    return out


def _collect_totales(sections: List[ResumenTable], otros: List[List[str]]) -> List[List[str]]:
    """
    Suma por CÓDIGO. Entrada: filas con columnas:
    [ITEM, CÓDIGO, MODELO, NOMBRE, DESCRIPCIÓN, ICC240, ICC480, REF, TORQUE]
    """
    agg: Dict[str, Dict[str, object]] = {}

    def add_row(row: List[str]) -> None:
        if len(row) < 2:
            return
        code = (row[1] or "").strip()
        if not code:
            return
        if code not in agg:
            agg[code] = {
                "codigo": code,
                "modelo": row[2] if len(row) > 2 else "",
                "nombre": row[3] if len(row) > 3 else "",
                "descripcion": row[4] if len(row) > 4 else "",
                "icc240": row[5] if len(row) > 5 else "",
                "icc480": row[6] if len(row) > 6 else "",
                "ref": row[7] if len(row) > 7 else "",
                "torque": row[8] if len(row) > 8 else "",
                "cantidad": 0,
            }
        agg[code]["cantidad"] = int(agg[code]["cantidad"]) + 1  # type: ignore

    for t in sections:
        for r in t.rows:
            add_row(r)
    for r in (otros or []):
        add_row(r)

    # orden por código
    ordered = sorted(agg.values(), key=lambda d: str(d["codigo"]))
    out: List[List[str]] = []
    for d in ordered:
        out.append([
            d["codigo"], d["modelo"], d["nombre"], d["descripcion"],
            d["icc240"], d["icc480"], d["ref"], d["torque"],
            str(d["cantidad"]),
        ])
    return out


# ----------------------------- API de exportación ----------------------------
def export_step4_excel_only(
    basedatos_path: Path | str,
    step2_state: Dict[str, Dict[str, str]],
    globs: Dict[str, object],
    out_dir: Path | str,
) -> Path:
    """
    Genera SOLO el Excel (hojas: Resumen, Totales, Borneras) con estilo claro.
    Devuelve la ruta del archivo.
    """
    book = Path(basedatos_path)
    engine = Step4Engine(book)
    res = engine.calcular(step2_state, globs)

    tables: List[ResumenTable] = res.get("tables_compresores", []) or []
    otros_rows: List[List[str]] = _ensure_item_col(res.get("otros_rows", []) or [])
    b_comp = res.get("borneras_compresores", {"fase": 0, "neutro": 0, "tierra": 0})
    b_otros = res.get("borneras_otros", {"fase": 0, "neutro": 0, "tierra": 0})
    b_total = res.get("borneras_total", {"fase": 0, "neutro": 0, "tierra": 0})

    # badges de cable para subtítulo por compresor
    cable_badges = build_cable_badges(step2_state, book)

    wb = Workbook()
    ws = wb.active
    ws.title = "RESUMEN"

    # ---- Encabezado del proyecto (print-friendly) ----
    meta = [
        ("PROYECTO",       str(globs.get("nombre_proyecto", ""))),
        ("CIUDAD",         str(globs.get("ciudad", ""))),
        ("RESPONSABLE",    str(globs.get("responsable", ""))),
        ("TENSIÓN ALIM.",  str(globs.get("t_alim", ""))),
        ("TENSIÓN CONTROL",str(globs.get("t_ctl", ""))),
        ("NORMA",          str(globs.get("norma_ap", ""))),
    ]
    row = _write_meta(ws, meta, start_row=1)

    # ---- Secciones por compresor ----
    headers = ["ITEM", "CÓDIGO", "MODELO", "NOMBRE", "DESCRIPCIÓN",
               "C 240 V (kA)", "C 480 V (kA)", "REFERENCIA", "TORQUE"]

    for t in tables:
        row = _write_section_header(ws, t.title, row)
        # subtítulo: compresor + corriente + cable
        parts: List[str] = []
        st = step2_state.get(t.comp_key, {}) or {}
        _, model, amps = (st.get("modelo", ""), st.get("modelo", ""), st.get("corriente", ""))
        if st.get("modelo") or st.get("marca"):
            parts.append(f"COMPRESOR: {st.get('modelo','')}".strip())
        if amps:
            parts.append(f"CORRIENTE: {amps} A")
        badge = cable_badges.get(t.comp_key)
        if badge:
            parts.append(badge.strip(" |"))
        row = _write_subtitle(ws, parts, row)
        row = _write_table(ws, headers, t.rows, row)

    # ---- Otros fijos ----
    if otros_rows:
        row = _write_section_header(ws, "OTROS ELEMENTOS (FIJOS)", row)
        row = _write_table(ws, headers, otros_rows, row)

    _autosize(ws)

    # ---- Totales (por código) ----
    ws2 = wb.create_sheet("TOTALES")
    tot_rows = _collect_totales(tables, otros_rows)
    headers_tot = ["CÓDIGO", "MODELO", "NOMBRE", "DESCRIPCIÓN",
                   "C 240 V (kA)", "C 480 V (kA)", "REFERENCIA", "TORQUE", "CANTIDAD"]
    _ = _write_table(ws2, headers_tot, tot_rows, 1)
    _autosize(ws2)

    # ---- Borneras ----
    ws3 = wb.create_sheet("BORNERAS")
    heads = ["TIPO", "TOTAL"]
    _ = _write_table(
        ws3, heads,
        [["FASE", b_total["fase"]], ["NEUTRO", b_total["neutro"]], ["TIERRA", b_total["tierra"]]],
        1,
    )
    _autosize(ws3)

    # ---- Guardar ----
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    project = _sanitize_name(str(globs.get("nombre_proyecto", ""))).upper()
    stamp = datetime.now().strftime("%Y%m%d")
    out_path = out_dir / f"{stamp}_{project}.xlsx"
    wb.save(out_path)
    return out_path


# ----------------------- snapshot de programación (.ecalc.json) ----------------
def _normalize_step2_for_json(step2: Dict[str, Dict[str, str]]) -> Dict[str, Dict[str, str]]:
    """Garantiza tipos JSON-friendly."""
    out: Dict[str, Dict[str, str]] = {}
    for k, v in (step2 or {}).items():
        rec: Dict[str, str] = {}
        if isinstance(v, dict):
            for kk, vv in v.items():
                rec[str(kk)] = "" if vv is None else str(vv)
        out[str(k)] = rec
    return out


def save_programacion_snapshot(
    step2_state: Dict[str, Dict[str, str]],
    globs: Dict[str, object],
    out_dir: Path | str,
    filename_prefix: str | None = None,
) -> Path:
    """
    Guarda un snapshot .ecalc.json con TODO lo necesario para recargar.
    Nombre: YYYYMMDD_<PROYECTO>.ecalc.json (o prefix si se indica)
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    project = _sanitize_name(str(globs.get("nombre_proyecto", ""))).upper()
    stamp = datetime.now().strftime("%Y%m%d")
    base = filename_prefix if filename_prefix else project
    fname = f"{stamp}_{base}.ecalc.json"
    out_path = out_dir / fname

    payload = {
        "globs": {
            "nombre_proyecto": str(globs.get("nombre_proyecto", "")),
            "ciudad": str(globs.get("ciudad", "")),
            "responsable": str(globs.get("responsable", "")),
            "t_alim": str(globs.get("t_alim", "")),
            "t_ctl": str(globs.get("t_ctl", "")),
            "norma_ap": str(globs.get("norma_ap", "")),
            "refrigerante": str(globs.get("refrigerante", "")),
            "marca_elem": str(globs.get("marca_elem", "")),
            "marca_variadores": str(globs.get("marca_variadores", "")),
            "tipo_compresores": str(globs.get("tipo_compresores", "")),
            "step3_state": globs.get("step3_state", {}),
        },
        "step2": _normalize_step2_for_json(step2_state),
    }

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    return out_path


def save_programacion_exact(
    step2_state: Dict[str, Dict[str, str]],
    globs: Dict[str, object],
    out_path: Path | str,
) -> Path:
    """
    Guarda la programación EXACTAMENTE en 'out_path' (p.ej. 'C:/.../MI_PROYECTO.ecalc.json').
    - Crea la carpeta si no existe.
    - Genera JSON correcto y sustituye si ya existe.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Asegurar extensión .ecalc.json
    if not out_path.name.lower().endswith(".ecalc.json"):
        out_path = out_path.with_name(out_path.stem + ".ecalc.json")

    # Generar snapshot temporal con el mismo "stem"
    tmp = save_programacion_snapshot(
        step2_state=step2_state,
        globs=globs,
        out_dir=out_path.parent,
        filename_prefix=out_path.stem,
    )

    # Reemplazar/renombrar al nombre exacto
    try:
        tmp.replace(out_path)  # atómico cuando es posible
    except Exception:
        data = tmp.read_bytes()
        out_path.write_bytes(data)
        try:
            tmp.unlink()
        except Exception:
            pass

    return out_path
