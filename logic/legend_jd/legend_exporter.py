from __future__ import annotations

from copy import copy
from pathlib import Path
import re
import os
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from typing import Any, Dict, Iterable, List, Tuple
from datetime import datetime

import json
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import unicodedata

try:
    from logic.legend.eev_calc import compute_eev
except Exception:
    compute_eev = None  # type: ignore

def _to_float(val: Any, default: float = 0.0) -> float:
    try:
        if isinstance(val, str):
            val = val.replace(",", ".").strip()
        return float(val)
    except Exception:
        return default


def _to_int(val: Any, default: int = 0) -> int:
    try:
        if isinstance(val, str):
            val = val.strip()
        return int(float(val))
    except Exception:
        return default


def _norm_label(text: Any) -> str:
    s = str(text or "").strip().upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace(":", "").replace("°", "").replace("º", "")
    s = " ".join(s.split())
    return s


def _find_label_cell(ws, label: str, max_row: int = 20) -> Tuple[int, int] | None:
    target = _norm_label(label)
    for r in range(1, max_row + 1):
        for c in range(1, ws.max_column + 1):
            if _norm_label(ws.cell(r, c).value) == target:
                return r, c
    return None


def _find_all_label_cells(ws, label: str, max_row: int = 200) -> List[Tuple[int, int]]:
    target = _norm_label(label)
    found: List[Tuple[int, int]] = []
    for r in range(1, min(max_row, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            if _norm_label(ws.cell(r, c).value) == target:
                found.append((r, c))
    return found


def _find_label_cell_in_rows(
    ws, label: str, start_row: int, end_row: int
) -> Tuple[int, int] | None:
    target = _norm_label(label)
    for r in range(start_row, min(end_row, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            if _norm_label(ws.cell(r, c).value) == target:
                return r, c
    return None


def _find_cell_contains(ws, text: str, max_row: int | None = None) -> Tuple[int, int] | None:
    target = _norm_label(text)
    max_r = max_row or ws.max_row
    for r in range(1, max_r + 1):
        for c in range(1, ws.max_column + 1):
            val = _norm_label(ws.cell(r, c).value)
            if val and target in val:
                return r, c
    return None


def _merged_anchor_cell(ws, row: int, col: int) -> Tuple[int, int]:
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row, rng.min_col
    return row, col


def _col_after_label(ws, row: int, col: int) -> Tuple[int, int]:
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return row, rng.max_col + 1
    return row, col + 1


def _find_header_row(ws, max_row: int = 30) -> int | None:
    for r in range(1, max_row + 1):
        row_vals = [_norm_label(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if "RAMAL" in row_vals and "EQUIPO" in row_vals and "USO" in row_vals:
            return r
    return None


def _find_comp_section_row(ws, max_row: int = 200) -> int | None:
    for r in range(1, min(max_row, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            val = _norm_label(ws.cell(r, c).value)
            if not val:
                continue
            if "COMPRESORES" in val:
                return r
    return None


def _clear_merges_in_rows(ws, min_row: int, max_row: int) -> None:
    if min_row > max_row:
        return
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= min_row and rng.max_row <= max_row:
            ws.unmerge_cells(str(rng))


def _clear_merges_in_cells(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for rng in list(ws.merged_cells.ranges):
        if rng.max_row < min_row or rng.min_row > max_row:
            continue
        if rng.max_col < min_col or rng.min_col > max_col:
            continue
        try:
            ws.unmerge_cells(str(rng))
        except KeyError:
            try:
                ws.merged_cells.ranges.remove(rng)
            except (ValueError, KeyError):
                pass

def _clear_merges_on_row(ws, row: int) -> None:
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= row <= rng.max_row:
            try:
                ws.unmerge_cells(str(rng))
            except KeyError:
                try:
                    ws.merged_cells.ranges.remove(rng)
                except (ValueError, KeyError):
                    pass


def _merge_safe(ws, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    if start_row > end_row or start_col > end_col:
        return
    if start_row == end_row and start_col == end_col:
        return
    rng = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    if any(str(r) == rng for r in ws.merged_cells.ranges):
        return
    ws.merge_cells(rng)


def _find_label_row(ws, label: str, max_row: int = 200) -> int | None:
    pos = _find_label_cell(ws, label, max_row=max_row)
    if pos:
        return pos[0]
    return None


def _comp_position_cols(ws, pos_row: int) -> List[Tuple[int, int]]:
    cols: List[Tuple[int, int]] = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(pos_row, c).value
        if isinstance(val, (int, float)):
            cols.append((_to_int(val), c))
    cols.sort(key=lambda t: t[0])
    return cols


def _merged_range_for_label(ws, row: int, label: str) -> Tuple[int, int] | None:
    target = _norm_label(label)
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row:
            cell_val = ws.cell(rng.min_row, rng.min_col).value
            if _norm_label(cell_val) == target:
                return rng.min_col, rng.max_col
    # fallback to exact cell in row
    for c in range(1, ws.max_column + 1):
        if _norm_label(ws.cell(row, c).value) == target:
            return c, c
    return None


def _expand_comp_models(items: Any) -> List[str]:
    models: List[str] = []
    if not isinstance(items, list):
        return models
    for it in items:
        if not isinstance(it, dict):
            continue
        model = str(it.get("model", "") or "").strip()
        qty = _to_int(it.get("n", 0), 0)
        if model and qty > 0:
            models.extend([model] * qty)
    return models


def _expand_comp_items(items: Any) -> List[Dict[str, Any]]:
    expanded: List[Dict[str, Any]] = []
    if not isinstance(items, list):
        return expanded
    for it in items:
        if not isinstance(it, dict):
            continue
        model = str(it.get("model", "") or "").strip()
        qty = _to_int(it.get("n", 0), 0)
        if not model or qty <= 0:
            continue
        for _ in range(qty):
            expanded.append(dict(it))
    return expanded


def _load_compresores_perf() -> Dict[str, Any]:
    for path in (
        Path("data/LEGEND/compresores_perf.json"),
        Path("data/legend/compresores_perf.json"),
    ):
        if path.exists():
            try:
                return json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                return {}
    return {}


def _norm_ref(text: Any) -> str:
    return _norm_label(text).replace("-", "")


def _parse_perf_points(points: Dict[str, Any]) -> List[Dict[str, Any]]:
    parsed: List[Dict[str, Any]] = []
    for key, val in points.items():
        if not isinstance(key, str):
            continue
        parts = key.split("|")
        if len(parts) < 3:
            continue
        try:
            ref = parts[0].strip()
            tcond = float(parts[1])
            tevap = float(parts[2])
        except Exception:
            continue
        parsed.append({"ref": ref, "tcond_f": tcond, "tevap_f": tevap, "data": val})
    return parsed


def _get_nearest_perf_point(
    perf_db: Dict[str, Any],
    brand: str,
    model: str,
    refrigerante: str,
    tcond_f: float,
    tevap_f: float,
) -> Dict[str, Any] | None:
    brands = perf_db.get("brands", {}) if isinstance(perf_db, dict) else {}
    brand_data = brands.get(brand, {}) if isinstance(brands, dict) else {}
    models = brand_data.get("models", {}) if isinstance(brand_data, dict) else {}
    model_data = models.get(model, {}) if isinstance(models, dict) else {}
    points = model_data.get("points", {}) if isinstance(model_data, dict) else {}
    parsed = _parse_perf_points(points) if isinstance(points, dict) else []
    if not parsed:
        return None
    ref_norm = _norm_ref(refrigerante)
    matches = []
    for p in parsed:
        pref = _norm_ref(p.get("ref", ""))
        if ref_norm and pref and ref_norm in pref:
            matches.append(p)
    candidates = matches if matches else parsed
    best = None
    best_score = None
    for p in candidates:
        score = abs(p["tcond_f"] - tcond_f) + abs(p["tevap_f"] - tevap_f)
        if best is None or best_score is None or score < best_score:
            best = p
            best_score = score
    return best.get("data") if best else None


def _compute_tevap_design(items: List[Dict[str, Any]]) -> float | None:
    vals: List[float] = []
    for it in items:
        load = _to_float(it.get("btu_hr", it.get("carga_btu_h", 0)) or 0)
        if load <= 0:
            continue
        tev = _to_float(it.get("tevap_f", it.get("tevap", 0)) or 0, default=None)  # type: ignore
        if tev is None:
            continue
        vals.append(tev)
    if not vals:
        return None
    return min(vals)


def _format_fecha_es(value: Any) -> str:
    months = [
        "ENERO",
        "FEBRERO",
        "MARZO",
        "ABRIL",
        "MAYO",
        "JUNIO",
        "JULIO",
        "AGOSTO",
        "SEPTIEMBRE",
        "OCTUBRE",
        "NOVIEMBRE",
        "DICIEMBRE",
    ]
    if isinstance(value, datetime):
        dt = value
    else:
        try:
            dt = datetime.fromisoformat(str(value))
        except Exception:
            dt = datetime.now()
    return f"{months[dt.month - 1]} {dt.day} DE {dt.year}"


def _write_next_label(
    ws, label: str, value: Any, number_format: str | None = None, uppercase: bool = False
) -> bool:
    pos = _find_label_cell(ws, label, max_row=ws.max_row)
    if not pos:
        return False
    row, col = _merged_anchor_cell(ws, pos[0], pos[1])
    tgt_row, tgt_col = _col_after_label(ws, row, col)
    tgt_row, tgt_col = _merged_anchor_cell(ws, tgt_row, tgt_col)
    cell = ws.cell(tgt_row, tgt_col)
    if uppercase and isinstance(value, str):
        value = value.upper()
    cell.value = value
    if number_format:
        cell.number_format = number_format
    return True


def _write_next_label_in_rows(
    ws,
    label: str,
    value: Any,
    start_row: int,
    end_row: int,
    number_format: str | None = None,
    uppercase: bool = False,
) -> bool:
    pos = _find_label_cell_in_rows(ws, label, start_row, end_row)
    if not pos:
        return False
    row, col = _merged_anchor_cell(ws, pos[0], pos[1])
    tgt_row, tgt_col = _col_after_label(ws, row, col)
    tgt_row, tgt_col = _merged_anchor_cell(ws, tgt_row, tgt_col)
    cell = ws.cell(tgt_row, tgt_col)
    if uppercase and isinstance(value, str):
        value = value.upper()
    cell.value = value
    if number_format:
        cell.number_format = number_format
    return True


def _write_baja_media(
    ws, label: str, baja: Any, media: Any, number_format: str | None = None
) -> bool:
    pos = _find_label_cell(ws, label, max_row=ws.max_row)
    if not pos:
        return False
    row, col = _merged_anchor_cell(ws, pos[0], pos[1])
    for idx, val in enumerate((baja, media), start=1):
        tgt_row, tgt_col = _merged_anchor_cell(ws, row, col + idx)
        cell = ws.cell(tgt_row, tgt_col)
        cell.value = val
        if number_format:
            cell.number_format = number_format
    return True


def _write_rh_box(ws, rh_bt: float, rh_mt: float, rh_total: float) -> None:
    pos = _find_label_cell(ws, "RH", max_row=ws.max_row)
    if not pos:
        return
    row, col = _merged_anchor_cell(ws, pos[0], pos[1])
    for offset, val in enumerate((rh_bt, rh_mt, rh_total), start=1):
        r, c = _merged_anchor_cell(ws, row + offset, col)
        cell = ws.cell(r, c)
        cell.value = int(round(val))
        cell.number_format = "#,##0"


def _compute_comp_totals(
    project_data: Dict[str, Any],
    perf_db: Dict[str, Any],
    tcond_f: float,
    tevap_bt: float | None,
    tevap_mt: float | None,
    refrigerante: str,
) -> Tuple[float, float, float, float]:
    comp = project_data.get("compressors", {}) if isinstance(project_data, dict) else {}
    brand = str(comp.get("brand", "") or "")
    def _sum_group(items: Any, tevap_design: float | None) -> Tuple[float, float]:
        if tevap_design is None or not brand:
            return 0.0, 0.0
        items_exp = _expand_comp_items(items)
        cap_total = 0.0
        rh_total = 0.0
        for it in items_exp:
            model = str(it.get("model", "") or "")
            if not model:
                continue
            perf = _get_nearest_perf_point(perf_db, brand, model, refrigerante, tcond_f, tevap_design)
            if not perf:
                continue
            cap = _to_float(perf.get("capacity_btu_h", perf.get("evaporator_capacity_btu_h", 0.0)), 0.0)
            rh = _to_float(perf.get("heat_rejection_btu_h", 0.0), 0.0)
            power_kw = _to_float(perf.get("consumption_kw", 0.0), 0.0)
            if not power_kw:
                power_kw = _to_float(perf.get("power_kw", 0.0), 0.0)
            if not power_kw:
                power_kw = _to_float(perf.get("power_w", 0.0), 0.0) / 1000.0
            if rh <= 0 and power_kw > 0:
                rh = cap + power_kw * 3412.142
            cap_total += cap
            rh_total += rh
        return cap_total, rh_total

    cap_bt, rh_bt = _sum_group(comp.get("bt", {}).get("items", []), tevap_bt)
    cap_mt, rh_mt = _sum_group(comp.get("mt", {}).get("items", []), tevap_mt)
    return cap_bt, cap_mt, rh_bt, rh_mt


def _find_label_row_any(ws, labels: List[str], max_row: int = 200) -> int | None:
    for label in labels:
        row = _find_label_row(ws, label, max_row=max_row)
        if row:
            return row
    return None


def _write_compresores(ws, project_data: Dict[str, Any]) -> None:
    comp = project_data.get("compressors", {}) if isinstance(project_data, dict) else {}
    if not isinstance(comp, dict):
        return
    pos_row = _find_label_row(ws, "POSICION DE COMPRESOR EN EL RACK")
    if pos_row is None:
        pos_row = _find_label_row(ws, "POSICION DE  COMPRESOR EN EL RACK")
    model_row = _find_label_row(ws, "MODELO DE COMPRESOR")
    if pos_row is None or model_row is None:
        return

    positions = _comp_position_cols(ws, pos_row)
    if not positions:
        return

    baja_range = _merged_range_for_label(ws, pos_row - 1, "BAJA")
    media_range = _merged_range_for_label(ws, pos_row - 1, "MEDIA")

    bt_cols: List[int] = []
    mt_cols: List[int] = []
    use_merged = False
    if baja_range and media_range and (baja_range != media_range):
        if (baja_range[1] - baja_range[0]) >= 1 or (media_range[1] - media_range[0]) >= 1:
            use_merged = True
    if use_merged:
        for pos, col in positions:
            if baja_range[0] <= col <= baja_range[1]:
                bt_cols.append(col)
            elif media_range[0] <= col <= media_range[1]:
                mt_cols.append(col)
    else:
        media_col = media_range[0] if media_range else None
        if media_col:
            for pos, col in positions:
                if col < media_col:
                    bt_cols.append(col)
                else:
                    mt_cols.append(col)
        else:
            for pos, col in positions:
                if pos <= 3:
                    bt_cols.append(col)
                else:
                    mt_cols.append(col)

    bt_items_exp = _expand_comp_items(comp.get("bt", {}).get("items", []))
    mt_items_exp = _expand_comp_items(comp.get("mt", {}).get("items", []))
    bt_models = [str(it.get("model", "") or "") for it in bt_items_exp]
    mt_models = [str(it.get("model", "") or "") for it in mt_items_exp]

    for idx, col in enumerate(bt_cols):
        val = bt_models[idx] if idx < len(bt_models) else ""
        _safe_set(ws, model_row, col, val)
    for idx, col in enumerate(mt_cols):
        val = mt_models[idx] if idx < len(mt_models) else ""
        _safe_set(ws, model_row, col, val)

    overflow_msgs: List[str] = []
    if bt_cols and len(bt_items_exp) > len(bt_cols):
        overflow_msgs.append("EXCEDE SLOTS BAJA")
    if mt_cols and len(mt_items_exp) > len(mt_cols):
        overflow_msgs.append("EXCEDE SLOTS MEDIA")
    if overflow_msgs:
        pos = _find_label_cell(ws, "NOTAS", max_row=ws.max_row)
        if pos:
            r, c = _col_after_label(ws, pos[0], pos[1])
            _safe_set(ws, r, c, "; ".join(overflow_msgs))

    labels = {
        "DEMAND": ["DEMAND COOLING / CIC", "DEMAND COOLING/CIC"],
        "VENTILADOR": ["VENTILADOR DE CABEZA"],
        "CONTACTOR": ["CONTACTOR"],
        "BREAKER": ["BREAKER"],
        "CONTROL": ["CONTROL DE CAPACIDAD"],
        "REGULADOR": ["REGULADOR DE NIVEL DE ACEITE"],
        "HP": ["HP NOMINAL"],
        "REF": ["REFRIGERANTE"],
        "TSUCCION": ["TEMPERATURA DE SUCCION", "TEMPERATURA DE SUCCIÓN"],
        "TCOND": ["TEMPERATURA DE CONDENSACION", "TEMPERATURA DE CONDENSACIÓN"],
        "CAP": ["CAPACIDAD  BTU/HR", "CAPACIDAD BTU/HR", "CAPACIDAD  BTU/HR"],
        "RH": ["RH  BTU/HR", "RH BTU/HR", "RH  BTU/HR"],
        "RLA": ["RLA"],
        "CONSUMO": ["CONSUMO EN PUNTO DE OPERACION", "CONSUMO EN PUNTO DE OPERACIÓN"],
        "RLA_TOTAL": ["RLA TOTAL"],
        "CONSUMO_TOTAL": ["CONSUMO TOTAL  PUNTO OPERACION", "CONSUMO TOTAL PUNTO OPERACION"],
    }
    row_map = {k: _find_label_row_any(ws, v) for k, v in labels.items()}

    perf_db = _load_compresores_perf()
    specs = project_data.get("specs", {}) if isinstance(project_data, dict) else {}
    tcond_f = _to_float(specs.get("tcond_f", 0) or 0)
    refrigerante = str(specs.get("refrigerante", "") or "").strip()
    bt_items = project_data.get("bt_items", []) if isinstance(project_data, dict) else []
    mt_items = project_data.get("mt_items", []) if isinstance(project_data, dict) else []
    tevap_bt = _compute_tevap_design(bt_items) if isinstance(bt_items, list) else None
    tevap_mt = _compute_tevap_design(mt_items) if isinstance(mt_items, list) else None

    def _control_cap(idx: int) -> str:
        if idx == 0:
            return "VARIADOR"
        if idx == 1:
            return "UNLOADER"
        return "NO"

    def _fill_group(cols: List[int], items_exp: List[Dict[str, Any]], tevap_design: float | None) -> Tuple[float, float]:
        rla_total = 0.0
        power_total = 0.0
        for idx, col in enumerate(cols):
            item = items_exp[idx] if idx < len(items_exp) else {}
            model = str(item.get("model", "") or "")
            perf = None
            if model and tcond_f and tevap_design is not None:
                perf = _get_nearest_perf_point(
                    perf_db, comp.get("brand", ""), model, refrigerante, tcond_f, tevap_design
                )
            cap = _to_float(perf.get("capacity_btu_h", 0) if isinstance(perf, dict) else 0)
            rh = _to_float(perf.get("heat_rejection_btu_h", 0) if isinstance(perf, dict) else 0)
            rla = _to_float(perf.get("rla", 0) if isinstance(perf, dict) else 0)
            power_kw = _to_float(perf.get("consumption_kw", 0) if isinstance(perf, dict) else 0)
            if not power_kw and isinstance(perf, dict):
                power_kw = _to_float(perf.get("power_kw", 0))
            if not power_kw and isinstance(perf, dict):
                power_kw = _to_float(perf.get("power_w", 0)) / 1000.0
            hp = _to_float(perf.get("hp_nominal", 0) if isinstance(perf, dict) else 0)

            ctrl = str(item.get("control_capacidad", "") or "").strip().upper()
            if not ctrl:
                ctrl = _control_cap(idx)
            demand = str(item.get("demand_cooling", "") or "").strip().upper()
            if not demand:
                demand = "SI" if ctrl != "NO" else "NO"
            vent = str(item.get("ventilador_cabeza", "") or "").strip().upper() or "SI"
            contactor = str(item.get("contactor", "") or "").strip().upper() or "SI"
            breaker = str(item.get("breaker", "") or "").strip().upper() or "SI"
            reg_aceite = str(item.get("regulador_aceite", "") or "").strip().upper() or "SI"
            if row_map["CONTROL"]:
                _safe_set(ws, row_map["CONTROL"], col, ctrl if model else "")
            if row_map["DEMAND"]:
                _safe_set(ws, row_map["DEMAND"], col, demand if model else "")
            if row_map["VENTILADOR"]:
                _safe_set(ws, row_map["VENTILADOR"], col, vent if model else "")
            if row_map["CONTACTOR"]:
                _safe_set(ws, row_map["CONTACTOR"], col, contactor if model else "")
            if row_map["BREAKER"]:
                _safe_set(ws, row_map["BREAKER"], col, breaker if model else "")
            if row_map["REGULADOR"]:
                _safe_set(ws, row_map["REGULADOR"], col, reg_aceite if model else "")
            if row_map["HP"]:
                _safe_set(ws, row_map["HP"], col, hp if model else "")
            if row_map["REF"]:
                _safe_set(ws, row_map["REF"], col, refrigerante if model else "")
            if row_map["TSUCCION"]:
                _safe_set(ws, row_map["TSUCCION"], col, tevap_design if model else "")
            if row_map["TCOND"]:
                _safe_set(ws, row_map["TCOND"], col, tcond_f if model else "")
            if row_map["CAP"]:
                _safe_set(ws, row_map["CAP"], col, int(round(cap)) if model else "")
            if row_map["RH"]:
                _safe_set(ws, row_map["RH"], col, int(round(rh)) if model else "")
            if row_map["RLA"]:
                _safe_set(ws, row_map["RLA"], col, round(rla, 2) if model else "")
            if row_map["CONSUMO"]:
                _safe_set(ws, row_map["CONSUMO"], col, round(power_kw, 2) if model else "")

            if model and rla:
                rla_total += rla
            if model and power_kw:
                power_total += power_kw
        return rla_total, power_total

    bt_rla_total, bt_power_total = _fill_group(bt_cols, bt_items_exp, tevap_bt)
    mt_rla_total, mt_power_total = _fill_group(mt_cols, mt_items_exp, tevap_mt)

    total_rla = bt_rla_total + mt_rla_total
    total_power = bt_power_total + mt_power_total
    if row_map["RLA_TOTAL"]:
        pos = _find_label_cell(ws, labels["RLA_TOTAL"][0], max_row=ws.max_row)
        if pos:
            r, c = _col_after_label(ws, pos[0], pos[1])
            _safe_set(ws, r, c, round(total_rla, 2) if total_rla else "")
    if row_map["CONSUMO_TOTAL"]:
        pos = _find_label_cell(ws, labels["CONSUMO_TOTAL"][0], max_row=ws.max_row)
        if pos:
            r, c = _col_after_label(ws, pos[0], pos[1])
            _safe_set(ws, r, c, round(total_power, 2) if total_power else "")


def _cleanup_invalid_merges(ws) -> None:
    seen = set()
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row > rng.max_row or rng.min_col > rng.max_col:
            ws.unmerge_cells(str(rng))
            continue
        key = str(rng)
        if key in seen:
            ws.unmerge_cells(key)
            continue
        seen.add(key)


def _restore_template_merges(
    ws,
    tpl_ranges: List[Tuple[int, int, int, int]],
    row_shift: int,
    min_row_tpl: int,
) -> None:
    if not tpl_ranges or row_shift == 0 and min_row_tpl <= 0:
        return
    max_tpl_row = max(r[2] for r in tpl_ranges)
    target_min = min_row_tpl + row_shift
    target_max = max_tpl_row + row_shift
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= target_min and rng.max_row <= target_max:
            try:
                ws.unmerge_cells(str(rng))
            except KeyError:
                try:
                    ws.merged_cells.ranges.remove(rng)
                except ValueError:
                    pass
    for r1, c1, r2, c2 in tpl_ranges:
        if r1 < min_row_tpl:
            continue
        _merge_safe(ws, r1 + row_shift, c1, r2 + row_shift, c2 + row_shift)


def _row_max_used_col(ws, row: int) -> int | None:
    for c in range(ws.max_column, 0, -1):
        val = ws.cell(row, c).value
        if val not in (None, ""):
            return c
    return None


def _repair_comp_block_merges(ws) -> None:
    sys_row = _find_label_row(ws, "SISTEMA PARALELO", max_row=ws.max_row)
    comp_row = _find_label_row(ws, "COMPRESORES Y ACESORIOS", max_row=ws.max_row)
    pos_row = _find_label_row(ws, "POSICION DE COMPRESOR EN EL RACK", max_row=ws.max_row)
    if not sys_row or not comp_row or not pos_row:
        return
    # locate BAJA/MEDIA columns in comp_row
    baja_col = None
    media_col = None
    for c in range(1, ws.max_column + 1):
        val = _norm_label(ws.cell(comp_row, c).value)
        if val == "BAJA":
            baja_col = c
        elif val == "MEDIA":
            media_col = c
    last_col = _row_max_used_col(ws, pos_row)
    if not last_col:
        last_col = ws.max_column
    # clear merges on comp block rows
    for r in (sys_row, comp_row):
        _clear_merges_on_row(ws, r)
    # rebuild merges
    if last_col and last_col > 1:
        _merge_safe(ws, sys_row, 1, sys_row, last_col)
    if baja_col and media_col and baja_col < media_col:
        _merge_safe(ws, comp_row, 1, comp_row, baja_col - 1)
        _merge_safe(ws, comp_row, baja_col, comp_row, media_col - 1)
        _merge_safe(ws, comp_row, media_col, comp_row, last_col)


def _repair_rh_merges(ws) -> None:
    cap_row = _find_label_row(ws, "CAP", max_row=ws.max_row)
    rh_pos = _find_label_cell(ws, "RH", max_row=ws.max_row)
    if not cap_row or not rh_pos:
        return
    rh_row, rh_col = rh_pos
    # clear merges in RH block area
    _clear_merges_in_cells(ws, rh_row, rh_row + 3, rh_col, rh_col + 1)
    for r in range(rh_row, rh_row + 4):
        _merge_safe(ws, r, rh_col, r, rh_col + 1)


def _prune_overlapping_merges(ws, min_row: int | None = None, max_row: int | None = None) -> None:
    def _area(rng) -> int:
        return (rng.max_row - rng.min_row + 1) * (rng.max_col - rng.min_col + 1)

    def _overlaps(a, b) -> bool:
        return not (
            a.max_row < b.min_row
            or a.min_row > b.max_row
            or a.max_col < b.min_col
            or a.min_col > b.max_col
        )

    ranges = list(ws.merged_cells.ranges)
    if min_row is not None and max_row is not None:
        ranges = [
            r
            for r in ranges
            if not (r.max_row < min_row or r.min_row > max_row)
        ]
    ranges = sorted(ranges, key=lambda r: (r.min_row, r.min_col, r.max_row, r.max_col))
    kept = []
    for rng in ranges:
        conflict = None
        for k in kept:
            if _overlaps(rng, k):
                conflict = k
                break
        if conflict is None:
            kept.append(rng)
            continue
        # si hay solapamiento, mantener el de mayor área
        if _area(rng) > _area(conflict):
            ws.unmerge_cells(str(conflict))
            kept.remove(conflict)
            kept.append(rng)
        else:
            ws.unmerge_cells(str(rng))


def _map_header_cols(ws, header_row: int) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        val = _norm_label(ws.cell(header_row, c).value)
        if not val:
            continue
        mapping.setdefault(val, c)
    return mapping


def _find_btu_col(ws, header_row: int, mapping: Dict[str, int]) -> int | None:
    col = _col_for(mapping, ["CARGA (BTU/HR)", "CARGA (BTU/H)", "CARGA (BTU/HR)"])
    if col:
        return col
    for c in range(1, ws.max_column + 1):
        val = _norm_label(ws.cell(header_row, c).value)
        if "CARGA" in val and "BTU" in val:
            return c
    return None


def _find_tevap_col(ws, header_row: int, mapping: Dict[str, int]) -> int | None:
    col = _col_for(mapping, ["TEVAP (F)", "TEVAP(°F)", "TEVAP"])
    if col:
        return col
    for c in range(1, ws.max_column + 1):
        val = _norm_label(ws.cell(header_row, c).value)
        if "TEVAP" in val:
            return c
    return None


def _force_total_value(ws, label: str, total: float, btu_col: int | None) -> None:
    if not btu_col:
        return
    pos = _find_label_cell(ws, label, max_row=ws.max_row)
    if not pos:
        return
    row = pos[0]
    _safe_set(ws, row, btu_col, int(round(total)))


def _find_subheader_cols(ws, header_row: int, look_ahead: int = 3) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for r in range(header_row + 1, header_row + 1 + look_ahead):
        for c in range(1, ws.max_column + 1):
            val = _norm_label(ws.cell(r, c).value)
            if val in ("CANTIDAD", "MODELO"):
                mapping[val] = c
    return mapping


def _col_for(mapping: Dict[str, int], labels: Iterable[str]) -> int | None:
    for lab in labels:
        key = _norm_label(lab)
        if key in mapping:
            return mapping[key]
    return None


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)


def _copy_row_style_from(ws_src, ws_dst, src_row: int, dst_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws_src.cell(src_row, col)
        dst = ws_dst.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)


def _restore_template_block(ws, tpl_ws, start_row_tpl: int, start_row_new: int) -> None:
    if start_row_tpl is None or start_row_new is None:
        return
    row_shift = start_row_new - start_row_tpl
    max_row_tpl = tpl_ws.max_row
    max_col_tpl = tpl_ws.max_column
    target_min = start_row_new
    target_max = start_row_new + (max_row_tpl - start_row_tpl)
    # clear merges in target area
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= target_min and rng.max_row <= target_max:
            try:
                ws.unmerge_cells(str(rng))
            except KeyError:
                try:
                    ws.merged_cells.ranges.remove(rng)
                except (ValueError, KeyError):
                    pass
    # copy rows (values + styles)
    for r in range(start_row_tpl, max_row_tpl + 1):
        dst_r = r + row_shift
        _copy_row_style_from(tpl_ws, ws, r, dst_r, max_col_tpl)
        _set_row_height(ws, dst_r, _row_height(tpl_ws, r))
        for c in range(1, max_col_tpl + 1):
            _safe_set(ws, dst_r, c, tpl_ws.cell(r, c).value)
    # restore merges from template in this block
    for rng in list(tpl_ws.merged_cells.ranges):
        if rng.min_row < start_row_tpl:
            continue
        _merge_safe(ws, rng.min_row + row_shift, rng.min_col, rng.max_row + row_shift, rng.max_col)


def _row_height(ws, row: int) -> float | None:
    return ws.row_dimensions[row].height


def _set_row_height(ws, row: int, height: float | None) -> None:
    if height:
        ws.row_dimensions[row].height = height


def _safe_set(ws, row: int, col: int, value: Any) -> None:
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        target = None
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                target = ws.cell(rng.min_row, rng.min_col)
                break
        if target is None:
            return
        cell = target
    cell.value = value


def _merge_span(ws, row: int, col: int) -> Tuple[int, int] | None:
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_col, rng.max_col
    return None


def _apply_borders(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="medium", color="000000")
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            left = thick if c == min_col else thin
            right = thick if c == max_col else thin
            top = thick if r == min_row else thin
            bottom = thick if r == max_row else thin
            ws.cell(r, c).border = Border(left=left, right=right, top=top, bottom=bottom)


def _ensure_sheet_drawing(sheet_xml: bytes) -> bytes:
    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ET.register_namespace("", ns_main)
    ET.register_namespace("r", ns_rel)
    root = ET.fromstring(sheet_xml)
    tag = f"{{{ns_main}}}drawing"
    has_drawing = root.find(tag) is not None
    if not has_drawing:
        drawing = ET.SubElement(root, tag)
        drawing.set(f"{{{ns_rel}}}id", "rId1")
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _ensure_sheet_rels(rels_xml: bytes | None) -> bytes:
    ns_pkg = "http://schemas.openxmlformats.org/package/2006/relationships"
    ET.register_namespace("", ns_pkg)
    if rels_xml:
        root = ET.fromstring(rels_xml)
    else:
        root = ET.Element(f"{{{ns_pkg}}}Relationships")
    has_rel = False
    for rel in root.findall(f"{{{ns_pkg}}}Relationship"):
        if rel.get("Type", "").endswith("/drawing"):
            has_rel = True
            break
    if not has_rel:
        rel = ET.SubElement(root, f"{{{ns_pkg}}}Relationship")
        rel.set("Id", "rId1")
        rel.set("Type", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing")
        rel.set("Target", "../drawings/drawing1.xml")
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _merge_content_types(out_xml: bytes, tpl_xml: bytes) -> bytes:
    ns = "http://schemas.openxmlformats.org/package/2006/content-types"
    ET.register_namespace("", ns)
    out_root = ET.fromstring(out_xml)
    tpl_root = ET.fromstring(tpl_xml)
    out_defaults = {(el.get("Extension"), el.get("ContentType")) for el in out_root.findall(f"{{{ns}}}Default")}
    out_overrides = {(el.get("PartName"), el.get("ContentType")) for el in out_root.findall(f"{{{ns}}}Override")}
    for el in tpl_root.findall(f"{{{ns}}}Default"):
        key = (el.get("Extension"), el.get("ContentType"))
        if key not in out_defaults:
            out_root.append(el)
            out_defaults.add(key)
    for el in tpl_root.findall(f"{{{ns}}}Override"):
        key = (el.get("PartName"), el.get("ContentType"))
        if key not in out_overrides:
            out_root.append(el)
            out_overrides.add(key)
    return ET.tostring(out_root, encoding="utf-8", xml_declaration=True)


def restore_template_assets(template_path: Path, output_path: Path, logo_override: Path | None = None) -> bool:
    try:
        with zipfile.ZipFile(template_path) as ztpl, zipfile.ZipFile(output_path) as zout:
            tpl_names = ztpl.namelist()
            media_names = [n for n in tpl_names if n.startswith("xl/media/")]
            drawing_names = [n for n in tpl_names if n.startswith("xl/drawings/")]
            sheet_xml = zout.read("xl/worksheets/sheet1.xml")
            new_sheet_xml = _ensure_sheet_drawing(sheet_xml)
            rels_path = "xl/worksheets/_rels/sheet1.xml.rels"
            rels_xml = zout.read(rels_path) if rels_path in zout.namelist() else None
            new_rels_xml = _ensure_sheet_rels(rels_xml)
            ct_xml = zout.read("[Content_Types].xml")
            tpl_ct_xml = ztpl.read("[Content_Types].xml")
            new_ct_xml = _merge_content_types(ct_xml, tpl_ct_xml)

            tmp_fd, tmp_name = tempfile.mkstemp(suffix=".xlsx")
            os.close(tmp_fd)
            Path(tmp_name).unlink(missing_ok=True)
            tmp_path = Path(tmp_name)
            with zipfile.ZipFile(tmp_path, "w") as znew:
                for item in zout.infolist():
                    name = item.filename
                    if name in ("xl/worksheets/sheet1.xml", rels_path, "[Content_Types].xml"):
                        continue
                    if name.startswith("xl/media/") or name.startswith("xl/drawings/"):
                        continue
                    znew.writestr(item, zout.read(name))
                znew.writestr("xl/worksheets/sheet1.xml", new_sheet_xml)
                znew.writestr(rels_path, new_rels_xml)
                znew.writestr("[Content_Types].xml", new_ct_xml)
                for name in media_names:
                    if logo_override and media_names and name == media_names[0]:
                        znew.writestr(name, logo_override.read_bytes())
                    else:
                        znew.writestr(name, ztpl.read(name))
                for name in drawing_names:
                    znew.writestr(name, ztpl.read(name))
        output_path.write_bytes(tmp_path.read_bytes())
        return True
    except Exception:
        return False


def _normalize_rows(items: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows = []
    for it in items:
        if not isinstance(it, dict):
            continue
        rows.append(it)
    return rows


def _sort_rows(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return sorted(items, key=lambda r: (_to_int(r.get("ramal", 0)), _to_int(r.get("loop", 0))))


def _normalize_deshielo_type(val: Any) -> str:
    s = _norm_label(val)
    if not s:
        return ""
    if "GAS CALIENTE" in s or "GASCALIENTE" in s:
        return "GAS CALIENTE"
    if "GAS TIBIO" in s or "GASTIBIO" in s:
        return "GAS TIBIO"
    if "ELECTR" in s:
        return "ELECTRICO"
    return s


def _format_deshielo(default_type: Any) -> str:
    dtype = _normalize_deshielo_type(default_type)
    if not dtype:
        return "DESHIELO"
    return f"DESHIELO {dtype}"


def _load_deshielo_por_uso() -> Dict[str, Dict[str, str]]:
    candidates = [
        Path("data/legend/deshielo_por_uso.json"),
        Path("data/LEGEND/deshielo_por_uso.json"),
    ]
    for path in candidates:
        if path.exists():
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
                if isinstance(data, dict):
                    return data
            except Exception:
                return {}
    return {}


def _load_eev_cost_profile() -> Dict[str, Any]:
    for path in (
        Path("data/LEGEND/eev_cost_profile.json"),
        Path("data/legend/eev_cost_profile.json"),
    ):
        if path.exists():
            try:
                return json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                return {}
    return {}


def _compute_eev_sets_for_export(project_data: Dict[str, Any], result: Dict[str, Any]) -> List[Dict[str, Any]]:
    profile = _load_eev_cost_profile()
    if not profile:
        return []
    overrides = project_data.get("eev_cost_overrides", {}) if isinstance(project_data, dict) else {}
    factor_default = _to_float(profile.get("factor_default", 0.0))
    factor_override = overrides.get("factor") if isinstance(overrides, dict) else None
    factor_val = _to_float(factor_override, factor_default) if factor_override is not None else factor_default
    parts = profile.get("parts", {}) if isinstance(profile, dict) else {}
    parts_override = overrides.get("parts_base_cost", {}) if isinstance(overrides, dict) else {}
    sets_cfg = profile.get("sets", {}) if isinstance(profile, dict) else {}
    currency = profile.get("currency", "") if isinstance(profile, dict) else ""
    package_counts = result.get("package_counts", {}) if isinstance(result, dict) else {}

    def _unit_cost_for_part(part_key: str) -> float:
        base_cost = _to_float(parts.get(part_key, {}).get("base_cost", 0.0))
        if isinstance(parts_override, dict) and part_key in parts_override:
            base_cost = _to_float(parts_override.get(part_key, base_cost))
        return base_cost * (1.0 + factor_val)

    def _set_unit_cost(set_key: str) -> float | None:
        cfg = sets_cfg.get(set_key, {}) if isinstance(sets_cfg, dict) else {}
        parts_list = cfg.get("parts", []) if isinstance(cfg, dict) else []
        if not parts_list:
            return None
        total = 0.0
        for part_key in parts_list:
            total += _unit_cost_for_part(str(part_key))
        return total

    order = ["VALVULA", "CONTROL", "SENSOR", "TRANSDUCTOR", "CAJAS", "SENSORES CO2"]
    rows: List[Dict[str, Any]] = []
    for key in order:
        qty = _to_int(package_counts.get(key, 0), 0)
        if key in ("CAJAS", "SENSORES CO2"):
            part_key = "CAJAS_ELECTRICAS" if key == "CAJAS" else "DGS_IR_CO2"
            unit_cost = _unit_cost_for_part(part_key)
        else:
            unit_cost = _set_unit_cost(key)
        total_cost = None if unit_cost is None else qty * unit_cost
        rows.append(
            {
                "category": key,
                "qty": qty,
                "unit_cost": unit_cost,
                "total_cost": total_cost,
                "currency": currency,
            }
        )
    return rows


def _write_eev_sheet(wb, project_data: Dict[str, Any], mt_ramal_offset: int) -> None:
    ws = wb["EEV"] if "EEV" in wb.sheetnames else wb.create_sheet("EEV")
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    if ws.max_column > 0:
        ws.delete_cols(1, ws.max_column)

    specs = project_data.get("specs", {}) if isinstance(project_data, dict) else {}
    if _norm_label(specs.get("expansion", "")) != "ELECTRONICA":
        return
    if not compute_eev:
        return

    bt_items = _normalize_rows(project_data.get("bt_items", []))
    mt_items = _normalize_rows(project_data.get("mt_items", []))
    result = compute_eev(project_data, bt_items, mt_items, mt_ramal_offset=mt_ramal_offset)
    detail_rows = result.get("detail_rows", []) if isinstance(result, dict) else []
    bom_rows = result.get("bom_rows", []) if isinstance(result, dict) else []
    set_rows = _compute_eev_sets_for_export(project_data, result if isinstance(result, dict) else {})
    if not detail_rows and not bom_rows and not set_rows:
        return

    header_fill = PatternFill("solid", fgColor="C6EFCE")
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    def _upper(val: Any) -> str:
        return str(val or "").upper()

    def _write_headers(row: int, headers: List[str], start_col: int = 1) -> None:
        for idx, h in enumerate(headers):
            cell = ws.cell(row, start_col + idx, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

    def _set_widths(widths: Dict[int, float]) -> None:
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    row = 1
    detail_headers = [
        "SUCCION",
        "LOOP",
        "RAMAL",
        "EQUIPO",
        "USO",
        "CARGA (BTU/HR)",
        "TEVAP (F)",
        "FAMILIA (EEV)",
        "ORIFICIO",
        "MODELO",
    ]
    max_detail_col = len(detail_headers)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_detail_col)
    title = ws.cell(row, 1, "EEV - EXPANSION ELECTRONICA")
    title.font = title_font
    title.alignment = left
    row += 2
    _write_headers(row, detail_headers)
    row += 1
    start_detail = row - 1
    for drow in detail_rows:
        ws.cell(row, 1, _upper(drow.get("suction", ""))).alignment = center
        ws.cell(row, 2, drow.get("loop", "")).alignment = center
        ws.cell(row, 3, drow.get("ramal", "")).alignment = center
        ws.cell(row, 4, _upper(drow.get("equipo", ""))).alignment = left
        ws.cell(row, 5, _upper(drow.get("uso", ""))).alignment = left
        cell_btu = ws.cell(row, 6, int(round(_to_float(drow.get("btu_hr", 0)))))
        cell_btu.number_format = "#,##0"
        cell_btu.alignment = right
        cell_t = ws.cell(row, 7, _to_float(drow.get("tevap_f", 0.0)))
        cell_t.number_format = "0.0"
        cell_t.alignment = right
        ws.cell(row, 8, _upper(drow.get("familia", ""))).alignment = center
        ws.cell(row, 9, _upper(drow.get("orifice", ""))).alignment = center
        ws.cell(row, 10, _upper(drow.get("model", ""))).alignment = left
        row += 1
    end_detail = row - 1
    _apply_borders(ws, start_detail, end_detail, 1, max_detail_col)

    row += 1
    bom_headers = ["MODELO", "DESCRIPCION", "CANTIDAD", "COSTO UNITARIO", "COSTO TOTAL", "MONEDA"]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(bom_headers))
    bom_title = ws.cell(row, 1, "RESUMEN EEV")
    bom_title.font = title_font
    bom_title.alignment = left
    row += 1
    _write_headers(row, bom_headers)
    row += 1
    start_bom = row - 1
    currency = result.get("cost_currency", "") if isinstance(result, dict) else ""
    for brow in bom_rows:
        ws.cell(row, 1, _upper(brow.get("model", ""))).alignment = left
        ws.cell(row, 2, _upper(brow.get("description", ""))).alignment = left
        cell_qty = ws.cell(row, 3, _to_int(brow.get("qty", 0)))
        cell_qty.alignment = right
        unit_cost = brow.get("unit_cost")
        total_cost = brow.get("total_cost")
        if unit_cost is not None:
            c_unit = ws.cell(row, 4, float(unit_cost))
            c_unit.number_format = "#,##0.00"
            c_unit.alignment = right
        if total_cost is not None:
            c_tot = ws.cell(row, 5, float(total_cost))
            c_tot.number_format = "#,##0.00"
            c_tot.alignment = right
        ws.cell(row, 6, _upper(brow.get("currency", currency))).alignment = center
        row += 1
    end_bom = row - 1
    _apply_borders(ws, start_bom, end_bom, 1, len(bom_headers))

    if set_rows:
        row += 1
        set_headers = ["CATEGORIA", "CANTIDAD", "COSTO UNITARIO", "COSTO TOTAL", "MONEDA"]
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(set_headers))
        set_title = ws.cell(row, 1, "PAQUETES")
        set_title.font = title_font
        set_title.alignment = left
        row += 1
        _write_headers(row, set_headers)
        row += 1
        start_set = row - 1
        total_set_cost = 0.0
        for srow in set_rows:
            ws.cell(row, 1, _upper(srow.get("category", ""))).alignment = left
            cell_qty = ws.cell(row, 2, _to_int(srow.get("qty", 0)))
            cell_qty.alignment = right
            unit_cost = srow.get("unit_cost")
            total_cost = srow.get("total_cost")
            if unit_cost is not None:
                c_unit = ws.cell(row, 3, float(unit_cost))
                c_unit.number_format = "#,##0.00"
                c_unit.alignment = right
            if total_cost is not None:
                c_tot = ws.cell(row, 4, float(total_cost))
                c_tot.number_format = "#,##0.00"
                c_tot.alignment = right
                total_set_cost += float(total_cost)
            ws.cell(row, 5, _upper(srow.get("currency", currency))).alignment = center
            row += 1
        ws.cell(row, 1, "TOTAL").font = header_font
        tot_cell = ws.cell(row, 4, total_set_cost)
        tot_cell.number_format = "#,##0.00"
        tot_cell.alignment = right
        ws.cell(row, 5, _upper(currency)).alignment = center
        end_set = row
        _apply_borders(ws, start_set, end_set, 1, len(set_headers))
        row += 1

    def _autofit_columns(min_col: int, max_col: int) -> None:
        for col_idx in range(min_col, max_col + 1):
            max_len = 0
            for r in range(1, ws.max_row + 1):
                val = ws.cell(r, col_idx).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            if max_len <= 0:
                continue
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    _autofit_columns(1, max_detail_col)


def _make_block_rows(
    items: List[Dict[str, Any]],
    default_deshielo: str,
    deshielo_map: Dict[str, str],
    ramal_offset: int = 0,
) -> Tuple[List[Dict[str, Any]], float]:
    rows = []
    total = 0.0
    for it in _sort_rows(items):
        btu_hr = int(round(_to_float(it.get("btu_hr", it.get("carga_btu_h", 0.0)))))
        total += btu_hr
        model = str(it.get("evap_modelo", "") or "")
        if model:
            model = re.sub(r"(?i)\s*-?\s*(DUAL|FRONTAL)\b", "", model)
            model = re.sub(r"\s{2,}", " ", model)
            model = model.strip(" -")
        uso = str(it.get("uso", "") or "")
        uso_key = _norm_label(uso)
        flag = deshielo_map.get(uso_key)
        if flag is not None:
            if _norm_label(flag) == "NO":
                deshielo_val = "DESHIELO POR TIEMPO"
            else:
                deshielo_val = _format_deshielo(default_deshielo)
        else:
            deshielo_val = str(it.get("deshielo", "") or "")
            if not deshielo_val:
                deshielo_val = "DESHIELO POR TIEMPO"
        ramal_val = _to_int(it.get("ramal", ""))
        if ramal_offset:
            ramal_val += ramal_offset
        rows.append(
            {
                "loop": _to_int(it.get("loop", "")),
                "ramal": ramal_val,
                "dim_ft": it.get("dim_ft", ""),
                "equipo": it.get("equipo", ""),
                "uso": uso,
                "btu_hr": btu_hr,
                "tevap": it.get("tevap_f", ""),
                "evap_qty": _to_int(it.get("evap_qty", "")),
                "evap_modelo": model,
                "deshielo": deshielo_val,
            }
        )
    return rows, total


def build_legend_workbook(template_path: Path, project_data: Dict[str, Any]):
    wb = load_workbook(template_path)
    ws = wb.active
    tpl_wb = load_workbook(template_path)
    tpl_ws = tpl_wb.active
    cap_row_tpl = _find_label_row(tpl_ws, "CAP")
    rh_row_tpl = _find_label_row(tpl_ws, "RH")
    block_row_tpl = None
    if cap_row_tpl or rh_row_tpl:
        block_row_tpl = min(r for r in (cap_row_tpl, rh_row_tpl) if r)
        if block_row_tpl > 1:
            block_row_tpl -= 1

    specs = project_data.get("specs", {}) if isinstance(project_data, dict) else {}

    def _write_spec(label: str, value: Any) -> None:
        pos = _find_label_cell(ws, label, max_row=6)
        if not pos:
            return
        row, col = pos
        row, col = _col_after_label(ws, row, col)
        _safe_set(ws, row, col, value)

    _write_spec("PROYECTO", specs.get("proyecto", ""))
    _write_spec("CIUDAD", specs.get("ciudad", ""))
    _write_spec("TIPO SISTEMA", specs.get("tipo_sistema", ""))
    _write_spec("VOLTAJE PRINCIPAL", specs.get("voltaje_principal", ""))
    _write_spec("VOLTAJE CONTROL", specs.get("voltaje_control", ""))
    _write_spec("T. CONDENSACION", specs.get("tcond_f", ""))
    _write_spec("REFRIGERANTE", specs.get("refrigerante", ""))
    _write_spec("CONTROLADOR", specs.get("controlador", ""))
    _write_spec("MEDIDAS", "FT")

    default_deshielo = specs.get("deshielos", "")
    deshielo_por_uso = _load_deshielo_por_uso()

    header_row = _find_header_row(ws) or 7
    header_map = _map_header_cols(ws, header_row)
    sub_map = _find_subheader_cols(ws, header_row)

    suction_col = _col_for(header_map, ["SUCCION"])
    loop_col = _col_for(header_map, ["LOOP"])
    ramal_col = _col_for(header_map, ["RAMAL"])
    dim_col = _col_for(header_map, ["DIM (FT)", "DIM"])
    equipo_col = _col_for(header_map, ["EQUIPO"])
    uso_col = _col_for(header_map, ["USO"])
    btu_col = _find_btu_col(ws, header_row, header_map)
    tevap_col = _find_tevap_col(ws, header_row, header_map)
    evap_hdr_col = _col_for(header_map, ["EVAPORADORES"])
    qty_col = sub_map.get("CANTIDAD") or evap_hdr_col
    model_col = sub_map.get("MODELO") or (evap_hdr_col + 1 if evap_hdr_col else None)
    deshielo_col = _col_for(header_map, ["DESHIELO"])
    equipo_span = _merge_span(ws, header_row, equipo_col) if equipo_col else None
    deshielo_span = _merge_span(ws, header_row, deshielo_col) if deshielo_col else None

    start_row = header_row + 2

    bt_items = _normalize_rows(project_data.get("bt_items", []))
    mt_items = _normalize_rows(project_data.get("mt_items", []))

    bt_count = _to_int(project_data.get("bt_ramales", 0))
    if bt_count <= 0:
        bt_count = max((_to_int(r.get("ramal", 0)) for r in bt_items), default=0)
    bt_rows, bt_total = _make_block_rows(
        bt_items, default_deshielo, deshielo_por_uso.get("BT", {}), ramal_offset=0
    )
    mt_rows, mt_total = _make_block_rows(
        mt_items, default_deshielo, deshielo_por_uso.get("MT", {}), ramal_offset=bt_count
    )
    needed_rows = len(bt_rows) + len(mt_rows) + 3
    if needed_rows < 1:
        needed_rows = 1

    comp_row = _find_comp_section_row(ws)
    reserva_pos = _find_label_cell(ws, "RESERVA %", max_row=200) or _find_label_cell(
        ws, "RESERVA%", max_row=200
    )
    summary_rows: List[int] = []
    for label in ("CAP", "CARGA", "RESERVA %", "RESERVA%", "RH"):
        pos = _find_label_cell(ws, label, max_row=200)
        if pos:
            summary_rows.append(pos[0])
    block_start = None
    if summary_rows:
        block_start = min(summary_rows) - 1
    if block_start is not None:
        block_start = max(start_row, block_start)
    anchor_row = block_start or comp_row

    if anchor_row and anchor_row > start_row:
        available_rows = anchor_row - start_row
        bt_spacer = True
        mt_spacer = True
        required_rows = (
            len(bt_rows)
            + 1
            + (1 if bt_spacer else 0)
            + len(mt_rows)
            + 1
            + (1 if mt_spacer else 0)
        )
        if available_rows < required_rows:
            ws.insert_rows(anchor_row, required_rows - available_rows)
        elif available_rows > required_rows:
            ws.delete_rows(start_row + required_rows, available_rows - required_rows)
    else:
        if ws.max_row >= start_row:
            ws.delete_rows(start_row, ws.max_row - start_row + 1)
        rows_to_insert = max(needed_rows, 0)
        if rows_to_insert:
            ws.insert_rows(start_row, rows_to_insert)
    # recalcular anchor_row luego de insertar/borrar filas
    if anchor_row:
        cap_row_new = _find_label_row(ws, "CAP", max_row=ws.max_row)
        rh_row_new = _find_label_row(ws, "RH", max_row=ws.max_row)
        if cap_row_new or rh_row_new:
            anchor_row = min(r for r in (cap_row_new, rh_row_new) if r) - 1
            if anchor_row < 1:
                anchor_row = 1
    # limpiar merges en el bloque de datos para evitar rangos superpuestos
    clear_end = (anchor_row - 1) if anchor_row and anchor_row > start_row else (start_row + needed_rows + 2)
    _clear_merges_in_rows(ws, start_row, min(clear_end, ws.max_row))

    data_style_row = start_row
    total_style_row = header_row
    max_col = ws.max_column
    data_row_height = _row_height(ws, data_style_row)
    total_row_height = _row_height(ws, total_style_row)

    cur = start_row
    total_ranges: List[Tuple[int, int, int]] = []
    spacer_rows: List[int] = []

    def _is_deshielo_por_tiempo(val: Any) -> bool:
        return _norm_label(val) in ("DESHIELO POR TIEMPO", "POR TIEMPO", "NO")

    def _set_block_label(label: str, start: int, end: int) -> None:
        if not suction_col or start > end:
            return
        _merge_safe(ws, start, suction_col, end, suction_col)
        cell = ws.cell(start, suction_col)
        cell.value = label
        cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

    def _merge_loop_block_empty(start: int, end: int) -> None:
        if not loop_col or start > end:
            return
        for r in range(start, end + 1):
            ws.cell(r, loop_col).value = ""
        if end > start:
            _merge_safe(ws, start, loop_col, end, loop_col)
        cell = ws.cell(start, loop_col)
        cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

    def _merge_loop_ranges(start: int, end: int) -> None:
        if not loop_col or start > end:
            return
        row = start
        while row <= end:
            val = ws.cell(row, loop_col).value
            if val is None or str(val).strip() == "":
                row += 1
                continue
            next_row = row + 1
            while next_row <= end:
                next_val = ws.cell(next_row, loop_col).value
                if next_val == val:
                    next_row += 1
                    continue
                break
            if next_row - row > 1:
                _merge_safe(ws, row, loop_col, next_row - 1, loop_col)
                cell = ws.cell(row, loop_col)
                cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
            row = next_row

    def _write_rows(rows: List[Dict[str, Any]]) -> None:
        nonlocal cur
        for row in rows:
            _copy_row_style(ws, data_style_row, cur, max_col)
            _set_row_height(ws, cur, data_row_height)
            if equipo_span and equipo_span[1] > equipo_span[0]:
                _merge_safe(ws, cur, equipo_span[0], cur, equipo_span[1])
            if deshielo_span and deshielo_span[1] > deshielo_span[0]:
                _merge_safe(ws, cur, deshielo_span[0], cur, deshielo_span[1])
            if loop_col:
                _safe_set(ws, cur, loop_col, row.get("loop", ""))
                lcell = ws.cell(cur, loop_col)
                lcell.alignment = lcell.alignment.copy(horizontal="center", vertical="center")
            if ramal_col:
                _safe_set(ws, cur, ramal_col, row.get("ramal", ""))
            if dim_col:
                _safe_set(ws, cur, dim_col, row.get("dim_ft", ""))
            if equipo_col:
                _safe_set(ws, cur, equipo_col, row.get("equipo", ""))
            if uso_col:
                _safe_set(ws, cur, uso_col, row.get("uso", ""))
            if btu_col:
                _safe_set(ws, cur, btu_col, int(round(_to_float(row.get("btu_hr", 0)))))
            if tevap_col:
                _safe_set(ws, cur, tevap_col, row.get("tevap", ""))
            if qty_col:
                _safe_set(ws, cur, qty_col, row.get("evap_qty", ""))
            if model_col:
                _safe_set(ws, cur, model_col, row.get("evap_modelo", ""))
            if deshielo_col:
                _safe_set(ws, cur, deshielo_col, row.get("deshielo", ""))
            is_pt = _is_deshielo_por_tiempo(row.get("deshielo", default_deshielo))
            if tevap_col:
                cell = ws.cell(cur, tevap_col)
                cell.font = cell.font.copy(color="000000" if is_pt else "FF0000")
            if deshielo_col:
                dcell = ws.cell(cur, deshielo_col)
                dcell.font = dcell.font.copy(color="000000" if is_pt else "FF0000")
            cur += 1

    def _write_total(label: str, total: float, *, add_spacer: bool) -> None:
        nonlocal cur
        _clear_merges_on_row(ws, cur)
        _copy_row_style(ws, total_style_row, cur, max_col)
        _set_row_height(ws, cur, total_row_height)
        label_col = suction_col or 1
        end_col = deshielo_span[1] if deshielo_span else (deshielo_col or max_col)
        if btu_col and btu_col > label_col + 1:
            _merge_safe(ws, cur, label_col, cur, btu_col - 1)
        if tevap_col and end_col and end_col >= tevap_col:
            _merge_safe(ws, cur, tevap_col, cur, end_col)
        _safe_set(ws, cur, label_col, label)
        if btu_col:
            _safe_set(ws, cur, btu_col, int(round(total)))
        clear_fill = PatternFill()
        for c in range(1, max_col + 1):
            ws.cell(cur, c).fill = clear_fill
        if btu_col:
            total_ranges.append((cur, label_col, btu_col - 1))
            total_ranges.append((cur, btu_col, btu_col))
        if tevap_col and end_col and end_col >= tevap_col:
            total_ranges.append((cur, tevap_col, end_col))
        if add_spacer:
            spacer_row = cur + 1
            _clear_merges_on_row(ws, spacer_row)
            spacer_rows.append(spacer_row)
            cur += 2
        else:
            cur += 1

    bt_start = cur
    _write_rows(bt_rows)
    bt_end = cur - 1
    if bt_rows:
        _set_block_label("BAJA", bt_start, bt_end)
        if _norm_label(specs.get("distribucion_tuberia", "")) == "AMERICANA":
            _merge_loop_block_empty(bt_start, bt_end)
        else:
            _merge_loop_ranges(bt_start, bt_end)
    _write_total("CARGA TOTAL BAJA", bt_total, add_spacer=True)
    mt_start = cur
    _write_rows(mt_rows)
    mt_end = cur - 1
    if mt_rows:
        _set_block_label("MEDIA", mt_start, mt_end)
        if _norm_label(specs.get("distribucion_tuberia", "")) == "AMERICANA":
            _merge_loop_block_empty(mt_start, mt_end)
        else:
            _merge_loop_ranges(mt_start, mt_end)
    _write_total("CARGA TOTAL MEDIA", mt_total, add_spacer=True if anchor_row else False)

    last_row = cur - 1
    if header_row and last_row >= header_row:
        _apply_borders(ws, header_row, last_row, 1, max_col)
        for row_idx, min_col, max_col in total_ranges:
            _apply_borders(ws, row_idx, row_idx, min_col, max_col)
        if spacer_rows:
            empty_border = Border()
            for row_idx in spacer_rows:
                for c in range(1, max_col + 1):
                    ws.cell(row_idx, c).border = empty_border
    if block_row_tpl:
        cap_row_new = _find_label_row(ws, "CAP", max_row=ws.max_row)
        rh_row_new = _find_label_row(ws, "RH", max_row=ws.max_row)
        if cap_row_new or rh_row_new:
            block_row_new = min(r for r in (cap_row_new, rh_row_new) if r)
            if block_row_new > 1:
                block_row_new -= 1
            _restore_template_block(ws, tpl_ws, block_row_tpl, block_row_new)
    _write_compresores(ws, project_data)
    # ------------------------ resumen CAP/CARGA/RESERVA + RH + bloques ------------------------
    try:
        perf_db = _load_compresores_perf()
        tcond_f = _to_float(specs.get("tcond_f", 0) or 0)
        refrigerante = str(specs.get("refrigerante", "") or "").strip()
        tevap_bt = _compute_tevap_design(bt_items) if isinstance(bt_items, list) else None
        tevap_mt = _compute_tevap_design(mt_items) if isinstance(mt_items, list) else None
        cap_bt, cap_mt, rh_bt, rh_mt = _compute_comp_totals(
            project_data, perf_db, tcond_f, tevap_bt, tevap_mt, refrigerante
        )
        carga_bt = bt_total
        carga_mt = mt_total
        reserva_bt_pct = (cap_bt - carga_bt) / carga_bt * 100.0 if carga_bt > 0 else 0.0
        reserva_mt_pct = (cap_mt - carga_mt) / carga_mt * 100.0 if carga_mt > 0 else 0.0

        _write_baja_media(ws, "CAP", int(round(cap_bt)), int(round(cap_mt)), number_format="#,##0")
        _write_baja_media(ws, "CARGA", int(round(carga_bt)), int(round(carga_mt)), number_format="#,##0")
        _write_baja_media(ws, "RESERVA %", reserva_bt_pct / 100.0, reserva_mt_pct / 100.0, number_format="0%")

        _write_rh_box(ws, rh_bt, rh_mt, rh_bt + rh_mt)

        # Bloque CONDENSADOR
        cond_header = _find_cell_contains(ws, "CONDENSADOR", max_row=ws.max_row)
        cond_start = cond_header[0] if cond_header else 1
        cond_end = cond_start + 20 if cond_header else ws.max_row
        _write_next_label_in_rows(ws, "REFR.", refrigerante, cond_start, cond_end, uppercase=True)
        _write_next_label_in_rows(ws, "T. COND", tcond_f, cond_start, cond_end, number_format="0")
        def _spec(*keys: str) -> Any:
            for k in keys:
                if k in specs:
                    return specs.get(k)
            return ""
        _write_next_label_in_rows(ws, "MODELO", _spec("condensador_modelo", "condenser_model"), cond_start, cond_end)
        _write_next_label_in_rows(ws, "CANTIDAD", _spec("condensador_cantidad", "condenser_qty"), cond_start, cond_end, number_format="0")
        _write_next_label_in_rows(ws, "CAPACIDAD", _spec("condensador_capacidad", "condenser_capacity"), cond_start, cond_end)
        _write_next_label_in_rows(ws, "RPM", _spec("condensador_rpm", "condenser_rpm"), cond_start, cond_end)
        _write_next_label_in_rows(ws, "DBA", _spec("condensador_dba", "condenser_dba"), cond_start, cond_end)
        _write_next_label_in_rows(ws, "DIMENSION", _spec("condensador_dimension", "condenser_dimension"), cond_start, cond_end)

        # Bloque WESTON
        west_header = _find_cell_contains(ws, "WESTON SAS - BOGOTA - COLOMBIA", max_row=ws.max_row)
        west_start = west_header[0] if west_header else 1
        west_end = west_start + 20 if west_header else ws.max_row
        _write_next_label_in_rows(ws, "CALCULO", specs.get("calculo", ""), west_start, west_end)
        reviso_val = specs.get("reviso", "")
        rev_cells = _find_all_label_cells(ws, "REVISO", max_row=ws.max_row)
        for pos in rev_cells:
            if pos[0] < west_start or pos[0] > west_end:
                continue
            r, c = _merged_anchor_cell(ws, pos[0], pos[1])
            tr, tc = _col_after_label(ws, r, c)
            tr, tc = _merged_anchor_cell(ws, tr, tc)
            ws.cell(tr, tc).value = reviso_val
        _write_next_label_in_rows(ws, "VENDEDOR", specs.get("vendedor", ""), west_start, west_end)
        _write_next_label_in_rows(ws, "CLIENTE", specs.get("cliente", ""), west_start, west_end)
        fecha_val = specs.get("fecha") or datetime.now()
        _write_next_label_in_rows(ws, "FECHA", _format_fecha_es(fecha_val), west_start, west_end, uppercase=True)
        _write_next_label_in_rows(ws, "ULTIMA REFORMA", specs.get("ultima_reforma", ""), west_start, west_end)
        _write_next_label_in_rows(ws, "VERSION", specs.get("version", ""), west_start, west_end)
    except Exception:
        pass
    _cleanup_invalid_merges(ws)
    _prune_overlapping_merges(ws, header_row, last_row)
    _force_total_value(ws, "CARGA TOTAL BAJA", bt_total, btu_col)
    _force_total_value(ws, "CARGA TOTAL MEDIA", mt_total, btu_col)
    ws.calculate_dimension()
    try:
        _write_eev_sheet(wb, project_data, bt_count)
    except Exception:
        pass
    return wb
