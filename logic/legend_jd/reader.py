from __future__ import annotations

import re
from typing import Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from .models import (
    EquipoCatalogItem,
    LegendConfig,
    PlantillaBOMItem,
    VariadorItem,
    WcrItem,
)


def _load_wb(path) -> Workbook:
    return load_workbook(path, keep_vba=True, data_only=False)


def _get_sheet(wb: Workbook, name: str = "INFO"):
    if name in wb.sheetnames:
        return wb[name]
    # fallback: first sheet
    return wb[wb.sheetnames[0]]


def read_info_config(wb: Workbook) -> LegendConfig:
    ws = _get_sheet(wb)
    mapping = {
        "PROYECTO": "proyecto",
        "CIUDAD": "ciudad",
        "TIPO DE SISTEMA": "tipo_sistema",
        "REFRIGERANTE": "refrigerante",
        "TCOND": "tcond",
        "TEVAP BT": "tevap_bt",
        "TEVAP MT": "tevap_mt",
        "MARCA COMPRESORES": "marca_compresores",
        "TIPO INSTALACION": "tipo_instalacion",
        "FACTOR DE SEGURIDAD": "factor_seguridad",
    }
    values = {}
    extras = {}
    # Leer rangos B2:C40 aprox (1-index)
    for row in range(2, 41):
        label = ws.cell(row=row, column=2).value
        val = ws.cell(row=row, column=3).value
        if label is None:
            continue
        key = str(label).strip().upper()
        if key in mapping:
            values[mapping[key]] = val
        else:
            extras[key] = val
    cfg = LegendConfig(**values)
    cfg.extras = extras
    return cfg


def _eval_formula_simple(value, ws):
    if isinstance(value, str) and value.startswith("="):
        m = re.match(r"=([A-Z]+\\d+)\\*(\\d+(?:\\.\\d+)?)", value.strip())
        if m:
            ref_cell, mult = m.groups()
            try:
                ref_val = ws[ref_cell].value
                ref_num = float(ref_val) if ref_val is not None else 0.0
                return ref_num * float(mult)
            except Exception:
                return value
    return value


def read_equipos(wb: Workbook) -> List[EquipoCatalogItem]:
    ws = _get_sheet(wb)
    items: List[EquipoCatalogItem] = []
    for row in ws.iter_rows(min_col=6, max_col=7):  # F/G
        equipo = row[0].value
        cap_cell = row[1].value
        if equipo is None:
            continue
        cap_val = _eval_formula_simple(cap_cell, ws)
        try:
            cap = float(cap_val)
        except Exception:
            continue
        items.append(EquipoCatalogItem(str(equipo), cap))
    return items


def read_usos(wb: Workbook) -> Dict[str, List[str]]:
    ws = _get_sheet(wb)
    usos = {"BT": [], "MT": []}
    markers = {"USO BT": "BT", "USO MT": "MT"}
    # Buscar marcadores
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if not isinstance(val, str):
                continue
            key = markers.get(val.strip().upper())
            if key:
                # Leer hacia abajo en la misma columna hasta encontrar celda vacia
                r = cell.row + 1
                while True:
                    v = ws.cell(row=r, column=cell.column).value
                    if v is None or str(v).strip() == "":
                        break
                    usos[key].append(str(v).strip())
                    r += 1
    return usos


def _read_pairs_by_marker(ws, marker_text: str) -> List[Tuple[str, float]]:
    results: List[Tuple[str, float]] = []
    for row in ws.iter_rows(min_col=12, max_col=12):  # col L
        val = row[0].value
        if isinstance(val, str) and marker_text.upper() in val.upper():
            start_row = row[0].row + 1
            cur = start_row
            while True:
                model = ws.cell(row=cur, column=12).value  # L
                power = ws.cell(row=cur, column=13).value  # M
                if model is None or str(model).strip() == "":
                    break
                try:
                    power_val = float(power)
                except Exception:
                    power_val = None
                if power_val is not None:
                    results.append((str(model), power_val))
                cur += 1
    return results


def read_variadores_fc102(wb: Workbook) -> List[VariadorItem]:
    ws = _get_sheet(wb)
    pairs = _read_pairs_by_marker(ws, "VARIADOR")
    return [VariadorItem(modelo=m, potencia=p) for m, p in pairs]


def read_wcr(wb: Workbook) -> List[WcrItem]:
    ws = _get_sheet(wb)
    pairs = _read_pairs_by_marker(ws, "WCR")
    return [WcrItem(modelo=m, capacidad=p) for m, p in pairs]


def read_plantillas_bom(wb: Workbook) -> Dict[str, List[PlantillaBOMItem]]:
    ws = _get_sheet(wb)
    plantillas = {
        "Rack Loop": [],
        "Minisistema": [],
        "Rack Americano": [],
    }
    # Asumimos pares de columnas: AC/AD, AE/AF, AG/AH
    col_pairs = [
        ("Rack Loop", ("AC", "AD")),
        ("Minisistema", ("AE", "AF")),
        ("Rack Americano", ("AG", "AH")),
    ]
    for nombre, (col_qty, col_desc) in col_pairs:
        start_row = 2
        while True:
            qty_val = ws[f"{col_qty}{start_row}"].value
            desc_val = ws[f"{col_desc}{start_row}"].value
            if desc_val is None or str(desc_val).strip() == "":
                # detener si ambas estan vacias
                if qty_val is None:
                    break
                start_row += 1
                continue
            try:
                qty = float(qty_val) if qty_val is not None else None
            except Exception:
                qty = None
            plantillas[nombre].append(PlantillaBOMItem(qty=qty, descripcion=str(desc_val)))
            start_row += 1
    return plantillas


if __name__ == "__main__":
    from .service import LegendJDService

    svc = LegendJDService()
    data = svc.load_all()
    cfg = data["config"]
    print(f"Proyecto: {cfg.proyecto} | Ciudad: {cfg.ciudad} | Refrigerante: {cfg.refrigerante} | Tcond: {cfg.tcond}")
    print(f"Equipos: {len(data['equipos'])} | Usos BT: {len(data['usos'].get('BT', []))} | Usos MT: {len(data['usos'].get('MT', []))}")
    print(f"Variadores: {len(data['variadores'])} | WCR: {len(data['wcr'])}")
    print({k: len(v) for k, v in data["plantillas"].items()})
