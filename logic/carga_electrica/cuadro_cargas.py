"""
Motor de calculo para el modulo de CARGA ELECTRICA.

Lee el Excel base (CUADRO DE CARGAS -INDIVIDUAL.xlsm o similar), toma la
hoja SELECCION como entrada, consulta las tablas MUEBLES y UNIDADES y genera
un cuadro de cargas por ramal, sin macros.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import unicodedata

import pandas as pd
from openpyxl import load_workbook
import math
import re


# --------------------------------------------------------------------------- #
# Normalizacion de textos
# --------------------------------------------------------------------------- #

def _norm(text: str) -> str:
    """Normaliza nombres: mayusculas, sin tildes, sin dobles espacios."""
    s = str(text or "").strip().upper()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = " ".join(s.split())
    return s


def _norm_key(text: str) -> str:
    """Normaliza para usar como llave: sin tildes y sin caracteres no alfanumericos."""
    import re
    s = _norm(text)
    return re.sub(r"[^A-Z0-9]", "", s)


def _as_int(val, default: int = 0) -> int:
    """Convierte a int manejando NaN y valores vacios."""
    num = pd.to_numeric(val, errors="coerce")
    if pd.isna(num):
        return default
    try:
        return int(num)
    except Exception:
        return default


def _as_float(val, default: float = 0.0) -> float:
    """Convierte a float limpiando unidades y evaluando expresiones simples."""
    if val is None:
        return default
    # NaN -> default
    if isinstance(val, float):
        try:
            import math
            if math.isnan(val):
                return default
        except Exception:
            pass
    if isinstance(val, (int, float)):
        try:
            return float(val)
        except Exception:
            return default
    s = str(val)
    s = s.replace("\u202f", " ").replace("\xa0", " ").replace(",", ".")
    s = s.replace(" ", "")
    if s.startswith("="):
        s = s[1:]
    # reemplazar funciones conocidas
    s = s.replace("SQRT", "math.sqrt")
    # permitir solo caracteres seguros y evaluar
    if re.fullmatch(r"[0-9A-Za-z_\\.\+\-\*\/\(\)]+", s):
        try:
            return float(eval(s, {"__builtins__": None, "math": math}))
        except Exception:
            pass
    # fallback: primera cifra que aparezca
    m = re.search(r"[-+]?[0-9]*\\.?[0-9]+", s)
    if m:
        try:
            return float(m.group(0))
        except Exception:
            return default
    return default


# --------------------------------------------------------------------------- #
# Modelos de datos
# --------------------------------------------------------------------------- #

@dataclass
class MuebleLoad:
    """Corrientes base de un mueble + en cuantas fases usa cada carga."""
    name: str
    dimension: str
    illumination: float = 0.0
    fans: float = 0.0
    res_desemp: float = 0.0
    res_antisudor: float = 0.0
    res_desague: float = 0.0
    res_calef: float = 0.0
    turb_camara: float = 0.0
    turb_puerta: float = 0.0
    motor_carro: float = 0.0
    res_descong: float = 0.0
    ilum_por_nivel: float = 0.0
    fases: Dict[str, int] = field(default_factory=dict)


@dataclass
class Unidad:
    codigo: str
    compresor_modelo: str
    compresor_amp: float
    compresor_fases: int
    condensador_amp: float
    condensador_fases: int
    refrigerante: str = ""


@dataclass
class SeleccionRamal:
    ramal: int
    equipo: str
    uso: str
    medida: str
    descongelamiento: str
    tiene_unidad: bool
    ref_unidad: str
    niveles: int


@dataclass
class CargaFase:
    """Carga desglosada en L1/L2/L3."""
    etiqueta: str
    l1: float = 0.0
    l2: float = 0.0
    l3: float = 0.0


@dataclass
class ResultadoRamal:
    ramal: SeleccionRamal
    cargas: List[CargaFase]

    @property
    def total(self) -> CargaFase:
        return CargaFase(
            etiqueta="TOTAL (A)",
            l1=sum(c.l1 for c in self.cargas),
            l2=sum(c.l2 for c in self.cargas),
            l3=sum(c.l3 for c in self.cargas),
        )


# --------------------------------------------------------------------------- #
# Lectura de catalogos
# --------------------------------------------------------------------------- #

def _load_muebles(book: Path) -> Dict[Tuple[str, str], MuebleLoad]:
    df = pd.read_excel(book, sheet_name="MUEBLES", engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    lookup = {_norm(c): c for c in df.columns}
    # Para recuperar formulas usamos openpyxl (sin data_only) y un workbook separado en data_only True
    wb_formulas = load_workbook(book, data_only=False)
    wb_values = load_workbook(book, data_only=True)
    ws_f = wb_formulas["MUEBLES"]
    ws_v = wb_values["MUEBLES"]

    def col(target: str) -> Optional[str]:
        return lookup.get(_norm(target))

    targets = {
        "NOMBRE": "name",
        "DIMENSION": "dimension",
        "ILUMINACION": "illumination",
        "VENTILADORES": "fans",
        "RESISTENCIA DESEMPANANTE": "res_desemp",
        "RESISTENCIA ANTISUDOR": "res_antisudor",
        "RESISTENCIA DESAGUE": "res_desague",
        "RESISTENCIA CALEFACTORA": "res_calef",
        "TURBINA CAMARA DE COMBUSTION": "turb_camara",
        "TURBINA PUERTA": "turb_puerta",
        "MOTOR CARRO PAN": "motor_carro",
        "RESISTENCIA DESCONGELAMIENTO": "res_descong",
        "ILUMINACION X ENTREPANO": "ilum_por_nivel",
        "FASES ILUMINACION": "f_ilum",
        "FASES VENTILADORES": "f_fans",
        "FASES RESISTENCIA DESEMPANANTE": "f_res_desemp",
        "FASE RESISTENCIA ANTISUDOR": "f_res_antisudor",
        "FASES RESISTENCIA DESAGUE": "f_res_desague",
        "FASES RESISTENCIA CALEFACTORA": "f_res_calef",
        "FASES TURBINA CAMARA DE COMB.": "f_turb_camara",
        "FASES TURBINA PUERTA": "f_turb_puerta",
        "FASE MOTOR CARRO PAN": "f_motor_carro",
        "FASE RESISTENCIA DESCONGELAMIENTO": "f_res_descong",
    }

    col_map: Dict[str, str] = {}
    for target, dest in targets.items():
        found = col(target)
        if found:
            col_map[dest] = found

    out: Dict[Tuple[str, str], MuebleLoad] = {}

    def _get_with_formula(idx: int, col_key: str, default=0.0):
        col_name = col_map.get(col_key, "")
        val = row.get(col_name, default)
        if pd.isna(val) or val == "":
            try:
                col_idx = df.columns.get_loc(col_name)
                # primero intento con el workbook data_only (por si ya viene el valor)
                cell_val = ws_v.cell(row=idx + 2, column=col_idx + 1).value
                if cell_val not in (None, ""):
                    val = cell_val
                else:
                    cell = ws_f.cell(row=idx + 2, column=col_idx + 1)
                    if cell.value not in (None, ""):
                        val = cell.value
            except Exception:
                pass
        return val
    for _, row in df.iterrows():
        raw_name = str(row.get(col_map.get("name", ""), ""))
        raw_dim = str(row.get(col_map.get("dimension", ""), ""))
        name = _norm_key(raw_name)
        dim = _norm_key(raw_dim)
        if not name or not dim:
            continue
        fases = {
            "illumination": _as_int(row.get(col_map.get("f_ilum", ""), 0), 0),
            "fans": _as_int(row.get(col_map.get("f_fans", ""), 0), 0),
            "res_desemp": _as_int(row.get(col_map.get("f_res_desemp", ""), 0), 0),
            "res_antisudor": _as_int(row.get(col_map.get("f_res_antisudor", ""), 0), 0),
            "res_desague": _as_int(row.get(col_map.get("f_res_desague", ""), 0), 0),
            "res_calef": _as_int(row.get(col_map.get("f_res_calef", ""), 0), 0),
            "turb_camara": _as_int(row.get(col_map.get("f_turb_camara", ""), 0), 0),
            "turb_puerta": _as_int(row.get(col_map.get("f_turb_puerta", ""), 0), 0),
            "motor_carro": _as_int(row.get(col_map.get("f_motor_carro", ""), 0), 0),
            "res_descong": _as_int(row.get(col_map.get("f_res_descong", ""), 0), 0),
        }
        out[(name, dim)] = MuebleLoad(
            name=name,
            dimension=dim,
            illumination=_as_float(_get_with_formula(_, "illumination", 0.0), 0.0),
            fans=_as_float(_get_with_formula(_, "fans", 0.0), 0.0),
            res_desemp=_as_float(_get_with_formula(_, "res_desemp", 0.0), 0.0),
            res_antisudor=_as_float(_get_with_formula(_, "res_antisudor", 0.0), 0.0),
            res_desague=_as_float(_get_with_formula(_, "res_desague", 0.0), 0.0),
            res_calef=_as_float(_get_with_formula(_, "res_calef", 0.0), 0.0),
            turb_camara=_as_float(_get_with_formula(_, "turb_camara", 0.0), 0.0),
            turb_puerta=_as_float(_get_with_formula(_, "turb_puerta", 0.0), 0.0),
            motor_carro=_as_float(_get_with_formula(_, "motor_carro", 0.0), 0.0),
            res_descong=_as_float(_get_with_formula(_, "res_descong", 0.0), 0.0),
            ilum_por_nivel=_as_float(_get_with_formula(_, "ilum_por_nivel", 0.0), 0.0),
            fases=fases,
        )
    return out


def _load_unidades(book: Path) -> Dict[str, Unidad]:
    df = pd.read_excel(book, sheet_name="UNIDADES", header=0)
    df.columns = [str(c).strip() for c in df.columns]
    lookup = {_norm(c): c for c in df.columns}

    def pick(targets: List[str], fallback_idx: Optional[int] = None) -> str:
        for t in targets:
            if _norm(t) in lookup:
                return lookup[_norm(t)]
        if fallback_idx is not None and fallback_idx < len(df.columns):
            return df.columns[fallback_idx]
        return df.columns[0]

    # Columnas: C codigo, D compresor modelo, E consumo, F condensador, G #PH, H #PHCON, I refrigerante
    c_codigo = pick(["COD UNIDAD", "CODIGO", "CODIGO UNIDAD"], 2 if len(df.columns) > 2 else 0)
    c_comp = pick(["COMPRESOR", "MODELO COMPRESOR"], 3 if len(df.columns) > 3 else 0)
    c_comp_amp = pick(["CONSUMO", "AMP COMPRESOR"], 4 if len(df.columns) > 4 else 0)
    c_cond_amp = pick(["CONDENSADOR", "AMP CONDENSADOR"], 5 if len(df.columns) > 5 else 0)
    c_ph = pick(["#PH", "PH COMPRESOR"], 6 if len(df.columns) > 6 else 0)
    c_phcon = pick(["#PHCON", "PH CONDENSADOR"], 7 if len(df.columns) > 7 else 0)
    c_refrig = pick(["REFRIGERANTE"], 8 if len(df.columns) > 8 else len(df.columns) - 1)

    out: Dict[str, Unidad] = {}
    for _, row in df.iterrows():
        cod_raw = row.get(c_codigo, "")
        cod = _norm(str(cod_raw))
        if not cod:
            continue
        out[cod] = Unidad(
            codigo=cod,
            compresor_modelo=str(row.get(c_comp, "")),
            compresor_amp=_as_float(row.get(c_comp_amp, 0.0), 0.0),
            compresor_fases=max(1, _as_int(row.get(c_ph, 0), 0)),
            condensador_amp=_as_float(row.get(c_cond_amp, 0.0), 0.0),
            condensador_fases=max(1, _as_int(row.get(c_phcon, 0), 0)),
            refrigerante=str(row.get(c_refrig, "")),
        )
    return out


def _parse_seleccion(book: Path) -> List[SeleccionRamal]:
    def _to_int(val, default: Optional[int] = 0) -> Optional[int]:
        num = pd.to_numeric(val, errors="coerce")
        if pd.isna(num):
            return default
        try:
            return int(num)
        except Exception:
            return default

    xls = pd.ExcelFile(book)
    sel_sheet = None
    for name in xls.sheet_names:
        if _norm(name) == "SELECCION":
            sel_sheet = name
            break
    sel_sheet = sel_sheet or xls.sheet_names[0]
    df = pd.read_excel(xls, sheet_name=sel_sheet, header=None, skiprows=1)

    # columnas segun descripcion: C..J -> indices 2..9
    df = df.iloc[:, 2:10]
    df.columns = [
        "ramal",
        "equipo",
        "uso",
        "medida",
        "descongelamiento",
        "tiene_unidad",
        "ref_unidad",
        "niveles",
    ]
    out: List[SeleccionRamal] = []
    for _, row in df.iterrows():
        ramal_num = pd.to_numeric(row["ramal"], errors="coerce")
        if pd.isna(ramal_num):
            continue
        ramal_val = _to_int(ramal_num, default=None)
        if ramal_val is None:
            continue
        out.append(
            SeleccionRamal(
                ramal=ramal_val,
                equipo=str(row["equipo"]).strip(),
                uso=str(row["uso"]).strip(),
                medida=str(row["medida"]).strip(),
                descongelamiento=str(row["descongelamiento"]).strip(),
                tiene_unidad=str(row["tiene_unidad"]).strip().upper() == "SI",
                ref_unidad=str(row["ref_unidad"] or "").strip(),
                niveles=_to_int(row.get("niveles"), default=0) or 0,
            )
        )
    return out


# --------------------------------------------------------------------------- #
# Calculo y balanceo por fases
# --------------------------------------------------------------------------- #

def _assign_balance(etq: str, amps: float, fases: int, tot: List[float],
                    lim_fases: int, cargas: List[CargaFase]) -> None:
    """
    Replica la logica de AsignarCarga del VBA:
    - monofasica: usa la fase mas liviana dentro del limite.
    - bifasica: usa las dos fases mas livianas dentro del limite.
    - trifasica: reparte en las tres.
    """
    if amps <= 0 or fases <= 0:
        return

    limit = max(1, min(lim_fases, 3))
    dist = [0.0, 0.0, 0.0]

    if fases >= 3:
        for i in range(3):
            tot[i] += amps
            dist[i] = amps
    elif fases == 2:
        allowed = list(range(limit))
        order = sorted(allowed, key=lambda idx: tot[idx])
        picks = order[:2] if len(order) >= 2 else order
        for idx in picks:
            tot[idx] += amps
            dist[idx] = amps
    else:  # monofasica
        allowed = list(range(limit))
        order = sorted(allowed, key=lambda idx: tot[idx])
        idx = order[0] if order else 0
        tot[idx] += amps
        dist[idx] = amps

    cargas.append(CargaFase(etq, dist[0], dist[1], dist[2]))


# --------------------------------------------------------------------------- #
# Motor principal
# --------------------------------------------------------------------------- #

def calcular_cuadro(book_path: Path, selection: Optional[List[SeleccionRamal]] = None) -> List[ResultadoRamal]:
    """
    Calcula el cuadro de cargas completo a partir de la hoja SELECCION
    o de un listado de seleccion entregado por la GUI.
    """
    book_path = Path(book_path)
    muebles = _load_muebles(book_path)
    unidades = _load_unidades(book_path)
    seleccion = selection if selection is not None else _parse_seleccion(book_path)

    resultados: List[ResultadoRamal] = []

    for sel in seleccion:
        key = (_norm_key(sel.equipo), _norm_key(sel.medida))
        mueble = muebles.get(key)
        cargas: List[CargaFase] = []

        # Buffers para repetir el orden del VBA
        items_3: List[Tuple[str, float, int]] = []
        items_2: List[Tuple[str, float, int]] = []
        items_1: List[Tuple[str, float, int]] = []
        lim_fases = 1
        tot = [0.0, 0.0, 0.0]

        def enqueue(etq: str, amps: float, fases: int) -> None:
            nonlocal lim_fases
            if amps <= 0:
                return
            fases = max(1, min(int(fases or 1), 3))
            lim_fases = max(lim_fases, fases)
            if fases >= 3:
                items_3.append((etq, amps, fases))
            elif fases == 2:
                items_2.append((etq, amps, fases))
            else:
                items_1.append((etq, amps, fases))

        if mueble:
            # Iluminacion (base + por nivel)
            ilum = mueble.illumination + mueble.ilum_por_nivel * max(sel.niveles, 0)
            enqueue("ILUMINACION", ilum, mueble.fases.get("illumination", 1))
            enqueue("VENTILADORES", mueble.fans, mueble.fases.get("fans", 1))

            # Resistencias solo si el descongelamiento es RESISTENCIAS
            descong_norm = _norm(sel.descongelamiento)
            if "RESIST" in descong_norm:
                enqueue("RESISTENCIAS", mueble.res_descong, mueble.fases.get("res_descong", 1))

            # Otras cargas auxiliares
            for etq, val, fas in [
                ("RESISTENCIA DESEMPANANTE", mueble.res_desemp, mueble.fases.get("res_desemp", 1)),
                ("RESISTENCIA ANTISUDOR", mueble.res_antisudor, mueble.fases.get("res_antisudor", 1)),
                ("RESISTENCIA DESAGUE", mueble.res_desague, mueble.fases.get("res_desague", 1)),
                ("RESISTENCIA CALEFACTORA", mueble.res_calef, mueble.fases.get("res_calef", 1)),
                ("OTRAS CARGAS", mueble.turb_camara, mueble.fases.get("turb_camara", 1)),
                ("OTRAS CARGAS", mueble.turb_puerta, mueble.fases.get("turb_puerta", 1)),
                ("OTRAS CARGAS", mueble.motor_carro, mueble.fases.get("motor_carro", 1)),
            ]:
                if _as_float(val, 0.0) > 0:
                    enqueue(etq, _as_float(val, 0.0), fas)

        # Unidad (compresor/condensador)
        if sel.tiene_unidad and sel.ref_unidad:
            uni = unidades.get(_norm(sel.ref_unidad))
            if uni:
                enqueue(f"COMPRESOR {sel.ref_unidad}", uni.compresor_amp, uni.compresor_fases or 3)
                enqueue(f"CONDENSADOR {sel.ref_unidad}", uni.condensador_amp, uni.condensador_fases or 3)

        # Asignar respetando el orden: trifasicas -> bifasicas -> monofasicas
        for etq, amps, fas in items_3:
            _assign_balance(etq, amps, fas, tot, lim_fases, cargas)
        for etq, amps, fas in items_2:
            _assign_balance(etq, amps, fas, tot, lim_fases, cargas)
        for etq, amps, fas in items_1:
            _assign_balance(etq, amps, fas, tot, lim_fases, cargas)

        resultados.append(ResultadoRamal(sel, cargas))

    return resultados


# --------------------------------------------------------------------------- #
# Exportar a Excel
# --------------------------------------------------------------------------- #

def escribir_cuadro(book_path: Path, resultados: List[ResultadoRamal], sheet_name: str = "CUADRO_GENERADO") -> None:
    """Escribe el cuadro en una hoja nueva (sobrescribe si existe)."""
    book_path = Path(book_path)
    wb = load_workbook(book_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    row = 1
    for res in resultados:
        sel = res.ramal
        titulo = f"RAMAL {sel.ramal}: {sel.equipo} {sel.uso}".strip()
        ws.cell(row=row, column=2, value=titulo)
        ws.cell(row=row, column=3, value="MODELO")
        ws.cell(row=row, column=5, value="L1")
        ws.cell(row=row, column=6, value="L2")
        ws.cell(row=row, column=7, value="L3")
        row += 1
        ws.cell(row=row, column=3, value=sel.medida)
        row += 1
        for carga in res.cargas:
            ws.cell(row=row, column=2, value=carga.etiqueta)
            ws.cell(row=row, column=5, value=round(carga.l1, 2))
            ws.cell(row=row, column=6, value=round(carga.l2, 2))
            ws.cell(row=row, column=7, value=round(carga.l3, 2))
            row += 1
        total = res.total
        ws.cell(row=row, column=2, value=total.etiqueta)
        ws.cell(row=row, column=5, value=round(total.l1, 2))
        ws.cell(row=row, column=6, value=round(total.l2, 2))
        ws.cell(row=row, column=7, value=round(total.l3, 2))
        row += 2  # espacio entre ramales

    wb.save(book_path)


def recalcular(book_path: Path, write_sheet: bool = True, sheet_name: str = "CUADRO_GENERADO",
               selection: Optional[List[SeleccionRamal]] = None) -> List[ResultadoRamal]:
    """
    Calcula y (opcionalmente) escribe el cuadro de cargas en el mismo Excel.
    Devuelve la lista de resultados por ramal.
    """
    resultados = calcular_cuadro(book_path, selection=selection)
    if write_sheet:
        escribir_cuadro(book_path, resultados, sheet_name=sheet_name)
    return resultados


__all__ = [
    "MuebleLoad",
    "Unidad",
    "SeleccionRamal",
    "CargaFase",
    "ResultadoRamal",
    "calcular_cuadro",
    "escribir_cuadro",
    "recalcular",
]
