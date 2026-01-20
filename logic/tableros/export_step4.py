# logic/export_step4.py
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple
import math
import json
import re
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .step4_engine import Step4Engine
from .opciones_co2_engine import ResumenTable
from .resumen_cables import build_cable_badges


# ----------------------------- helpers de estilo -----------------------------
_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

F_WHITE = PatternFill("solid", fgColor="FFFFFF")
F_GREY = PatternFill("solid", fgColor="F2F4F7")      # gris muy claro
F_GREY2 = PatternFill("solid", fgColor="E9EDF3")     # gris claro para cabeceras

FONT_TITLE = Font(name="Calibri", bold=True, size=12, color="000000")
FONT_HEADER = Font(name="Calibri", bold=True, size=11, color="000000")
FONT_CELL = Font(name="Calibri", size=11, color="000000")


def _autosize(ws) -> None:
    """Autoajuste simple por ancho de texto."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = max(10, min(90, max_len + 2))


def _write_meta(ws, meta: List[Tuple[str, str]], start_row: int = 1) -> int:
    """Escribe bloques clave/valor (PROYECTO, CIUDAD, …) con estilo claro."""
    r = start_row
    for label, value in meta:
        c1 = ws.cell(row=r, column=1, value=label)
        c2 = ws.cell(row=r, column=2, value=value)
        for c in (c1, c2):
            c.fill = F_GREY
            c.font = FONT_CELL if c is c2 else FONT_HEADER
            c.alignment = Alignment(vertical="center")
            c.border = _BORDER
        r += 1
    return r + 1  # deja una fila en blanco


def _write_section_header(ws, title: str, row: int) -> int:
    c = ws.cell(row=row, column=1, value=title)
    c.font = FONT_TITLE
    c.fill = F_GREY2
    c.alignment = Alignment(vertical="center")
    c.border = _BORDER
    # banda hasta la J (10 columnas)
    for j in range(2, 11):
        cc = ws.cell(row=row, column=j, value=None)
        cc.fill = F_GREY2
        cc.border = _BORDER
    return row + 1


def _write_subtitle(ws, parts: List[str], row: int) -> int:
    """Línea de subtítulo (compresor / corriente / cable / cap / ICOND) en negro."""
    text = "  |  ".join([p for p in parts if p])
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_CELL  # NEGRO
    c.alignment = Alignment(vertical="center")
    c.border = _BORDER
    for j in range(2, 11):
        cc = ws.cell(row=row, column=j)
        cc.border = _BORDER
    return row + 1


def _write_table(ws, headers: List[str], rows: List[List[str]], row: int) -> int:
    # cabecera
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=j, value=h)
        cell.font = FONT_HEADER
        cell.fill = F_GREY2
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
    row += 1

    # filas
    for rdata in rows:
        for j, val in enumerate(rdata, start=1):
            cell = ws.cell(row=row, column=j, value=val)
            cell.font = FONT_CELL
            cell.fill = F_WHITE
            cell.alignment = Alignment(vertical="center")
            cell.border = _BORDER
        row += 1
    return row + 1  # espacio


def _write_corriente_total(ws, data: Dict[str, object], row: int) -> int:
    """
    Escribe el cuadro de corriente total (mayor +25% + suma restantes).
    Espera campos: 'mayor', 'ajuste_mayor', 'suma_restantes', 'total', 'suma_simple' y 'breaker' (dict).
    """
    ws.cell(row=row, column=1, value="CORRIENTE TOTAL DEL SISTEMA").font = FONT_TITLE
    row += 1
    # Tabla con cada compresor
    headers = ["COMPRESOR", "CORRIENTE (A)", "CORR. AJUSTADA (A)"]
    comp_rows = []
    comps = data.get("comp_detalles") if isinstance(data, dict) else None
    if isinstance(comps, list):
        for d in comps:
            key = str(d.get("comp_key", "")).upper()
            modelo = str(d.get("modelo", "") or "").strip()
            if modelo:
                key = f"{key} - {modelo}".upper()
            amps = float(d.get("amps", 0.0))
            adj = amps * 1.25 if d.get("ajustado") else amps
            comp_rows.append([key, amps, adj])
    if comp_rows:
        row = _write_table(ws, headers, comp_rows, row)
    # Resumen de totales
    headers2 = ["MÉTRICA", "VALOR (A)"]
    row = _write_table(ws, headers2,
                       [["TOTAL AJUSTADO", data.get("total", 0.0)],
                        ["TOTAL SIN AJUSTE", data.get("suma_simple", 0.0)]],
                       row)
    brk = data.get("breaker", {}) if isinstance(data, dict) else {}
    info_txt = brk.get("motivo", "No se encontró breaker >= corriente calculada")
    if brk.get("found"):
        modelo = (brk.get("modelo") or "").strip()
        info_txt = f"BREAKER TOTALIZADOR: {modelo} ({brk.get('amp',0)} A)"
        if brk.get("codigo"):
            info_txt += f"  - CÓDIGO: {brk.get('codigo')}"
    ws.cell(row=row, column=1, value=info_txt).font = FONT_HEADER
    return row + 2


def _sanitize_name(name: str) -> str:
    keep = "".join(ch for ch in (name or "") if ch.isalnum() or ch in (" ", "-", "_"))
    return keep.strip().replace(" ", "_") or "Proyecto"


def _ensure_item_col(rows: List[List[str]]) -> List[List[str]]:
    """
    Si una fila trae 8 columnas (sin ITEM), usa la REFERENCIA como ITEM.
    Formato de 8 cols esperado (sin ITEM):
      [CÓDIGO, MODELO, NOMBRE, DESCRIPCIÓN, ICC240, ICC480, REFERENCIA, TORQUE]
    """
    out: List[List[str]] = []
    for r in rows or []:
        if len(r) == 8:
            item = r[6] if len(r) > 6 else ""
            out.append([item] + r)
        else:
            out.append(r)
    return out


def _inject_marca(rows: List[List[str]], marca_default: str = "", brand_map: Dict[str, str] | None = None) -> List[List[str]]:
    """
    Inserta la columna MARCA entre DESCRIPCIÓN e ICC240.
    Estructura de salida: [ITEM, CÓDIGO, MODELO, NOMBRE, DESCRIPCIÓN, MARCA, ICC240, ICC480, REF, TORQUE]
    """
    new_rows: List[List[str]] = []
    for r in rows or []:
        item = codigo = modelo = nombre = desc = marca = icc240 = icc480 = ref = torque = ""
        if len(r) >= 10:
            item, codigo, modelo, nombre, desc, marca, icc240, icc480, ref, torque = r[:10]
        elif len(r) >= 9:
            item, codigo, modelo, nombre, desc, icc240, icc480, ref, torque = r[:9]
            marca = marca_default
        else:
            # relleno mínimo si viene más corto
            rr = list(r) + [""] * (9 - len(r))
            if len(rr) >= 9:
                item, codigo, modelo, nombre, desc, icc240, icc480, ref, torque = rr[:9]
                marca = marca_default
        code_clean = str(codigo or "").strip()
        try:
            f = float(code_clean)
            if f.is_integer():
                code_clean = str(int(f))
        except Exception:
            pass
        marca_lookup = (brand_map or {}).get(code_clean, "")
        marca_final = marca_lookup if marca_lookup else marca_default
        new_rows.append([item, codigo, modelo, nombre, desc, marca_final, icc240, icc480, ref, torque])
    return new_rows


def _build_brand_map(book: Path) -> Dict[str, str]:
    """Construye un mapa código->marca leyendo hojas con columnas CODIGO y MARCA."""
    brand_map: Dict[str, str] = {}
    paths = [book]
    # intentar también BD_TABLEROS.xlsx si existe en la misma carpeta
    extra = book.parent / "BD_TABLEROS.xlsx"
    if extra.exists():
        paths.append(extra)
    try:
        import pandas as pd  # se usa solo aquí para el Excel
        for path in paths:
            try:
                xls = pd.ExcelFile(path)
            except Exception:
                continue
            for sheet in xls.sheet_names:
                try:
                    df = pd.read_excel(path, sheet_name=sheet)
                except Exception:
                    continue
                def norm_col(c: str) -> str:
                    import unicodedata, re
                    cc = unicodedata.normalize("NFD", c)
                    cc = "".join(ch for ch in cc if unicodedata.category(ch) != "Mn")
                    cc = re.sub(r"\s+", " ", cc)
                    return cc.strip().upper()

                cols = {norm_col(c): c for c in df.columns if isinstance(c, str)}
                if not cols:
                    continue
                if ("CODIGO" in cols or "CÓDIGO" in cols):
                    c_code = cols.get("CODIGO") or cols.get("CÓDIGO")
                    c_brand = cols.get("MARCA")  # algunas hojas sí tienen

                    def _norm_code(val) -> str:
                        s = str(val).strip()
                        try:
                            f = float(s)
                            if f.is_integer():
                                s = str(int(f))
                        except Exception:
                            pass
                        return s

                    for _, row in df[[c_code]].dropna(subset=[c_code]).iterrows():
                        val = row[c_code]
                        code = _norm_code(val)
                        if not code:
                            continue
                        marca = ""
                        if c_brand and c_brand in df.columns:
                            try:
                                marca = str(df.loc[_, c_brand]).strip()
                            except Exception:
                                marca = ""
                        if not marca:
                            marca = norm_col(sheet)  # usa el nombre de la hoja como marca
                        if code and marca and code not in brand_map:
                            brand_map[code] = marca
    except Exception:
        pass
    return brand_map


# ----------------------------- inventario ------------------------------------
def _load_inventario_map(inv_path: Path) -> Dict[str, Dict[str, float | str]]:
    """
    Lee data/inventarios.xlsx (Hoja1) y devuelve un mapa codigo->{disp, solic}.
    Columnas esperadas (case-insensitive):
      - Referencia (código)
      - Cant. disponible
      - Cant. solic. compra
    Si el archivo no existe o falla la lectura, retorna {}.
    """
    inv: Dict[str, Dict[str, float]] = {}
    if not inv_path.exists():
        return inv
    try:
        df = pd.read_excel(inv_path, sheet_name=0)
    except Exception:
        return inv

    # normalizar nombres de columnas
    def norm_col(c: str) -> str:
        import unicodedata, re
        cc = unicodedata.normalize("NFD", str(c))
        cc = "".join(ch for ch in cc if unicodedata.category(ch) != "Mn")
        cc = re.sub(r"\s+", " ", cc)
        return cc.strip().upper()

    cols = {norm_col(c): c for c in df.columns if isinstance(c, str)}
    c_ref = cols.get("REFERENCIA")
    c_desc = cols.get("DESC. ITEM") or cols.get("DESC ITEM")
    c_um = cols.get("U.M. INVENT.") or cols.get("U.M. INVENT") or cols.get("U.M. INVENTARIO")
    c_disp = cols.get("CANT. DISPONIBLE") or cols.get("CANT DISPONIBLE")
    c_solic = cols.get("CANT. SOLIC. COMPRA") or cols.get("CANT SOLIC COMPRA")
    if not c_ref:
        return inv

    for _, row in df.iterrows():
        code = str(row.get(c_ref, "")).strip()
        if not code:
            continue
        try:
            code_num = float(code)
            if code_num.is_integer():
                code = str(int(code_num))
        except Exception:
            pass
        desc = str(row.get(c_desc, "")).strip() if c_desc else ""
        um = str(row.get(c_um, "")).strip() if c_um else ""
        disp = row.get(c_disp, 0) if c_disp else 0
        solic = row.get(c_solic, 0) if c_solic else 0
        try:
            disp_f = float(disp)
        except Exception:
            disp_f = 0.0
        try:
            solic_f = float(solic)
        except Exception:
            solic_f = 0.0
        cur = inv.get(code, {"disp": 0.0, "solic": 0.0, "desc": "", "um": ""})
        if desc and not cur.get("desc"):
            cur["desc"] = desc
        if um and not cur.get("um"):
            cur["um"] = um
        cur["disp"] += disp_f
        cur["solic"] += solic_f
        inv[code] = cur
    return inv


def _write_inventario_sheet(wb: Workbook, tot_rows: List[List[str]], inv_map: Dict[str, Dict[str, float | str]]) -> None:
    """
    Agrega hoja INVENTARIO con comparativo requerido vs disponible/solicitado.
    tot_rows: filas [codigo, modelo, nombre, desc, marca, icc240, icc480, ref, torque, cantidad]
    """
    ws = wb.create_sheet("INVENTARIO")
    headers = ["CÓDIGO", "MODELO", "NOMBRE", "REQUERIDO", "DISPONIBLE", "SOLICITADO", "STOCK TOTAL", "FALTANTE"]
    _write_table(ws, headers, [], 1)  # escribe cabecera; luego rellenamos manualmente

    # agregar filas debajo de la cabecera
    row = 2
    for r in tot_rows:
        if len(r) < 10:
            continue
        code, modelo, nombre = r[0], r[1], r[2]
        try:
            req = float(r[9])
        except Exception:
            req = 0.0
        if req <= 0:
            continue
        inv_info = inv_map.get(str(code).strip(), {"disp": 0.0, "solic": 0.0})
        disp = inv_info.get("disp", 0.0) or 0.0
        solic = inv_info.get("solic", 0.0) or 0.0
        stock = disp + solic
        falt = max(req - stock, 0.0)
        data_row = [code, modelo, nombre, req, disp, solic, stock, falt]
        for j, val in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=j, value=val)
            cell.font = FONT_CELL
            cell.fill = F_WHITE
            cell.border = _BORDER
        row += 1

    _autosize(ws)


def _as_num(value: object) -> object:
    try:
        f = float(value)
        if f.is_integer():
            return int(f)
        return f
    except Exception:
        return value


def _norm_text(s: object) -> str:
    v = "" if s is None else str(s)
    v = v.strip().upper()
    v = v.translate(str.maketrans("ÁÉÍÓÚÜÑ", "AEIOUUN"))
    v = v.replace("  ", " ")
    return v


def _norm_code(code: object) -> str:
    s = "" if code is None else str(code).strip()
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
    except Exception:
        pass
    return s


def _step3_yes(step3: Dict[str, object], label: str) -> bool:
    if not isinstance(step3, dict):
        return False
    want = _norm_text(label)
    for k, v in step3.items():
        if _norm_text(k) == want:
            if isinstance(v, dict) and "value" in v:
                v = v["value"]
            vv = _norm_text(v)
            return vv in ("SI", "S", "TRUE", "1")
    return False


def _qty_from_notes(notes: str, total_comp: int, total_borneras: int) -> int:
    if not notes:
        return 0
    t = _norm_text(notes)
    nums = [int(n) for n in re.findall(r"\d+", t)]
    if "BORNER" in t:
        if 20 in nums:
            factor = nums[-1] if nums else 1
            return int(math.ceil(total_borneras / 20.0) * factor)
        return nums[-1] if nums else 1
    if "COMPRESOR" in t:
        factor = nums[-1] if nums else 1
        return max(total_comp, 0) * factor
    if "PROYECTO" in t:
        return nums[-1] if nums else 1
    return nums[-1] if nums else 1


def _load_reglas_listado(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    try:
        df = pd.read_excel(path, sheet_name=0)
    except Exception:
        return []
    cols = {str(c).strip().lower(): c for c in df.columns}
    def col(name: str) -> str | None:
        return cols.get(name.lower())
    c_code = col("codigo")
    c_name = col("nombre")
    c_um = col("um")
    c_rule = col("regla")
    c_param = col("parametro")
    c_notes = col("notas")
    if not c_code:
        return []
    rows: List[Dict[str, str]] = []
    for _, row in df.iterrows():
        code = _norm_code(row.get(c_code, ""))
        if not code:
            continue
        rows.append({
            "codigo": code,
            "nombre": str(row.get(c_name, "") or "").strip(),
            "um": str(row.get(c_um, "") or "").strip(),
            "regla": str(row.get(c_rule, "") or "").strip(),
            "parametro": str(row.get(c_param, "") or "").strip(),
            "notas": str(row.get(c_notes, "") or "").strip(),
        })
    return rows


def _apply_reglas_listado(
    rules: List[Dict[str, str]],
    step3: Dict[str, object],
    norma_ap: str,
    total_comp: int,
    total_borneras: int,
) -> Dict[str, Dict[str, object]]:
    extras: Dict[str, Dict[str, object]] = {}
    norma_u = _norm_text(norma_ap)
    for r in rules:
        regla = _norm_text(r.get("regla"))
        if not regla:
            continue
        param = _norm_text(r.get("parametro"))
        cond = False
        if regla in ("IEC", "UL"):
            cond = regla == norma_u
        else:
            cond = _step3_yes(step3, regla)

        if param == "PONE":
            if not cond:
                continue
        elif param == "QUITA":
            if cond:
                continue
        else:
            if not cond:
                continue

        qty = _qty_from_notes(r.get("notas", ""), total_comp, total_borneras)
        if qty <= 0:
            continue
        code = _norm_code(r.get("codigo"))
        if not code:
            continue
        cur = extras.get(code, {"qty": 0, "nombre": r.get("nombre", ""), "um": r.get("um", "")})
        cur["qty"] = int(cur.get("qty", 0)) + int(qty)
        if not cur.get("nombre"):
            cur["nombre"] = r.get("nombre", "")
        if not cur.get("um"):
            cur["um"] = r.get("um", "")
        extras[code] = cur
    return extras


def _write_listado_costos_sheet(
    wb: Workbook,
    tot_rows: List[List[str]],
    inv_map: Dict[str, Dict[str, float | str]],
    step2_state: Dict[str, Dict[str, str]],
    globs: Dict[str, object],
    res: Dict[str, object],
) -> None:
    """
    Hoja tipo "LISTADO" para costos, similar a LISTADO DE MATERIALES.
    Columnas: CODIGO, NOMBRE, U.M., Cant. requerida, Disponible, Cant. solic. compra, (saldo).
    Inserta dos filas en blanco cada 36 ítems para impresión.
    """
    ws = wb.create_sheet("LISTADO", 1)
    headers = ["CODIGO", "NOMBRE", "U.M.", "Cant. requerida", "Disponible", "Cant. solic. compra", ""]

    # base desde tot_rows
    listado: Dict[str, Dict[str, object]] = {}
    for r in tot_rows:
        if len(r) < 10:
            continue
        code = _norm_code(r[0])
        if not code:
            continue
        try:
            req = float(r[-1])
        except Exception:
            req = 0.0
        if req <= 0:
            continue
        inv = inv_map.get(code, {})
        desc = str(inv.get("desc") or r[2] or r[3] or "").strip()
        um = str(inv.get("um") or "").strip()
        cur = listado.get(code, {"req": 0.0, "desc": desc, "um": um})
        cur["req"] = float(cur.get("req", 0.0)) + req
        if not cur.get("desc"):
            cur["desc"] = desc
        if not cur.get("um"):
            cur["um"] = um
        listado[code] = cur

    # aplicar reglas desde excel (si existe)
    rules_path = Path("data/tableros_electricos/reglas_listado_costos_base.xlsx")
    rules = _load_reglas_listado(rules_path)
    step3 = (globs.get("step3_state") or globs.get("step3") or {}) if isinstance(globs, dict) else {}
    norma_ap = str(globs.get("norma_ap", "")).strip() if isinstance(globs, dict) else ""
    total_comp = len(step2_state or {})
    b_total = res.get("borneras_total", {}) if isinstance(res, dict) else {}
    total_borneras = int(b_total.get("fase", 0) or 0) + int(b_total.get("neutro", 0) or 0) + int(b_total.get("tierra", 0) or 0)
    extras = _apply_reglas_listado(rules, step3, norma_ap, total_comp, total_borneras)
    for code, data in extras.items():
        cur = listado.get(code, {"req": 0.0, "desc": data.get("nombre", ""), "um": data.get("um", "")})
        cur["req"] = float(cur.get("req", 0.0)) + float(data.get("qty", 0) or 0)
        if not cur.get("desc"):
            cur["desc"] = data.get("nombre", "")
        if not cur.get("um"):
            cur["um"] = data.get("um", "")
        listado[code] = cur

    # construir filas ordenadas
    def _sort_key(x: str) -> tuple:
        try:
            return (0, int(float(x)))
        except Exception:
            return (1, x)

    rows: List[List[object]] = []
    count = 0
    for code in sorted(listado.keys(), key=_sort_key):
        cur = listado[code]
        req = float(cur.get("req", 0.0) or 0.0)
        if req <= 0:
            continue
        inv = inv_map.get(code, {})
        disp = float(inv.get("disp", 0.0) or 0.0)
        solic = float(inv.get("solic", 0.0) or 0.0)
        saldo = (disp + solic) - req
        rows.append([
            code,
            str(cur.get("desc", "") or "").strip(),
            str(cur.get("um", "") or "").strip(),
            _as_num(req),
            _as_num(disp),
            _as_num(solic),
            _as_num(saldo),
        ])
        count += 1
        if count % 36 == 0:
            rows.append([""] * len(headers))
            rows.append([""] * len(headers))

    _write_table(ws, headers, rows, 1)
    _autosize(ws)


def _collect_totales(rows: List[List[str]], marca_default: str = "") -> List[List[str]]:
    """
    Suma por CÓDIGO. Entrada: filas con columnas ya normalizadas:
    [ITEM, CÓDIGO, MODELO, NOMBRE, DESCRIPCIÓN, MARCA, ICC240, ICC480, REF, TORQUE]
    """
    agg: Dict[str, Dict[str, object]] = {}

    def add_row(row: List[str]) -> None:
        if len(row) < 2:
            return
        code = (row[1] or "").strip()
        if not code:
            return
        if code not in agg:
            agg[code] = {
                "codigo": code,
                "modelo": row[2] if len(row) > 2 else "",
                "nombre": row[3] if len(row) > 3 else "",
                "descripcion": row[4] if len(row) > 4 else "",
                "marca": row[5] if len(row) > 5 else marca_default,
                "icc240": row[6] if len(row) > 6 else "",
                "icc480": row[7] if len(row) > 7 else "",
                "ref": row[8] if len(row) > 8 else "",
                "torque": row[9] if len(row) > 9 else "",
                "cantidad": 0,
            }
        agg[code]["cantidad"] = int(agg[code]["cantidad"]) + 1  # type: ignore

    for r in (rows or []):
        add_row(r)

    # orden por código
    ordered = sorted(agg.values(), key=lambda d: str(d["codigo"]))
    out: List[List[str]] = []
    for d in ordered:
        out.append([
            d["codigo"], d["modelo"], d["nombre"], d["descripcion"],
            d.get("marca", marca_default),
            d["icc240"], d["icc480"], d["ref"], d["torque"],
            str(d["cantidad"]),
        ])
    return out


# ----------------------------- API de exportación ----------------------------
def export_step4_excel_only(
    basedatos_path: Path | str,
    step2_state: Dict[str, Dict[str, str]],
    globs: Dict[str, object],
    out_dir: Path | str,
) -> Path:
    """
    Genera SOLO el Excel (hojas: Resumen, Totales, Borneras) con estilo claro.
    Devuelve la ruta del archivo.
    """
    book = Path(basedatos_path)
    engine = Step4Engine(book)
    res = engine.calcular(step2_state, globs)

    tables: List[ResumenTable] = res.get("tables_compresores", []) or []
    otros_rows: List[List[str]] = _ensure_item_col(res.get("otros_rows", []) or [])
    corriente_total = res.get("corriente_total", {})
    b_comp = res.get("borneras_compresores", {"fase": 0, "neutro": 0, "tierra": 0})
    b_otros = res.get("borneras_otros", {"fase": 0, "neutro": 0, "tierra": 0})
    b_total = res.get("borneras_total", {"fase": 0, "neutro": 0, "tierra": 0})

    # badges de cable para subtítulo por compresor
    cable_badges = build_cable_badges(step2_state, book)

    brand_map = _build_brand_map(book)

    wb = Workbook()
    ws = wb.active
    ws.title = "RESUMEN"

    # ---- Encabezado del proyecto (print-friendly) ----
    meta = [
        ("PROYECTO",       str(globs.get("nombre_proyecto", ""))),
        ("CIUDAD",         str(globs.get("ciudad", ""))),
        ("RESPONSABLE",    str(globs.get("responsable", ""))),
        ("TENSIÓN ALIM.",  str(globs.get("t_alim", ""))),
        ("TENSIÓN CONTROL",str(globs.get("t_ctl", ""))),
        ("NORMA",          str(globs.get("norma_ap", ""))),
    ]
    row = _write_meta(ws, meta, start_row=1)

    all_rows_for_totals: List[List[str]] = []

    # ---- Secciones por compresor ----
    headers = ["ITEM", "CÓDIGO", "MODELO", "NOMBRE", "DESCRIPCIÓN",
               "MARCA", "C 240 V (kA)", "C 480 V (kA)", "REFERENCIA", "TORQUE"]

    for t in tables:
        row = _write_section_header(ws, t.title, row)
        # subtítulo: compresor + corriente + cable
        parts: List[str] = []
        st = step2_state.get(t.comp_key, {}) or {}
        _, model, amps = (st.get("modelo", ""), st.get("modelo", ""), st.get("corriente", ""))
        if st.get("modelo") or st.get("marca"):
            parts.append(f"COMPRESOR: {st.get('modelo','')}".strip())
        if amps:
            parts.append(f"CORRIENTE: {amps} A")
        badge = cable_badges.get(t.comp_key)
        if badge:
            parts.append(badge.strip(" |"))
        row = _write_subtitle(ws, parts, row)
        # inyecta marca para cada fila (si no viene, usa marca global)
        marca_global = str(globs.get("marca_elem", "")).strip()
        filas = _inject_marca(t.rows, marca_global, brand_map)
        all_rows_for_totals.extend(filas)
        row = _write_table(ws, headers, filas, row)

    # ---- Otros fijos ----
    if otros_rows:
        row = _write_section_header(ws, "OTROS ELEMENTOS (FIJOS)", row)
        # aplicar marca global también aquí
        marca_global = str(globs.get("marca_elem", "")).strip()
        otros_fixed = _inject_marca(otros_rows, marca_global, brand_map)
        all_rows_for_totals.extend(otros_fixed)
        row = _write_table(ws, headers, otros_fixed, row)

    # ---- Corriente total del sistema ----
    if corriente_total:
        detalle = corriente_total.get("detalle", {}) if isinstance(corriente_total, dict) else {}
        if not detalle and isinstance(corriente_total, dict):
            detalle = {k: v for k, v in corriente_total.items() if isinstance(v, (int, float))}
        breaker_info = corriente_total.get("breaker") if isinstance(corriente_total, dict) else {}
        if not breaker_info:
            breaker_info = res.get("breaker_total", {})
        comp_det = corriente_total.get("comp_detalles", []) if isinstance(corriente_total, dict) else []
        data_corr = {**detalle, "breaker": breaker_info, "comp_detalles": comp_det}
        row = _write_corriente_total(ws, data_corr, row)

    _autosize(ws)

    # ---- Totales (por código) ----
    ws2 = wb.create_sheet("TOTALES")
    marca_global = str(globs.get("marca_elem", "")).strip() or "SIN MARCA"
    tot_rows = _collect_totales(all_rows_for_totals, marca_global)
    # inventario local
    inv_path = Path("data/tableros_electricos/inventarios.xlsx")
    inv_map = _load_inventario_map(inv_path)
    _write_listado_costos_sheet(wb, tot_rows, inv_map, step2_state, globs, res)

    headers_tot = ["CÓDIGO", "MODELO", "NOMBRE", "DESCRIPCIÓN", "MARCA",
                   "C 240 V (kA)", "C 480 V (kA)", "REFERENCIA", "TORQUE", "CANTIDAD",
                   "DISPONIBLE", "SOLICITADO", "STOCK TOTAL", "FALTANTE"]

    rows_tot_inv: List[List[object]] = []
    for r in tot_rows:
        code = str(r[0])
        qty = 0.0
        try:
            qty = float(r[-1])
        except Exception:
            qty = 0.0
        inv_info = inv_map.get(code, {"disp": 0.0, "solic": 0.0})
        disp = inv_info.get("disp", 0.0) or 0.0
        solic = inv_info.get("solic", 0.0) or 0.0
        stock = disp + solic
        falt = max(qty - stock, 0.0)
        rows_tot_inv.append(r + [disp, solic, stock, falt])

    _ = _write_table(ws2, headers_tot, rows_tot_inv, 1)
    _autosize(ws2)

    # ---- Borneras ----
    ws3 = wb.create_sheet("BORNERAS")
    heads = ["TIPO", "TOTAL"]
    _ = _write_table(
        ws3, heads,
        [["FASE", b_total["fase"]], ["NEUTRO", b_total["neutro"]], ["TIERRA", b_total["tierra"]]],
        1,
    )
    _autosize(ws3)

    # ---- Guardar ----
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    project = _sanitize_name(str(globs.get("nombre_proyecto", ""))).upper()
    stamp = datetime.now().strftime("%Y%m%d")
    out_path = out_dir / f"{stamp}_{project}.xlsx"
    wb.save(out_path)
    return out_path


# ----------------------- snapshot de programación (.ecalc.json) ----------------
def _normalize_step2_for_json(step2: Dict[str, Dict[str, str]]) -> Dict[str, Dict[str, str]]:
    """Garantiza tipos JSON-friendly."""
    out: Dict[str, Dict[str, str]] = {}
    for k, v in (step2 or {}).items():
        if isinstance(v, dict):
            rec: Dict[str, str] = {}
            for kk, vv in v.items():
                rec[str(kk)] = "" if vv is None else str(vv)
            out[str(k)] = rec
        else:
            # si no es dict, guarda el valor como string plano
            out[str(k)] = {"value": "" if v is None else str(v)}
    return out


def save_programacion_snapshot(
    step2_state: Dict[str, Dict[str, str]],
    globs: Dict[str, object],
    out_dir: Path | str,
    filename_prefix: str | None = None,
) -> Path:
    """
    Guarda un snapshot .ecalc.json con TODO lo necesario para recargar.
    Nombre: YYYYMMDD_<PROYECTO>.ecalc.json (o prefix si se indica)
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    project = _sanitize_name(str(globs.get("nombre_proyecto", ""))).upper()
    stamp = datetime.now().strftime("%Y%m%d")
    base = filename_prefix if filename_prefix else project
    fname = f"{stamp}_{base}.ecalc.json"
    out_path = out_dir / fname

    # contar compresores por prefijo para ayudar al reconstruir
    keys_upper = [str(k).upper() for k in (step2_state or {})]
    # Convención interna: G# baja, B# media (y algunos proyectos antiguos usan M#)
    nb = sum(1 for k in keys_upper if k.startswith("G"))
    nm = sum(1 for k in keys_upper if k.startswith(("B", "M")))
    # Paralelos: en algunas cargas vienen con prefijo F o P
    np = sum(1 for k in keys_upper if k.startswith(("F", "P")))

    payload = {
        "globs": {
            "nombre_proyecto": str(globs.get("nombre_proyecto", "")),
            "ciudad": str(globs.get("ciudad", "")),
            "responsable": str(globs.get("responsable", "")),
            "t_alim": str(globs.get("t_alim", "")),
            "t_ctl": str(globs.get("t_ctl", "")),
            "norma_ap": str(globs.get("norma_ap", "")),
            "refrigerante": str(globs.get("refrigerante", "")),
            "marca_elem": str(globs.get("marca_elem", "")),
            "marca_variadores": str(globs.get("marca_variadores", "")),
            "tipo_compresores": str(globs.get("tipo_compresores", "")),
            "n_comp_baja": nb,
            "n_comp_media": nm,
            "n_comp_paralelo": np,
            "step3_state": globs.get("step3_state", {}),
        },
        "step2": step2_state,
    }

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=str)

    return out_path


def save_programacion_exact(
    step2_state: Dict[str, Dict[str, str]],
    globs: Dict[str, object],
    out_path: Path | str,
) -> Path:
    """
    Guarda la programación EXACTAMENTE en 'out_path' (p.ej. 'C:/.../MI_PROYECTO.ecalc.json').
    - Crea la carpeta si no existe.
    - Genera JSON correcto y sustituye si ya existe.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Asegurar extensión .ecalc.json
    if not out_path.name.lower().endswith(".ecalc.json"):
        out_path = out_path.with_name(out_path.stem + ".ecalc.json")

    # Generar snapshot temporal con el mismo "stem"
    tmp = save_programacion_snapshot(
        step2_state=step2_state,
        globs=globs,
        out_dir=out_path.parent,
        filename_prefix=out_path.stem,
    )

    # Reemplazar/renombrar al nombre exacto
    try:
        tmp.replace(out_path)  # atómico cuando es posible
    except Exception:
        data = tmp.read_bytes()
        out_path.write_bytes(data)
        try:
            tmp.unlink()
        except Exception:
            pass

    return out_path
